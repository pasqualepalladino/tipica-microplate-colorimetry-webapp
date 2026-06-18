# -*- coding: utf-8 -*-

import os
import time
import csv
import math
import json
import re
import cv2
import numpy as np
from .config_io import load_all_config
from .roi_geometry import _build_well_bottom_masks_v12, _overlay_projected_mouth_circles
from .roi_io import _bilinear_grid_from_four_corners, _load_fourcorner_geometry
from .roi_overlay import _build_plate_overlay
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import textwrap
from pathlib import Path
from matplotlib.ticker import FixedLocator, FormatStrFormatter
from matplotlib.lines import Line2D


# ---------------------------
# Embedded plate geometry from user-provided datasheets
# ---------------------------
PLATE_GEOMETRY_DB = {
    (2, 3): {
        "name": "6-well standard flat-bottom cell-culture plate",
        "footprint_length_mm": 127.76,
        "footprint_width_mm": 85.48,
        "pitch_mm": 39.12,
        "well_depth_mm_nominal": 17.5,
        "mouth_diam_mm_nominal": 35.0,
        "floor_diam_mm_nominal": 35.0,
        "standard_basis": "approximate standard flat-bottom 6-well geometry; verify vendor-specific dimensions for quantitative path-length work",
    },
    (3, 4): {
        "name": "12-well standard flat-bottom cell-culture plate",
        "footprint_length_mm": 127.76,
        "footprint_width_mm": 85.48,
        "pitch_mm": 26.0,
        "well_depth_mm_nominal": 17.5,
        "mouth_diam_mm_nominal": 22.1,
        "floor_diam_mm_nominal": 22.1,
        "standard_basis": "approximate standard flat-bottom 12-well geometry; verify vendor-specific dimensions for quantitative path-length work",
    },
    (4, 6): {
        "name": "24-well standard flat-bottom cell-culture plate",
        "footprint_length_mm": 127.76,
        "footprint_width_mm": 85.48,
        "pitch_mm": 19.3,
        "well_depth_mm_nominal": 17.5,
        "mouth_diam_mm_nominal": 15.6,
        "floor_diam_mm_nominal": 15.6,
        "standard_basis": "approximate standard flat-bottom 24-well geometry; verify vendor-specific dimensions for quantitative path-length work",
    },
    (6, 8): {
        "name": "48-well standard flat-bottom cell-culture plate",
        "footprint_length_mm": 127.76,
        "footprint_width_mm": 85.48,
        "pitch_mm": 13.0,
        "well_depth_mm_nominal": 17.5,
        "mouth_diam_mm_nominal": 11.0,
        "floor_diam_mm_nominal": 11.0,
        "standard_basis": "approximate standard flat-bottom 48-well geometry; verify vendor-specific dimensions for quantitative path-length work",
    },
    (8, 12): {
        "name": "96-well standard F-bottom / flat-bottom",
        "footprint_length_mm": 127.76,
        "footprint_width_mm": 85.48,
        "footprint_length_mm_range": (127.76, 127.80),
        "footprint_width_mm_range": (85.48, 85.50),
        "plate_height_mm_nominal": 14.4,
        "plate_height_mm_range": (14.2, 14.6),
        "a1_row_offset_mm": 11.24,
        "a1_row_offset_mm_range": (11.18, 11.24),
        "a1_col_offset_mm": 14.38,
        "a1_col_offset_mm_range": (14.29, 14.38),
        "pitch_mm": 9.0,
        "pitch_mm_range": (9.0, 9.02),
        "well_depth_mm": 10.9,
        "well_depth_mm_range": (10.67, 10.90),
        "mouth_diam_mm": 6.90,
        "mouth_diam_mm_range": (6.69, 6.96),
        "floor_diam_mm": 6.48,
        "floor_diam_mm_range": (6.35, 6.58),
        "outer_diam_mm": 7.75,
        "inner_diam_mm": 6.90,
        "bridge_width_mm": 0.60,
        "extra_optical_margin_mm": 0.20,
        "well_bottom_elevation_mm_nominal": 3.6,
        "well_bottom_elevation_mm_range": (3.5, 3.7),
        "flange_or_skirt_height_mm": 2.5,
    },
    (16, 24): {
        "name": "384-well standard F-bottom / flat-bottom",
        "footprint_length_mm": 127.76,
        "footprint_width_mm": 85.48,
        "plate_height_mm_nominal": 14.2,
        "a1_row_offset_mm": 8.99,
        "a1_col_offset_mm": 12.13,
        "pitch_mm": 4.5,
        "well_depth_mm_nominal": 11.5,
        "mouth_diam_mm_nominal": 3.66,
        "floor_diam_mm_nominal": 3.05,
        "flange_or_skirt_height_mm": 2.5,
    },
    (32, 48): {
        "name": "1536-well standard F-bottom / flat-bottom",
        "footprint_length_mm": 127.76,
        "footprint_width_mm": 85.48,
        "plate_height_mm_nominal": 10.4,
        "a1_row_offset_mm": 7.87,
        "a1_col_offset_mm": 11.01,
        "pitch_mm": 2.25,
        "well_depth_mm_nominal": 4.9,
        "mouth_diam_mm_nominal": 1.75,
        "floor_diam_mm_nominal": 1.56,
        "flange_or_skirt_height_mm": 2.0,
    },
}


def _get_plate_geometry(nrow, ncol):
    geom = PLATE_GEOMETRY_DB.get((int(nrow), int(ncol)))
    if geom is None:
        return {
            "name": f"generic {nrow}x{ncol}",
            "pitch_mm": 9.0 if int(nrow) == 8 and int(ncol) == 12 else 1.0,
            "mouth_diam_mm": 6.90 if int(nrow) == 8 and int(ncol) == 12 else 1.0,
            "floor_diam_mm": 6.48 if int(nrow) == 8 and int(ncol) == 12 else 0.9,
            "well_depth_mm": 10.9 if int(nrow) == 8 and int(ncol) == 12 else 1.0,
            "outer_diam_mm": 7.75 if int(nrow) == 8 and int(ncol) == 12 else 1.1,
            "inner_diam_mm": 6.90 if int(nrow) == 8 and int(ncol) == 12 else 1.0,
            "bridge_width_mm": 0.60 if int(nrow) == 8 and int(ncol) == 12 else 0.05,
            "extra_optical_margin_mm": 0.20 if int(nrow) == 8 and int(ncol) == 12 else 0.02,
            "standard_basis": "generic geometry; not ANSI/SLAS-certified",
        }
    out = dict(geom)
    out.setdefault("standard_basis", "ANSI/SLAS-compatible flat-bottom geometry; vendor dimensions may vary by plate family")
    return out


def _geometry_numeric(geom, *keys):
    """Return the first finite positive numeric value among several geometry keys."""
    for key in keys:
        try:
            value = float((geom or {}).get(key, np.nan))
        except Exception:
            value = np.nan
        if np.isfinite(value) and value > 0:
            return value
    return np.nan


def _compute_pathlength_payload_from_volume(volume_ul, geom):
    """Estimate optical path length from liquid volume and flat-bottom well geometry.

    For flat-bottom wells, l = V / A. Because 1 uL = 1 mm^3, l is first
    obtained in mm and then converted to cm for Beer-Lambert-like use.
    """
    try:
        vol_ul = float(volume_ul)
    except Exception:
        vol_ul = np.nan
    if not np.isfinite(vol_ul) or vol_ul <= 0:
        return {
            "liquid_volume_ul": np.nan,
            "path_length": np.nan,
            "path_length_mm": np.nan,
            "path_length_cm": np.nan,
            "path_length_source": "not_available",
            "well_bottom_diam_mm": np.nan,
            "well_bottom_area_mm2": np.nan,
            "plate_geometry_assumption": "not_available",
            "plate_geometry_name": str((geom or {}).get("name", "")),
            "path_length_warning": "liquid volume not configured",
        }

    d_mm = _geometry_numeric(geom, "floor_diam_mm", "floor_diam_mm_nominal", "well_bottom_width_mm", "bottom_diam_mm")
    if not np.isfinite(d_mm) or d_mm <= 0:
        return {
            "liquid_volume_ul": vol_ul,
            "path_length": np.nan,
            "path_length_mm": np.nan,
            "path_length_cm": np.nan,
            "path_length_source": "not_available",
            "well_bottom_diam_mm": np.nan,
            "well_bottom_area_mm2": np.nan,
            "plate_geometry_assumption": "flat-bottom geometry required; bottom diameter unavailable",
            "plate_geometry_name": str((geom or {}).get("name", "")),
            "path_length_warning": "bottom diameter not available for path-length calculation",
        }

    area_mm2 = float(math.pi * (0.5 * d_mm) ** 2)
    l_mm = float(vol_ul / area_mm2)
    l_cm = float(l_mm / 10.0)
    well_depth_mm = _geometry_numeric(geom, "well_depth_mm", "well_depth_mm_nominal")
    warning = ""
    if np.isfinite(well_depth_mm) and l_mm > well_depth_mm:
        warning = "computed liquid height exceeds nominal well depth"

    return {
        "liquid_volume_ul": vol_ul,
        "path_length": l_cm,
        "path_length_mm": l_mm,
        "path_length_cm": l_cm,
        "path_length_source": "calculated_from_volume_and_flat_bottom_geometry",
        "well_bottom_diam_mm": d_mm,
        "well_bottom_area_mm2": area_mm2,
        "plate_geometry_assumption": str((geom or {}).get("standard_basis", "ANSI/SLAS-compatible flat-bottom geometry assumed")),
        "plate_geometry_name": str((geom or {}).get("name", "")),
        "path_length_warning": warning,
    }


def _unit_label_to_molar_factor(unit_label):
    """Return how many mol/L correspond to one displayed concentration unit."""
    text = str(unit_label or "M").replace("μ", "µ").strip()
    base = text.split()[0] if text else "M"
    factors = {"M": 1.0, "mM": 1e-3, "µM": 1e-6, "uM": 1e-6, "nM": 1e-9}
    factor = factors.get(base, 1.0)
    if "10^" in text:
        try:
            exp_txt = text.split("10^", 1)[1].strip()
            factor *= 10.0 ** float(exp_txt)
        except Exception:
            pass
    return float(factor) if np.isfinite(factor) and factor > 0 else 1.0


def to_linear_rgb01(img_bgr, gamma=2.2):
    x = img_bgr.astype(np.float32) / 255.0
    x = np.clip(x, 0.0, 1.0)
    return np.power(x, gamma)


def grid_centers_affine(origin, vrow, vcol, ncol, nrow):
    centers = np.zeros((nrow, ncol, 2), dtype=np.float32)
    for r in range(nrow):
        for c in range(ncol):
            centers[r, c] = origin + r * vrow + c * vcol
    return centers


def _make_run_dir(image_path):
    base = os.path.splitext(os.path.basename(image_path))[0]
    out_root = os.path.join(os.path.dirname(image_path), f"OUTPUT_{base}")
    os.makedirs(out_root, exist_ok=True)
    ts = time.strftime("%Y%m%d_%H%M%S")
    run_dir = os.path.join(out_root, f"RUN_{ts}")
    os.makedirs(run_dir, exist_ok=True)
    return run_dir, base


def _order_quad_tl_tr_br_bl(quad):
    q = np.array(quad, dtype=np.float32)
    s = q.sum(axis=1)
    d = q[:, 0] - q[:, 1]

    tl = q[np.argmin(s)]
    br = q[np.argmax(s)]
    tr = q[np.argmin(d)]
    bl = q[np.argmax(d)]

    return np.array([tl, tr, br, bl], dtype=np.float32)


def _warp_to_canonical(img, quad_img, size, interp=cv2.INTER_LINEAR):
    src = _order_quad_tl_tr_br_bl(quad_img)
    dst = np.array(
        [[0, 0], [size - 1, 0], [size - 1, size - 1], [0, size - 1]],
        dtype=np.float32
    )
    H = cv2.getPerspectiveTransform(src, dst)
    out = cv2.warpPerspective(img, H, (size, size), flags=interp)
    return out, H


def _warp_mask_from_canonical(mask_can, quad_img, out_shape):
    h, w = out_shape[:2]
    src = np.array(
        [[0, 0], [mask_can.shape[1] - 1, 0],
         [mask_can.shape[1] - 1, mask_can.shape[0] - 1], [0, mask_can.shape[0] - 1]],
        dtype=np.float32
    )
    dst = _order_quad_tl_tr_br_bl(quad_img)
    H = cv2.getPerspectiveTransform(src, dst)
    out = cv2.warpPerspective(mask_can, H, (w, h), flags=cv2.INTER_NEAREST)
    return out


def _cell_polygon_mask(quad_img, out_shape):
    h, w = out_shape[:2]
    m = np.zeros((h, w), dtype=np.uint8)
    q = _order_quad_tl_tr_br_bl(quad_img).astype(np.int32)
    cv2.fillConvexPoly(m, q, 255)
    return m


def _masked_otsu_threshold(gray_u8, mask_u8):
    vals = gray_u8[mask_u8 > 0]
    if vals.size < 32:
        return 255

    hist = cv2.calcHist([vals], [0], None, [256], [0, 256]).ravel().astype(np.float64)
    total = hist.sum()
    if total <= 0:
        return 255

    prob = hist / total
    omega = np.cumsum(prob)
    mu = np.cumsum(prob * np.arange(256))
    mu_t = mu[-1]

    denom = omega * (1.0 - omega)
    denom[denom <= 1e-15] = np.nan
    sigma_b2 = ((mu_t * omega - mu) ** 2) / denom

    if np.all(np.isnan(sigma_b2)):
        return int(np.median(vals))

    return int(np.nanargmax(sigma_b2))


def _stable_percentile(x, q):
    arr = np.asarray(x)
    if arr.size == 0:
        return np.nan
    try:
        return float(np.percentile(arr, q, method="nearest"))
    except TypeError:
        return float(np.percentile(arr, q, interpolation="nearest"))


def _keep_component_near_center(mask_u8):
    nlab, lab, stats, cent = cv2.connectedComponentsWithStats(mask_u8, connectivity=8)
    if nlab <= 1:
        return mask_u8

    h, w = mask_u8.shape[:2]
    cx0 = 0.5 * (w - 1)
    cy0 = 0.5 * (h - 1)

    best_idx = -1
    best_score = -1e30

    for i in range(1, nlab):
        area = float(stats[i, cv2.CC_STAT_AREA])
        if area < 8:
            continue

        cx, cy = cent[i]
        d2 = (cx - cx0) ** 2 + (cy - cy0) ** 2
        score = area - 0.08 * d2 + 1e-9 * float(i)
        if score > best_score:
            best_score = score
            best_idx = i

    if best_idx < 0:
        return np.zeros_like(mask_u8)

    out = np.zeros_like(mask_u8)
    out[lab == best_idx] = 255
    return out


def _fit_local_scales(centers, r, c):
    c00 = centers[r, c]
    c01 = centers[r, c + 1]
    c10 = centers[r + 1, c]

    sx = float(np.linalg.norm(c01 - c00))
    sy = float(np.linalg.norm(c10 - c00))
    return max(1e-6, sx), max(1e-6, sy)


def _canonical_model_bg_mask(
    size,
    r_forbidden_x,
    r_forbidden_y,
    bridge_half_w,
    bridge_half_h,
    central_guard_frac=0.05,
    text_exclusion_frac_x=0.16,
    text_exclusion_frac_y=0.09,
):
    mask = np.full((size, size), 255, dtype=np.uint8)

    rx = max(1, int(round(r_forbidden_x * size)))
    ry = max(1, int(round(r_forbidden_y * size)))
    bw = max(1, int(round(bridge_half_w * size)))
    bh = max(1, int(round(bridge_half_h * size)))

    corners = [
        (0, 0),
        (size - 1, 0),
        (size - 1, size - 1),
        (0, size - 1),
    ]

    for cx, cy in corners:
        cv2.ellipse(mask, (cx, cy), (rx, ry), 0, 0, 360, 0, -1)

    cv2.rectangle(mask, (int(0.5 * size - rx), 0), (int(0.5 * size + rx), min(size - 1, bh)), 0, -1)
    cv2.rectangle(mask, (int(0.5 * size - rx), max(0, size - 1 - bh)), (int(0.5 * size + rx), size - 1), 0, -1)
    cv2.rectangle(mask, (0, int(0.5 * size - ry)), (min(size - 1, bw), int(0.5 * size + ry)), 0, -1)
    cv2.rectangle(mask, (max(0, size - 1 - bw), int(0.5 * size - ry)), (size - 1, int(0.5 * size + ry)), 0, -1)

    # Exclude the inter-well printed labels located near the center of each side of the quadrilateral cell.
    tx = max(2, int(round(text_exclusion_frac_x * size)))
    ty = max(2, int(round(text_exclusion_frac_y * size)))
    cx = int(round(0.5 * size))
    cy = int(round(0.5 * size))
    # top / bottom text areas
    cv2.rectangle(mask, (max(0, cx - tx), max(0, int(0.15 * size) - ty)),
                        (min(size - 1, cx + tx), min(size - 1, int(0.15 * size) + ty)), 0, -1)
    cv2.rectangle(mask, (max(0, cx - tx), max(0, int(0.85 * size) - ty)),
                        (min(size - 1, cx + tx), min(size - 1, int(0.85 * size) + ty)), 0, -1)
    # left / right text areas
    cv2.rectangle(mask, (max(0, int(0.15 * size) - ty), max(0, cy - tx)),
                        (min(size - 1, int(0.15 * size) + ty), min(size - 1, cy + tx)), 0, -1)
    cv2.rectangle(mask, (max(0, int(0.85 * size) - ty), max(0, cy - tx)),
                        (min(size - 1, int(0.85 * size) + ty), min(size - 1, cy + tx)), 0, -1)

    g = max(1, int(round(central_guard_frac * size)))
    cv2.rectangle(mask, (int(0.5 * size - g), int(0.5 * size - g)), (int(0.5 * size + g), int(0.5 * size + g)), 255, -1)
    return mask


def _refine_bg_mask_in_canonical(gray_can_u8, model_bg_can_u8, erode_px=1):
    gray_blur = cv2.GaussianBlur(gray_can_u8, (5, 5), 0)

    vals = gray_blur[model_bg_can_u8 > 0]
    if vals.size < 32:
        return model_bg_can_u8.copy()

    thr_otsu = _masked_otsu_threshold(gray_blur, model_bg_can_u8)
    p55 = int(round(_stable_percentile(vals, 55)))
    p60 = int(round(_stable_percentile(vals, 60)))

    thr = max(thr_otsu, p55)
    thr = min(thr, p60 + 6)

    cand = np.zeros_like(model_bg_can_u8)
    cand[(model_bg_can_u8 > 0) & (gray_blur >= thr)] = 255

    # Exclude pixels that are too dark or too bright to be plausible inter-well background.
    vals_model = gray_blur[model_bg_can_u8 > 0]
    if vals_model.size > 0:
        dark_thr = float(_stable_percentile(vals_model, 30))
        bright_thr = float(_stable_percentile(vals_model, 95))
        cand[gray_blur < dark_thr] = 0
        cand[gray_blur > bright_thr] = 0

        # Remove text-like dark details using a local black-hat response.
        bh = cv2.morphologyEx(gray_blur, cv2.MORPH_BLACKHAT,
                              cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (9, 9)))
        bh_thr = float(_stable_percentile(bh[model_bg_can_u8 > 0], 80))
        cand[bh > bh_thr] = 0

    k3 = cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (3, 3))
    k5 = cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (5, 5))

    cand = cv2.morphologyEx(cand, cv2.MORPH_OPEN, k3, iterations=1)
    cand = cv2.morphologyEx(cand, cv2.MORPH_CLOSE, k5, iterations=1)
    cand = _keep_component_near_center(cand)

    if np.count_nonzero(cand) < 20:
        cand = model_bg_can_u8.copy()

    if erode_px > 0:
        ke = cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (2 * erode_px + 1, 2 * erode_px + 1))
        cand = cv2.erode(cand, ke, iterations=1)

    return cand


def _build_global_exclusion_mask(centers, img_shape, exclusion_r_px):
    h, w = img_shape[:2]
    mask = np.zeros((h, w), dtype=np.uint8)
    rr = max(1, int(round(exclusion_r_px)))
    for r in range(centers.shape[0]):
        for c in range(centers.shape[1]):
            cx, cy = centers[r, c]
            cv2.circle(mask, (int(round(cx)), int(round(cy))), rr, 255, -1)
    return mask


def build_bg_masks_physical(
    centers,
    img_bgr,
    well_r_px,
    pitch_mm=9.0,
    inner_diam_mm=6.90,
    outer_diam_mm=7.75,
    bridge_width_mm=0.60,
    extra_optical_margin_mm=0.20,
    canonical_size=220,
    erode_px=1,
):
    h, w = img_bgr.shape[:2]
    gray_u8 = cv2.cvtColor(img_bgr, cv2.COLOR_BGR2GRAY)

    nrow, ncol = centers.shape[:2]
    per_cell_masks = []
    union_mask = np.zeros((h, w), dtype=np.uint8)

    px_per_mm = float(well_r_px) / (0.5 * inner_diam_mm)
    forbidden_radius_mm = 0.5 * outer_diam_mm + extra_optical_margin_mm

    global_exclusion_radius_px = px_per_mm * forbidden_radius_mm
    global_exclusion_mask = _build_global_exclusion_mask(centers=centers, img_shape=img_bgr.shape, exclusion_r_px=global_exclusion_radius_px)

    for r in range(nrow - 1):
        row_masks = []
        for c in range(ncol - 1):
            c00 = centers[r, c]
            c01 = centers[r, c + 1]
            c10 = centers[r + 1, c]
            c11 = centers[r + 1, c + 1]
            quad = np.array([c00, c01, c11, c10], dtype=np.float32)

            sx_px, sy_px = _fit_local_scales(centers, r, c)
            px_per_mm_x = sx_px / pitch_mm
            px_per_mm_y = sy_px / pitch_mm

            r_forbidden_x = (forbidden_radius_mm * px_per_mm_x) / sx_px
            r_forbidden_y = (forbidden_radius_mm * px_per_mm_y) / sy_px
            bridge_half_w = 0.5 * (bridge_width_mm * px_per_mm_x) / sx_px
            bridge_half_h = 0.5 * (bridge_width_mm * px_per_mm_y) / sy_px

            r_forbidden_x = float(np.clip(r_forbidden_x, 0.18, 0.42))
            r_forbidden_y = float(np.clip(r_forbidden_y, 0.18, 0.42))
            bridge_half_w = float(np.clip(bridge_half_w, 0.01, 0.06))
            bridge_half_h = float(np.clip(bridge_half_h, 0.01, 0.06))

            model_bg_can = _canonical_model_bg_mask(
                size=canonical_size,
                r_forbidden_x=r_forbidden_x,
                r_forbidden_y=r_forbidden_y,
                bridge_half_w=bridge_half_w,
                bridge_half_h=bridge_half_h,
                central_guard_frac=0.05,
            )

            gray_can, _ = _warp_to_canonical(gray_u8, quad, canonical_size, interp=cv2.INTER_LINEAR)
            refined_bg_can = _refine_bg_mask_in_canonical(gray_can_u8=gray_can, model_bg_can_u8=model_bg_can, erode_px=erode_px)

            bg_img = _warp_mask_from_canonical(refined_bg_can, quad, img_bgr.shape)
            cell_img = _cell_polygon_mask(quad, img_bgr.shape)
            bg_img = cv2.bitwise_and(bg_img, cell_img)
            bg_img = cv2.bitwise_and(bg_img, cv2.bitwise_not(global_exclusion_mask))

            row_masks.append(bg_img)
            union_mask = cv2.bitwise_or(union_mask, bg_img)

        per_cell_masks.append(row_masks)

    return per_cell_masks, union_mask, global_exclusion_mask


def _robust_filter_mask_by_intensity(img_bgr, mask_u8, k_mad=2.5, min_pixels=30):
    if np.count_nonzero(mask_u8) < min_pixels:
        return mask_u8.copy()

    gray = cv2.cvtColor(img_bgr, cv2.COLOR_BGR2GRAY).astype(np.float32)
    vals = gray[mask_u8 > 0]
    if vals.size < min_pixels:
        return mask_u8.copy()

    med = float(np.median(vals))
    mad = float(np.median(np.abs(vals - med)))
    if mad < 1e-6:
        return mask_u8.copy()

    sigma_rob = 1.4826 * mad
    lo = med - k_mad * sigma_rob
    hi = med + k_mad * sigma_rob

    out = np.zeros_like(mask_u8)
    sel = (mask_u8 > 0) & (gray >= lo) & (gray <= hi)
    out[sel] = 255
    return out


def _extract_bg_masks_statistical(per_cell_masks, img_bgr, k_mad=2.5, open_close=True):
    nrow_m1 = len(per_cell_masks)
    ncol_m1 = len(per_cell_masks[0]) if nrow_m1 > 0 else 0
    out_masks = [[None for _ in range(ncol_m1)] for _ in range(nrow_m1)]

    for r in range(nrow_m1):
        for c in range(ncol_m1):
            m = per_cell_masks[r][c]
            mf = _robust_filter_mask_by_intensity(img_bgr, m, k_mad=k_mad, min_pixels=30)
            if open_close and np.count_nonzero(mf) > 0:
                k3 = cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (3, 3))
                k5 = cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (5, 5))
                mf = cv2.morphologyEx(mf, cv2.MORPH_OPEN, k3, iterations=1)
                mf = cv2.morphologyEx(mf, cv2.MORPH_CLOSE, k5, iterations=1)
            out_masks[r][c] = mf

    return out_masks


def _mask_centroid(mask_u8):
    ys, xs = np.where(mask_u8 > 0)
    if xs.size == 0:
        return None
    return float(np.mean(xs)), float(np.mean(ys))


def _extract_bg_samples(per_cell_masks_stat, img_bgr):
    samples = []
    img = img_bgr.astype(np.float32)

    for r in range(len(per_cell_masks_stat)):
        for c in range(len(per_cell_masks_stat[r])):
            m = per_cell_masks_stat[r][c]
            area = int(np.count_nonzero(m))
            if area < 30:
                continue

            ctr = _mask_centroid(m)
            if ctr is None:
                continue

            x, y = ctr
            vals = img[m > 0]
            if vals.shape[0] < 30:
                continue

            b_med = float(np.median(vals[:, 0]))
            g_med = float(np.median(vals[:, 1]))
            r_med = float(np.median(vals[:, 2]))

            samples.append({
                "cell_r": int(r), "cell_c": int(c), "x": x, "y": y, "area": area,
                "B": b_med, "G": g_med, "R": r_med,
            })
    return samples


def _poly2_design(xn, yn):
    return np.column_stack([np.ones_like(xn), xn, yn, xn * yn, xn * xn, yn * yn])


def _fit_poly2_robust(x, y, z, max_iter=1, clip_k=2.5):
    x = np.asarray(x, dtype=np.float64)
    y = np.asarray(y, dtype=np.float64)
    z = np.asarray(z, dtype=np.float64)

    order = np.lexsort((z, y, x))
    x = x[order]
    y = y[order]
    z = z[order]

    x0 = float(np.mean(x))
    y0 = float(np.mean(y))
    sx = float(np.std(x))
    sy = float(np.std(y))
    if sx < 1e-9:
        sx = 1.0
    if sy < 1e-9:
        sy = 1.0

    xn = (x - x0) / sx
    yn = (y - y0) / sy
    keep = np.isfinite(z).copy()
    if np.count_nonzero(keep) < 6:
        raise RuntimeError("Not enough valid BG samples for quadratic fit.")

    coef = None
    for _ in range(max_iter):
        A = _poly2_design(xn[keep], yn[keep])
        zz = z[keep]
        if A.shape[0] < 6:
            break
        coef, _, _, _ = np.linalg.lstsq(A, zz, rcond=None)
        pred = A @ coef
        res = zz - pred
        med = float(np.median(res))
        mad = float(np.median(np.abs(res - med)))
        sigma = max(1.4826 * mad, 1e-9)
        keep_local = np.abs(res - med) <= clip_k * sigma
        if np.all(keep_local):
            break
        idx = np.where(keep)[0]
        new_keep = keep.copy()
        new_keep[idx] = keep_local
        if np.array_equal(new_keep, keep):
            break
        keep = new_keep

    if coef is None:
        A = _poly2_design(xn, yn)
        coef, _, _, _ = np.linalg.lstsq(A, z, rcond=None)

    return {"coef": coef, "x0": x0, "y0": y0, "sx": sx, "sy": sy}


def _eval_poly2_model(model, x, y):
    x = np.asarray(x, dtype=np.float64)
    y = np.asarray(y, dtype=np.float64)
    xn = (x - model["x0"]) / model["sx"]
    yn = (y - model["y0"]) / model["sy"]
    A = _poly2_design(xn, yn)
    return A @ model["coef"]


def _predict_bg_for_wells(centers, model_B, model_G, model_R):
    rows = []
    nrow, ncol = centers.shape[:2]
    for r in range(nrow):
        for c in range(ncol):
            x = float(centers[r, c, 0])
            y = float(centers[r, c, 1])
            rows.append({
                "well_r": int(r),
                "well_c": int(c),
                "x": x,
                "y": y,
                "B_bg": float(_eval_poly2_model(model_B, [x], [y])[0]),
                "G_bg": float(_eval_poly2_model(model_G, [x], [y])[0]),
                "R_bg": float(_eval_poly2_model(model_R, [x], [y])[0]),
            })
    return rows


def _write_bg_samples_csv(path, samples):
    with open(path, "w", newline="", encoding="utf-8") as f:
        wr = csv.writer(f, delimiter=';')
        wr.writerow(["cell_r", "cell_c", "x", "y", "area", "B_med", "G_med", "R_med"])
        for s in samples:
            wr.writerow([s["cell_r"], s["cell_c"], s["x"], s["y"], s["area"], s["B"], s["G"], s["R"]])


def _write_well_bg_csv(path, rows):
    with open(path, "w", newline="", encoding="utf-8") as f:
        wr = csv.writer(f, delimiter=';')
        wr.writerow(["well_r", "well_c", "x", "y", "B_bg", "G_bg", "R_bg"])
        for s in rows:
            wr.writerow([s["well_r"], s["well_c"], s["x"], s["y"], s["B_bg"], s["G_bg"], s["R_bg"]])


def _extract_annular_bg_rows(annular_masks, img_bgr, centers):
    img = img_bgr.astype(np.float32)
    rows = []
    nrow = len(annular_masks)
    ncol = len(annular_masks[0]) if nrow > 0 else 0
    for r in range(nrow):
        for c in range(ncol):
            m = annular_masks[r][c]
            if m is None:
                continue
            sel = (m > 0)
            n = int(np.count_nonzero(sel))
            if n <= 0:
                rows.append({
                    "well_r": int(r), "well_c": int(c),
                    "x": float(centers[r, c, 0]), "y": float(centers[r, c, 1]),
                    "B_bg": np.nan, "G_bg": np.nan, "R_bg": np.nan,
                    "n_annular": 0,
                })
                continue
            vals = img[sel]
            rows.append({
                "well_r": int(r), "well_c": int(c),
                "x": float(centers[r, c, 0]), "y": float(centers[r, c, 1]),
                "B_bg": float(np.median(vals[:, 0])),
                "G_bg": float(np.median(vals[:, 1])),
                "R_bg": float(np.median(vals[:, 2])),
                "n_annular": n,
            })
    return rows


def _build_bg_method_comparison_rows(raw_rows_interstitial, raw_rows_annular, metadata_rows):
    ann_map = {str(r.get("Well", "")): dict(r) for r in (raw_rows_annular or [])}
    rows = []
    type_cal = {"CAL", "STD", "STANDARD", "CALIBRATION", "C"}
    for r in (raw_rows_interstitial or []):
        well = str(r.get("Well", ""))
        a = ann_map.get(well, {})
        meta_type = str(r.get("Type", "")).upper()
        conc = float(r.get("Conc", np.nan)) if np.isfinite(r.get("Conc", np.nan)) else np.nan
        is_zero_cal = int(meta_type in type_cal and np.isfinite(conc) and abs(conc) <= 1e-12)
        row = {
            "Well": well, "ID": r.get("ID", ""), "Type": r.get("Type", ""), "Conc": r.get("Conc", np.nan), "DF": r.get("DF", np.nan),
            "ZeroCal": is_zero_cal,
        }
        for suffix in ["Blue", "Green", "Red"]:
            row[f"MeanW_{suffix}"] = r.get(f"MeanW_{suffix}", np.nan)
            for family, src in [("interstitial", r), ("annular", a)]:
                for key in [f"MeanBG_{suffix}", f"Signal_{suffix}", f"PAbs_{suffix}"]:
                    row[f"{family}_{key}"] = src.get(key, np.nan)
            v1 = row.get(f"interstitial_MeanBG_{suffix}", np.nan)
            v2 = row.get(f"annular_MeanBG_{suffix}", np.nan)
            row[f"DeltaBG_{suffix}"] = (float(v2) - float(v1)) if np.isfinite(v1) and np.isfinite(v2) else np.nan
            row[f"RatioBG_{suffix}"] = (float(v2) / float(v1)) if np.isfinite(v1) and np.isfinite(v2) and abs(float(v1)) > 1e-15 else np.nan
        rows.append(row)
    return rows


def _build_bg_method_summary_rows(bg_method_rows):
    out = []
    for method in ["interstitial", "annular"]:
        for suffix in ["Blue", "Green", "Red"]:
            bg_vals = np.asarray([float(r.get(f"{method}_MeanBG_{suffix}", np.nan)) for r in (bg_method_rows or []) if np.isfinite(r.get(f"{method}_MeanBG_{suffix}", np.nan))], dtype=np.float64)
            sig_zero = np.asarray([float(r.get(f"{method}_Signal_{suffix}", np.nan)) for r in (bg_method_rows or []) if int(r.get("ZeroCal", 0)) == 1 and np.isfinite(r.get(f"{method}_Signal_{suffix}", np.nan))], dtype=np.float64)
            pabs_zero = np.asarray([float(r.get(f"{method}_PAbs_{suffix}", np.nan)) for r in (bg_method_rows or []) if int(r.get("ZeroCal", 0)) == 1 and np.isfinite(r.get(f"{method}_PAbs_{suffix}", np.nan))], dtype=np.float64)
            meanw = np.asarray([float(r.get(f"MeanW_{suffix}", np.nan)) for r in (bg_method_rows or []) if np.isfinite(r.get(f"MeanW_{suffix}", np.nan))], dtype=np.float64)
            bg_corr = np.nan
            if bg_vals.size and meanw.size and bg_vals.size == meanw.size:
                try:
                    bg_corr = float(np.corrcoef(bg_vals, meanw)[0,1])
                except Exception:
                    bg_corr = np.nan
            out.append({
                "Method": method,
                "Channel": suffix,
                "n_wells": int(bg_vals.size),
                "n_zero_cal": int(sig_zero.size),
                "BG_median": float(np.median(bg_vals)) if bg_vals.size else np.nan,
                "BG_sd": float(np.std(bg_vals, ddof=1)) if bg_vals.size > 1 else np.nan,
                "Signal_zero_median": float(np.median(sig_zero)) if sig_zero.size else np.nan,
                "Signal_zero_sd": float(np.std(sig_zero, ddof=1)) if sig_zero.size > 1 else np.nan,
                "PAbs_zero_median": float(np.median(pabs_zero)) if pabs_zero.size else np.nan,
                "PAbs_zero_sd": float(np.std(pabs_zero, ddof=1)) if pabs_zero.size > 1 else np.nan,
                "Corr_BG_vs_MeanW": bg_corr,
            })
    return out


def _write_generic_csv(path, rows):
    if not path or not rows:
        return
    headers = list(rows[0].keys())
    with open(path, 'w', newline='', encoding='utf-8') as f:
        wr = csv.writer(f, delimiter=';')
        wr.writerow(headers)
        for row in rows:
            wr.writerow([row.get(h, '') for h in headers])


def _make_bg_method_comparison_png(path, image_basename, summary_rows):
    if not path or not summary_rows:
        return
    channels = ["Blue", "Green", "Red"]
    methods = ["interstitial", "annular"]
    metric_keys = ["Signal_zero_median", "Signal_zero_sd", "PAbs_zero_sd", "Corr_BG_vs_MeanW"]
    metric_titles = ["Zero-cal signal median", "Zero-cal signal SD", "Zero-cal PAbs SD", "Corr(BG, MeanW)"]
    fig, axes = plt.subplots(2, 2, figsize=(10.5, 7.5), dpi=220)
    axes = axes.ravel()
    width = 0.36
    x = np.arange(len(channels), dtype=np.float64)
    for ax, key, title in zip(axes, metric_keys, metric_titles):
        for i, method in enumerate(methods):
            vals = []
            for ch in channels:
                row = next((r for r in summary_rows if str(r.get("Method")) == method and str(r.get("Channel")) == ch), None)
                vals.append(float(row.get(key, np.nan)) if row is not None and np.isfinite(row.get(key, np.nan)) else np.nan)
            ax.bar(x + (i - 0.5) * width, vals, width=width, label=method)
        if key.endswith("median"):
            ax.axhline(0.0, linestyle='--', linewidth=1.0)
        if key.startswith("Corr_"):
            ax.axhline(0.0, linestyle='--', linewidth=1.0)
        ax.set_xticks(x)
        ax.set_xticklabels(channels)
        ax.set_title(title)
        ax.grid(True, axis='y', alpha=0.25)
    axes[0].legend(loc='best', frameon=False)
    fig.suptitle(f'BG method comparison - {image_basename}')
    fig.tight_layout()
    fig.savefig(path, dpi=220, bbox_inches='tight')
    plt.close(fig)


# ---------------------------
# ROI model: nominal geometry with weak BG correction and fast local fitting
# ---------------------------

def _safe_percentile(x, q, default=np.nan):
    x = np.asarray(x)
    if x.size == 0:
        return default
    return float(_stable_percentile(x, q))


def _channel_summary(vals):
    vals = np.asarray(vals, dtype=np.float32)
    if vals.size == 0:
        return {
            "mean": np.nan, "median": np.nan, "sd": np.nan,
            "p10": np.nan, "p25": np.nan, "p50": np.nan, "p75": np.nan, "p90": np.nan,
            "iqr": np.nan,
        }
    p10 = _safe_percentile(vals, 10)
    p25 = _safe_percentile(vals, 25)
    p50 = _safe_percentile(vals, 50)
    p75 = _safe_percentile(vals, 75)
    p90 = _safe_percentile(vals, 90)
    return {
        "mean": float(np.mean(vals)),
        "median": float(np.median(vals)),
        "sd": float(np.std(vals, ddof=1)) if vals.size > 1 else 0.0,
        "p10": p10, "p25": p25, "p50": p50, "p75": p75, "p90": p90,
        "iqr": float(p75 - p25),
    }


def _robust_inner_mask(mask_u8, erosion_px=4):
    if np.count_nonzero(mask_u8) == 0:
        return np.zeros_like(mask_u8)
    k = max(1, int(erosion_px))
    ker = cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (2 * k + 1, 2 * k + 1))
    inner = cv2.erode(mask_u8, ker, iterations=1)
    if np.count_nonzero(inner) < 16:
        inner = mask_u8.copy()
    return inner




def _largest_component(mask_u8):
    nlab, lab, stats, _ = cv2.connectedComponentsWithStats(mask_u8, connectivity=8)
    if nlab <= 1:
        return mask_u8
    best = 1 + int(np.argmax(stats[1:, cv2.CC_STAT_AREA]))
    out = np.zeros_like(mask_u8)
    out[lab == best] = 255
    return out


def _fuzzy_liquid_mask_within_roi(img_bgr, roi_mask_u8, center_xy=None, tol_gray=10, tol_sat=20):
    """
    Stronger fuzzy-like contiguous selection inside ROI.
    Seed is taken from the median-like gray/saturation values within the ROI,
    then only the connected component overlapping the seed point is kept.
    """
    h, w = roi_mask_u8.shape[:2]
    if np.count_nonzero(roi_mask_u8) == 0:
        return roi_mask_u8.copy()

    gray = cv2.cvtColor(img_bgr, cv2.COLOR_BGR2GRAY)
    hsv = cv2.cvtColor(img_bgr, cv2.COLOR_BGR2HSV)
    sat = hsv[:, :, 1]

    ys, xs = np.where(roi_mask_u8 > 0)
    if xs.size == 0:
        return roi_mask_u8.copy()

    # Choose a seed point near the ROI centroid, but derive target values from ROI medians.
    if center_xy is None:
        cx = int(round(float(np.mean(xs))))
        cy = int(round(float(np.mean(ys))))
    else:
        cx = int(round(np.clip(center_xy[0], 0, w - 1)))
        cy = int(round(np.clip(center_xy[1], 0, h - 1)))

    seed_gray = float(np.median(gray[roi_mask_u8 > 0]))
    seed_sat = float(np.median(sat[roi_mask_u8 > 0]))

    cand = np.zeros_like(roi_mask_u8)
    ok = (
        (roi_mask_u8 > 0)
        & (np.abs(gray.astype(np.float32) - seed_gray) <= tol_gray)
        & (np.abs(sat.astype(np.float32) - seed_sat) <= tol_sat)
    )
    cand[ok] = 255

    # Remove isolated debris and keep only the connected region overlapping the seed.
    k3 = cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (3, 3))
    cand = cv2.morphologyEx(cand, cv2.MORPH_OPEN, k3, iterations=1)
    cand = cv2.morphologyEx(cand, cv2.MORPH_CLOSE, k3, iterations=1)

    nlab, lab, stats, _ = cv2.connectedComponentsWithStats(cand, connectivity=8)
    if nlab <= 1:
        return roi_mask_u8.copy()

    seed_label = int(lab[cy, cx])
    out = np.zeros_like(roi_mask_u8)
    if seed_label > 0:
        out[lab == seed_label] = 255
    else:
        # choose the component whose mean position is closest to the ROI centroid
        roi_cx = float(np.mean(xs))
        roi_cy = float(np.mean(ys))
        best_i = -1
        best_score = 1e18
        for i in range(1, nlab):
            area = int(stats[i, cv2.CC_STAT_AREA])
            if area < 20:
                continue
            yi, xi = np.where(lab == i)
            if xi.size == 0:
                continue
            dx = float(np.mean(xi)) - roi_cx
            dy = float(np.mean(yi)) - roi_cy
            score = dx * dx + dy * dy
            if score < best_score:
                best_score = score
                best_i = i
        if best_i > 0:
            out[lab == best_i] = 255
        else:
            return roi_mask_u8.copy()

    if np.count_nonzero(out) < max(20, int(0.20 * np.count_nonzero(roi_mask_u8))):
        return roi_mask_u8.copy()
    return out

def _compute_well_robust_statistics(img_bgr, roi_masks, highlight_gray_threshold=245, trim_bright_q=88.0, trim_dark_q=8.0, return_debug=False):
    img = img_bgr.astype(np.float32)
    gray = cv2.cvtColor(img_bgr, cv2.COLOR_BGR2GRAY).astype(np.float32)
    b = img[:, :, 0]
    g = img[:, :, 1]
    r = img[:, :, 2]
    purple = 0.5 * (r + b) - g
    lab_u8 = cv2.cvtColor(img_bgr, cv2.COLOR_BGR2LAB).astype(np.float32)
    lab_L = lab_u8[:, :, 0] * (100.0 / 255.0)
    lab_a = lab_u8[:, :, 1] - 128.0
    lab_b = lab_u8[:, :, 2] - 128.0

    nrow = len(roi_masks)
    ncol = len(roi_masks[0]) if nrow > 0 else 0
    rows = []
    debug_rows = []

    for rr in range(nrow):
        for cc in range(ncol):
            m = roi_masks[rr][cc]
            if m is None:
                continue

            full_mask = (m > 0)
            n_full = int(np.count_nonzero(full_mask))
            if n_full == 0:
                rows.append({
                    "well_r": rr, "well_c": cc, "n_roi": 0, "n_core": 0, "n_used": 0,
                    "used_fraction": 0.0, "highlight_fraction_roi": 0.0, "highlight_fraction_core": 0.0,
                    "is_image_quality_warning": 1, "warning_reason": "empty_roi",
                })
                if return_debug:
                    debug_rows.append({
                        "row": rr,
                        "col": cc,
                        "roi_u8": m.copy().astype(np.uint8),
                        "used_u8": np.zeros_like(m, dtype=np.uint8),
                        "bright_excluded_u8": np.zeros_like(m, dtype=np.uint8),
                        "dark_excluded_u8": np.zeros_like(m, dtype=np.uint8),
                    })
                continue

            inner_u8 = _robust_inner_mask(m, erosion_px=4)
            core_mask = (inner_u8 > 0)
            n_core = int(np.count_nonzero(core_mask))

            gray_core = gray[core_mask]
            if gray_core.size == 0:
                core_mask = full_mask.copy()
                gray_core = gray[core_mask]
                n_core = int(np.count_nonzero(core_mask))

            ctr = _mask_centroid(inner_u8)
            fuzzy_u8 = _fuzzy_liquid_mask_within_roi(img_bgr, inner_u8, ctr)
            fuzzy_mask = (fuzzy_u8 > 0)
            if int(np.count_nonzero(fuzzy_mask)) > 30:
                core_mask = fuzzy_mask

            gray_core = gray[core_mask]
            n_core = int(np.count_nonzero(core_mask))

            hi_thr = _safe_percentile(gray_core, trim_bright_q, default=255.0)
            lo_thr = _safe_percentile(gray_core, trim_dark_q, default=0.0)

            used_mask = core_mask.copy()
            n_used = int(np.count_nonzero(used_mask))
            fallback_no_dark = False
            fallback_full_core = False
            if n_used < max(20, int(0.35 * n_core)):
                used_mask = core_mask & (gray <= hi_thr)
                n_used = int(np.count_nonzero(used_mask))
                fallback_no_dark = True
            if n_used < 12:
                used_mask = core_mask.copy()
                n_used = int(np.count_nonzero(used_mask))
                fallback_full_core = True

            vals_B = b[used_mask]
            vals_G = g[used_mask]
            vals_R = r[used_mask]
            vals_gray = gray[used_mask]
            vals_purple = purple[used_mask]
            vals_L = lab_L[used_mask]
            vals_a = lab_a[used_mask]
            vals_b = lab_b[used_mask]

            sB = _channel_summary(vals_B)
            sG = _channel_summary(vals_G)
            sR = _channel_summary(vals_R)
            sGray = _channel_summary(vals_gray)
            sPurple = _channel_summary(vals_purple)
            sL = _channel_summary(vals_L)
            sa = _channel_summary(vals_a)
            sb = _channel_summary(vals_b)

            hi_roi = float(np.mean(gray[full_mask] >= highlight_gray_threshold)) if n_full > 0 else 0.0
            hi_core = float(np.mean(gray[core_mask] >= highlight_gray_threshold)) if n_core > 0 else 0.0
            used_fraction = float(n_used / max(1, n_core))

            bright_mask = core_mask & (gray > hi_thr)
            n_bright_excl = int(np.count_nonzero(bright_mask))
            bright_excl_fraction = float(n_bright_excl / max(1, n_core))
            if n_bright_excl > 0:
                bright_vals = gray[bright_mask].astype(np.float64)
                bright_mean_gray = float(np.mean(bright_vals))
                bright_excess_mean = float(np.mean(np.maximum(bright_vals - float(hi_thr), 0.0)))
            else:
                bright_mean_gray = np.nan
                bright_excess_mean = 0.0
            highlight_index = float(bright_excl_fraction * bright_excess_mean)

            warning_reason = []
            if hi_core > 0.02:
                warning_reason.append("core_highlights")
            if hi_roi > 0.05:
                warning_reason.append("roi_highlights")
            if used_fraction < 0.55:
                warning_reason.append("strong_trimming")
            if sGray["sd"] > 20.0:
                warning_reason.append("high_sd")

            rows.append({
                "well_r": rr, "well_c": cc,
                "n_roi": n_full, "n_core": n_core, "n_used": n_used,
                "used_fraction": used_fraction,
                "highlight_fraction_roi": hi_roi,
                "highlight_fraction_core": hi_core,
                "B_mean": sB["mean"], "B_median": sB["median"], "B_sd": sB["sd"],
                "B_p10": sB["p10"], "B_p25": sB["p25"], "B_p50": sB["p50"], "B_p75": sB["p75"], "B_p90": sB["p90"], "B_iqr": sB["iqr"],
                "G_mean": sG["mean"], "G_median": sG["median"], "G_sd": sG["sd"],
                "G_p10": sG["p10"], "G_p25": sG["p25"], "G_p50": sG["p50"], "G_p75": sG["p75"], "G_p90": sG["p90"], "G_iqr": sG["iqr"],
                "R_mean": sR["mean"], "R_median": sR["median"], "R_sd": sR["sd"],
                "R_p10": sR["p10"], "R_p25": sR["p25"], "R_p50": sR["p50"], "R_p75": sR["p75"], "R_p90": sR["p90"], "R_iqr": sR["iqr"],
                "Gray_mean": sGray["mean"], "Gray_median": sGray["median"], "Gray_sd": sGray["sd"],
                "Gray_p10": sGray["p10"], "Gray_p25": sGray["p25"], "Gray_p50": sGray["p50"], "Gray_p75": sGray["p75"], "Gray_p90": sGray["p90"], "Gray_iqr": sGray["iqr"],
                "Purple_mean": sPurple["mean"], "Purple_median": sPurple["median"], "Purple_sd": sPurple["sd"],
                "Purple_p10": sPurple["p10"], "Purple_p25": sPurple["p25"], "Purple_p50": sPurple["p50"], "Purple_p75": sPurple["p75"], "Purple_p90": sPurple["p90"], "Purple_iqr": sPurple["iqr"],
                "L_mean": sL["mean"], "L_median": sL["median"], "L_sd": sL["sd"],
                "L_p10": sL["p10"], "L_p25": sL["p25"], "L_p50": sL["p50"], "L_p75": sL["p75"], "L_p90": sL["p90"], "L_iqr": sL["iqr"],
                "a_mean": sa["mean"], "a_median": sa["median"], "a_sd": sa["sd"],
                "a_p10": sa["p10"], "a_p25": sa["p25"], "a_p50": sa["p50"], "a_p75": sa["p75"], "a_p90": sa["p90"], "a_iqr": sa["iqr"],
                "b_mean": sb["mean"], "b_median": sb["median"], "b_sd": sb["sd"],
                "b_p10": sb["p10"], "b_p25": sb["p25"], "b_p50": sb["p50"], "b_p75": sb["p75"], "b_p90": sb["p90"], "b_iqr": sb["iqr"],
                "BrightExcludedFraction": bright_excl_fraction,
                "BrightExcludedMeanGray": bright_mean_gray,
                "BrightExcessMeanGray": bright_excess_mean,
                "HighlightIndex": highlight_index,
                "is_image_quality_warning": int(len(warning_reason) > 0),
                "warning_reason": ",".join(warning_reason),
            })

            if return_debug:
                bright_mask = core_mask & (gray > hi_thr)
                dark_mask = core_mask & (gray < lo_thr)
                if fallback_no_dark:
                    dark_mask = np.zeros_like(core_mask, dtype=bool)
                if fallback_full_core:
                    bright_mask = np.zeros_like(core_mask, dtype=bool)
                    dark_mask = np.zeros_like(core_mask, dtype=bool)
                debug_rows.append({
                    "row": rr,
                    "col": cc,
                    "roi_u8": (m > 0).astype(np.uint8) * 255,
                    "used_u8": used_mask.astype(np.uint8) * 255,
                    "bright_excluded_u8": bright_mask.astype(np.uint8) * 255,
                    "dark_excluded_u8": dark_mask.astype(np.uint8) * 255,
                })

    if return_debug:
        return rows, debug_rows
    return rows

def _write_well_statistics_csv(path, rows):
    if not rows:
        return
    keys = list(rows[0].keys())
    with open(path, "w", newline="", encoding="utf-8") as f:
        wr = csv.writer(f, delimiter=';')
        wr.writerow(keys)
        for row in rows:
            wr.writerow([row.get(k, "") for k in keys])


def _write_well_bottom_csv(path, shapes):
    with open(path, "w", newline="", encoding="utf-8") as f:
        wr = csv.writer(f, delimiter=';')
        wr.writerow([
            "well_r", "well_c", "cx", "cy", "local_pitch_px", "px_per_mm", "cyl_r_bg",
            "mouth_r_geom", "floor_r_geom", "mouth_cx", "mouth_cy", "mouth_r", "mouth_score",
            "floor_cx", "floor_cy", "floor_r", "floor_score", "shift_px"
        ])
        for r in range(len(shapes)):
            for c in range(len(shapes[r])):
                s = shapes[r][c]
                if s is None:
                    continue
                wr.writerow([
                    r, c, s["cx"], s["cy"], s["local_pitch_px"], s["px_per_mm"], s["cyl_r_bg"],
                    s["mouth_r_geom"], s["floor_r_geom"], s["mouth_cx"], s["mouth_cy"], s["mouth_r"], s["mouth_score"],
                    s["floor_cx"], s["floor_cy"], s["floor_r"], s.get("floor_score", ""), s["shift_px"]
                ])


def _write_plate_geometry_csv(path, geom):
    with open(path, "w", newline="", encoding="utf-8") as f:
        wr = csv.writer(f, delimiter=';')
        wr.writerow(["key", "value"])
        for k in sorted(geom.keys()):
            wr.writerow([k, geom[k]])






def _estimate_alpha_from_bg_highlight(well_stats_rows, well_bg_rows):
    """
    Estimate the illumination-compensation exponent alpha from the relation
    between predicted background intensity and highlight response inside wells.

    The model assumes H ~ BG**alpha, where H is the mean excess brightness of
    pixels excluded as overly bright. Alpha is estimated from a log-log fit
    across wells within a single image and clamped to a conservative range.
    """
    if not well_stats_rows or not well_bg_rows:
        return 1.0

    bg_map = {(int(r["well_r"]), int(r["well_c"])): r for r in well_bg_rows}
    x_vals = []
    y_vals = []

    for row in well_stats_rows:
        key = (int(row.get("well_r", -1)), int(row.get("well_c", -1)))
        bg = bg_map.get(key)
        if bg is None:
            continue
        bg_mean = float(np.nanmean([bg.get("B_bg", np.nan), bg.get("G_bg", np.nan), bg.get("R_bg", np.nan)]))
        h_excess = float(row.get("BrightExcessMeanGray", np.nan))
        if np.isfinite(bg_mean) and bg_mean > 0 and np.isfinite(h_excess) and h_excess > 0:
            x_vals.append(np.log(bg_mean))
            y_vals.append(np.log(h_excess))

    if len(x_vals) < 10:
        return 1.0

    x = np.asarray(x_vals, dtype=np.float64)
    y = np.asarray(y_vals, dtype=np.float64)
    A = np.column_stack([x, np.ones_like(x)])
    coef, _, _, _ = np.linalg.lstsq(A, y, rcond=None)
    alpha = float(coef[0])
    if not np.isfinite(alpha):
        return 1.0
    return float(np.clip(alpha, 0.3, 1.2))



def _stabilize_alpha_with_quality(alpha_raw, well_stats_rows):
    """
    Stabilize alpha using image-internal quality descriptors.
    This tempers over-correction when the highlight-derived estimate is noisy.
    """
    if not np.isfinite(alpha_raw):
        alpha_raw = 1.0

    if not well_stats_rows:
        return float(np.clip(alpha_raw, 0.3, 1.2))

    used_med = float(np.nanmedian(np.asarray([r.get("used_fraction", np.nan) for r in well_stats_rows], dtype=np.float64)))
    bright_frac_med = float(np.nanmedian(np.asarray([r.get("BrightExcludedFraction", np.nan) for r in well_stats_rows], dtype=np.float64)))
    highlight_idx_med = float(np.nanmedian(np.asarray([r.get("HighlightIndex", np.nan) for r in well_stats_rows], dtype=np.float64)))

    if not np.isfinite(used_med):
        used_med = 0.85
    if not np.isfinite(bright_frac_med):
        bright_frac_med = 0.10
    if not np.isfinite(highlight_idx_med):
        highlight_idx_med = 0.20

    q_used = np.clip((used_med - 0.55) / 0.30, 0.0, 1.0)
    q_bright = np.clip(1.0 - bright_frac_med / 0.18, 0.0, 1.0)
    q_high = np.clip(1.0 - highlight_idx_med / 0.45, 0.0, 1.0)
    quality = 0.45 * q_used + 0.30 * q_bright + 0.25 * q_high

    alpha_eff = 0.72 * alpha_raw + 0.28 * (alpha_raw * quality + (1.0 - quality) * min(alpha_raw, 0.55))
    return float(np.clip(alpha_eff, 0.3, 1.2))

def _row_label_from_index(idx):
    idx = int(idx)
    label = ""
    while True:
        idx, rem = divmod(idx, 26)
        label = chr(ord("A") + rem) + label
        if idx == 0:
            return label
        idx -= 1


def _row_index_from_label(label):
    s = ''.join(ch for ch in str(label).strip().upper() if ch.isalpha())
    if not s:
        raise ValueError("invalid row label")
    out = 0
    for ch in s:
        out = out * 26 + (ord(ch) - ord("A") + 1)
    return out - 1


def _well_name_from_indices(rr, cc):
    return f"{_row_label_from_index(rr)}{int(cc) + 1}"


def _safe_gamma_linearize_scalar(x, gamma=2.2):
    if x is None:
        return np.nan
    xv = float(x)
    if not np.isfinite(xv):
        return np.nan
    if xv > 1.5:
        xv = xv / 255.0
    xv = float(np.clip(xv, 1e-9, 1.0))
    return float(xv ** gamma)


def _parse_float_or_nan(v):
    try:
        return float(v)
    except Exception:
        return np.nan


def _num_or_nan(v):
    try:
        x = float(v)
    except Exception:
        return np.nan
    return x if np.isfinite(x) else np.nan


def _cfg_optional_float(cfg, key):
    try:
        value = float((cfg or {}).get(key, np.nan))
    except Exception:
        return np.nan
    return value if np.isfinite(value) else np.nan

def _effective_path_length(path_length):
    try:
        value = float(path_length)
    except Exception:
        value = np.nan
    return value if np.isfinite(value) and value > 0 else 1.0


def _epsilon_concentration_from_pabs(pabs_value, pabs_sd, dilution_factor, epsilon, path_length, unit_label="M"):
    """Estimate concentration from PAbs using epsilon in M-1 cm-1.

    The Beer-Lambert-like calculation returns mol/L first:
        C_M = PAbs / (epsilon * path_length_cm)
    The result is then converted to the concentration unit selected in the UI.
    """
    eps = _num_or_nan(epsilon)
    path_eff = _effective_path_length(path_length)
    unit_factor_molar = _unit_label_to_molar_factor(unit_label)
    if not np.isfinite(eps) or eps <= 0:
        return np.nan, np.nan, path_eff, False
    denom = float(eps * path_eff)
    if not np.isfinite(denom) or abs(denom) <= 1e-15:
        return np.nan, np.nan, path_eff, False
    yv = _num_or_nan(pabs_value)
    dfv = _num_or_nan(dilution_factor)
    if not np.isfinite(dfv) or abs(dfv) <= 1e-15:
        dfv = 1.0
    c_dil_m = (yv / denom) if np.isfinite(yv) else np.nan
    c_orig_m = (dfv * c_dil_m) if np.isfinite(c_dil_m) else np.nan
    c_orig = (c_orig_m / unit_factor_molar) if np.isfinite(c_orig_m) else np.nan
    ysd = _num_or_nan(pabs_sd)
    c_dil_sd_m = abs(ysd / denom) if np.isfinite(ysd) else np.nan
    c_orig_sd_m = (abs(dfv) * c_dil_sd_m) if np.isfinite(c_dil_sd_m) else np.nan
    c_orig_sd = (c_orig_sd_m / unit_factor_molar) if np.isfinite(c_orig_sd_m) else np.nan
    return c_orig, c_orig_sd, path_eff, True


def _count_metadata_types(metadata_rows):
    type_cal = {"CAL", "STD", "STANDARD", "CALIBRATION", "C"}
    type_stdadd = {"A", "SA", "STDADD", "STANDARD_ADDITION", "ADDITION"}
    type_unk = {"UNK", "UNKNOWN", "U"}
    counts = {"calibration": 0, "stdadd": 0, "unknown": 0}
    for row in metadata_rows or []:
        typ = str(row.get("Type", "")).strip().upper()
        if typ in type_cal:
            counts["calibration"] += 1
        elif typ in type_stdadd:
            counts["stdadd"] += 1
        elif typ in type_unk:
            counts["unknown"] += 1
    return counts


def _compute_empty_well_qc_status(empty_well_payload, stored_empty_comparison_rows=None):
    stored_empty_comparison_rows = stored_empty_comparison_rows or []
    med_ratios = []
    robust_sds = []
    for label in ["Red", "Green", "Blue"]:
        item = (empty_well_payload or {}).get(label, {}) if isinstance(empty_well_payload, dict) else {}
        robust_sd = _num_or_nan(item.get("robust_sd", np.nan))
        if np.isfinite(robust_sd):
            robust_sds.append(robust_sd)
    for row in stored_empty_comparison_rows:
        ratio = _num_or_nan(row.get("Ratio_median", np.nan))
        if np.isfinite(ratio) and ratio > 0:
            med_ratios.append(abs(math.log(ratio)))
    drift_score = float(np.mean(med_ratios)) if med_ratios else np.nan
    sd_med = float(np.nanmedian(np.asarray(robust_sds, dtype=np.float64))) if robust_sds else np.nan
    if med_ratios and drift_score > 0.20:
        status = "warning"
    elif med_ratios and drift_score > 0.10:
        status = "watch"
    elif np.isfinite(sd_med) and sd_med > 0.06:
        status = "watch"
    elif robust_sds:
        status = "ok"
    else:
        status = "not_available"
    return {
        "status": status,
        "empty_drift_score": drift_score,
        "empty_robust_sd_median": sd_med,
        "n_empty_channels": len(robust_sds),
    }


def _build_reliability_payload(metadata_counts, fit_rows, fit_options, plate_qc_payload, empty_qc_payload, selection_info, cfg):
    metadata_counts = metadata_counts or {}
    fit_rows = fit_rows or []
    fit_options = fit_options or {}
    plate_qc_payload = plate_qc_payload or {}
    empty_qc_payload = empty_qc_payload or {}
    selection_info = selection_info or {}
    cfg = cfg or {}

    has_calibration_rows = any(str(r.get("FitType", "")) == "Calibration" and np.isfinite(r.get("m", np.nan)) for r in fit_rows)
    has_stdadd = any(str(r.get("FitType", "")) == "StdAdd" and np.isfinite(r.get("C0", np.nan)) for r in fit_rows)
    has_unknown_quant = any(str(r.get("FitType", "")) in {"UnknownFromCal", "UnknownFromEpsilon"} and np.isfinite(r.get("C0", np.nan)) for r in fit_rows)
    has_unknown_input = int(metadata_counts.get("unknown", 0) or 0) > 0
    epsilon = _cfg_optional_float(cfg, "epsilon")
    path_length = _cfg_optional_float(cfg, "path_length")
    liquid_volume_ul = _cfg_optional_float(cfg, "liquid_volume_ul")
    path_length_mm = _cfg_optional_float(cfg, "path_length_mm")
    well_bottom_diam_mm = _cfg_optional_float(cfg, "well_bottom_diam_mm_for_pathlength")
    well_bottom_area_mm2 = _cfg_optional_float(cfg, "well_bottom_area_mm2_for_pathlength")
    path_length_source = str((cfg or {}).get("path_length_source", ""))
    plate_geometry_name = str((cfg or {}).get("plate_geometry_name", ""))
    plate_geometry_assumption = str((cfg or {}).get("plate_geometry_assumption", ""))
    epsilon_valid = np.isfinite(epsilon) and epsilon > 0
    path_valid = np.isfinite(path_length) and path_length > 0
    use_stored = bool(fit_options.get("use_stored_calibration", False))

    notes = []
    reasons = []
    score = 50.0
    quantification_available = bool(has_stdadd or has_unknown_quant or (has_calibration_rows and int(metadata_counts.get("calibration", 0) or 0) > 0))

    if has_calibration_rows and int(metadata_counts.get("calibration", 0) or 0) > 0:
        score += 30.0
        notes.append("intraplate calibration present")
        reasons.append("calibration and samples can be compared within the same plate")
    elif use_stored and has_unknown_input:
        score -= 8.0
        notes.append("stored calibration used")
        reasons.append("stored calibration is less reliable than intraplate calibration")
    elif has_unknown_input and not has_unknown_quant:
        notes.append("unknown-only fallback")
        reasons.append("no valid calibration available for concentration estimation")
    elif has_unknown_input and has_unknown_quant and (not has_calibration_rows) and (not use_stored):
        notes.append("epsilon quantification used")
        reasons.append("concentration estimated from PAbs with epsilon because no calibration was available")
        score -= 18.0

    if use_stored:
        score -= 10.0
    if str(plate_qc_payload.get("status", "OK")).upper() != "OK":
        score -= 18.0
        reasons.append("plate QC warnings present")
    crit = int(plate_qc_payload.get("critical_wells", 0) or 0)
    total = max(1, int(plate_qc_payload.get("total_wells", 0) or 0))
    if crit > 0:
        score -= min(20.0, 30.0 * crit / total)
        reasons.append("critical wells detected by optical QC")
    used_fracs = [float(r.get("UsedFraction", np.nan)) for r in (fit_rows or []) if np.isfinite(r.get("UsedFraction", np.nan))]
    used_med = float(np.nanmedian(np.asarray(used_fracs, dtype=np.float64))) if used_fracs else np.nan
    if np.isfinite(used_med) and used_med < 0.70:
        score -= 12.0
        reasons.append("strong trimming or low used fraction")

    empty_status = str(empty_qc_payload.get("status", "not_available"))
    empty_drift_score = _num_or_nan(empty_qc_payload.get("empty_drift_score", np.nan))
    if empty_status == "warning":
        score -= 12.0
        reasons.append("empty-well drift indicates limited comparability")
    elif empty_status == "watch":
        score -= 6.0
        reasons.append("empty-well QC suggests moderate drift")
    if empty_status != "not_available":
        notes.append(f"empty-well QC: {empty_status}")

    ranking = list((selection_info or {}).get("ranking", []) or [])
    finite_scores = [_num_or_nan(r.get("Score", np.nan)) for r in ranking if np.isfinite(_num_or_nan(r.get("Score", np.nan)))]
    if len(finite_scores) >= 2:
        best = max(finite_scores)
        second = sorted(finite_scores, reverse=True)[1]
        gap = abs(best - second) / max(abs(best), 1e-12)
        if gap < 0.10:
            score -= 6.0
            reasons.append("method ranking is not strongly separated")
    best_mode = str(((selection_info or {}).get("best", {}) or {}).get("Mode", ""))
    if best_mode == "unavailable":
        score -= 15.0
        reasons.append("no quantitative method was ranked as available")

    if epsilon_valid:
        notes.append("epsilon configured")
    if path_valid:
        if path_length_source:
            notes.append(f"path length {path_length_source}")
        else:
            notes.append("path length configured")

    if has_unknown_input and (not has_calibration_rows) and (not use_stored) and (not epsilon_valid):
        quantification_available = False
        confidence_class = "LOW"
        score = min(score, 15.0)
        reasons.append("unknown-only fallback without calibration or valid epsilon")
    elif has_unknown_input and (not has_calibration_rows) and (not use_stored) and has_unknown_quant:
        quantification_available = True
        score = min(score, 42.0)
        confidence_class = "LOW"
        reasons.append("epsilon-based quantification without calibration has intrinsically limited reliability")
    elif not quantification_available:
        confidence_class = "NOT QUANTIFIABLE"
        score = min(score, 10.0)
    else:
        score = float(np.clip(score, 0.0, 100.0))
        if score >= 75.0:
            confidence_class = "HIGH"
        elif score >= 45.0:
            confidence_class = "MEDIUM"
        else:
            confidence_class = "LOW"

    if not reasons:
        reasons.append("no major reliability penalty detected")

    quantification_status = "available" if quantification_available else "not available"
    return {
        "reliability_score": float(np.clip(score, 0.0, 100.0)),
        "confidence_class": confidence_class,
        "quantification_available": bool(quantification_available),
        "quantification_status": quantification_status,
        "notes": notes,
        "reason": "; ".join(dict.fromkeys(reasons)),
        "empty_drift_score": empty_drift_score,
        "empty_qc_status": empty_status,
        "epsilon": epsilon,
        "path_length": path_length,
        "liquid_volume_ul": liquid_volume_ul,
        "path_length_mm": path_length_mm,
        "path_length_source": path_length_source,
        "well_bottom_diam_mm": well_bottom_diam_mm,
        "well_bottom_area_mm2": well_bottom_area_mm2,
        "plate_geometry_name": plate_geometry_name,
        "plate_geometry_assumption": plate_geometry_assumption,
        "epsilon_valid": bool(epsilon_valid),
        "path_length_valid": bool(path_valid),
    }


def _isfinite_num(v):
    return np.isfinite(_num_or_nan(v))


def _detect_csv_delimiter(path):
    with open(path, "r", encoding="utf-8") as f:
        head = f.read(4096)
    if head.count(";") >= head.count(","):
        return ";"
    return ","




def _load_well_metadata(cfg, nrow, ncol):
    """
    Load well metadata with this priority:
    1) cfg["data"] from the UI
    2) external CSV explicitly referenced in cfg
    3) return [] if no valid metadata are available

    Supported UI formats:
    - {(r, c): (conc, type, df, id), ...}
    - {"A1": {"Conc": ..., "Type": ..., "DF": ..., "ID": ...}, ...}
    """
    rows = []

    def _is_well_name(s):
        s = str(s).strip()
        if len(s) < 2:
            return False
        if not s[0].isalpha():
            return False
        try:
            int(s[1:])
            return True
        except Exception:
            return False

    def _well_to_rc(well):
        well = str(well).strip()
        m = re.match(r"^([A-Za-z]+)(\d+)$", well)
        if not m:
            raise ValueError("invalid well name")
        rr = _row_index_from_label(m.group(1))
        cc = int(m.group(2)) - 1
        return rr, cc

    cfg_data = cfg.get("data", None)
    if isinstance(cfg_data, dict) and len(cfg_data) > 0:
        for key, item in cfg_data.items():
            # UI tuple format
            if isinstance(item, tuple) and len(item) == 4:
                try:
                    rr, cc = key
                    rr = int(rr)
                    cc = int(cc)
                    conc, typ, df, id_c = item
                except Exception:
                    continue
                if rr < 0 or rr >= int(nrow) or cc < 0 or cc >= int(ncol):
                    continue
                rows.append({
                    "well_r": rr,
                    "well_c": cc,
                    "Well": _well_name_from_indices(rr, cc),
                    "ID": str(id_c).strip(),
                    "Type": str(typ).strip(),
                    "Conc": _parse_float_or_nan(conc),
                    "DF": _parse_float_or_nan(df),
                })
                continue

            if not isinstance(item, dict):
                continue

            well = (
                item.get("Well")
                or item.get("well")
                or item.get("Name")
                or item.get("name")
                or item.get("Label")
                or item.get("label")
                or (key if _is_well_name(key) else "")
            )
            well = str(well).strip()
            if not _is_well_name(well):
                continue

            try:
                rr, cc = _well_to_rc(well)
            except Exception:
                continue
            if rr < 0 or rr >= int(nrow) or cc < 0 or cc >= int(ncol):
                continue

            row = {
                "well_r": rr,
                "well_c": cc,
                "Well": well,
                "ID": str(item.get("ID", item.get("id", ""))).strip(),
                "Type": str(item.get("Type", item.get("type", ""))).strip(),
                "Conc": _parse_float_or_nan(item.get("Conc", item.get("conc", np.nan))),
                "DF": _parse_float_or_nan(item.get("DF", item.get("df", np.nan))),
            }

            meaningful = (
                len(row["ID"]) > 0
                or len(row["Type"]) > 0
                or np.isfinite(row["Conc"])
                or np.isfinite(row["DF"])
            )
            if meaningful:
                rows.append(row)

        if len(rows) > 0:
            rows.sort(key=lambda d: (d["well_r"], d["well_c"]))
            return rows

    meta_path = cfg.get("well_metadata_csv") or cfg.get("plate_map_csv") or cfg.get("layout_csv")
    if meta_path:
        delim = _detect_csv_delimiter(meta_path)
        with open(meta_path, "r", encoding="utf-8") as f:
            rd = csv.DictReader(f, delimiter=delim)
            for row in rd:
                well = str(row.get("Well", row.get("well", ""))).strip()
                if len(well) < 2:
                    continue
                try:
                    m = re.match(r"^([A-Za-z]+)(\d+)$", well)
                    if not m:
                        continue
                    rr = _row_index_from_label(m.group(1))
                    cc = int(m.group(2)) - 1
                except Exception:
                    continue
                if rr < 0 or rr >= int(nrow) or cc < 0 or cc >= int(ncol):
                    continue
                rows.append({
                    "well_r": rr,
                    "well_c": cc,
                    "Well": well,
                    "ID": str(row.get("ID", row.get("id", ""))).strip(),
                    "Type": str(row.get("Type", row.get("type", ""))).strip(),
                    "Conc": _parse_float_or_nan(row.get("Conc", row.get("conc", np.nan))),
                    "DF": _parse_float_or_nan(row.get("DF", row.get("df", np.nan))),
                })
        return rows

    return []



def _median_and_robust_sd(values):
    arr = np.asarray([float(v) for v in values if np.isfinite(v)], dtype=np.float64)
    if arr.size == 0:
        return np.nan, np.nan, 0, np.nan
    med = float(np.median(arr))
    mad = float(np.median(np.abs(arr - med)))
    sd = float(1.4826 * mad)
    return med, sd, int(arr.size), mad


def _build_raw_report_rows(well_stats_rows, well_bg_rows, metadata_rows, alpha_bg=1.0, stored_calibration_bundle=None):
    """
    Build analytical rows from well statistics and per-well background fit.

    RGB pseudo-absorbance is fixed to full background normalization:
        Signal = -log10(W / BG)
    """

    meta_map = {(int(r["well_r"]), int(r["well_c"])): r for r in metadata_rows}
    stats_map = {(int(r["well_r"]), int(r["well_c"])): r for r in well_stats_rows}
    bg_map = {(int(r["well_r"]), int(r["well_c"])): r for r in well_bg_rows}

    type_cal = {"CAL", "STD", "STANDARD", "CALIBRATION", "C"}
    type_blank = {"BLK", "BLANK", "EMPTY", "E"}

    def _meanw_from_stats(s, ch):
        if ch == "Blue":
            return _safe_gamma_linearize_scalar(s.get("B_p50", s.get("B_median")))
        if ch == "Green":
            return _safe_gamma_linearize_scalar(s.get("G_p50", s.get("G_median")))
        return _safe_gamma_linearize_scalar(s.get("R_p50", s.get("R_median")))

    def _meanbg_from_bg(bg, ch):
        if ch == "Blue":
            return _safe_gamma_linearize_scalar(bg.get("B_bg"))
        if ch == "Green":
            return _safe_gamma_linearize_scalar(bg.get("G_bg"))
        return _safe_gamma_linearize_scalar(bg.get("R_bg"))

    # Build a temporary table first, then derive plate references from it.
    tmp_rows = []
    for key, meta in meta_map.items():
        srow = stats_map.get(key)
        bg = bg_map.get(key)
        if srow is None or bg is None:
            continue
        row = {
            "Row": _row_label_from_index(int(meta["well_r"])),
            "Col": int(meta["well_c"]) + 1,
            "Well": meta["Well"],
            "ID": str(meta.get("ID", "")),
            "Type": str(meta.get("Type", "")),
            "Conc": float(meta.get("Conc", np.nan)),
            "DF": float(meta.get("DF", np.nan)),
            "L": float(srow.get("L_median", np.nan)),
            "a": float(srow.get("a_median", np.nan)),
            "b": float(srow.get("b_median", np.nan)),
            "ImageWarning": int(srow.get("is_image_quality_warning", 0)),
            "WarningReason": str(srow.get("warning_reason", "")),
            "UsedFraction": float(srow.get("used_fraction", np.nan)),
            "BrightExcludedFraction": float(srow.get("BrightExcludedFraction", np.nan)),
            "BrightExcludedMeanGray": float(srow.get("BrightExcludedMeanGray", np.nan)),
            "BrightExcessMeanGray": float(srow.get("BrightExcessMeanGray", np.nan)),
            "HighlightIndex": float(srow.get("HighlightIndex", np.nan)),
        }
        for ch in ["Blue", "Green", "Red"]:
            row[f"MeanW_{ch}"] = _meanw_from_stats(srow, ch)
            row[f"MeanBG_{ch}"] = _meanbg_from_bg(bg, ch)
        tmp_rows.append(row)

    def _median_channel(rows, field):
        vals = np.asarray([_num_or_nan(r.get(field, np.nan)) for r in rows], dtype=np.float64)
        vals = vals[np.isfinite(vals) & (vals > 0)]
        return float(np.nanmedian(vals)) if vals.size else np.nan

    zero_cal_rows = [
        r for r in tmp_rows
        if str(r.get("Type", "")).strip().upper() in type_cal
        and np.isfinite(r.get("Conc", np.nan))
        and abs(float(r.get("Conc", np.nan))) <= 1e-12
    ]
    blank_rows = [r for r in tmp_rows if str(r.get("Type", "")).strip().upper() in type_blank]

    stored_rgb_ref = {}
    if isinstance(stored_calibration_bundle, dict):
        stored_rgb_ref = stored_calibration_bundle.get("rgb_reference", {}) or {}

    def _stored_rgb_ref_value(ch):
        try:
            item = stored_rgb_ref.get(ch, {}) if isinstance(stored_rgb_ref, dict) else {}
            if isinstance(item, dict):
                v = float(item.get("value", np.nan))
            else:
                v = float(item)
            return v if np.isfinite(v) and v > 0 else np.nan
        except Exception:
            return np.nan

    refs = {}
    for ch in ["Blue", "Green", "Red"]:
        w0 = _median_channel(zero_cal_rows, f"MeanW_{ch}")
        w0_source = "zero_calibration" if np.isfinite(w0) else "unavailable"
        if not np.isfinite(w0):
            w0 = _stored_rgb_ref_value(ch)
            w0_source = "stored_calibration" if np.isfinite(w0) else "unavailable"

        wref_source = "zero_calibration" if np.isfinite(w0) else "unavailable"
        wref = w0
        if not np.isfinite(wref):
            wref = _median_channel(blank_rows, f"MeanW_{ch}")
            wref_source = "blank_empty"
        if not np.isfinite(wref):
            # Stored fallback is deliberately explicit, for inter-plate unknown-only runs.
            wref = _stored_rgb_ref_value(ch)
            wref_source = "stored_calibration" if np.isfinite(wref) else "unavailable"
        refs[ch] = {"W0": w0, "W0Source": w0_source, "UnusedPlateReference": wref, "UnusedPlateReferenceSource": wref_source}

    def _signal_from_t(t):
        if not np.isfinite(t) or t <= 0.0:
            return np.nan
        return float(-np.log10(t))

    rows = []
    for row in tmp_rows:
        out = dict(row)
        for ch in ["Blue", "Green", "Red"]:
            meanw = _num_or_nan(out.get(f"MeanW_{ch}", np.nan))
            meanbg = _num_or_nan(out.get(f"MeanBG_{ch}", np.nan))
            w0 = refs[ch]["W0"]
            wref = refs[ch]["UnusedPlateReference"]

            t_bg = (meanw / meanbg) if np.isfinite(meanw) and np.isfinite(meanbg) and meanw > 0 and meanbg > 0 else np.nan
            out[f"PAbs_{ch}"] = _signal_from_t(t_bg)

            t_signal = t_bg
            ref_value = meanbg

            out[f"SignalT_{ch}"] = t_signal
            out[f"SignalRef_{ch}"] = ref_value
            out[f"Signal_{ch}"] = _signal_from_t(t_signal)
        rows.append(out)

    rows.sort(key=lambda d: (str(d["ID"]), float(d["DF"]), str(d["Row"]), int(d["Col"])))
    return rows

def _apply_plate_local_cielab_reference(raw_rows):
    """
    Re-center CIELAB variables using the zero-concentration calibration points
    present in THE SAME plate.

    This keeps the correction plate-local and avoids importing a global CIELAB
    reference from another plate.

    Rule:
    - use only calibration rows (Type in C/CAL/CALIBRATION/STD/STANDARD)
    - use only rows with Conc == 0
    - compute plate-local medians of L, a, b
    - recompute DeltaL, Deltaa, Deltab, DeltaE_ab, DeltaE_ab_chroma
    """
    if not raw_rows:
        return raw_rows

    cal_types = {"C", "CAL", "CALIBRATION", "STD", "STANDARD"}

    ref_rows = []
    for row in raw_rows:
        typ = str(row.get("Type", "")).strip().upper()
        conc = row.get("Conc", np.nan)
        try:
            conc = float(conc)
        except Exception:
            conc = np.nan

        if typ in cal_types and np.isfinite(conc) and abs(conc) < 1e-15:
            L = float(row.get("L", np.nan))
            a = float(row.get("a", np.nan))
            b = float(row.get("b", np.nan))
            if np.isfinite(L) and np.isfinite(a) and np.isfinite(b):
                ref_rows.append((L, a, b))

    if not ref_rows:
        return raw_rows

    ref_arr = np.asarray(ref_rows, dtype=np.float64)
    L_ref = float(np.nanmedian(ref_arr[:, 0]))
    a_ref = float(np.nanmedian(ref_arr[:, 1]))
    b_ref = float(np.nanmedian(ref_arr[:, 2]))

    out = []
    for row in raw_rows:
        new_row = dict(row)

        L = float(new_row.get("L", np.nan))
        a = float(new_row.get("a", np.nan))
        b = float(new_row.get("b", np.nan))

        if np.isfinite(L) and np.isfinite(a) and np.isfinite(b):
            dL = L - L_ref
            da = a - a_ref
            db = b - b_ref

            new_row["L_ref"] = L_ref
            new_row["a_ref"] = a_ref
            new_row["b_ref"] = b_ref
            new_row["CIELAB_ref_source"] = "plate_zero_calibration"

            new_row["DeltaL"] = float(dL)
            new_row["Deltaa"] = float(da)
            new_row["Deltab"] = float(db)
            new_row["DeltaE_ab"] = float(np.sqrt(dL * dL + da * da + db * db))
            new_row["DeltaE_ab_chroma"] = float(np.sqrt(da * da + db * db))

        out.append(new_row)

    return out

def _choose_best_channel_from_fits(fit_rows):
    best_score = -1e12
    best_channel = None

    for r in fit_rows:
        if r["FitType"] != "Calibration":
            continue
        m = r.get("m", np.nan)
        r2 = r.get("R2", np.nan)
        if not np.isfinite(m) or not np.isfinite(r2):
            continue
        score = r2 * abs(m)
        if score > best_score:
            best_score = score
            best_channel = r["Channel"]

    return best_channel or "Red"



def _compute_group_qc_flags(rows):
    used = [float(r.get("UsedFraction", np.nan)) for r in rows if np.isfinite(r.get("UsedFraction", np.nan))]
    used_med = float(np.nanmedian(np.asarray(used, dtype=np.float64))) if len(used) else np.nan

    n_rep = int(len(rows))
    n_warn = 0
    n_crit = 0

    for r in rows:
        warn = int(r.get("ImageWarning", 0)) > 0
        used_fraction = float(r.get("UsedFraction", np.nan))
        warning_reason = str(r.get("WarningReason", ""))

        well_critical = (warning_reason == "empty_roi") or (np.isfinite(used_fraction) and used_fraction < 0.40)
        if warn:
            n_warn += 1
        if well_critical:
            n_crit += 1

    f_warn = float(n_warn / n_rep) if n_rep > 0 else np.nan
    f_crit = float(n_crit / n_rep) if n_rep > 0 else np.nan
    warn_any = int(n_warn > 0)

    critical = int(np.isfinite(f_crit) and f_crit >= 0.50)
    flagged = int((n_warn > 0) or (n_crit > 0) or (np.isfinite(used_med) and used_med < 0.70))

    return {
        "UsedFraction_median": used_med,
        "ImageWarning_any": warn_any,
        "QCFlagged": flagged,
        "QCCritical": critical,
        "NReplicates": n_rep,
        "NWellWarnings": n_warn,
        "NWellCritical": n_crit,
        "FracWellWarnings": f_warn,
        "FracWellCritical": f_crit,
    }


def _group_conc_key(value, typ=None):
    typ_u = str(typ or "").upper()
    if typ_u in {"UNK", "UNKNOWN", "U"}:
        return "__UNKNOWN__"
    try:
        v = float(value)
    except Exception:
        return np.nan
    return v


def _select_cielab_reference_rows(raw_rows):
    type_cal = {"CAL", "STD", "STANDARD", "CALIBRATION", "C"}
    type_blank = {"BLK", "BLANK", "EMPTY", "E"}

    usable = [
        r for r in raw_rows
        if np.isfinite(r.get("L", np.nan)) and np.isfinite(r.get("a", np.nan)) and np.isfinite(r.get("b", np.nan))
    ]
    if not usable:
        return [], "unavailable"

    blank_rows = [r for r in usable if str(r.get("Type", "")).upper() in type_blank]
    if blank_rows:
        return blank_rows, "blank_rows"

    cal_rows = [
        r for r in usable
        if str(r.get("Type", "")).upper() in type_cal and np.isfinite(r.get("Conc", np.nan))
    ]
    zero_rows = [r for r in cal_rows if abs(float(r.get("Conc", np.nan))) <= 1e-6]
    if zero_rows:
        return zero_rows, "zero_calibration"

    if cal_rows:
        min_conc = min(float(r.get("Conc", np.nan)) for r in cal_rows if np.isfinite(r.get("Conc", np.nan)))
        ref_rows = [r for r in cal_rows if np.isfinite(r.get("Conc", np.nan)) and abs(float(r.get("Conc", np.nan)) - min_conc) <= 1e-12]
        if ref_rows:
            return ref_rows, "lowest_calibration"

    return usable, "global_median"


def _extract_cielab_reference_payload(raw_rows):
    rows = [dict(r) for r in (raw_rows or [])]
    ref_rows, ref_source = _select_cielab_reference_rows(rows)
    if not ref_rows:
        return {"L_ref": np.nan, "a_ref": np.nan, "b_ref": np.nan, "source": "unavailable"}
    L0 = float(np.nanmedian(np.asarray([r.get("L", np.nan) for r in ref_rows], dtype=np.float64)))
    a0 = float(np.nanmedian(np.asarray([r.get("a", np.nan) for r in ref_rows], dtype=np.float64)))
    b0 = float(np.nanmedian(np.asarray([r.get("b", np.nan) for r in ref_rows], dtype=np.float64)))
    return {"L_ref": L0, "a_ref": a0, "b_ref": b0, "source": ref_source}


def _get_cielab_reference_from_bundle(stored_calibration_bundle):
    if not isinstance(stored_calibration_bundle, dict):
        return None
    payload = stored_calibration_bundle.get("cielab_reference", {}) or {}
    L0 = _num_or_nan(payload.get("L_ref", np.nan))
    a0 = _num_or_nan(payload.get("a_ref", np.nan))
    b0 = _num_or_nan(payload.get("b_ref", np.nan))
    source = str(payload.get("source", "stored_calibration"))
    if np.isfinite(L0) and np.isfinite(a0) and np.isfinite(b0):
        return {"L_ref": float(L0), "a_ref": float(a0), "b_ref": float(b0), "source": source or "stored_calibration"}
    return None


def _augment_raw_rows_with_deltae(raw_rows, stored_calibration_bundle=None):
    rows = [dict(r) for r in (raw_rows or [])]
    stored_ref = _get_cielab_reference_from_bundle(stored_calibration_bundle)
    if stored_ref is not None:
        L0 = float(stored_ref.get("L_ref", np.nan))
        a0 = float(stored_ref.get("a_ref", np.nan))
        b0 = float(stored_ref.get("b_ref", np.nan))
        ref_source = str(stored_ref.get("source", "stored_calibration"))
    else:
        ref_rows, ref_source = _select_cielab_reference_rows(rows)
        if not ref_rows:
            for row in rows:
                row["DeltaL"] = np.nan
                row["Deltaa"] = np.nan
                row["Deltab"] = np.nan
                row["DeltaE_ab"] = np.nan
                row["DeltaE_ab_chroma"] = np.nan
                row["CIELAB_ref_source"] = "unavailable"
                row["L_ref"] = np.nan
                row["a_ref"] = np.nan
                row["b_ref"] = np.nan
            return rows
        L0 = float(np.nanmedian(np.asarray([r.get("L", np.nan) for r in ref_rows], dtype=np.float64)))
        a0 = float(np.nanmedian(np.asarray([r.get("a", np.nan) for r in ref_rows], dtype=np.float64)))
        b0 = float(np.nanmedian(np.asarray([r.get("b", np.nan) for r in ref_rows], dtype=np.float64)))

    for row in rows:
        L = float(row.get("L", np.nan)) if np.isfinite(row.get("L", np.nan)) else np.nan
        a = float(row.get("a", np.nan)) if np.isfinite(row.get("a", np.nan)) else np.nan
        b = float(row.get("b", np.nan)) if np.isfinite(row.get("b", np.nan)) else np.nan
        if np.isfinite(L) and np.isfinite(a) and np.isfinite(b) and np.isfinite(L0) and np.isfinite(a0) and np.isfinite(b0):
            dL = float(L - L0)
            da = float(a - a0)
            db = float(b - b0)
            dEab = float(np.sqrt(dL * dL + da * da + db * db))
            dEch = float(np.sqrt(da * da + db * db))
        else:
            dL = np.nan
            da = np.nan
            db = np.nan
            dEab = np.nan
            dEch = np.nan
        row["DeltaL"] = dL
        row["Deltaa"] = da
        row["Deltab"] = db
        row["DeltaE_ab"] = dEab
        row["DeltaE_ab_chroma"] = dEch
        row["CIELAB_ref_source"] = ref_source
        row["L_ref"] = L0
        row["a_ref"] = a0
        row["b_ref"] = b0
    return rows


def _build_summary_rows(raw_rows):
    groups = {}
    for row in raw_rows:
        conc_key = _group_conc_key(row.get("Conc", np.nan), row.get("Type", ""))
        key = (str(row["ID"]), float(row["DF"]), str(row["Type"]), conc_key)
        groups.setdefault(key, []).append(row)

    out = []
    def _sort_key(k):
        conc_sort = k[3]
        if isinstance(conc_sort, str):
            conc_sort = np.inf
        return (k[0], k[1], conc_sort)

    summary_channels = [
        "MeanW_Red", "MeanW_Green", "MeanW_Blue",
        "MeanBG_Red", "MeanBG_Green", "MeanBG_Blue",
        "SignalT_Red", "SignalT_Green", "SignalT_Blue",
        "Signal_Red", "Signal_Green", "Signal_Blue",
        "PAbs_Red", "PAbs_Green", "PAbs_Blue",
        "L", "a", "b",
        "DeltaL", "Deltaa", "Deltab",
        "DeltaE_ab", "DeltaE_ab_chroma",
    ]

    for key in sorted(groups.keys(), key=_sort_key):
        rows = groups[key]
        conc_vals = [float(r.get("Conc", np.nan)) for r in rows if np.isfinite(r.get("Conc", np.nan))]
        conc_out = float(np.nanmedian(np.asarray(conc_vals, dtype=np.float64))) if len(conc_vals) else np.nan
        out_row = {
            "ID": key[0],
            "DF": key[1],
            "Type": key[2],
            "Conc": conc_out,
        }
        for ch in summary_channels:
            med, sd, count, mad = _median_and_robust_sd([r.get(ch, np.nan) for r in rows])
            out_row[f"{ch}_median"] = med
            out_row[f"{ch}__mad"] = mad
            out_row[f"{ch}_count"] = count
            out_row[f"{ch}_sd"] = sd
        ref_sources = [str(r.get("CIELAB_ref_source", "")) for r in rows if str(r.get("CIELAB_ref_source", ""))]
        out_row["CIELAB_ref_source"] = ref_sources[0] if ref_sources else ""
        _lref_vals = np.asarray([_num_or_nan(r.get("L_ref", np.nan)) for r in rows], dtype=np.float64) if rows else np.asarray([], dtype=np.float64)
        _lref_vals = _lref_vals[np.isfinite(_lref_vals)]
        out_row["L_ref"] = float(np.nanmedian(_lref_vals)) if _lref_vals.size else np.nan
        _aref_vals = np.asarray([_num_or_nan(r.get("a_ref", np.nan)) for r in rows], dtype=np.float64) if rows else np.asarray([], dtype=np.float64)
        _aref_vals = _aref_vals[np.isfinite(_aref_vals)]
        out_row["a_ref"] = float(np.nanmedian(_aref_vals)) if _aref_vals.size else np.nan
        _bref_vals = np.asarray([_num_or_nan(r.get("b_ref", np.nan)) for r in rows], dtype=np.float64) if rows else np.asarray([], dtype=np.float64)
        _bref_vals = _bref_vals[np.isfinite(_bref_vals)]
        out_row["b_ref"] = float(np.nanmedian(_bref_vals)) if _bref_vals.size else np.nan
        qc = _compute_group_qc_flags(rows)
        out_row.update(qc)
        out.append(out_row)
    return out


def _fit_line_with_covariance(x, y, w=None, force_zero=False):
    """Robust linear fit by IRLS using residual-based Huber weights only.

    The input argument w is accepted for backward compatibility but is
    deliberately ignored: this simplified version does not use SD weighting.
    All finite points are included; points farther from the fitted line receive
    lower robust weights instead of being excluded.
    """
    x = np.asarray(x, dtype=np.float64)
    y = np.asarray(y, dtype=np.float64)
    keep = np.isfinite(x) & np.isfinite(y)
    x = x[keep]
    y = y[keep]

    if x.size < 2:
        return None

    order = np.lexsort((y, x))
    x = x[order]
    y = y[order]

    def _solve(weight_vec):
        weight_vec = np.asarray(weight_vec, dtype=np.float64)
        weight_vec = np.where(np.isfinite(weight_vec) & (weight_vec > 0), weight_vec, 1.0)
        sw = np.sqrt(weight_vec)
        if force_zero:
            A = x[:, None]
        else:
            A = np.column_stack([x, np.ones_like(x)])
        Aw = A * sw[:, None]
        yw = y * sw
        coef, _, _, _ = np.linalg.lstsq(Aw, yw, rcond=None)
        if force_zero:
            m = float(coef[0])
            q = 0.0
        else:
            m = float(coef[0])
            q = float(coef[1])
        return m, q, A

    robust_w = np.ones_like(x, dtype=np.float64)
    m = np.nan
    q = 0.0
    for _ in range(12):
        m, q, A = _solve(robust_w)
        yhat_iter = m * x + q
        resid_iter = y - yhat_iter
        med = float(np.nanmedian(resid_iter)) if resid_iter.size else 0.0
        mad = float(np.nanmedian(np.abs(resid_iter - med))) if resid_iter.size else 0.0
        sigma = 1.4826 * mad
        if not np.isfinite(sigma) or sigma <= 1e-12:
            sigma = float(np.sqrt(np.nanmean(resid_iter ** 2))) if resid_iter.size else 0.0
        if not np.isfinite(sigma) or sigma <= 1e-12:
            break
        c = 1.5 * sigma
        abs_r = np.abs(resid_iter - med)
        new_w = np.ones_like(robust_w)
        mask = abs_r > c
        new_w[mask] = c / np.maximum(abs_r[mask], 1e-12)
        new_w = np.clip(new_w, 1e-6, 1.0)
        if np.allclose(new_w, robust_w, rtol=1e-5, atol=1e-7):
            robust_w = new_w
            break
        robust_w = new_w

    m, q, A = _solve(robust_w)
    yhat = m * x + q
    resid = y - yhat

    sse = float(np.sum(resid ** 2))
    sst = float(np.sum((y - np.mean(y)) ** 2))
    r2 = float(1.0 - sse / sst) if sst > 1e-15 else np.nan
    rmse = float(np.sqrt(np.mean(resid ** 2))) if resid.size > 0 else np.nan

    if force_zero:
        dof = max(1, x.size - 1)
        xtx = float(np.sum(robust_w * x * x))
        sigma2 = float(np.sum(robust_w * resid ** 2) / dof)
        if xtx > 1e-15:
            cov = np.array([[sigma2 / xtx, 0.0], [0.0, 0.0]], dtype=np.float64)
        else:
            cov = np.full((2, 2), np.nan, dtype=np.float64)
    else:
        if x.size > 2:
            dof = max(1, x.size - 2)
            sigma2 = float(np.sum(robust_w * resid ** 2) / dof)
            try:
                xtx_inv = np.linalg.inv(A.T @ (robust_w[:, None] * A))
                cov = sigma2 * xtx_inv
            except Exception:
                cov = np.full((2, 2), np.nan, dtype=np.float64)
        else:
            cov = np.full((2, 2), np.nan, dtype=np.float64)

    return {
        "n_points": int(x.size),
        "m": float(m),
        "q": float(q),
        "R2": r2,
        "RMSE": rmse,
        "cov_mq": cov,
        "x": x,
        "y": y,
        "w": robust_w,
        "yhat": yhat,
        "fit_method": "robust_irls_no_sd_weighting",
    }

def _stdadd_c0_sd_from_fit(fit):
    if fit is None:
        return np.nan, np.nan
    m = float(fit["m"])
    q = float(fit["q"])
    if not np.isfinite(m) or abs(m) < 1e-15 or not np.isfinite(q):
        return np.nan, np.nan

    x0 = -q / m
    cov = fit["cov_mq"]
    if cov is None or not np.isfinite(cov).all():
        return float(-x0), np.nan

    d_dm = q / (m * m)
    d_dq = -1.0 / m
    var_x0 = d_dm * d_dm * cov[0, 0] + d_dq * d_dq * cov[1, 1] + 2.0 * d_dm * d_dq * cov[0, 1]
    var_x0 = max(float(var_x0), 0.0)
    return float(-x0), float(np.sqrt(var_x0))

def _estimate_sigma_for_lod(summary_rows, base):
    """
    Noise estimate for LOD/LOQ.

    Preferred rule:
    use the median SD across calibration points, because the zero point can be
    degenerate for centered variables (for example Deltaa, Deltab, DeltaE_ab_chroma).

    Fallback:
    if the global median calibration SD is unavailable, use the SD at zero concentration.

    Returns
    -------
    (sigma_value, sigma_source)
    """
    type_cal = {"CAL", "STD", "STANDARD", "CALIBRATION", "C"}

    cal_points = [
        r for r in summary_rows
        if str(r.get("Type", "")).upper() in type_cal
        and np.isfinite(r.get(f"{base}_median", np.nan))
        and np.isfinite(r.get("Conc", np.nan))
    ]
    if len(cal_points) == 0:
        return np.nan, "unavailable"

    sigma_vals = [
        float(r.get(f"{base}_sd", np.nan))
        for r in cal_points
        if np.isfinite(r.get(f"{base}_sd", np.nan))
    ]
    if len(sigma_vals) > 0:
        sigma_med = float(np.nanmedian(np.asarray(sigma_vals, dtype=np.float64)))
        if np.isfinite(sigma_med) and sigma_med > 0:
            return sigma_med, "median_calibration_sd"

    zero_points = [
        r for r in cal_points
        if abs(float(r.get("Conc", np.nan))) <= 1e-12
        and np.isfinite(r.get(f"{base}_sd", np.nan))
    ]
    if len(zero_points) > 0:
        vals = np.asarray([float(r[f"{base}_sd"]) for r in zero_points], dtype=np.float64)
        vals = vals[np.isfinite(vals)]
        if vals.size > 0:
            return float(np.nanmedian(vals)), "blank_zero_calibration"

    return np.nan, "unavailable"

def _get_calibration_reference_from_summary(summary_rows, base):
    """Return the reference response used to center calibration-only fits.

    Priority:
    1) calibration blank/zero-concentration point
    2) lowest finite calibration concentration
    Returns (y_ref, ref_source).
    """
    type_cal = {"CAL", "STD", "STANDARD", "CALIBRATION", "C"}
    cal_points = [
        r for r in summary_rows
        if str(r.get("Type", "")).upper() in type_cal
        and np.isfinite(r.get(f"{base}_median", np.nan))
        and np.isfinite(r.get("Conc", np.nan))
    ]
    if len(cal_points) == 0:
        return np.nan, "unavailable"

    zero_points = [
        r for r in cal_points
        if abs(float(r.get("Conc", np.nan))) <= 1e-12
    ]
    if len(zero_points) > 0:
        vals = np.asarray([float(r.get(f"{base}_median", np.nan)) for r in zero_points], dtype=np.float64)
        vals = vals[np.isfinite(vals)]
        if vals.size > 0:
            return float(np.nanmedian(vals)), "zero_calibration"

    min_conc = min(float(r.get("Conc", np.nan)) for r in cal_points if np.isfinite(r.get("Conc", np.nan)))
    ref_rows = [
        r for r in cal_points
        if np.isfinite(r.get("Conc", np.nan)) and abs(float(r.get("Conc", np.nan)) - min_conc) <= 1e-12
    ]
    vals = np.asarray([float(r.get(f"{base}_median", np.nan)) for r in ref_rows], dtype=np.float64)
    vals = vals[np.isfinite(vals)]
    if vals.size > 0:
        return float(np.nanmedian(vals)), "lowest_calibration"

    return np.nan, "unavailable"


def _center_summary_rows_for_calibration(summary_rows, base):
    """Return shallow copies with calibration-point responses centered to the
    selected reference, plus the reference metadata.

    Standard-addition and unknown rows are left unchanged here. They are handled
    later when projected through calibration.
    """
    y_ref, ref_source = _get_calibration_reference_from_summary(summary_rows, base)
    out = []
    type_cal = {"CAL", "STD", "STANDARD", "CALIBRATION", "C"}
    for row in summary_rows:
        rr = dict(row)
        if str(rr.get("Type", "")).upper() in type_cal and np.isfinite(y_ref):
            yv = rr.get(f"{base}_median", np.nan)
            if np.isfinite(yv):
                rr[f"{base}_median"] = float(yv) - float(y_ref)
        rr[f"{base}_cal_ref"] = float(y_ref) if np.isfinite(y_ref) else np.nan
        rr[f"{base}_cal_ref_source"] = ref_source
        out.append(rr)
    return out, y_ref, ref_source


def _channel_stats_from_summary(summary_rows, base):
    type_stdadd = {"A", "SA", "STDADD", "STANDARD_ADDITION", "ADDITION"}

    centered_summary_rows, y_ref, ref_source = _center_summary_rows_for_calibration(summary_rows, base)
    type_cal = {"CAL", "STD", "STANDARD", "CALIBRATION", "C"}
    cal_points = [
        r for r in centered_summary_rows
        if str(r.get("Type", "")).upper() in type_cal and np.isfinite(r.get(f"{base}_median", np.nan)) and np.isfinite(r.get("Conc", np.nan))
    ]
    cal_fit = None
    sigma_cal, sigma_source = _estimate_sigma_for_lod(summary_rows, base)
    if len(cal_points) >= 3:
        x = [r["Conc"] for r in cal_points]
        y = [r[f"{base}_median"] for r in cal_points]
        cal_fit = _fit_line_with_covariance(x, y, force_zero=False)

    groups = {}
    for row in summary_rows:
        if str(row.get("Type", "")).upper() not in type_stdadd:
            continue
        yv = row.get(f"{base}_median", np.nan)
        xv = row.get("Conc", np.nan)
        if not (np.isfinite(yv) and np.isfinite(xv)):
            continue
        groups.setdefault((str(row["ID"]), float(row["DF"])), []).append(row)

    std_fits = []
    for key, rows in groups.items():
        rows_sorted = sorted(rows, key=lambda r: float(r["Conc"]))
        x = np.asarray([float(r["Conc"]) for r in rows_sorted], dtype=np.float64)
        y = np.asarray([float(r[f"{base}_median"]) for r in rows_sorted], dtype=np.float64)
        fit = _fit_line_with_covariance(x, y)
        if fit is not None:
            fit = dict(fit)
            fit["ID"] = key[0]
            fit["DF"] = key[1]
            std_fits.append(fit)

    slope_agreements = []
    if cal_fit is not None and len(std_fits) > 0 and np.isfinite(cal_fit.get("m", np.nan)):
        for f in std_fits:
            mcal = abs(float(cal_fit["m"]))
            madd = abs(float(f["m"]))
            denom = max(mcal, madd, 1e-12)
            slope_agreements.append(min(mcal, madd) / denom)
    slope_agreement = float(np.mean(slope_agreements)) if len(slope_agreements) > 0 else np.nan

    r2_cal = float(cal_fit["R2"]) if cal_fit is not None and np.isfinite(cal_fit["R2"]) else np.nan
    std_r2_vals = [f["R2"] for f in std_fits if np.isfinite(f.get("R2", np.nan))]
    r2_std = float(np.mean(std_r2_vals)) if len(std_r2_vals) > 0 else np.nan
    snr = float(abs(cal_fit["m"]) / sigma_cal) if cal_fit is not None and np.isfinite(sigma_cal) and sigma_cal > 0 else np.nan
    loq = float(10.0 * sigma_cal / abs(cal_fit["m"])) if cal_fit is not None and np.isfinite(sigma_cal) and sigma_cal > 0 and abs(float(cal_fit["m"])) > 1e-15 else np.nan
    score = np.nan
    if np.isfinite(r2_cal) and np.isfinite(r2_std) and np.isfinite(slope_agreement):
        score = _compute_balanced_score(r2_cal, r2_std, slope_agreement, loq=loq)

    return {
        "cal_fit": cal_fit,
        "std_fits": std_fits,
        "r2_cal": r2_cal,
        "r2_std": r2_std,
        "slope_agreement": slope_agreement,
        "score": score,
        "snr": snr,
        "loq": loq,
        "sigma_cal": sigma_cal,
        "sigma_source": sigma_source,
        "calibration_ref": y_ref,
        "calibration_ref_source": ref_source,
    }

def _compute_balanced_score(r2_cal, r2_std, slope_agreement, snr=None, loq=None):
    """Common global score used for channel/method ranking.

    Primary calibration + standard-addition score:
        GlobalScore = slope_agreement^2 * sqrt(R2_cal * R2_std) * (1 / LOQ)

    LOQ is used when available because it combines calibration sensitivity and
    noise in a standard analytical parameter. SNR, clipping, expected/recovery
    values are not used in the score. The snr argument is retained only for
    backward compatibility with existing calls.
    """
    return _compute_fit_base_score(r2_cal, r2_std, slope_agreement, loq=loq)


def _compute_fit_base_score(r2_cal, r2_std, slope_agreement, loq=None):
    """Fit-quality score with optional LOQ-dependent weighting.

    If LOQ is finite and positive:
        slope_agreement^2 * sqrt(R2_cal * R2_std) * (1 / LOQ)

    If LOQ is unavailable, the function returns the fit-only component for
    backward compatibility and for diagnostic methods lacking a valid LOQ.
    """
    if not (np.isfinite(r2_cal) and np.isfinite(r2_std) and np.isfinite(slope_agreement)):
        return np.nan
    base = float((float(slope_agreement) ** 2) * math.sqrt(max(float(r2_cal), 0.0) * max(float(r2_std), 0.0)))
    loq_v = _num_or_nan(loq)
    if np.isfinite(loq_v) and loq_v > 0:
        return float(base / loq_v)
    return base

def _ranking_should_use_snr(candidate_rows):
    """Global ranking never uses SNR.

    SNR is an RGB/calibration diagnostic and is not available in a comparable
    form for all CIELAB-derived methods. It is therefore excluded from the
    common ranking score.
    """
    return False


def _score_candidate_for_ranking(row, use_snr=False):
    """Return the common global ranking score for one method row.

    calibration + standard addition:
        slope_agreement^2 * sqrt(R2_cal * R2_std)

    calibration only and std-add only fallbacks are retained for diagnostic rows,
    but they are not directly comparable to full calibration+std-add methods.
    """
    r2_cal = _num_or_nan(row.get("R2_cal", np.nan))
    r2_std = _num_or_nan(row.get("R2_std", row.get("R2_std_mean", np.nan)))
    slope = _num_or_nan(row.get("SlopeAgreement", np.nan))
    mode = str(row.get("Mode", row.get("RankMode", "")))
    if mode == "calibration_plus_stdadd" or (np.isfinite(r2_cal) and np.isfinite(r2_std) and np.isfinite(slope)):
        loq = _num_or_nan(row.get("LOQ", row.get("loq", np.nan)))
        sc = _compute_fit_base_score(r2_cal, r2_std, slope, loq=loq)
        return float(sc) if np.isfinite(sc) and sc > 0 else 0.0
    score = _num_or_nan(row.get("Score", np.nan))
    return float(score) if np.isfinite(score) and score > 0 else 0.0




def _cfg_bool(cfg, key, default=False):
    try:
        v = (cfg or {}).get(key, default)
    except Exception:
        return bool(default)
    if isinstance(v, str):
        return v.strip().lower() in {"1", "true", "yes", "y", "on"}
    return bool(v)


def _large_odd_kernel(size_hint, min_size=101, max_size=501):
    k = int(round(size_hint))
    k = max(int(min_size), min(int(max_size), k))
    if k % 2 == 0:
        k += 1
    return max(3, k)


def _compute_initial_image_qc_payload(img_bgr):
    """Image-level QC from a single image.

    This is independent from calibration/standard-addition results and only
    reports quality alerts. The image is never modified.
    """
    if img_bgr is None:
        return {
            "initial_image_qc": "FAIL",
            "image_qc_class": "non_correctable",
            "image_qc_messages": "image could not be opened",
        }

    h, w = img_bgr.shape[:2]
    img_u8 = img_bgr.astype(np.uint8, copy=False)
    img_f = img_u8.astype(np.float32)
    gray = cv2.cvtColor(img_u8, cv2.COLOR_BGR2GRAY)
    hsv = cv2.cvtColor(img_u8, cv2.COLOR_BGR2HSV)

    max_side = max(int(w), int(h))
    approx_pitch = min(float(w) / 13.0, float(h) / 9.0)
    approx_floor_diam = 0.72 * approx_pitch
    approx_roi_pixels = float(math.pi * (0.5 * approx_floor_diam) ** 2)

    g_small = gray
    if max_side > 1200:
        scale = 1200.0 / float(max_side)
        g_small = cv2.resize(
            gray,
            (max(1, int(round(w * scale))), max(1, int(round(h * scale)))),
            interpolation=cv2.INTER_AREA,
        )
    try:
        blur_score = float(cv2.Laplacian(g_small, cv2.CV_64F).var())
    except Exception:
        blur_score = np.nan

    sat_any = np.any(img_u8 >= 250, axis=2)
    sat_all = np.all(img_u8 >= 250, axis=2)
    clip_low_any = np.any(img_u8 <= 3, axis=2)
    specular = (gray >= 245) & (hsv[:, :, 1] <= 35)

    saturation_fraction = float(np.mean(sat_any))
    saturation_all_channels_fraction = float(np.mean(sat_all))
    clip_low_fraction = float(np.mean(clip_low_any))
    specular_fraction = float(np.mean(specular))

    ch_mean = np.mean(img_f.reshape(-1, 3), axis=0)
    ch_std = np.std(img_f.reshape(-1, 3), axis=0)
    dead_channel = bool(np.any(ch_mean < 8.0) or np.any(ch_std < 1.5))

    k = _large_odd_kernel(min(h, w) / 5.0, min_size=151, max_size=601)
    try:
        slow = cv2.GaussianBlur(gray.astype(np.float32), (k, k), 0)
        med_slow = float(np.median(slow))
        flatfield_span = float((np.percentile(slow, 95) - np.percentile(slow, 5)) / max(med_slow, 1e-6))
    except Exception:
        flatfield_span = np.nan

    # Conservative thresholds. Saturation/specular clipping destroys information;
    # slow gradients and global WB are reported as quality warnings.
    messages = []
    destructive = False
    borderline = False
    quality_warning = False

    if dead_channel:
        destructive = True
        messages.append("dead channel")
    if saturation_fraction > 0.003 or saturation_all_channels_fraction > 0.0008:
        destructive = True
        messages.append("saturation")
    elif saturation_fraction > 0.0005:
        borderline = True
        messages.append("borderline saturation")
    if specular_fraction > 0.003:
        destructive = True
        messages.append("specular reflections")
    elif specular_fraction > 0.0005:
        borderline = True
        messages.append("borderline specular reflections")
    if np.isfinite(flatfield_span) and flatfield_span > 0.18:
        quality_warning = True
        messages.append("slow illumination/background gradient")
    if np.isfinite(blur_score) and blur_score < 35.0:
        borderline = True
        messages.append("borderline blur")
    if max_side < 900 or approx_roi_pixels < 120:
        destructive = True
        messages.append("resolution too low")

    if destructive:
        initial_qc = "FAIL"
        image_class = "non_correctable"
    elif quality_warning:
        initial_qc = "WARNING"
        image_class = "quality_warning"
    elif borderline:
        initial_qc = "WARNING"
        image_class = "usable_with_warnings"
    else:
        initial_qc = "OK"
        image_class = "good"

    return {
        "initial_image_qc": initial_qc,
        "image_qc_class": image_class,
        "image_qc_messages": "; ".join(dict.fromkeys(messages)) if messages else "",
        "flatfield_span": flatfield_span,
        "specular_fraction": specular_fraction,
        "saturation_fraction": saturation_fraction,
        "saturation_all_channels_fraction": saturation_all_channels_fraction,
        "clip_low_fraction": clip_low_fraction,
        "dead_channel": int(dead_channel),
        "blur_score_global": blur_score,
        "approx_well_pitch_px_pre": approx_pitch,
        "approx_roi_pixels_pre": approx_roi_pixels,
    }


def _select_best_channel(summary_rows):
    mapping = [("Signal_Red", "Signal_Red"), ("Signal_Green", "Signal_Green"), ("Signal_Blue", "Signal_Blue")]
    ranking = []
    for label, base in mapping:
        st = _channel_stats_from_summary(summary_rows, base)
        has_cal = st.get("cal_fit") is not None and np.isfinite(st.get("r2_cal", np.nan))
        has_std = len(st.get("std_fits", [])) > 0 and np.isfinite(st.get("r2_std", np.nan))

        if has_cal and has_std:
            mode = "calibration_plus_stdadd"
            score = _compute_balanced_score(st.get("r2_cal", np.nan), st.get("r2_std", np.nan), st.get("slope_agreement", np.nan), loq=st.get("loq", np.nan))
            std_best = None
            std_fits = [f for f in st.get("std_fits", []) if np.isfinite(f.get("R2", np.nan))]
            if len(std_fits) > 0:
                std_best = max(std_fits, key=lambda f: f.get("R2", -np.inf))
            c0_sd = np.nan
            if std_best is not None:
                try:
                    _, c0_sd = _stdadd_c0_sd_from_fit(std_best)
                except Exception:
                    c0_sd = np.nan
            ranking.append({
                "Channel": label,
                "Base": base,
                "Mode": mode,
                "R2_cal": st["r2_cal"],
                "R2_std": st["r2_std"],
                "SlopeAgreement": st["slope_agreement"],
                "Score": score,
                "SNR": st["snr"],
                "LOQ": st.get("loq", np.nan),
                "|mstd|": float(abs(std_best["m"])) if std_best is not None and np.isfinite(std_best.get("m", np.nan)) else np.nan,
                "C0_sd": c0_sd,
                "cal_fit": st["cal_fit"],
                "std_fits": st["std_fits"],
            })
        elif has_std:
            mode = "stdadd_only"
            std_best = None
            std_fits = [f for f in st.get("std_fits", []) if np.isfinite(f.get("R2", np.nan))]
            if len(std_fits) > 0:
                std_best = max(std_fits, key=lambda f: f.get("R2", -np.inf))
            if std_best is not None and np.isfinite(std_best.get("R2", np.nan)) and np.isfinite(std_best.get("m", np.nan)):
                score = float((std_best["R2"] ** 2) * abs(std_best["m"]))
                try:
                    _, c0_sd = _stdadd_c0_sd_from_fit(std_best)
                except Exception:
                    c0_sd = np.nan
            else:
                score = np.nan
                c0_sd = np.nan
            ranking.append({
                "Channel": label,
                "Base": base,
                "Mode": mode,
                "R2_cal": st["r2_cal"],
                "R2_std": st["r2_std"],
                "SlopeAgreement": st["slope_agreement"],
                "Score": score,
                "SNR": st["snr"],
                "LOQ": st.get("loq", np.nan),
                "|mstd|": float(abs(std_best["m"])) if std_best is not None and np.isfinite(std_best.get("m", np.nan)) else np.nan,
                "C0_sd": c0_sd,
                "cal_fit": st["cal_fit"],
                "std_fits": st["std_fits"],
            })
        elif has_cal:
            mode = "calibration_only"
            cal_fit = st.get("cal_fit")
            if cal_fit is not None and np.isfinite(cal_fit.get("R2", np.nan)) and np.isfinite(cal_fit.get("m", np.nan)):
                score = float((cal_fit["R2"] ** 2) * abs(cal_fit["m"]))
            else:
                score = np.nan
            ranking.append({
                "Channel": label,
                "Base": base,
                "Mode": mode,
                "R2_cal": st["r2_cal"],
                "R2_std": st["r2_std"],
                "SlopeAgreement": st["slope_agreement"],
                "Score": score,
                "SNR": st["snr"],
                "LOQ": st.get("loq", np.nan),
                "|mstd|": np.nan,
                "C0_sd": np.nan,
                "cal_fit": st["cal_fit"],
                "std_fits": st["std_fits"],
            })
        else:
            ranking.append({
                "Channel": label,
                "Base": base,
                "Mode": "unavailable",
                "R2_cal": st["r2_cal"],
                "R2_std": st["r2_std"],
                "SlopeAgreement": st["slope_agreement"],
                "Score": np.nan,
                "SNR": st["snr"],
                "LOQ": st.get("loq", np.nan),
                "|mstd|": np.nan,
                "C0_sd": np.nan,
                "cal_fit": st["cal_fit"],
                "std_fits": st["std_fits"],
            })

    ranking_sorted = sorted(ranking, key=lambda d: (-d["Score"] if np.isfinite(d.get("Score", np.nan)) else np.inf))
    best = ranking_sorted[0]

    reason = [
        f"Selected channel: {best['Channel'].upper()}",
        "Selection criterion:",
        "Balanced fit-quality ranking:",
        "if calibration + std add: GlobalScore = (slope_agreement^2) x sqrt(R²cal x R²std) x (1/LOQ); SNR, clipping and expected values are not used",
        "if std add only: score = (R²std^2) x |mstd|",
        "if calibration only: score = (R²cal^2) x |mcal|",
        "slope_agreement = min(|mcal|,|madd|)/max(|mcal|,|madd|)",
        "",
        "Ranking:",
    ]
    for r in ranking_sorted:
        score_txt = f"{_num_or_nan(r.get('Score', np.nan)):.4f}" if _isfinite_num(r.get('Score', np.nan)) else "nan"
        r2cal_txt = f"{_num_or_nan(r.get('R2_cal', np.nan)):.4f}" if _isfinite_num(r.get('R2_cal', np.nan)) else "nan"
        r2std_txt = f"{_num_or_nan(r.get('R2_std', np.nan)):.4f}" if _isfinite_num(r.get('R2_std', np.nan)) else "nan"
        slope_txt = f"{_num_or_nan(r.get('SlopeAgreement', np.nan)):.4f}" if _isfinite_num(r.get('SlopeAgreement', np.nan)) else "nan"
        snr_txt = f"{_num_or_nan(r.get('SNR', np.nan)):.3f}" if _isfinite_num(r.get('SNR', np.nan)) else "nan"
        mstd_txt = f"{_num_or_nan(r.get('|mstd|', np.nan)):.5f}" if _isfinite_num(r.get('|mstd|', np.nan)) else "nan"
        loq_txt = f"{_num_or_nan(r.get('LOQ', np.nan)):.5g}" if _isfinite_num(r.get('LOQ', np.nan)) else "nan"
        c0sd_txt = f"{_num_or_nan(r.get('C0_sd', np.nan)):.3f}" if _isfinite_num(r.get('C0_sd', np.nan)) else "nan"
        reason.append(f"{r['Channel']}: mode={r['Mode']}  score={score_txt}  R²cal={r2cal_txt}  R²std={r2std_txt}  slope={slope_txt}  LOQ={loq_txt}  |mstd|={mstd_txt}  SNR={snr_txt}  C0sd={c0sd_txt}")

    return best["Channel"], {"best": best, "ranking": ranking_sorted, "reason_lines": reason}


def _get_stored_selected_channel(stored_calibration_bundle):
    if not isinstance(stored_calibration_bundle, dict):
        return None

    def _normalize_to_rgb_signal(name):
        s = str(name or "").strip()
        if s in {"Red", "Green", "Blue"}:
            return f"Signal_{s}"
        if s in {"Signal_Red", "Signal_Green", "Signal_Blue"}:
            return s
        # Legacy stored calibration files may contain PAbs_* as selected channel.
        # The current RGB analytical family is Signal_*; for full_bg the numerical
        # definition is equivalent, while other modes require a regenerated stored calibration.
        if s in {"PAbs_Red", "PAbs_Green", "PAbs_Blue"}:
            return "Signal_" + s.split("_", 1)[1]
        return None

    selected_channel = _normalize_to_rgb_signal(stored_calibration_bundle.get("selected_channel"))
    if selected_channel is not None:
        return selected_channel
    selection_info = stored_calibration_bundle.get("selection_info", {}) or {}
    best = selection_info.get("best", {}) or {}
    best_channel = _normalize_to_rgb_signal(best.get("Channel"))
    if best_channel is not None:
        return best_channel
    return None

def _get_stored_calibration_path(image_path, cfg):
    p = cfg.get("stored_calibration_path", None)
    if p:
        return str(p)
    return str(Path(image_path).with_name("stored_calibration.json"))


def _load_stored_calibration_bundle(path):
    p = Path(path)
    if not p.exists():
        return None
    try:
        with open(p, "r", encoding="utf-8") as f:
            data = json.load(f)
        if not isinstance(data, dict):
            return None
        return data
    except Exception:
        return None


def _json_safe(obj):
    if isinstance(obj, dict):
        return {str(k): _json_safe(v) for k, v in obj.items()}
    if isinstance(obj, (list, tuple)):
        return [_json_safe(v) for v in obj]
    if isinstance(obj, np.ndarray):
        return obj.tolist()
    if isinstance(obj, np.floating):
        v = float(obj)
        return v if np.isfinite(v) else None
    if isinstance(obj, np.integer):
        return int(obj)
    if obj is None:
        return None
    if isinstance(obj, float):
        return obj if np.isfinite(obj) else None
    if isinstance(obj, (str, int, bool)):
        return obj
    return str(obj)




def _sanitize_rgb_references_in_lines(lines):
    out = []
    for line in lines or []:
        s = str(line)
        s = s.replace("selected single RGB channel", "selected single channel")
        s = s.replace("selected RGB channel", "selected channel")
        s = s.replace("single RGB channel", "single channel")
        s = s.replace("RGB channel", "channel")
        out.append(s)
    return out


def _calibration_points_from_plot_payload(plot_payload):
    out = {}
    if not isinstance(plot_payload, dict):
        return out
    for label, payload in plot_payload.items():
        pts = []
        for p in payload.get("calibration_points", []) or []:
            pts.append({
                "x": float(p.get("x", np.nan)) if np.isfinite(p.get("x", np.nan)) else None,
                "y": float(p.get("y", np.nan)) if np.isfinite(p.get("y", np.nan)) else None,
                "yerr": float(p.get("yerr", np.nan)) if np.isfinite(p.get("yerr", np.nan)) else None,
                "excluded": bool(p.get("excluded", False)),
            })
        out[label] = pts
    return out


def _empty_bg_rows_for_wells(centers):
    rows = []
    nrow, ncol = centers.shape[:2]
    for r in range(nrow):
        for c in range(ncol):
            x = float(centers[r, c, 0])
            y = float(centers[r, c, 1])
            rows.append({
                "well_r": int(r),
                "well_c": int(c),
                "x": x,
                "y": y,
                "B_bg": 1.0,
                "G_bg": 1.0,
                "R_bg": 1.0,
            })
    return rows




def _get_stored_blank_value(stored_calibration_bundle, base):
    if not isinstance(stored_calibration_bundle, dict):
        return np.nan, np.nan

    blank_info = stored_calibration_bundle.get("blank_info", {}) or {}
    row = blank_info.get(base, {}) or {}

    blank = row.get("blank", None)
    blank_sd = row.get("blank_sd", None)

    try:
        blank = float(blank) if blank is not None else np.nan
    except Exception:
        blank = np.nan

    try:
        blank_sd = float(blank_sd) if blank_sd is not None else np.nan
    except Exception:
        blank_sd = np.nan

    return blank, blank_sd

def _apply_blank_mode_to_rows(raw_rows, blank_info, blank_mode):
    blank_mode = str(blank_mode or "none").strip().lower()
    if blank_mode not in {"none", "both", "calibration_only", "sam_only"}:
        blank_mode = "none"
    if blank_mode == "none":
        return raw_rows

    type_cal = {"CAL", "STD", "STANDARD", "CALIBRATION", "C"}
    type_stdadd = {"A", "SA", "STDADD", "STANDARD_ADDITION", "ADDITION"}

    out = []
    for r in raw_rows:
        rr = dict(r)
        rtype = str(rr.get("Type", "")).upper()
        apply_here = (
            (blank_mode == "both") or
            (blank_mode == "calibration_only" and rtype in type_cal) or
            (blank_mode == "sam_only" and rtype in type_stdadd)
        )
        if apply_here:
            for _base in ["Signal_Red", "Signal_Green", "Signal_Blue"]:
                _b = blank_info.get(_base, {}).get("blank", np.nan)
                _v = rr.get(_base, np.nan)
                if np.isfinite(_v) and np.isfinite(_b):
                    rr[_base] = float(_v) - float(_b)
        out.append(rr)
    return out


def _compute_raw_rows_for_background_mode(raw_report_rows_base, use_background_subtraction, blank_mode="none", stored_calibration_bundle=None):
    raw_rows = [dict(r) for r in raw_report_rows_base]
    blank_info = {}
    for _base in ["Signal_Red", "Signal_Green", "Signal_Blue"]:
        if use_background_subtraction:
            _b, _sd = _compute_blank_from_raw(raw_rows, _base)
        else:
            _b, _sd = np.nan, np.nan
        blank_info[_base] = {"blank": _b, "blank_sd": _sd}
    raw_rows = _apply_blank_mode_to_rows(raw_rows, blank_info, blank_mode)
    return raw_rows, blank_info


def _compute_raw_rows_for_stored_background_mode(raw_report_rows_base, use_background_subtraction, blank_mode="none", stored_calibration_bundle=None):
    raw_rows = [dict(r) for r in raw_report_rows_base]
    blank_info = {}
    for _base in ["Signal_Red", "Signal_Green", "Signal_Blue"]:
        if use_background_subtraction:
            _b, _sd = _get_stored_blank_value(stored_calibration_bundle, _base)
        else:
            _b, _sd = np.nan, np.nan
        blank_info[_base] = {"blank": _b, "blank_sd": _sd}
    raw_rows = _apply_blank_mode_to_rows(raw_rows, blank_info, blank_mode)
    return raw_rows, blank_info

def _save_stored_calibration_bundle(path, image_basename, unit_label, selected_channel, fit_rows, selection_info, blank_info=None, plot_payload=None, empty_well_payload=None, empty_well_rows=None, well_bg_rows=None, bg_samples=None, bg_models=None, fit_rows_cielab=None, plot_payload_cielab=None, fit_rows_deltae=None, plot_payload_deltae=None, cielab_reference=None, rgb_reference=None):
    calibration_points_by_channel = {}
    if isinstance(plot_payload, dict):
        for ch, payload in plot_payload.items():
            pts = payload.get("calibration_points", []) or []
            calibration_points_by_channel[ch] = []
            for p in pts:
                xv = p.get("x", np.nan)
                yv = p.get("y", np.nan)
                if not (np.isfinite(xv) and np.isfinite(yv)):
                    continue
                yerr = p.get("yerr", np.nan)
                calibration_points_by_channel[ch].append({
                    "x": float(xv),
                    "y": float(yv),
                    "yerr": float(yerr) if np.isfinite(yerr) else None,
                    "excluded": bool(p.get("excluded", False)),
                })

    if isinstance(plot_payload_cielab, dict):
        for ch, payload in plot_payload_cielab.items():
            pts = payload.get("calibration_points", []) or []
            calibration_points_by_channel[ch] = []
            for p in pts:
                xv = p.get("x", np.nan)
                yv = p.get("y", np.nan)
                if not (np.isfinite(xv) and np.isfinite(yv)):
                    continue
                yerr = p.get("yerr", np.nan)
                calibration_points_by_channel[ch].append({
                    "x": float(xv),
                    "y": float(yv),
                    "yerr": float(yerr) if np.isfinite(yerr) else None,
                    "excluded": bool(p.get("excluded", False)),
                })

    if isinstance(plot_payload_deltae, dict):
        for ch, payload in plot_payload_deltae.items():
            pts = payload.get("calibration_points", []) or []
            calibration_points_by_channel[ch] = []
            for p in pts:
                xv = p.get("x", np.nan)
                yv = p.get("y", np.nan)
                if not (np.isfinite(xv) and np.isfinite(yv)):
                    continue
                yerr = p.get("yerr", np.nan)
                calibration_points_by_channel[ch].append({
                    "x": float(xv),
                    "y": float(yv),
                    "yerr": float(yerr) if np.isfinite(yerr) else None,
                    "excluded": bool(p.get("excluded", False)),
                })

    channels = {}
    for row in list(fit_rows or []) + list(fit_rows_cielab or []) + list(fit_rows_deltae or []):
        if row.get("FitType") != "Calibration":
            continue
        ch = row.get("Channel")
        channels[ch] = {
            "Channel": ch,
            "FitType": "Calibration",
            "n_points": int(row.get("n_points", 0) or 0),
            "m": float(row["m"]) if np.isfinite(row.get("m", np.nan)) else None,
            "q": float(row["q"]) if np.isfinite(row.get("q", np.nan)) else None,
            "R2": float(row["R2"]) if np.isfinite(row.get("R2", np.nan)) else None,
            "RMSE": float(row["RMSE"]) if np.isfinite(row.get("RMSE", np.nan)) else None,
            "sigma_cal": float(row["sigma_cal"]) if np.isfinite(row.get("sigma_cal", np.nan)) else None,
            "sigma_source": row.get("sigma_source", "unavailable"),
            "SNR": float(row["SNR"]) if np.isfinite(row.get("SNR", np.nan)) else None,
            "LOD": float(row["LOD"]) if np.isfinite(row.get("LOD", np.nan)) else None,
            "LOQ": float(row["LOQ"]) if np.isfinite(row.get("LOQ", np.nan)) else None,
            "a": float(row["a"]) if np.isfinite(row.get("a", np.nan)) else None,
            "b": float(row["b"]) if np.isfinite(row.get("b", np.nan)) else None,
            "c": float(row["c"]) if np.isfinite(row.get("c", np.nan)) else None,
            "S0_calibration": float(row["S0_calibration"]) if np.isfinite(row.get("S0_calibration", np.nan)) else None,
            "NClipPoints": int(row.get("NClipPoints", 0) or 0) if np.isfinite(row.get("NClipPoints", np.nan)) else None,
            "ClipX": row.get("ClipX", ""),
            "ClipDelta": row.get("ClipDelta", ""),
            "rgb_low_signal_correction": _json_safe(row.get("rgb_low_signal_correction", {}) or {}),
            "calibration_points": calibration_points_by_channel.get(ch, []),
        }

    bundle = {
        "image_basename": image_basename,
        "unit_label": unit_label,
        "selected_channel": str(selected_channel),
        "selection_info": _json_safe(selection_info),
        "blank_info": _json_safe(blank_info or {}),
        "channels": channels,
        "empty_well_payload": _json_safe(empty_well_payload or {}),
        "empty_well_rows": _json_safe(empty_well_rows or []),
        "well_bg_rows": _json_safe(well_bg_rows or []),
        "bg_samples": _json_safe(bg_samples or []),
        "bg_models": _json_safe(bg_models or {}),
        "cielab_reference": _json_safe(cielab_reference or {}),
        "rgb_reference": _json_safe(rgb_reference or {}),
    }
    with open(path, "w", encoding="utf-8") as f:
        json.dump(bundle, f, indent=2)


def _bg_model_to_jsonable(model):
    if not isinstance(model, dict):
        return {}
    return {
        "coef": [float(v) for v in np.asarray(model.get("coef", []), dtype=np.float64).ravel()],
        "x0": float(model.get("x0", np.nan)) if np.isfinite(model.get("x0", np.nan)) else None,
        "y0": float(model.get("y0", np.nan)) if np.isfinite(model.get("y0", np.nan)) else None,
        "sx": float(model.get("sx", np.nan)) if np.isfinite(model.get("sx", np.nan)) else None,
        "sy": float(model.get("sy", np.nan)) if np.isfinite(model.get("sy", np.nan)) else None,
    }


def _bg_model_from_jsonable(payload):
    if not isinstance(payload, dict):
        return None
    coef = np.asarray(payload.get("coef", []), dtype=np.float64)
    if coef.size != 6:
        return None
    vals = {
        "coef": coef,
        "x0": float(payload.get("x0", np.nan)),
        "y0": float(payload.get("y0", np.nan)),
        "sx": float(payload.get("sx", np.nan)),
        "sy": float(payload.get("sy", np.nan)),
    }
    if not all(np.isfinite(vals[k]) for k in ["x0", "y0", "sx", "sy"]):
        return None
    return vals


def _stored_bg_surface_rows_for_wells(centers, stored_calibration_bundle):
    if not isinstance(stored_calibration_bundle, dict):
        return []
    bg_models = stored_calibration_bundle.get("bg_models", {}) or {}
    model_B = _bg_model_from_jsonable(bg_models.get("B", {}))
    model_G = _bg_model_from_jsonable(bg_models.get("G", {}))
    model_R = _bg_model_from_jsonable(bg_models.get("R", {}))
    if model_B is not None and model_G is not None and model_R is not None:
        return _predict_bg_for_wells(centers, model_B, model_G, model_R)
    return list(stored_calibration_bundle.get("well_bg_rows", []) or [])


def _build_interplate_bg2d_correction_rows(raw_rows, current_well_bg_rows, stored_calibration_bundle, centers):
    """
    Diagnostic-only inter-plate correction based on the 2D BG surfaces reconstructed
    from inter-well spaces.

    The official analytical pipeline is NOT modified. This function only reports, for
    each well, the raw value, the stored/current BG surfaces evaluated at that well,
    the local scale factor, and the corrected relative intensity signal in the signal domain:
        Signal_corr = ((Signal_raw + 1) * scale_factor) - 1
    """
    stored_well_bg_rows = _stored_bg_surface_rows_for_wells(centers, stored_calibration_bundle)
    current_bg_map = {(int(r.get("well_r", -1)), int(r.get("well_c", -1))): dict(r) for r in current_well_bg_rows}
    stored_bg_map = {(int(r.get("well_r", -1)), int(r.get("well_c", -1))): dict(r) for r in stored_well_bg_rows}

    out = []
    for row in raw_rows:
        try:
            rr = _row_index_from_label(str(row.get("Row", "A")))
            cc = int(row.get("Col", 1)) - 1
        except Exception:
            continue
        cur_bg = current_bg_map.get((rr, cc))
        sto_bg = stored_bg_map.get((rr, cc))
        if cur_bg is None or sto_bg is None:
            continue

        out_row = {
            "Row": row.get("Row"),
            "Col": row.get("Col"),
            "Well": row.get("Well"),
            "ID": row.get("ID"),
            "Type": row.get("Type"),
            "Conc": row.get("Conc"),
            "DF": row.get("DF"),
        }

        for label, suffix in [("Blue", "Blue"), ("Green", "Green"), ("Red", "Red")]:
            bg_field = f"{label[0]}_bg" if label != "Green" else "G_bg"
            if label == "Blue":
                bg_field = "B_bg"
            elif label == "Red":
                bg_field = "R_bg"
            meanw_key = f"MeanW_{suffix}"
            abs_key = f"Signal_{suffix}"

            bg_cur_raw = float(cur_bg.get(bg_field, np.nan)) if np.isfinite(cur_bg.get(bg_field, np.nan)) else np.nan
            bg_sto_raw = float(sto_bg.get(bg_field, np.nan)) if np.isfinite(sto_bg.get(bg_field, np.nan)) else np.nan
            bg_cur_lin = _safe_gamma_linearize_scalar(bg_cur_raw)
            bg_sto_lin = _safe_gamma_linearize_scalar(bg_sto_raw)
            meanw_raw = float(row.get(meanw_key, np.nan)) if np.isfinite(row.get(meanw_key, np.nan)) else np.nan
            scale_factor = (bg_sto_lin / bg_cur_lin) if np.isfinite(bg_sto_lin) and np.isfinite(bg_cur_lin) and bg_cur_lin > 0 else np.nan
            meanw_corr = (meanw_raw * scale_factor) if np.isfinite(meanw_raw) and np.isfinite(scale_factor) else np.nan
            signal_raw = float(row.get(abs_key, np.nan)) if np.isfinite(row.get(abs_key, np.nan)) else np.nan
            signal_corr = (((signal_raw + 1.0) * scale_factor) - 1.0) if np.isfinite(signal_raw) and np.isfinite(scale_factor) else np.nan

            out_row[f"BG_{suffix}_current_raw"] = bg_cur_raw
            out_row[f"BG_{suffix}_stored_raw"] = bg_sto_raw
            out_row[f"BG_{suffix}_current_lin"] = bg_cur_lin
            out_row[f"BG_{suffix}_stored_lin"] = bg_sto_lin
            out_row[f"ScaleFactor_{suffix}"] = scale_factor
            out_row[f"ScaleFactorShift_{suffix}"] = (scale_factor - 1.0) if np.isfinite(scale_factor) else np.nan
            out_row[f"MeanW_{suffix}_raw"] = meanw_raw
            out_row[f"MeanW_{suffix}_corr"] = meanw_corr
            out_row[f"Signal_{suffix}_raw"] = signal_raw
            out_row[f"Signal_{suffix}_corr"] = signal_corr

        out.append(out_row)
    out.sort(key=lambda d: (str(d.get("ID", "")), float(d.get("DF", np.nan)) if np.isfinite(d.get("DF", np.nan)) else np.inf, str(d.get("Row", "")), int(d.get("Col", 0))))
    return out


def _write_bg2d_correction_csv(path, rows):
    if not rows:
        return
    headers = list(rows[0].keys())
    with open(path, "w", newline="", encoding="utf-8") as f:
        wr = csv.writer(f, delimiter=';')
        wr.writerow(headers)
        for row in rows:
            wr.writerow([row.get(h, "") for h in headers])


def _append_bg2d_correction_to_workbook(path, rows, stored_calibration_bundle):
    from openpyxl import load_workbook
    wb = load_workbook(path)
    if "13_BG2D_CORRECTION" in wb.sheetnames:
        del wb["13_BG2D_CORRECTION"]
    ws = wb.create_sheet("13_BG2D_CORRECTION")

    stored_name = ""
    if isinstance(stored_calibration_bundle, dict):
        stored_name = str(stored_calibration_bundle.get("image_basename", ""))

    intro = [
        "Inter-plate 2D BG correction based on the reconstructed inter-well background surface.",
        "Diagnostic only: official fitting/results are unchanged.",
        "Corrected signal is computed as Signal_corr = ((Signal_raw + 1) * scale_factor) - 1.",
        f"Stored calibration source: {stored_name}",
        "",
    ]
    for i, line in enumerate(intro, start=1):
        ws.cell(i, 1, line)

    if rows:
        headers = list(rows[0].keys())
        _write_table(ws, 7, headers, rows)
    _autosize_worksheet(ws)
    wb.save(path)




def _build_bg2d_corrected_raw_rows(raw_rows, bg2d_correction_rows):
    """
    Build a parallel RAW table where analytical signal fields are replaced by the
    BG2D-corrected values when available. The original official pipeline is left intact.
    """
    corr_map = {str(r.get("Well", "")): dict(r) for r in (bg2d_correction_rows or [])}
    out = []
    for row in raw_rows:
        rr = dict(row)
        corr = corr_map.get(str(rr.get("Well", "")), {})
        for suffix in ["Blue", "Green", "Red"]:
            key_corr = f"Signal_{suffix}_corr"
            key_abs = f"Signal_{suffix}"
            v = corr.get(key_corr, np.nan)
            if np.isfinite(v):
                rr[key_abs] = float(v)
        out.append(rr)
    return out


def _summarize_bg2d_scale_factors(bg2d_correction_rows):
    rows = []
    for suffix in ["Blue", "Green", "Red"]:
        vals = np.asarray([float(r.get(f"ScaleFactor_{suffix}", np.nan)) for r in (bg2d_correction_rows or []) if np.isfinite(r.get(f"ScaleFactor_{suffix}", np.nan))], dtype=np.float64)
        dlogs = np.asarray([float(r.get(f"ScaleFactorShift_{suffix}", np.nan)) for r in (bg2d_correction_rows or []) if np.isfinite(r.get(f"ScaleFactorShift_{suffix}", np.nan))], dtype=np.float64)
        rows.append({
            "Channel": suffix,
            "n": int(vals.size),
            "ScaleFactor_mean": float(np.mean(vals)) if vals.size else np.nan,
            "ScaleFactor_median": float(np.median(vals)) if vals.size else np.nan,
            "ScaleFactor_sd": float(np.std(vals, ddof=1)) if vals.size > 1 else np.nan,
            "ScaleFactorShift_mean": float(np.mean(dlogs)) if dlogs.size else np.nan,
            "ScaleFactorShift_median": float(np.median(dlogs)) if dlogs.size else np.nan,
            "ScaleFactorShift_sd": float(np.std(dlogs, ddof=1)) if dlogs.size > 1 else np.nan,
        })
    return rows


def _append_bg2d_corrected_analysis_to_workbook(path, raw_rows_corr, summary_rows_corr, fit_rows_corr, stored_calibration_bundle=None, scale_summary_rows=None):
    from openpyxl import load_workbook
    wb = load_workbook(path)
    for name in ["14_BG2D_SCALE_SUMMARY", "15_BG2D_SUMMARY", "16_BG2D_FITTING"]:
        if name in wb.sheetnames:
            del wb[name]
    ws0 = wb.create_sheet("14_BG2D_SCALE_SUMMARY")
    ws1 = wb.create_sheet("15_BG2D_SUMMARY")
    ws2 = wb.create_sheet("16_BG2D_FITTING")

    stored_name = ""
    if isinstance(stored_calibration_bundle, dict):
        stored_name = str(stored_calibration_bundle.get("image_basename", ""))

    intro0 = [
        "BG2D inter-plate scale summary from reconstructed inter-well background surfaces.",
        f"Stored calibration source: {stored_name}",
        "",
    ]
    for i, line in enumerate(intro0, start=1):
        ws0.cell(i, 1, line)
    if scale_summary_rows:
        _write_table(ws0, 5, list(scale_summary_rows[0].keys()), scale_summary_rows)

    intro1 = [
        "BG2D-corrected analytical summary (parallel analysis; official pipeline unchanged).",
        f"Stored calibration source: {stored_name}",
        "Relative intensity signal fields were replaced by Signal_*_corr derived from the BG2D scale factor.",
        "",
    ]
    for i, line in enumerate(intro1, start=1):
        ws1.cell(i, 1, line)
    if summary_rows_corr:
        summary_headers = [
            "ID", "DF", "Type", "Conc",
            "Signal_Red_median", "Signal_Red_sd",
            "Signal_Green_median", "Signal_Green_sd",
            "Signal_Blue_median", "Signal_Blue_sd",
            "UsedFraction_median", "ImageWarning_any", "QCFlagged", "QCCritical",
            "NReplicates", "NWellWarnings", "NWellCritical", "FracWellWarnings", "FracWellCritical"
        ]
        _write_table(ws1, 6, summary_headers, summary_rows_corr)

    intro2 = [
        "BG2D-corrected fitting results (parallel analysis; official pipeline unchanged).",
        f"Stored calibration source: {stored_name}",
        "Unknown and standard-addition values are computed from BG2D-corrected signal.",
        "",
    ]
    for i, line in enumerate(intro2, start=1):
        ws2.cell(i, 1, line)
    if fit_rows_corr:
        fit_headers = [
            "Channel", "FitType", "ID", "DF", "n_points", "m", "q", "R2", "RMSE",
            "sigma_cal", "sigma_source", "SNR", "LOD", "LOQ", "C0", "C0_sd",
            "beta_k", "bias_index_k"
        ]
        _write_table(ws2, 6, fit_headers, fit_rows_corr)

    for ws in [ws0, ws1, ws2]:
        _autosize_worksheet(ws)
    wb.save(path)


def _build_bg_channel_model_rows(raw_rows, fit_rows, expected_refs=None):
    """
    Build diagnostic model-input rows combining background descriptors and channel-derived
    concentration estimates for non-calibration groups.

    This does not alter the official pipeline. It exports per-group predictors that can be
    merged across plates to fit external BG + channel correction models.
    """
    expected_refs = expected_refs or []
    type_skip = {"CAL", "STD", "STANDARD", "CALIBRATION", "C", "BLK", "BLANK", "EMPTY", "E"}
    groups = {}
    for row in (raw_rows or []):
        typ = str(row.get("Type", "")).upper()
        if typ in type_skip:
            continue
        key = (str(row.get("ID", "")), float(row.get("DF", np.nan)) if np.isfinite(row.get("DF", np.nan)) else np.nan, typ)
        groups.setdefault(key, []).append(row)

    fit_pref = {}
    for row in (fit_rows or []):
        ftype = str(row.get("FitType", ""))
        if ftype not in {"UnknownFromCal", "StdAdd"}:
            continue
        ch = str(row.get("Channel", ""))
        key = (ch, str(row.get("ID", "")), float(row.get("DF", np.nan)) if np.isfinite(row.get("DF", np.nan)) else np.nan)
        prev = fit_pref.get(key)
        if prev is None:
            fit_pref[key] = row
        elif str(prev.get("FitType", "")) != "UnknownFromCal" and ftype == "UnknownFromCal":
            fit_pref[key] = row

    out = []
    for (sample_id, df_val, typ), rows in sorted(groups.items(), key=lambda x: (x[0][0], x[0][1] if np.isfinite(x[0][1]) else np.inf, x[0][2])):
        out_row = {
            "ID": sample_id,
            "DF": df_val,
            "Type": typ,
            "n_wells": int(len(rows)),
        }
        for key in [
            "PAbs_Blue", "PAbs_Green", "PAbs_Red",
            "MeanBG_Blue", "MeanBG_Green", "MeanBG_Red",
            "MeanW_Blue", "MeanW_Green", "MeanW_Red",
            "L", "a", "b", "DeltaL", "Deltaa", "Deltab", "DeltaE_ab", "DeltaE_ab_chroma",
            "UsedFraction", "BrightExcludedFraction", "HighlightIndex",
        ]:
            vals = np.asarray([float(r.get(key, np.nan)) for r in rows if np.isfinite(r.get(key, np.nan))], dtype=np.float64)
            out_row[f"{key}_median"] = float(np.median(vals)) if vals.size else np.nan
            out_row[f"{key}_sd"] = float(np.std(vals, ddof=1)) if vals.size > 1 else np.nan

        for ch in ["PAbs_Blue", "PAbs_Green", "PAbs_Red", "L", "a", "b", "DeltaL", "Deltaa", "Deltab", "DeltaE", "DeltaE_chroma"]:
            fit = fit_pref.get((ch, sample_id, df_val))
            if fit is None:
                out_row[f"Estimate_{ch}"] = np.nan
                out_row[f"EstimateSD_{ch}"] = np.nan
                out_row[f"EstimateSource_{ch}"] = ""
            else:
                out_row[f"Estimate_{ch}"] = float(fit.get("C0", np.nan)) if np.isfinite(fit.get("C0", np.nan)) else np.nan
                out_row[f"EstimateSD_{ch}"] = float(fit.get("C0_sd", np.nan)) if np.isfinite(fit.get("C0_sd", np.nan)) else np.nan
                out_row[f"EstimateSource_{ch}"] = str(fit.get("FitType", ""))

        for i_ref, ref in enumerate(expected_refs, start=1):
            if not _reference_matches_fit_row(ref, {"ID": sample_id}):
                continue
            ref_label = _format_reference_label(ref, i_ref)
            ref_safe = _expected_ref_key_label(ref_label, i_ref)
            ref_val = float(ref.get("value", np.nan)) if np.isfinite(ref.get("value", np.nan)) else np.nan
            ref_sd = _num_or_nan(ref.get("sd", np.nan))
            out_row[f"reference_label_{ref_safe}"] = ref_label
            out_row[f"reference_value_{ref_safe}"] = ref_val
            out_row[f"reference_sd_{ref_safe}"] = ref_sd
            for ch in ["PAbs_Blue", "PAbs_Green", "PAbs_Red"]:
                est = out_row.get(f"Estimate_{ch}", np.nan)
                est_sd = out_row.get(f"EstimateSD_{ch}", np.nan)
                delta = (est - ref_val) if np.isfinite(est) and np.isfinite(ref_val) else np.nan
                z = (delta / ref_sd) if np.isfinite(delta) and np.isfinite(ref_sd) and ref_sd > 0 else np.nan
                zc = (delta / np.sqrt(ref_sd * ref_sd + est_sd * est_sd)) if np.isfinite(delta) and np.isfinite(ref_sd) and ref_sd > 0 and np.isfinite(est_sd) and est_sd >= 0 else np.nan
                out_row[f"delta_{ch}_{ref_safe}"] = delta
                out_row[f"z_{ch}_{ref_safe}"] = z
        out.append(out_row)
    return out


def _append_bg_channel_model_to_workbook(path, rows, expected_refs=None):
    if not rows:
        return
    from openpyxl import load_workbook
    wb = load_workbook(path)
    if '17_BG_CHANNEL_MODEL' in wb.sheetnames:
        del wb['17_BG_CHANNEL_MODEL']
    ws = wb.create_sheet('17_BG_CHANNEL_MODEL')
    lines = [
        'Diagnostic BG + channel model inputs.',
        'This sheet does not modify official fitting or reported results.',
        'It exports per-group predictors to build external correction models across plates.',
        'Recommended primary predictors: Estimate_PAbs_Blue, Estimate_PAbs_Red, MeanBG_Blue_median, MeanBG_Red_median.',
        '',
    ]
    for i, text in enumerate(lines, start=1):
        ws.cell(i, 1, text)
    headers = list(rows[0].keys())
    _write_table(ws, len(lines) + 2, headers, rows)
    _autosize_worksheet(ws)
    wb.save(path)

def _fit_from_stored_channel_row(row):
    if not isinstance(row, dict):
        return None
    m = row.get("m", None)
    q = row.get("q", None)
    if m is None or q is None:
        return None
    try:
        m = float(m)
        q = float(q)
    except Exception:
        return None
    return {
        "n_points": int(row.get("n_points", 0) or 0),
        "m": m,
        "q": q,
        "R2": float(row["R2"]) if row.get("R2", None) is not None else np.nan,
        "RMSE": float(row["RMSE"]) if row.get("RMSE", None) is not None else np.nan,
        "cov_mq": np.full((2, 2), np.nan, dtype=np.float64),
        "x": np.asarray([], dtype=np.float64),
        "y": np.asarray([], dtype=np.float64),
        "w": None,
        "yhat": np.asarray([], dtype=np.float64),
    }




def _build_empty_well_payload(empty_well_rows):
    out = {}
    for label, base in [("Red", "Signal_Red"), ("Green", "Signal_Green"), ("Blue", "Signal_Blue")]:
        vals = np.asarray([float(r.get(base, np.nan)) for r in empty_well_rows if np.isfinite(r.get(base, np.nan))], dtype=np.float64)
        if vals.size == 0:
            out[label] = {"n": 0, "mean": np.nan, "median": np.nan, "sd": np.nan}
        else:
            med, rob_sd, n, mad = _median_and_robust_sd(vals)
            out[label] = {
                "n": int(vals.size),
                "mean": float(np.mean(vals)),
                "median": med,
                "sd": float(np.std(vals, ddof=1)) if vals.size > 1 else np.nan,
                "robust_sd": rob_sd,
                "mad": mad,
            }
    return out


def _compare_empty_well_payloads(current_payload, stored_calibration_bundle):
    stored_payload = {}
    if isinstance(stored_calibration_bundle, dict):
        stored_payload = stored_calibration_bundle.get("empty_well_payload", {}) or {}

    rows = []
    for label in ["Red", "Green", "Blue"]:
        cur = current_payload.get(label, {}) if isinstance(current_payload, dict) else {}
        sto = stored_payload.get(label, {}) if isinstance(stored_payload, dict) else {}
        cur_med = float(cur.get("median", np.nan)) if np.isfinite(cur.get("median", np.nan)) else np.nan
        sto_med = float(sto.get("median", np.nan)) if np.isfinite(sto.get("median", np.nan)) else np.nan
        cur_mean = float(cur.get("mean", np.nan)) if np.isfinite(cur.get("mean", np.nan)) else np.nan
        sto_mean = float(sto.get("mean", np.nan)) if np.isfinite(sto.get("mean", np.nan)) else np.nan

        delta_median = cur_med - sto_med if np.isfinite(cur_med) and np.isfinite(sto_med) else np.nan
        delta_mean = cur_mean - sto_mean if np.isfinite(cur_mean) and np.isfinite(sto_mean) else np.nan
        ratio_median = (cur_med / sto_med) if np.isfinite(cur_med) and np.isfinite(sto_med) and abs(sto_med) > 1e-15 else np.nan
        ratio_mean = (cur_mean / sto_mean) if np.isfinite(cur_mean) and np.isfinite(sto_mean) and abs(sto_mean) > 1e-15 else np.nan

        rows.append({
            "Channel": label,
            "Current_n": int(cur.get("n", 0) or 0),
            "Stored_n": int(sto.get("n", 0) or 0),
            "Current_mean": cur_mean,
            "Stored_mean": sto_mean,
            "Delta_mean": delta_mean,
            "Ratio_mean": ratio_mean,
            "Current_median": cur_med,
            "Stored_median": sto_med,
            "Delta_median": delta_median,
            "Ratio_median": ratio_median,
            "Current_sd": float(cur.get("sd", np.nan)) if np.isfinite(cur.get("sd", np.nan)) else np.nan,
            "Stored_sd": float(sto.get("sd", np.nan)) if np.isfinite(sto.get("sd", np.nan)) else np.nan,
            "Current_robust_sd": float(cur.get("robust_sd", np.nan)) if np.isfinite(cur.get("robust_sd", np.nan)) else np.nan,
            "Stored_robust_sd": float(sto.get("robust_sd", np.nan)) if np.isfinite(sto.get("robust_sd", np.nan)) else np.nan,
        })
    return rows


def _write_stored_empty_comparison_csv(path, rows):
    headers = [
        "Channel", "Current_n", "Stored_n",
        "Current_mean", "Stored_mean", "Delta_mean", "Ratio_mean",
        "Current_median", "Stored_median", "Delta_median", "Ratio_median",
        "Current_sd", "Stored_sd", "Current_robust_sd", "Stored_robust_sd",
    ]
    with open(path, "w", newline="", encoding="utf-8") as f:
        wr = csv.writer(f, delimiter=';')
        wr.writerow(headers)
        for row in rows:
            wr.writerow([row.get(h, "") for h in headers])


def _append_stored_empty_comparison_to_workbook(path, comparison_rows, stored_calibration_bundle):
    from openpyxl import load_workbook

    wb = load_workbook(path)
    if "09_STORED_CAL_EMPTY" in wb.sheetnames:
        del wb["09_STORED_CAL_EMPTY"]
    ws = wb.create_sheet("09_STORED_CAL_EMPTY")

    stored_name = ""
    if isinstance(stored_calibration_bundle, dict):
        stored_name = str(stored_calibration_bundle.get("image_basename", ""))

    intro = [
        "Stored-calibration empty-well comparison (diagnostic only; no automatic rescaling applied).",
        f"Stored calibration source: {stored_name}",
        "",
    ]
    for i, line in enumerate(intro, start=1):
        ws.cell(i, 1, line)

    headers = [
        "Channel", "Current_n", "Stored_n",
        "Current_mean", "Stored_mean", "Delta_mean", "Ratio_mean",
        "Current_median", "Stored_median", "Delta_median", "Ratio_median",
        "Current_sd", "Stored_sd", "Current_robust_sd", "Stored_robust_sd",
    ]
    _write_table(ws, 5, headers, comparison_rows or [])
    _autosize_worksheet(ws)
    wb.save(path)





def _fit_affine_transfer_from_paired_empty_wells(current_empty_rows, stored_calibration_bundle):
    """
    Diagnostic-only inter-plate transfer model from paired empty wells.

    Preferred mode:
        fit current ~= a * stored + b
    using empty wells matched by Well name across the stored calibration plate and
    the current plate.

    Fallback mode:
        if paired wells are insufficient, use summary statistics with a fixed to 1.0
        and b estimated from the median shift.
    """
    stored_rows = []
    if isinstance(stored_calibration_bundle, dict):
        stored_rows = stored_calibration_bundle.get("empty_well_rows", []) or []

    current_map = {str(r.get("Well", "")).strip(): dict(r) for r in current_empty_rows if str(r.get("Well", "")).strip()}
    stored_map = {str(r.get("Well", "")).strip(): dict(r) for r in stored_rows if str(r.get("Well", "")).strip()}

    rows = []
    for label, base in [("Red", "Signal_Red"), ("Green", "Signal_Green"), ("Blue", "Signal_Blue")]:
        paired_x = []
        paired_y = []
        paired_wells = []
        for well in sorted(set(current_map.keys()) & set(stored_map.keys())):
            x = float(stored_map[well].get(base, np.nan))
            y = float(current_map[well].get(base, np.nan))
            if np.isfinite(x) and np.isfinite(y):
                paired_x.append(x)
                paired_y.append(y)
                paired_wells.append(well)

        mode = "paired_wells" if len(paired_x) >= 4 else "summary_fallback"
        a = np.nan
        b = np.nan
        r2 = np.nan
        n_pairs = int(len(paired_x))
        well_pairs = ", ".join(paired_wells)

        if len(paired_x) >= 4:
            fit = _fit_line_with_covariance(np.asarray(paired_x, dtype=np.float64), np.asarray(paired_y, dtype=np.float64), w=None)
            if fit is not None and np.isfinite(fit.get("m", np.nan)) and np.isfinite(fit.get("q", np.nan)):
                a = float(fit["m"])
                b = float(fit["q"])
                r2 = float(fit["R2"]) if np.isfinite(fit.get("R2", np.nan)) else np.nan

        if not np.isfinite(a) or not np.isfinite(b):
            current_payload = _build_empty_well_payload(current_empty_rows)
            stored_payload = {}
            if isinstance(stored_calibration_bundle, dict):
                stored_payload = stored_calibration_bundle.get("empty_well_payload", {}) or {}
            cur = current_payload.get(label, {}) if isinstance(current_payload, dict) else {}
            sto = stored_payload.get(label, {}) if isinstance(stored_payload, dict) else {}
            cur_med = float(cur.get("median", np.nan)) if np.isfinite(cur.get("median", np.nan)) else np.nan
            sto_med = float(sto.get("median", np.nan)) if np.isfinite(sto.get("median", np.nan)) else np.nan
            a = 1.0 if (np.isfinite(cur_med) and np.isfinite(sto_med)) else np.nan
            b = (cur_med - sto_med) if np.isfinite(cur_med) and np.isfinite(sto_med) else np.nan

        rows.append({
            "Channel": label,
            "Model": "current ~= a*stored + b",
            "Mode": mode,
            "a": a,
            "b": b,
            "R2": r2,
            "n_pairs": n_pairs,
            "Well_pairs": well_pairs,
            "Suggested_inverse_transform": f"A_corr = (A - ({b:.6g})) / ({a:.6g})" if np.isfinite(a) and np.isfinite(b) and abs(a) > 1e-15 else "",
        })
    return rows


def _write_empty_transfer_model_csv(path, rows):
    headers = [
        "Channel", "Model", "Mode", "a", "b", "R2", "n_pairs",
        "Well_pairs", "Suggested_inverse_transform",
    ]
    with open(path, "w", newline="", encoding="utf-8") as f:
        wr = csv.writer(f, delimiter=';')
        wr.writerow(headers)
        for row in rows or []:
            wr.writerow([row.get(h, "") for h in headers])


def _append_empty_transfer_model_to_workbook(path, transfer_rows, stored_calibration_bundle):
    from openpyxl import load_workbook

    wb = load_workbook(path)
    if "10_EMPTY_TRANSFER_MODEL" in wb.sheetnames:
        del wb["10_EMPTY_TRANSFER_MODEL"]
    ws = wb.create_sheet("10_EMPTY_TRANSFER_MODEL")

    stored_name = ""
    if isinstance(stored_calibration_bundle, dict):
        stored_name = str(stored_calibration_bundle.get("image_basename", ""))

    intro = [
        "Empty-based inter-plate transfer model (diagnostic only; no automatic correction applied).",
        "Model convention: current ~= a * stored + b",
        "Preferred mode: paired empty wells matched by well name; fallback: median-shift summary model.",
        f"Stored calibration source: {stored_name}",
        "",
    ]
    for i, line in enumerate(intro, start=1):
        ws.cell(i, 1, line)

    headers = [
        "Channel", "Model", "Mode", "a", "b", "R2", "n_pairs",
        "Well_pairs", "Suggested_inverse_transform",
    ]
    _write_table(ws, 7, headers, transfer_rows or [])
    _autosize_worksheet(ws)
    wb.save(path)




def _fit_bg_transfer_model(current_well_bg_rows, stored_calibration_bundle):
    """
    Diagnostic-only inter-plate BG-vs-BG comparison using per-well predicted background.
    Model convention: current_bg ~= a * stored_bg + b
    """
    stored_rows = []
    if isinstance(stored_calibration_bundle, dict):
        stored_rows = stored_calibration_bundle.get("well_bg_rows", []) or []

    current_map = {
        (int(r.get("well_r", -1)), int(r.get("well_c", -1))): dict(r)
        for r in current_well_bg_rows
        if np.isfinite(r.get("well_r", np.nan)) and np.isfinite(r.get("well_c", np.nan))
    }
    stored_map = {
        (int(r.get("well_r", -1)), int(r.get("well_c", -1))): dict(r)
        for r in stored_rows
        if np.isfinite(r.get("well_r", np.nan)) and np.isfinite(r.get("well_c", np.nan))
    }

    paired_rows = []
    summary_rows = []
    channel_map = [("Blue", "B_bg"), ("Green", "G_bg"), ("Red", "R_bg")]

    for label, field in channel_map:
        xs = []
        ys = []
        for key in sorted(set(current_map.keys()) & set(stored_map.keys())):
            cur = float(current_map[key].get(field, np.nan))
            sto = float(stored_map[key].get(field, np.nan))
            if np.isfinite(cur) and np.isfinite(sto):
                rr, cc = key
                well = _well_name_from_indices(rr, cc)
                xs.append(sto)
                ys.append(cur)
                paired_rows.append({
                    "Channel": label,
                    "Well": well,
                    "stored_bg": sto,
                    "current_bg": cur,
                    "delta_bg": float(cur - sto),
                    "ratio_bg": float(cur / sto) if abs(sto) > 1e-15 else np.nan,
                })

        fit = None
        if len(xs) >= 4:
            fit = _fit_line_with_covariance(np.asarray(xs, dtype=np.float64), np.asarray(ys, dtype=np.float64), w=None)

        a = float(fit["m"]) if fit is not None and np.isfinite(fit.get("m", np.nan)) else np.nan
        b = float(fit["q"]) if fit is not None and np.isfinite(fit.get("q", np.nan)) else np.nan
        r2 = float(fit["R2"]) if fit is not None and np.isfinite(fit.get("R2", np.nan)) else np.nan
        rmse = float(fit["RMSE"]) if fit is not None and np.isfinite(fit.get("RMSE", np.nan)) else np.nan

        dx = np.asarray(xs, dtype=np.float64)
        dy = np.asarray(ys, dtype=np.float64)
        delta_med = float(np.nanmedian(dy - dx)) if dx.size > 0 else np.nan
        ratio_med = float(np.nanmedian(dy / dx)) if dx.size > 0 and np.all(np.abs(dx[np.isfinite(dx)]) > 1e-15) else np.nan

        summary_rows.append({
            "Channel": label,
            "Model": "current_bg ~= a*stored_bg + b",
            "a": a,
            "b": b,
            "R2": r2,
            "RMSE": rmse,
            "n_pairs": int(len(xs)),
            "delta_bg_median": delta_med,
            "ratio_bg_median": ratio_med,
            "Suggested_inverse_transform": f"BG_corr = (BG - ({b:.6g})) / ({a:.6g})" if np.isfinite(a) and np.isfinite(b) and abs(a) > 1e-15 else "",
        })

    return {"summary_rows": summary_rows, "paired_rows": paired_rows}


def _write_bg_transfer_summary_csv(path, rows):
    headers = [
        "Channel", "Model", "a", "b", "R2", "RMSE", "n_pairs",
        "delta_bg_median", "ratio_bg_median", "Suggested_inverse_transform",
    ]
    with open(path, "w", newline="", encoding="utf-8") as f:
        wr = csv.writer(f, delimiter=';')
        wr.writerow(headers)
        for row in rows or []:
            wr.writerow([row.get(h, "") for h in headers])


def _write_bg_transfer_pairs_csv(path, rows):
    headers = ["Channel", "Well", "stored_bg", "current_bg", "delta_bg", "ratio_bg"]
    with open(path, "w", newline="", encoding="utf-8") as f:
        wr = csv.writer(f, delimiter=';')
        wr.writerow(headers)
        for row in rows or []:
            wr.writerow([row.get(h, "") for h in headers])


def _append_bg_transfer_to_workbook(path, bg_transfer_payload, stored_calibration_bundle):
    from openpyxl import load_workbook

    wb = load_workbook(path)
    for name in ["11_BG_TRANSFER_MODEL", "12_BG_TRANSFER_PAIRS"]:
        if name in wb.sheetnames:
            del wb[name]

    ws = wb.create_sheet("11_BG_TRANSFER_MODEL")
    ws2 = wb.create_sheet("12_BG_TRANSFER_PAIRS")

    stored_name = ""
    if isinstance(stored_calibration_bundle, dict):
        stored_name = str(stored_calibration_bundle.get("image_basename", ""))

    intro = [
        "Inter-plate BG-vs-BG transfer model (diagnostic only; no automatic correction applied).",
        "Model convention: current_bg ~= a * stored_bg + b",
        f"Stored calibration source: {stored_name}",
        "",
    ]
    for i, line in enumerate(intro, start=1):
        ws.cell(i, 1, line)

    headers1 = [
        "Channel", "Model", "a", "b", "R2", "RMSE", "n_pairs",
        "delta_bg_median", "ratio_bg_median", "Suggested_inverse_transform",
    ]
    _write_table(ws, 6, headers1, bg_transfer_payload.get("summary_rows", []) or [])

    intro2 = [
        "Paired predicted background values by well (diagnostic only).",
        "",
    ]
    for i, line in enumerate(intro2, start=1):
        ws2.cell(i, 1, line)

    headers2 = ["Channel", "Well", "stored_bg", "current_bg", "delta_bg", "ratio_bg"]
    _write_table(ws2, 4, headers2, bg_transfer_payload.get("paired_rows", []) or [])

    _autosize_worksheet(ws)
    _autosize_worksheet(ws2)
    wb.save(path)


def _prompt_fitting_options(summary_rows, cfg=None):
    cfg = cfg or {}
    print("\nFITTING OPTIONS")
    use_stored_calibration = bool(cfg.get("use_stored_calibration", False))
    use_background_subtraction = True
    show_background_mode_in_report = bool(cfg.get("show_background_mode_in_report", True))
    blank_mode = "both"
    print(f"Automatic mode: use_stored_calibration={use_stored_calibration}, use_background_subtraction={use_background_subtraction}; all points used; robust fit by IRLS, no SD weighting; RGB signal=full_bg.")
    return {
        "use_stored_calibration": use_stored_calibration,
        "use_background_subtraction": use_background_subtraction,
        "show_background_mode_in_report": show_background_mode_in_report,
        "blank_mode": blank_mode,
        "bg_correction_alpha": np.nan,
    }

def _row_weight(row, base):
    return 1.0


def _build_fitting_rows(summary_rows, selected_channel, fit_options=None, stored_calibration_bundle=None):
    if fit_options is None:
        fit_options = {}

    channels = [
        # --- RGB pseudo-absorbance fixed to full_bg ---
        ("Signal_Red", "Signal_Red"),
        ("Signal_Green", "Signal_Green"),
        ("Signal_Blue", "Signal_Blue"),

        # --- CIELAB ABSOLUTE ---
        ("L", "L"),
        ("a", "a"),
        ("b", "b"),

        # --- CIELAB DIFFERENCES ---
        ("DeltaL", "DeltaL"),
        ("Deltaa", "Deltaa"),
        ("Deltab", "Deltab"),

        # --- COLOR DISTANCE ---
        ("DeltaE", "DeltaE_ab"),
        ("DeltaE_chroma", "DeltaE_ab_chroma"),
    ]

    use_stored_calibration = bool(fit_options.get("use_stored_calibration", False))
    stored_channels = {}
    if isinstance(stored_calibration_bundle, dict):
        stored_channels = stored_calibration_bundle.get("channels", {}) or {}
    fit_rows = []
    plot_payload = {}

    type_cal = {"CAL", "STD", "STANDARD", "CALIBRATION", "C"}
    type_stdadd = {"A", "SA", "STDADD", "STANDARD_ADDITION", "ADDITION"}
    type_unk = {"UNK", "UNKNOWN", "U"}

    for label, base in channels:
        centered_summary_rows, cal_ref_value, cal_ref_source = _center_summary_rows_for_calibration(summary_rows, base)
        plot_payload[label] = {
            "calibration_points": [],
            "stdadd_groups": [],
            "unknown_points": [],
            "excluded_points": [],
            "calibration_ref": cal_ref_value,
            "calibration_ref_source": cal_ref_source,
        }

        cal_all = [
            r for r in centered_summary_rows
            if str(r.get("Type", "")).upper() in type_cal and np.isfinite(r.get(f"{base}_median", np.nan)) and np.isfinite(r.get("Conc", np.nan))
        ]
        cal_used = list(cal_all)
        cal_exc = []
        plot_payload[label]["excluded_points"].extend([
            {"x": float(r["Conc"]), "y": float(r[f"{base}_median"])} for r in cal_exc
        ])

        cal_fit = None
        rgb_low_signal_mapper = None
        sigma_cal, sigma_source = _estimate_sigma_for_lod(cal_used, base)
        if use_stored_calibration and label in stored_channels:
            cal_fit = _fit_from_stored_channel_row(stored_channels.get(label))
            stored_row = stored_channels.get(label, {})
            sigma_cal = float(stored_row["sigma_cal"]) if stored_row.get("sigma_cal", None) is not None else np.nan
            sigma_source = stored_row.get("sigma_source", "stored_calibration")
            if base.startswith("Signal_"):
                rgb_low_signal_mapper = _rgb_low_signal_mapper_from_payload(stored_row.get("rgb_low_signal_correction", {}) or {})
                if cal_fit is not None and hasattr(rgb_low_signal_mapper, "correction_payload"):
                    cal_fit["rgb_low_signal_correction"] = rgb_low_signal_mapper.correction_payload
                    cal_fit["S0_calibration"] = rgb_low_signal_mapper.correction_payload.get("S0_calibration", np.nan)
                    cal_fit["ClipX"] = ",".join(str(v) for v in rgb_low_signal_mapper.correction_payload.get("clip_x", []))
                    cal_fit["ClipDelta"] = ",".join(str(v) for v in rgb_low_signal_mapper.correction_payload.get("clip_delta", []))
                    cal_fit["NClipPoints"] = rgb_low_signal_mapper.correction_payload.get("n_clip_points", 0)
        elif len(cal_used) >= 3:
            if base.startswith("Signal_"):
                cal_fit, rgb_low_signal_mapper = _build_rgb_low_signal_correction(cal_used, base, fit_options)
            if cal_fit is None:
                x = np.asarray([float(r["Conc"]) for r in cal_used], dtype=np.float64)
                y = np.asarray([float(r[f"{base}_median"]) for r in cal_used], dtype=np.float64)
                w = np.asarray([_row_weight(r, base) for r in cal_used], dtype=np.float64)
                cal_fit = _fit_line_with_covariance(x, y, w=w, force_zero=False)

        fit_rows.append({
            "Channel": label, "FitType": "Calibration", "ID": None, "DF": None,
            "n_points": cal_fit["n_points"] if cal_fit else len(cal_used),
            "m": cal_fit["m"] if cal_fit else np.nan,
            "q": cal_fit["q"] if cal_fit else np.nan,
            "R2": cal_fit["R2"] if cal_fit else np.nan,
            "RMSE": cal_fit["RMSE"] if cal_fit else np.nan,
            "sigma_cal": sigma_cal,
            "sigma_source": sigma_source,
            "SNR": float(abs(cal_fit["m"]) / sigma_cal) if cal_fit and np.isfinite(sigma_cal) and sigma_cal > 0 else np.nan,
            "LOD": float(3.0 * sigma_cal / abs(cal_fit["m"])) if cal_fit and np.isfinite(sigma_cal) and sigma_cal > 0 and abs(cal_fit["m"]) > 1e-15 else np.nan,
            "LOQ": float(10.0 * sigma_cal / abs(cal_fit["m"])) if cal_fit and np.isfinite(sigma_cal) and sigma_cal > 0 and abs(cal_fit["m"]) > 1e-15 else np.nan,
            "C0": np.nan,
            "C0_sd": np.nan,
            "C0": np.nan,
            "C0_sd": np.nan,
            "beta_k": np.nan,
            "bias_index_k": np.nan,
            "S0_calibration": cal_fit.get("S0_calibration", np.nan) if cal_fit else np.nan,
            "S0_applied": cal_fit.get("S0_calibration", np.nan) if cal_fit else np.nan,
            "NClipPoints": cal_fit.get("NClipPoints", np.nan) if cal_fit else np.nan,
            "ClipX": cal_fit.get("ClipX", "") if cal_fit else "",
            "ClipDelta": cal_fit.get("ClipDelta", "") if cal_fit else "",
            "rgb_low_signal_correction": cal_fit.get("rgb_low_signal_correction", {}) if cal_fit else {},
            "__cov_mq": cal_fit["cov_mq"] if cal_fit else None,
        })

        if use_stored_calibration and label in stored_channels:
            stored_pts = stored_channels.get(label, {}).get("calibration_points", []) or []
            plot_payload[label]["calibration_points"] = [{
                "x": float(p.get("x", np.nan)),
                "y": float(p.get("y", np.nan)),
                "yerr": float(p.get("yerr", np.nan)) if p.get("yerr", None) is not None else np.nan,
                "excluded": bool(p.get("excluded", False)),
            } for p in stored_pts if np.isfinite(p.get("x", np.nan)) and np.isfinite(p.get("y", np.nan))]
        else:
            _cal_plot_points = []
            for r in cal_used:
                _yv = float(r[f"{base}_median"])
                if base.startswith("Signal_") and rgb_low_signal_mapper is not None:
                    _xv_tmp = float(r["Conc"])
                    if abs(_xv_tmp) <= 1e-12:
                        _yv = 0.0
                    else:
                        _mapped = rgb_low_signal_mapper(np.asarray([_yv], dtype=np.float64), np.asarray([_xv_tmp], dtype=np.float64))[0]
                        if np.isfinite(_mapped):
                            _yv = float(_mapped)
                _cal_plot_points.append({
                    "x": float(r["Conc"]),
                    "y": _yv,
                    "yerr": float(r.get(f"{base}_sd", np.nan)),
                    "excluded": False,
                })
            plot_payload[label]["calibration_points"] = _cal_plot_points
        plot_payload[label]["calibration_fit"] = cal_fit

        stdadd_all = [
            r for r in summary_rows
            if str(r.get("Type", "")).upper() in type_stdadd and np.isfinite(r.get(f"{base}_median", np.nan)) and np.isfinite(r.get("Conc", np.nan))
        ]
        groups = {}
        for row in stdadd_all:
            groups.setdefault((str(row["ID"]), float(row["DF"])), []).append(row)

        for (sample_id, dilution_factor), rows in sorted(groups.items(), key=lambda kv: (kv[0][0], kv[0][1])):
            rows_sorted = sorted(rows, key=lambda r: float(r["Conc"]))
            used_rows = list(rows_sorted)
            exc_rows = []
            plot_payload[label]["excluded_points"].extend([
                {"x": float(r["Conc"]), "y": float(r[f"{base}_median"])} for r in exc_rows
            ])

            fit = None
            c0_orig = np.nan
            c0_orig_sd = np.nan
            if len(used_rows) >= 2:
                x = np.asarray([float(r["Conc"]) for r in used_rows], dtype=np.float64)
                y = np.asarray([float(r[f"{base}_median"]) for r in used_rows], dtype=np.float64)
                if base.startswith("Signal_") and rgb_low_signal_mapper is not None:
                    y_mapped = rgb_low_signal_mapper(y, x)
                    finite_map = np.isfinite(y_mapped)
                    if np.any(finite_map):
                        y = np.where(finite_map, y_mapped, y)
                w = np.asarray([_row_weight(r, base) for r in used_rows], dtype=np.float64)
                fit = _fit_line_with_covariance(x, y, w=w)
                c0_dil, c0_dil_sd = _stdadd_c0_sd_from_fit(fit)
                c0_orig_stdadd = float(dilution_factor * c0_dil) if np.isfinite(c0_dil) else np.nan
                c0_orig_stdadd_sd = float(dilution_factor * c0_dil_sd) if np.isfinite(c0_dil_sd) else np.nan

                # Standard addition is quantified from its own fitted line,
                # consistently with the plotted intercept: C0 = DF * q_std / m_std.
                c0_orig = c0_orig_stdadd
                c0_orig_sd = c0_orig_stdadd_sd

            fit_rows.append({
                "Channel": label, "FitType": "StdAdd", "ID": sample_id, "DF": dilution_factor,
                "n_points": fit["n_points"] if fit else len(used_rows),
                "m": fit["m"] if fit else np.nan,
                "q": fit["q"] if fit else np.nan,
                "R2": fit["R2"] if fit else np.nan,
                "RMSE": fit["RMSE"] if fit else np.nan,
                "sigma_cal": np.nan,
                "SNR": np.nan,
                "LOD": np.nan,
                "LOQ": np.nan,
                "C0": c0_orig,
                "C0_sd": c0_orig_sd,
                "C0_stdadd_only": c0_orig_stdadd if 'c0_orig_stdadd' in locals() else np.nan,
                "C0_stdadd_only_sd": c0_orig_stdadd_sd if 'c0_orig_stdadd_sd' in locals() else np.nan,
                "C0_source": "stdadd_intercept",
                "beta_k": np.nan,
                "bias_index_k": np.nan,
                "S0_calibration": cal_fit.get("S0_calibration", np.nan) if cal_fit else np.nan,
                "S0_applied": float(cal_fit.get("S0_calibration", np.nan)) if cal_fit and np.isfinite(cal_fit.get("S0_calibration", np.nan)) else np.nan,
                "NClipPoints": cal_fit.get("NClipPoints", np.nan) if cal_fit else np.nan,
                "ClipX": cal_fit.get("ClipX", "") if cal_fit else "",
                "ClipDelta": cal_fit.get("ClipDelta", "") if cal_fit else "",
                "__cov_mq": fit["cov_mq"] if fit else None,
            })
            plot_payload[label]["stdadd_groups"].append({
                "ID": sample_id,
                "DF": dilution_factor,
                "x": np.asarray([float(r["Conc"]) for r in used_rows], dtype=np.float64),
                "y": y if 'y' in locals() and y is not None else np.asarray([float(r[f"{base}_median"]) for r in used_rows], dtype=np.float64),
                "yerr": np.asarray([float(r.get(f"{base}_sd", np.nan)) for r in used_rows], dtype=np.float64),
                "fit": fit,
                "c0_orig": c0_orig,
                "c0_orig_sd": c0_orig_sd,
                "c0_orig_stdadd_only": c0_orig_stdadd if 'c0_orig_stdadd' in locals() else np.nan,
                "c0_orig_stdadd_only_sd": c0_orig_stdadd_sd if 'c0_orig_stdadd_sd' in locals() else np.nan,
                "c0_source": "stdadd_intercept",
            })

        unk_points = [
            r for r in summary_rows
            if str(r.get("Type", "")).upper() in type_unk and np.isfinite(r.get(f"{base}_median", np.nan))
        ]
        epsilon = _num_or_nan(fit_options.get("epsilon", np.nan))
        path_length_cfg = _num_or_nan(fit_options.get("path_length", np.nan))
        path_length_eff = _effective_path_length(path_length_cfg)
        epsilon_mode_ok = bool((label.startswith("PAbs_") or label.startswith("Signal_")) and np.isfinite(epsilon) and epsilon > 0)
        if cal_fit is not None and len(unk_points) > 0 and abs(cal_fit["m"]) > 1e-15:
            for row in unk_points:
                y_obs = float(row[f"{base}_median"])
                if base.startswith("Signal_") and rgb_low_signal_mapper is not None:
                    _mapped = rgb_low_signal_mapper(np.asarray([y_obs], dtype=np.float64))[0]
                    if np.isfinite(_mapped):
                        y_obs = float(_mapped)
                y_obs_cal = float(y_obs - cal_ref_value) if np.isfinite(cal_ref_value) else y_obs
                c_dil = (y_obs_cal - cal_fit["q"]) / cal_fit["m"]
                df_val = float(row["DF"]) if np.isfinite(row.get("DF", np.nan)) else np.nan
                c_orig = df_val * c_dil if np.isfinite(df_val) else c_dil
                sigma_y = float(row.get(f"{base}_sd", np.nan))
                c_dil_sd = float(abs(sigma_y / cal_fit["m"])) if np.isfinite(sigma_y) else np.nan
                c_sd = (df_val * c_dil_sd) if np.isfinite(df_val) and np.isfinite(c_dil_sd) else c_dil_sd
                n_rep = int(row.get("NReplicates", 1))
                fit_rows.append({
                    "Channel": label, "FitType": "UnknownFromCal", "ID": str(row["ID"]), "DF": df_val,
                    "n_points": n_rep, "m": cal_fit["m"], "q": cal_fit["q"], "R2": cal_fit["R2"], "RMSE": cal_fit["RMSE"],
                    "sigma_cal": np.nan, "sigma_source": "unavailable", "SNR": np.nan, "LOD": np.nan, "LOQ": np.nan,
                    "C0": c_orig, "C0_sd": c_sd,
                    "beta_k": np.nan,
                    "bias_index_k": np.nan,
                    "UsedFraction": float(row.get("UsedFraction_median", np.nan)),
                    "__cov_mq": None,
                })
                plot_payload[label]["unknown_points"].append({
                    "ID": str(row["ID"]),
                    "DF": df_val,
                    "n_points": n_rep,
                    "x": float(c_dil) if np.isfinite(c_dil) else np.nan,
                    "xerr": float(c_dil_sd) if np.isfinite(c_dil_sd) else np.nan,
                    "y": y_obs_cal,
                    "yerr": sigma_y if np.isfinite(sigma_y) else np.nan,
                    "y_raw": y_obs,
                    "c_orig": float(c_orig) if np.isfinite(c_orig) else np.nan,
                    "c_orig_sd": float(c_sd) if np.isfinite(c_sd) else np.nan,
                    "source": "calibration",
                })
        elif len(unk_points) > 0 and epsilon_mode_ok:
            for row in unk_points:
                y_obs = float(row[f"{base}_median"])
                sigma_y = float(row.get(f"{base}_sd", np.nan))
                df_val = float(row.get("DF", np.nan)) if np.isfinite(row.get("DF", np.nan)) else np.nan
                c_orig, c_sd, path_length_eff, _ = _epsilon_concentration_from_pabs(
                    pabs_value=y_obs,
                    pabs_sd=sigma_y,
                    dilution_factor=df_val,
                    epsilon=epsilon,
                    path_length=path_length_cfg,
                    unit_label=str(fit_options.get("unit_label", "M")),
                )
                n_rep = int(row.get("NReplicates", 1))
                fit_rows.append({
                    "Channel": label, "FitType": "UnknownFromEpsilon", "ID": str(row.get("ID", "")), "DF": df_val,
                    "n_points": n_rep, "m": np.nan, "q": np.nan, "R2": np.nan, "RMSE": np.nan,
                    "sigma_cal": np.nan, "sigma_source": "epsilon", "SNR": np.nan, "LOD": np.nan, "LOQ": np.nan,
                    "C0": c_orig, "C0_sd": c_sd, "beta_k": np.nan, "bias_index_k": np.nan,
                    "UsedFraction": float(row.get("UsedFraction_median", np.nan)),
                    "epsilon": float(epsilon),
                    "path_length": float(path_length_eff),
                    "path_length_mm": float(fit_options.get("path_length_mm", np.nan)) if np.isfinite(fit_options.get("path_length_mm", np.nan)) else np.nan,
                    "liquid_volume_ul": float(fit_options.get("liquid_volume_ul", np.nan)) if np.isfinite(fit_options.get("liquid_volume_ul", np.nan)) else np.nan,
                    "path_length_source": str(fit_options.get("path_length_source", "")),
                    "Status": "Concentration estimated from PAbs with epsilon and calculated path length",
                    "__cov_mq": None,
                })
                c_dil = (c_orig / df_val) if np.isfinite(c_orig) and np.isfinite(df_val) and abs(df_val) > 1e-15 else c_orig
                c_dil_sd = (c_sd / abs(df_val)) if np.isfinite(c_sd) and np.isfinite(df_val) and abs(df_val) > 1e-15 else c_sd
                plot_payload[label]["unknown_points"].append({
                    "ID": str(row.get("ID", "")),
                    "DF": df_val,
                    "n_points": n_rep,
                    "x": float(c_dil) if np.isfinite(c_dil) else np.nan,
                    "xerr": float(c_dil_sd) if np.isfinite(c_dil_sd) else np.nan,
                    "y": y_obs,
                    "yerr": sigma_y if np.isfinite(sigma_y) else np.nan,
                    "y_raw": y_obs,
                    "c_orig": float(c_orig) if np.isfinite(c_orig) else np.nan,
                    "c_orig_sd": float(c_sd) if np.isfinite(c_sd) else np.nan,
                    "source": "epsilon",
                })
        elif len(unk_points) > 0:
            for row in unk_points:
                fit_rows.append({
                    "Channel": label, "FitType": "UnknownOnly", "ID": str(row.get("ID", "")), "DF": float(row.get("DF", np.nan)) if np.isfinite(row.get("DF", np.nan)) else np.nan,
                    "n_points": int(row.get("NReplicates", 1)), "m": np.nan, "q": np.nan, "R2": np.nan, "RMSE": np.nan,
                    "sigma_cal": np.nan, "sigma_source": "unavailable", "SNR": np.nan, "LOD": np.nan, "LOQ": np.nan,
                    "C0": np.nan, "C0_sd": np.nan, "beta_k": np.nan, "bias_index_k": np.nan,
                    "UsedFraction": float(row.get("UsedFraction_median", np.nan)),
                    "Status": "Concentration not estimated",
                    "__cov_mq": None,
                })

    return fit_rows, plot_payload









def _rgb_low_signal_mapper_from_payload(payload):
    """Rebuild the explicit RGB low-signal correction stored with calibration.

    The correction is additive and concentration-indexed, but clipping is applied
    only to points that are below the calibration clipping threshold at the same
    concentration. Points already above threshold are left unchanged.
    """
    payload = payload or {}
    s0 = _num_or_nan(payload.get("S0_calibration", payload.get("s0_calibration", 0.0)))
    if not np.isfinite(s0):
        s0 = 0.0
    x_clip = np.asarray(payload.get("clip_x", []), dtype=np.float64)
    d_clip = np.asarray(payload.get("clip_delta", []), dtype=np.float64)
    y_threshold = np.asarray(payload.get("clip_y_expected", []), dtype=np.float64)
    if y_threshold.size != x_clip.size:
        y_threshold = np.full_like(x_clip, np.nan, dtype=np.float64)
    finite = np.isfinite(x_clip) & np.isfinite(d_clip)
    x_clip = x_clip[finite]
    d_clip = d_clip[finite]
    y_threshold = y_threshold[finite]
    if x_clip.size > 0:
        order = np.argsort(x_clip)
        x_clip = x_clip[order]
        d_clip = d_clip[order]
        y_threshold = y_threshold[order]

    def mapper(y, x=None):
        arr = np.asarray(y, dtype=np.float64)
        out = arr.copy() + float(s0)
        if x is not None and x_clip.size > 0:
            xx = np.asarray(x, dtype=np.float64)
            if xx.shape != arr.shape:
                xx = np.broadcast_to(xx, arr.shape)
            delta = np.zeros_like(arr, dtype=np.float64)
            valid_clip = np.isfinite(x_clip) & np.isfinite(d_clip) & (d_clip > 0)
            for xv, dv in zip(x_clip[valid_clip], d_clip[valid_clip]):
                hit = np.isfinite(xx) & (np.abs(xx - float(xv)) <= 1e-12)
                if np.any(hit):
                    delta = np.where(hit, float(dv), delta)
            if np.any(delta > 0):
                out = out + delta
        return out

    mapper.correction_payload = {
        "S0_calibration": float(s0),
        "clip_x": [float(v) for v in x_clip.tolist()],
        "clip_delta": [float(v) for v in d_clip.tolist()],
        "clip_y_expected": [float(v) for v in y_threshold.tolist()],
        "n_clip_points": int(np.count_nonzero(d_clip > 0)) if d_clip.size else 0,
    }
    return mapper


def _build_rgb_low_signal_correction(cal_used, base, fit_options):
    """Build an explicit channel-specific RGB correction from calibration only.

    Correct sequence:
    1) estimate S0 from raw calibration responses, before forcing the calibration
       through zero;
    2) add S0 to calibration responses;
    3) fit the shifted calibration with forced zero intercept;
    4) compute ClipDelta(x) on the shifted calibration as the positive residual
       below the forced-zero calibration line, using the calibration SD as the
       decision threshold when available;
    5) store S0 and ClipDelta(x), so the same channel/x correction can be
       applied later to standard-addition or stored-calibration data.
    """
    if not str(base).startswith("Signal_"):
        return None, None

    rows = list(cal_used or [])
    if len(rows) < 3:
        return None, None

    x_all = np.asarray([float(r.get("Conc", np.nan)) for r in rows], dtype=np.float64)
    y_raw = np.asarray([float(r.get(f"{base}_median", np.nan)) for r in rows], dtype=np.float64)
    y_sd = np.asarray([float(r.get(f"{base}_sd", np.nan)) for r in rows], dtype=np.float64)
    w_all = np.asarray([_row_weight(r, base) for r in rows], dtype=np.float64)

    keep_all = np.isfinite(x_all) & np.isfinite(y_raw)
    if w_all is not None:
        keep_all = keep_all & np.isfinite(w_all) & (w_all > 0)

    x_all = x_all[keep_all]
    y_raw = y_raw[keep_all]
    y_sd = y_sd[keep_all]
    w_all = w_all[keep_all] if w_all is not None else None

    if x_all.size < 3:
        return None, None

    # 1) S0 is estimated from raw calibration data, before any zero-forced fit
    #    and before any clipping correction. It is a channel-wide vertical
    #    offset, therefore it must not be inferred only from x == 0.
    #    We use the lowest finite raw calibration response as the low-end
    #    offset estimate: if the raw calibration floor is negative, all data
    #    from the same channel are translated upward by that amount.
    zero_mask = np.abs(x_all) <= 1e-12
    y_raw_finite = y_raw[np.isfinite(y_raw)]
    if y_raw_finite.size:
        s0_calibration = max(0.0, -float(np.nanmin(y_raw_finite)))
    else:
        s0_calibration = 0.0

    # 2) Offset-correct calibration.
    y_shifted = y_raw + float(s0_calibration)

    # 3) Fit shifted calibration with forced zero intercept.
    fit0 = _fit_line_with_covariance(x_all, y_shifted, w=w_all, force_zero=True)
    if fit0 is None or not np.isfinite(fit0.get("m", np.nan)):
        return None, None

    m0 = float(fit0["m"])
    y_expected = m0 * x_all

    # 4) Compute clipping correction after offset correction and zero-forced fit.
    #    SD is used only as a decision threshold. If SD is unavailable, threshold
    #    is zero and the correction is the positive deficit itself.
    deficit = y_expected - y_shifted
    sd_thr = np.where(np.isfinite(y_sd) & (y_sd > 0), y_sd, 0.0)

    delta_clip = np.zeros_like(y_shifted, dtype=np.float64)
    clip_mask = np.isfinite(deficit) & (deficit > sd_thr)
    delta_clip[clip_mask] = deficit[clip_mask]

    y_corrected = y_shifted + delta_clip

    # The calibration fit itself remains forced through zero. This affects only
    # the calibration fit, not the stored S0/ClipDelta transferred to std add.
    y_for_fit = y_corrected.copy()
    y_for_fit[zero_mask] = 0.0

    fit_final = _fit_line_with_covariance(x_all, y_for_fit, w=w_all, force_zero=True)
    if fit_final is None:
        fit_final = fit0

    payload = {
        "S0_calibration": float(s0_calibration),
        "clip_x": [float(v) for v in x_all.tolist()],
        "clip_delta": [float(v) for v in delta_clip.tolist()],
        "clip_y_observed": [float(v) for v in y_raw.tolist()],
        "clip_y_shifted": [float(v) for v in y_shifted.tolist()],
        "clip_y_expected": [float(v) for v in y_expected.tolist()],
        "clip_y_corrected": [float(v) for v in y_corrected.tolist()],
        "clip_sd": [float(v) if np.isfinite(v) else None for v in y_sd.tolist()],
        "clip_sd_threshold": [float(v) for v in sd_thr.tolist()],
        "n_clip_points": int(np.count_nonzero(delta_clip > 0)),
    }

    fit_final["rgb_low_signal_correction"] = payload
    fit_final["S0_calibration"] = float(s0_calibration)
    fit_final["ClipX"] = ",".join(f"{v:.6g}" for v in x_all)
    fit_final["ClipDelta"] = ",".join(f"{v:.5g}" for v in delta_clip)
    fit_final["NClipPoints"] = int(np.count_nonzero(delta_clip > 0))
    mapper = _rgb_low_signal_mapper_from_payload(payload)
    return fit_final, mapper

def _corrected_c0_from_calibration_slope(cal_fit, std_fit, dilution_factor=1.0):
    """
    Compute corrected C0 using the intercept from standard addition and the
    slope from calibration:

        C0 derived from standard addition intercept and calibration slope

    Uncertainty propagation uses only the uncertainty of q_std if available
    in std_fit["cov_mq"]; uncertainty on m_cal is ignored here to keep the
    correction conservative and simple. If covariance is unavailable, SD is NaN.
    """
    if cal_fit is None or std_fit is None:
        return np.nan, np.nan

    m_cal = float(cal_fit.get("m", np.nan))
    q_std = float(std_fit.get("q", np.nan))
    if not np.isfinite(m_cal) or not np.isfinite(q_std) or abs(m_cal) < 1e-15:
        return np.nan, np.nan

    # Standard addition with x = added concentration:
    # y = m * x + q, x_intercept = -q/m, C0_diluted = -x_intercept = q/m.
    # When a stored calibration is applied, use the stored calibration slope
    # and the standard-addition intercept.
    c0_corr = q_std / m_cal
    c0_corr = float(dilution_factor) * c0_corr

    cov = std_fit.get("cov_mq", None)
    if cov is None or not np.isfinite(cov).all():
        return float(c0_corr), np.nan

    var_q = float(cov[1, 1]) if np.isfinite(cov[1, 1]) else np.nan
    if not np.isfinite(var_q) or var_q < 0:
        return float(c0_corr), np.nan

    sd = np.sqrt(var_q) / abs(m_cal)
    sd = float(dilution_factor) * float(sd)
    return float(c0_corr), float(sd)


def _channel_bias_metrics(cal_fit, std_fit):
    """
    Return channel-specific proportional-bias metrics:
        beta = m_std / m_cal
        bias_index = |beta - 1|
    """
    if cal_fit is None or std_fit is None:
        return np.nan, np.nan

    m_cal = float(cal_fit.get("m", np.nan))
    m_std = float(std_fit.get("m", np.nan))
    if not np.isfinite(m_cal) or not np.isfinite(m_std) or abs(m_cal) < 1e-15:
        return np.nan, np.nan

    beta = m_std / m_cal
    bias_index = abs(beta - 1.0)
    return float(beta), float(bias_index)


def _augment_fit_rows_with_bias_metrics(fit_rows):
    """
    Post-process fit_rows and add:
        beta_k
        bias_index_k
    """
    out = []
    cal_map = {}
    for row in fit_rows:
        if row.get("FitType") == "Calibration":
            cal_map[row.get("Channel")] = row

    for row in fit_rows:
        r = dict(row)
        r["beta_k"] = np.nan
        r["bias_index_k"] = np.nan

        if r.get("FitType") == "StdAdd":
            cal_row = cal_map.get(r.get("Channel"))
            cal_fit = None
            if cal_row is not None:
                cal_fit = {"m": cal_row.get("m", np.nan)}

            std_fit = {
                "m": r.get("m", np.nan),
                "q": r.get("q", np.nan),
                "cov_mq": r.get("__cov_mq", None),
            }

            beta, bias_index = _channel_bias_metrics(cal_fit, std_fit)
            r["beta_k"] = beta
            r["bias_index_k"] = bias_index

            c0_corr, c0_corr_sd = _corrected_c0_from_calibration_slope(
                cal_fit=cal_fit,
                std_fit=std_fit,
                dilution_factor=r.get("DF", 1.0),
            )

        out.append(r)
    return out




def _make_corrected_stdadd_fit(cal_fit, std_fit):
    """
    Build a corrected standard-addition line using the calibration slope
    and the standard-addition intercept:

        y = m_cal * x + q_std

    This corresponds to:
        C0 line estimated from q_std / m_cal
    """
    if cal_fit is None or std_fit is None:
        return None

    m_cal = float(cal_fit.get("m", np.nan))
    q_std = float(std_fit.get("q", np.nan))
    if not np.isfinite(m_cal) or not np.isfinite(q_std) or abs(m_cal) < 1e-15:
        return None

    out = dict(std_fit)
    out["m_corr"] = m_cal
    out["q_corr"] = q_std
    # Positive diluted concentration corresponding to the x-intercept.
    out["C0_line"] = float(q_std / m_cal)
    return out

def _autosize_worksheet(ws, extra=2):
    """Conservative autosize: avoid huge columns caused by paths/long notes."""
    widths = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            text = str(cell.value)
            # Long text fields should wrap instead of making giant columns.
            sample = max((len(x) for x in text.splitlines()), default=len(text))
            widths[cell.column] = max(widths.get(cell.column, 0), min(sample, 48))
    for col_idx, width in widths.items():
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = min(34, max(8, width + extra))
    # Freeze the header row for all data sheets.
    try:
        ws.freeze_panes = "A2"
    except Exception:
        pass


def _excel_cell_value(value, header=None):
    """Return a real Excel numeric value whenever possible; never write NaN.

    Identifiers and categorical columns are kept as text even if they look
    numeric, because values such as sample ID=1 are labels, not quantities.
    """
    if value is None:
        return ""
    h = str(header or "").strip().lower()
    text_headers = {
        "row", "well", "id", "type", "field", "notes", "meaning", "formula",
        "unit", "where used", "shown when", "component", "family", "method",
        "rankmode", "estimate_source", "sigma_source", "cielab_ref_source",
        "selected method from rank", "selected family", "ranking mode",
        "confidence class", "empty-well qc status", "quantification",
        "reference label", "status notes"
    }
    if h in text_headers or h.startswith("expected_label"):
        txt = str(value)
        return "" if txt.strip().lower() in {"nan", "none"} else txt
    if isinstance(value, (int, np.integer)):
        return int(value)
    if isinstance(value, (float, np.floating)):
        return float(value) if np.isfinite(value) else ""
    try:
        txt = str(value).strip()
        if txt.lower() in {"", "nan", "none"}:
            return ""
        x = float(txt)
        if np.isfinite(x):
            return float(x)
    except Exception:
        pass
    txt = str(value)
    if txt.strip().lower() in {"nan", "none"}:
        return ""
    return value


def _excel_number_format(header):
    h = str(header or "").lower()
    if h in {"rank", "selected", "col", "imagewarning", "qcflagged", "qccritical", "n_points", "nreplicates", "nwellwarnings", "nwellcritical", "nclippoints", "n_stdadd", "n_unknown"}:
        return "0"
    if h == "df":
        return "0.###"
    if h == "conc":
        return "0.###"
    if h in {"r2", "r2_cal", "r2_std", "r2_std_mean"}:
        return "0.0000"
    if "score" in h or "slopeagreement" in h or "bias_index" in h or "beta" in h or "rel_error" in h:
        return "0.0000"
    if "recovery" in h:
        return "0"
    if h in {"m", "q", "m_cal", "m_std_mean", "slope", "intercept"}:
        return "0.00000"
    if h in {"lod", "loq", "sigma_cal", "rmse", "snr"}:
        return "0.000"
    if h in {"c0", "c0_sd", "c0_mean", "c0_median", "c0_sd_median", "estimate_value", "estimate_sd"} or "expected_value" in h or "expected_sd" in h or "delta_expected" in h or "estimate_for_expected" in h:
        return "0.###"
    if "fraction" in h or "ratio" in h or "usedfraction" in h or "drift" in h or "span" in h:
        return "0.0000"
    if h in {"value", "sd"}:
        return "0.####"
    if h.startswith(("meanw", "meanbg", "signalt", "pseudoabs", "pabs", "signal")):
        return "0.0000"
    if h.startswith("delta") or h in {"l", "a", "b"} or h.startswith(("l_", "a_", "b_")):
        return "0.00"
    return "0.####"


def _write_table(ws, start_row, headers, rows, comments=None):
    thin = Side(style="thin", color="C8C8C8")
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    comments = comments or {}
    for cc, h in enumerate(headers, start=1):
        cell = ws.cell(start_row, cc, h)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.border = Border(top=thin, bottom=thin)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        note = comments.get(h)
        if note:
            cell.comment = Comment(str(note), "SAM analyzer")
    row_idx = start_row + 1
    for row in rows or []:
        for cc, h in enumerate(headers, start=1):
            val = _excel_cell_value(row.get(h, ""), h)
            cell = ws.cell(row_idx, cc, val)
            cell.alignment = Alignment(vertical="top", wrap_text=False)
            if isinstance(val, (int, float)) and not isinstance(val, bool):
                cell.number_format = _excel_number_format(h)
                cell.alignment = Alignment(vertical="top", horizontal="right", wrap_text=False)
            if isinstance(cell.value, str) and len(cell.value) > 60:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        row_idx += 1
    # No worksheet autofilter/dropdown menus: reports are static and traceable.
    return row_idx

def _rows_with_aliases(rows, alias_map):
    out = []
    for row in rows or []:
        rr = dict(row)
        for new_key, old_key in (alias_map or {}).items():
            if new_key not in rr and old_key in rr:
                rr[new_key] = rr.get(old_key)
        out.append(rr)
    return out

def _display_rows_with_pabs_names(rows):
    out = []
    for row in rows or []:
        rr = dict(row)
        for key in ["Channel", "Method", "Base", "Component", "Selected method from rank", "SelectedMethod"]:
            if key in rr and isinstance(rr.get(key), str):
                rr[key] = rr[key].replace("Signal_", "PAbs_")
        out.append(rr)
    return out



def _format_qc_value(value):
    if value is None:
        return ""
    try:
        v = float(value)
        if np.isfinite(v):
            return float(v)
    except Exception:
        pass
    return value


def _build_image_qc_metadata_rows(image_qc_info=None):
    info = image_qc_info or {}
    rows = []
    def add(field, value, notes=""):
        if value is None or value == "":
            return
        rows.append({"Field": field, "Value": _format_qc_value(value), "Notes": notes})

    add("Analysis image source", info.get("analysis_image_source"), "original or in-memory resized image used for all measurements")
    add("Original image path", info.get("original_image_path"), "source file selected by the user")
    add("Original image size", info.get("original_image_size"), "width x height in pixels, when known")
    add("Analysis image size", info.get("analysis_image_size"), "width x height in pixels actually analyzed")
    add("Resize scale", info.get("resize_scale"), "analysis_width / original_width, when a resized image is used")
    add("Initial image QC", info.get("initial_image_qc"), "Rule-based result: FAIL if destructive=True; WARNING if quality_warning=True or borderline=True; otherwise OK. No image correction is applied.")
    add("Initial image QC class", info.get("image_qc_class"), "non_correctable if destructive=True; quality_warning if flatfield_span > 0.18 and no destructive defect; usable_with_warnings if only borderline defects; good if no flags.")
    add("Initial image QC messages", info.get("image_qc_messages"), "Messages list the threshold rules triggered by the measured fields below.")
    add("Image QC decision rules", "thresholds", "destructive = dead_channel OR saturation_fraction>0.003 OR saturation_all_channels_fraction>0.0008 OR specular_fraction>0.003 OR max_side<900 OR approx_roi_pixels<120; borderline = saturation_fraction>0.0005 OR specular_fraction>0.0005 OR blur_score<35; quality_warning = flatfield_span>0.18.")
    add("Flat-field span", info.get("flatfield_span"), "Flatfield_span = (P95 - P5) / median of the slow illumination field estimated from the grayscale image. Larger values indicate stronger illumination/background gradient.")
    add("Specular fraction", info.get("specular_fraction"), "fraction of bright low-saturation pixels")
    add("Saturation fraction", info.get("saturation_fraction"), "fraction of pixels with at least one saturated channel")
    add("Dead channel", info.get("dead_channel"), "1 if a channel is near-dead by mean/variance checks")
    add("Approx well pitch", info.get("approx_well_pitch_px"), "median distance between adjacent well centers, in pixels")
    add("Approx ROI pixels/well", info.get("approx_roi_pixels_per_well"), "median number of pixels used in the final floor ROI mask per well")
    add("Blur score", info.get("blur_score"), "variance of Laplacian of the analysis image; lower values indicate stronger blur")
    return rows


def _compute_image_qc_info(img_bgr, centers=None, well_bottom_masks=None, image_path=None, cfg=None):
    cfg = cfg or {}
    h, w = img_bgr.shape[:2]
    original_w = cfg.get("original_width", cfg.get("source_width", cfg.get("input_width", None)))
    original_h = cfg.get("original_height", cfg.get("source_height", cfg.get("input_height", None)))
    try:
        original_w_f = float(original_w) if original_w is not None else np.nan
        original_h_f = float(original_h) if original_h is not None else np.nan
    except Exception:
        original_w_f = np.nan
        original_h_f = np.nan
    if not np.isfinite(original_w_f) or not np.isfinite(original_h_f):
        original_w_f = float(w)
        original_h_f = float(h)

    scale = float(w / original_w_f) if np.isfinite(original_w_f) and original_w_f > 0 else np.nan
    source = "original"
    if np.isfinite(scale) and abs(scale - 1.0) > 1e-3:
        source = "in-memory resized from original"

    pitch_vals = []
    try:
        cc = np.asarray(centers, dtype=np.float64)
        if cc.ndim == 3:
            if cc.shape[1] > 1:
                d = np.linalg.norm(cc[:, 1:, :] - cc[:, :-1, :], axis=2).ravel()
                pitch_vals.extend([float(x) for x in d if np.isfinite(x) and x > 0])
            if cc.shape[0] > 1:
                d = np.linalg.norm(cc[1:, :, :] - cc[:-1, :, :], axis=2).ravel()
                pitch_vals.extend([float(x) for x in d if np.isfinite(x) and x > 0])
    except Exception:
        pass
    pitch = float(np.nanmedian(pitch_vals)) if pitch_vals else np.nan

    roi_counts = []
    try:
        for row in well_bottom_masks or []:
            for m in row:
                if m is not None:
                    roi_counts.append(float(np.count_nonzero(m)))
    except Exception:
        roi_counts = []
    roi_pix = float(np.nanmedian(roi_counts)) if roi_counts else np.nan

    try:
        gray = cv2.cvtColor(img_bgr, cv2.COLOR_BGR2GRAY)
        blur_score = float(cv2.Laplacian(gray, cv2.CV_64F).var())
    except Exception:
        blur_score = np.nan

    initial_qc = "OK"
    if (np.isfinite(blur_score) and blur_score < 50.0) or (np.isfinite(roi_pix) and roi_pix < 500.0) or (np.isfinite(pitch) and pitch < 30.0):
        initial_qc = "WARNING"

    extra = {}
    try:
        extra = dict((cfg or {}).get("__image_qc_payload", {}) or {})
    except Exception:
        extra = {}

    out = {
        "analysis_image_source": source,
        "original_image_path": str(image_path or ""),
        "original_image_size": f"{int(round(original_w_f))} x {int(round(original_h_f))} px" if np.isfinite(original_w_f) and np.isfinite(original_h_f) else "",
        "analysis_image_size": f"{int(w)} x {int(h)} px",
        "resize_scale": scale if np.isfinite(scale) else np.nan,
        "initial_image_qc": initial_qc,
        "approx_well_pitch_px": pitch,
        "approx_roi_pixels_per_well": roi_pix,
        "blur_score": blur_score,
    }
    # Image QC/classification is computed before analysis. The image is never modified.
    out.update(extra)
    if not out.get("initial_image_qc"):
        out["initial_image_qc"] = initial_qc
    if not np.isfinite(_num_or_nan(out.get("blur_score", np.nan))):
        out["blur_score"] = blur_score
    return out


def _reference_delta_threshold(ref_value, ref_sd=np.nan):
    try:
        v = abs(float(ref_value))
    except Exception:
        v = np.nan
    try:
        sd = abs(float(ref_sd))
    except Exception:
        sd = np.nan
    candidates = []
    if np.isfinite(sd) and sd > 0:
        candidates.append(3.0 * sd)
    if np.isfinite(v) and v > 0:
        candidates.append(0.50 * v)
    if not candidates:
        return np.nan
    return float(max(candidates))


def _comparison_low_reliability(row, expected_refs=None):
    slope = _num_or_nan(row.get("SlopeAgreement", np.nan))
    r2_cal = _num_or_nan(row.get("R2_cal", np.nan))
    r2_std = _num_or_nan(row.get("R2_std_mean", np.nan))
    if np.isfinite(slope) and slope < 0.50:
        return True
    if (np.isfinite(r2_cal) and r2_cal < 0.80) or (np.isfinite(r2_std) and r2_std < 0.80):
        return True
    expected_refs = expected_refs or []
    best_abs_delta = np.nan
    best_thr = np.nan
    for i_ref, ref in enumerate(expected_refs, start=1):
        ref_label = _format_reference_label(ref, i_ref)
        safe = _expected_ref_key_label(ref_label, i_ref)
        delta = _num_or_nan(row.get(f"delta_expected_{safe}", np.nan))
        ref_val = _num_or_nan(ref.get("value", np.nan))
        ref_sd = _num_or_nan(ref.get("sd", np.nan))
        thr = _reference_delta_threshold(ref_val, ref_sd)
        if np.isfinite(delta) and np.isfinite(thr):
            if not np.isfinite(best_abs_delta) or abs(delta) < best_abs_delta:
                best_abs_delta = abs(delta)
                best_thr = thr
    if np.isfinite(best_abs_delta) and np.isfinite(best_thr) and best_abs_delta > best_thr:
        return True
    return False

def _build_matrix_diagnostic_rows(fit_rows):
    cal_map = {}
    rows = []
    for row in fit_rows:
        if row.get("FitType") == "Calibration":
            cal_map[row.get("Channel")] = row
    for row in fit_rows:
        if row.get("FitType") != "StdAdd":
            continue
        ch = row.get("Channel")
        cal = cal_map.get(ch, {})
        m_cal = float(cal.get("m", np.nan)) if cal else np.nan
        q_cal = float(cal.get("q", np.nan)) if cal else np.nan
        r2_cal = float(cal.get("R2", np.nan)) if cal else np.nan
        sigma_cal = float(cal.get("sigma_cal", np.nan)) if cal else np.nan
        m_std = float(row.get("m", np.nan)) if np.isfinite(row.get("m", np.nan)) else np.nan
        q_std = float(row.get("q", np.nan)) if np.isfinite(row.get("q", np.nan)) else np.nan
        r2_std = float(row.get("R2", np.nan)) if np.isfinite(row.get("R2", np.nan)) else np.nan
        slope_ratio = (m_std / m_cal) if np.isfinite(m_std) and np.isfinite(m_cal) and abs(m_cal) > 1e-15 else np.nan
        slope_delta = (m_std - m_cal) if np.isfinite(m_std) and np.isfinite(m_cal) else np.nan
        intercept_delta = (q_std - q_cal) if np.isfinite(q_std) and np.isfinite(q_cal) else np.nan
        snr_std_proxy = (abs(m_std) / sigma_cal) if np.isfinite(m_std) and np.isfinite(sigma_cal) and sigma_cal > 0 else np.nan
        blue_threshold_like_active = 1 if (str(ch) == "Blue" and np.isfinite(snr_std_proxy) and snr_std_proxy >= 3.0) else 0
        df_val = float(row.get("DF", np.nan)) if np.isfinite(row.get("DF", np.nan)) else np.nan
        inv_df = (1.0 / df_val) if np.isfinite(df_val) and abs(df_val) > 1e-15 else np.nan
        log10_inv_df = float(np.log10(inv_df)) if np.isfinite(inv_df) and inv_df > 0 else np.nan
        rows.append({
            "Channel": ch,
            "ID": row.get("ID"),
            "DF": row.get("DF"),
            "DF_numeric": df_val,
            "MatrixLevel_proxy_1_over_DF": inv_df,
            "MatrixLevel_index_1_over_DF": log10_inv_df,
            "m_cal": m_cal,
            "q_cal": q_cal,
            "R2_cal": r2_cal,
            "sigma_cal": sigma_cal,
            "m_std": m_std,
            "q_std": q_std,
            "R2_std": r2_std,
            "delta_m": slope_delta,
            "ratio_m": slope_ratio,
            "delta_q": intercept_delta,
            "SNR_std_proxy": snr_std_proxy,
            "Blue_threshold_like_active": blue_threshold_like_active,
            "C0": float(row.get("C0", np.nan)) if np.isfinite(row.get("C0", np.nan)) else np.nan,
            "C0_sd": float(row.get("C0_sd", np.nan)) if np.isfinite(row.get("C0_sd", np.nan)) else np.nan,
        })
    rows.sort(key=lambda d: (str(d.get("Channel", "")), float(d.get("DF_numeric", np.nan)) if np.isfinite(d.get("DF_numeric", np.nan)) else np.inf, str(d.get("ID", ""))))
    return rows


def _report_mode_flags(fit_rows):
    fit_rows = fit_rows or []
    has_calibration = any(str(r.get("FitType", "")) == "Calibration" and np.isfinite(r.get("m", np.nan)) for r in fit_rows)
    has_stdadd = any(str(r.get("FitType", "")) == "StdAdd" and np.isfinite(r.get("C0", np.nan)) for r in fit_rows)
    has_unknown = any(str(r.get("FitType", "")) in {"UnknownFromCal", "UnknownFromEpsilon"} and np.isfinite(r.get("C0", np.nan)) for r in fit_rows)
    has_unknown_only = any(str(r.get("FitType", "")) == "UnknownOnly" for r in fit_rows)
    return {
        "has_calibration": has_calibration,
        "has_stdadd": has_stdadd,
        "has_unknown": has_unknown,
        "has_unknown_only": has_unknown_only,
        "calibration_only": has_calibration and (not has_stdadd) and (not has_unknown),
    }



def _finalize_legend_rows(rows):
    """Normalize legend rows without changing analytical calculations.

    Later curated rows intentionally replace earlier provisional/legacy rows.
    This keeps the exported LEGENDS sheets aligned with the final output names
    rather than with internal or vestigial implementation names.
    """
    merged = {}
    order = []
    for row in rows or []:
        rr = dict(row)
        if "Field" in rr and "Term" not in rr:
            rr["Term"] = rr.pop("Field")
        term = str(rr.get("Term", "")).strip()
        if not term:
            continue
        rr["Term"] = term
        key = term.lower()
        if key not in merged:
            order.append(key)
            merged[key] = {"Term": term}
        for col in ["Meaning", "Formula", "Unit", "Where used", "Shown when", "Notes"]:
            val = rr.get(col, "")
            if val is not None and str(val).strip() != "":
                merged[key][col] = val
            else:
                merged[key].setdefault(col, "")
    out = []
    for key in order:
        rr = merged[key]
        for col in ["Meaning", "Formula", "Unit", "Where used", "Shown when", "Notes"]:
            rr.setdefault(col, "")
        note = str(rr.get("Notes", "")).strip()
        meaning = str(rr.get("Meaning", "")).strip().lower()
        formula = str(rr.get("Formula", "")).strip().lower()
        if note and (note.lower() == meaning or note.lower() == formula):
            rr["Notes"] = ""
        out.append(rr)
    return sorted(out, key=lambda r: str(r.get("Term", "")).lower())


def _build_legends_rows(unit_label, mode_flags=None):
    mode_flags = mode_flags or {}
    always = "always"
    cal_only = "calibration only"
    stdadd_present = "stdadd present"
    unknown_present = "unknown/CRM present"
    cielab_enabled = "CIELAB enabled"
    rows = [
        {"Field": "PAbs", "Meaning": "RGB pseudo-absorbance", "Formula": "PAbs = log10(BG/W) = -log10(W/BG)", "Unit": "dimensionless", "Where used": "RAW, REPLICATES_MEAN, FITTING, METHOD_COMPARISON, FIGURE_RGB", "Shown when": always, "Notes": "Fixed full-background normalization. In the source code some legacy internal arrays may still be named Signal_*; user-facing outputs report PAbs_* only."},
        {"Field": "PAbs_Red", "Meaning": "RGB pseudo-absorbance (red channel)", "Formula": "log10(BG_red/W_red) = -log10(W_red/BG_red)", "Unit": "dimensionless", "Where used": "RAW, REPLICATES_MEAN, FITTING", "Shown when": always, "Notes": "Image-derived pseudo-absorbance with fixed full-background normalization."},
        {"Field": "PAbs_Green", "Meaning": "RGB pseudo-absorbance (green channel)", "Formula": "log10(BG_green/W_green) = -log10(W_green/BG_green)", "Unit": "dimensionless", "Where used": "RAW, REPLICATES_MEAN, FITTING", "Shown when": always, "Notes": "Image-derived pseudo-absorbance with fixed full-background normalization."},
        {"Field": "PAbs_Blue", "Meaning": "RGB pseudo-absorbance (blue channel)", "Formula": "log10(BG_blue/W_blue) = -log10(W_blue/BG_blue)", "Unit": "dimensionless", "Where used": "RAW, REPLICATES_MEAN, FITTING", "Shown when": always, "Notes": "Image-derived pseudo-absorbance with fixed full-background normalization."},
        {"Field": "MeanW_*", "Meaning": "Linearized median well intensity for the channel", "Formula": "median ROI channel intensity after gamma linearization", "Unit": "dimensionless", "Where used": "RAW", "Shown when": always, "Notes": "Computed from selected well ROI pixels."},
        {"Field": "MeanBG_*", "Meaning": "Linearized local background intensity for the channel", "Formula": "background surface evaluated at the well and gamma linearized", "Unit": "dimensionless", "Where used": "RAW", "Shown when": always, "Notes": "Derived from inter-well background reconstruction."},
        {"Field": "SignalT_*", "Meaning": "Transmittance-like intensity ratio before logarithm", "Formula": "MeanW_* / MeanBG_*", "Unit": "dimensionless", "Where used": "RAW", "Shown when": always, "Notes": "PAbs_* = -log10(SignalT_*)."},
        {"Field": "RGB linearization", "Meaning": "Gamma-domain RGB values are converted to a linear intensity-like scale for RGB PAbs calculations", "Formula": "I_lin = (I/255)^γ, γ ≈ 2.2", "Unit": "dimensionless", "Where used": "RAW PAbs fields", "Shown when": always, "Notes": "Used for MeanW_*, MeanBG_*, SignalT_* and PAbs_*; no image correction is applied."},
        {"Field": "RGB to CIELAB", "Meaning": "Diagnostic conversion from image RGB to CIE L*a*b*", "Formula": "OpenCV BGR→Lab for 8-bit sRGB-like images; reference form: linearized sRGB → XYZ using the standard sRGB-to-XYZ matrix and D65 reference white → CIE L*a*b*", "Unit": "dimensionless", "Where used": "RAW, REPLICATES_MEAN, CIELAB diagnostics", "Shown when": cielab_enabled, "Notes": "Diagnostic descriptor, not the primary RGB quantitative output. References: OpenCV color conversions; IEC 61966-2-1:1999 sRGB; CIE 1976 L*a*b*; CIE standard illuminant D65."},
        {"Field": "CIELAB (L*, a*, b*)", "Meaning": "CIE 1976 color coordinates derived from the RGB image", "Formula": "L* = 116 f(Y/Yn) - 16; a* = 500[f(X/Xn) - f(Y/Yn)]; b* = 200[f(Y/Yn) - f(Z/Zn)]", "Unit": "dimensionless", "Where used": "RAW, REPLICATES_MEAN, CIELAB diagnostics", "Shown when": cielab_enabled, "Notes": "D65 is the reference white used by the sRGB/OpenCV conversion convention. References: CIE 1976 L*a*b* and CIE standard illuminant D65."},
        {"Field": "L", "Meaning": "CIELAB lightness", "Formula": "median ROI value after RGB → CIELAB conversion", "Unit": "dimensionless", "Where used": "RAW, SUMMARY", "Shown when": cielab_enabled, "Notes": ""},
        {"Field": "a", "Meaning": "CIELAB green–red axis", "Formula": "median ROI value after RGB → CIELAB conversion", "Unit": "dimensionless", "Where used": "RAW, SUMMARY", "Shown when": cielab_enabled, "Notes": ""},
        {"Field": "b", "Meaning": "CIELAB blue–yellow axis", "Formula": "median ROI value after RGB → CIELAB conversion", "Unit": "dimensionless", "Where used": "RAW, SUMMARY", "Shown when": cielab_enabled, "Notes": ""},
        {"Field": "DeltaL", "Meaning": "Difference in CIELAB lightness relative to the selected reference", "Formula": "L - L_ref", "Unit": "dimensionless", "Where used": "RAW, SUMMARY", "Shown when": cielab_enabled, "Notes": "Reference source is reported in CIELAB_ref_source."},
        {"Field": "Deltaa", "Meaning": "Difference in CIELAB a* relative to the selected reference", "Formula": "a - a_ref", "Unit": "dimensionless", "Where used": "RAW, SUMMARY", "Shown when": cielab_enabled, "Notes": "Reference source is reported in CIELAB_ref_source."},
        {"Field": "Deltab", "Meaning": "Difference in CIELAB b* relative to the selected reference", "Formula": "b - b_ref", "Unit": "dimensionless", "Where used": "RAW, SUMMARY", "Shown when": cielab_enabled, "Notes": "Reference source is reported in CIELAB_ref_source."},
        {"Field": "DeltaE_ab", "Meaning": "Total CIELAB color difference", "Formula": "sqrt(DeltaL^2 + Deltaa^2 + Deltab^2)", "Unit": "dimensionless", "Where used": "RAW, SUMMARY, FITTING", "Shown when": cielab_enabled, "Notes": "CIE 1976 L*a*b* color-difference form; reference source is reported in CIELAB_ref_source."},
        {"Field": "DeltaE_ab_chroma", "Meaning": "Chromatic CIELAB difference without the lightness term", "Formula": "sqrt(Deltaa^2 + Deltab^2)", "Unit": "dimensionless", "Where used": "RAW, SUMMARY", "Shown when": cielab_enabled, "Notes": "Reference source is reported in CIELAB_ref_source."},
        {"Field": "CIELAB_ref_source", "Meaning": "Source of the CIELAB reference used for Delta variables", "Formula": "plate_zero_calibration/zero_calibration = median L,a,b of calibration rows with Conc = 0; fallback sources are listed when zero is unavailable", "Unit": "text", "Where used": "RAW, REPLICATES_MEAN", "Shown when": cielab_enabled, "Notes": ""},
        {"Field": "UsedFraction", "Meaning": "Fraction of ROI core pixels retained after filtering", "Formula": "n_used / n_core", "Unit": "dimensionless", "Where used": "RAW, SUMMARY", "Shown when": always, "Notes": "Lower values indicate stronger filtering. Closer to 1 means less aggressive trimming."},
        {"Field": "IRLS", "Meaning": "Iteratively reweighted least-squares robust linear regression with residual-based weights", "Formula": "minimize sum_i w_i (y_i - (m x_i + q))^2; w_i is updated iteratively from residual magnitude", "Unit": "dimensionless", "Where used": "FIGURE_RGB, FIGURE_CIELAB_DELTAE, FITTING, METHOD_COMPARISON", "Shown when": always, "Notes": "Reference: Huber, P. J. (1964), Robust Estimation of a Location Parameter. Implementation: custom NumPy-based IRLS using repeated least-squares solves."},
        {"Field": "m", "Meaning": "Slope of linear fit", "Formula": "y = m x + q", "Unit": f"signal / {unit_label}", "Where used": "FITTING", "Shown when": always, "Notes": "Calibration and standard-addition fits use the same notation."},
        {"Field": "q", "Meaning": "Intercept of linear fit", "Formula": "y = m x + q", "Unit": "signal", "Where used": "FITTING", "Shown when": always, "Notes": "For standard addition, C0 is derived from m and q."},
        {"Field": "R2", "Meaning": "Coefficient of determination", "Formula": "1 - SSE/SST", "Unit": "dimensionless", "Where used": "FITTING", "Shown when": always, "Notes": "Closer to 1 indicates better fit."},
        {"Field": "RMSE", "Meaning": "Root mean squared error", "Formula": "sqrt(mean(residual^2))", "Unit": "signal or concentration", "Where used": "FITTING", "Shown when": always, "Notes": "Interpret in the units of the fitted response."},
        {"Field": "sigma_cal", "Meaning": "Noise estimate for calibration", "Formula": "SD at zero concentration or median calibration SD", "Unit": "signal", "Where used": "FITTING", "Shown when": cal_only + ' or calibration present', "Notes": "Source is reported in sigma_source."},
        {"Field": "SNR", "Meaning": "Slope-to-noise ratio", "Formula": "|m| / sigma_cal", "Unit": "dimensionless", "Where used": "FITTING", "Shown when": cal_only + ' or calibration present', "Notes": "Used only for calibration rows."},
        {"Field": "LOD", "Meaning": "Limit of detection", "Formula": "3 × σ / |m|", "Unit": unit_label, "Where used": "FITTING, OVERVIEW", "Shown when": cal_only + ' or calibration present', "Notes": "Shown only for calibration rows."},
        {"Field": "LOQ", "Meaning": "Limit of quantification", "Formula": "10 × σ / |m|", "Unit": unit_label, "Where used": "FITTING, OVERVIEW", "Shown when": cal_only + ' or calibration present', "Notes": "Shown only for calibration rows."},
        {"Field": "C0", "Meaning": "Concentration in original sample", "Formula": "standard addition: DF × (q/m); unknown/CRM from calibration: DF × ((y − q)/m); unknown from epsilon: DF × PAbs/(epsilon × l_cm), converted from M to the selected unit", "Unit": unit_label, "Where used": "FITTING, OVERVIEW", "Shown when": "stdadd present or unknown/CRM present", "Notes": "Same symbol is used for standard addition and unknown/CRM results; FitType defines the method."},
        {"Field": "C0_sd", "Meaning": "Uncertainty on C0", "Formula": "Fit covariance for calibration/std-add; for epsilon mode, propagated PAbs replicate SD divided by epsilon × l_cm and converted to the selected unit", "Unit": unit_label, "Where used": "FITTING, OVERVIEW", "Shown when": "stdadd present or unknown/CRM present", "Notes": "Reported in the same concentration units as C0."},
        {"Field": "epsilon", "Meaning": "User-configured Beer-Lambert-like proportionality coefficient for PAbs quantification", "Formula": "PAbs = epsilon × l_cm × C_M", "Unit": "M-1 cm-1", "Where used": "FITTING, OVERVIEW", "Shown when": "unknown-only epsilon mode", "Notes": "C_M is mol/L. The analyzer converts the calculated concentration to the selected output unit."},
        {"Field": "liquid_volume_ul", "Meaning": "Liquid volume loaded into each well", "Formula": "user input", "Unit": "uL", "Where used": "OVERVIEW, FITTING", "Shown when": "epsilon mode configured", "Notes": "Used only to calculate path length for flat-bottom wells."},
        {"Field": "path_length", "Meaning": "Estimated optical path length", "Formula": "l_cm = (V_uL / A_bottom_mm2) / 10", "Unit": "cm", "Where used": "OVERVIEW, FITTING", "Shown when": "epsilon mode configured", "Notes": "Assumes 1 uL = 1 mm^3 and a flat-bottom well with nominal bottom area."},
        {"Field": "path_length_mm", "Meaning": "Estimated liquid height", "Formula": "l_mm = V_uL / A_bottom_mm2", "Unit": "mm", "Where used": "OVERVIEW, FITTING", "Shown when": "epsilon mode configured", "Notes": "Converted to cm before Beer-Lambert-like concentration calculation."},
        {"Field": "well_bottom_area_mm2", "Meaning": "Nominal flat-bottom well area used for path-length calculation", "Formula": "pi × (d_bottom/2)^2", "Unit": "mm^2", "Where used": "OVERVIEW", "Shown when": "epsilon mode configured", "Notes": "Derived from the selected plate geometry."},
        {"Field": "SlopeAgreement", "Meaning": "Agreement between calibration and standard-addition slopes", "Formula": "min(|m_cal|, |m_std|) / max(|m_cal|, |m_std|)", "Unit": "dimensionless", "Where used": "METHOD_COMPARISON, Score", "Shown when": stdadd_present, "Notes": "Equals 1 when slopes are identical in magnitude; lower values indicate poorer agreement between calibration and standard addition."},
        {"Field": "BiasIndex", "Meaning": "Relative slope bias between standard addition and calibration", "Formula": "|β − 1|, where β = m_std / m_cal", "Unit": "dimensionless", "Where used": "METHOD_COMPARISON", "Shown when": stdadd_present, "Notes": "Equals 0 when the standard-addition slope equals the calibration slope. Larger values indicate stronger slope bias."},
        {"Field": "Reliability criteria", "Meaning": "Rule used to flag a method as low reliability in method comparison", "Formula": "low if slope agreement < 0.5 OR R² < 0.8 OR |Δ| > threshold", "Unit": "rule", "Where used": "METHOD_COMPARISON", "Shown when": stdadd_present, "Notes": "The Δ threshold is max(3×reference SD, 50% of the reference value), evaluated against the nearest available reference."},
        {"Field": "Score", "Meaning": "Common method-comparison score", "Formula": "For calibration+standard addition rows: slope_agreement^2 × sqrt(R²_cal × R²_std). Calibration-only and standard-addition-only fallback rows are assigned their own ComparableGroup and are not directly ranked against full rows.", "Unit": "dimensionless", "Where used": "OVERVIEW, METHOD_COMPARISON, METHOD_COMPARISON.png", "Shown when": always, "Notes": "Expected/reference values, recovery, SNR and clipping are excluded from the score."},
        {"Field": "ComparableGroup", "Meaning": "Set of methods scored with the same formula", "Formula": "calibration_plus_stdadd, calibration_only, stdadd_only, or not_ranked", "Unit": "text", "Where used": "METHOD_COMPARISON", "Shown when": always, "Notes": "Scores are directly comparable only within the same group. The group with the largest CommonFactorsN is used for Selected/Rank."},
        {"Field": "CommonFactorsN", "Meaning": "Number of common factors used in the score", "Formula": "3 for R²_cal, R²_std, slope agreement; 1 for calibration-only or stdadd-only fallbacks", "Unit": "integer", "Where used": "METHOD_COMPARISON", "Shown when": always, "Notes": "Used to avoid comparing scores obtained from different formulas."},
        {"Field": "ScoreFormula", "Meaning": "Formula used to compute Score", "Formula": "text descriptor", "Unit": "text", "Where used": "METHOD_COMPARISON", "Shown when": always, "Notes": "Documents which score formula was applied to the row."},
        {"Field": "Selected", "Meaning": "Workbook-level selected method flag", "Formula": "1 for the highest-ranked row within the most informative ComparableGroup, otherwise 0", "Unit": "0/1", "Where used": "METHOD_COMPARISON", "Shown when": always, "Notes": "Selection in the workbook is derived from method-comparison rank and not from expected/recovery."},
        {"Field": "Rank", "Meaning": "Workbook-level rank of each method", "Formula": "1 = highest Score within the selected ComparableGroup", "Unit": "integer", "Where used": "METHOD_COMPARISON", "Shown when": always, "Notes": "Rows outside the selected ComparableGroup are not assigned a direct rank."},
        {"Field": "QCFlagged", "Meaning": "Group-level optical quality warning", "Formula": "1 if any replicate has ImageWarning, any critical replicate, or median UsedFraction < 0.70", "Unit": "0/1", "Where used": "REPLICATES_MEAN", "Shown when": always, "Notes": "Warning flag; points remain reported and used."},
        {"Field": "QCCritical", "Meaning": "Group-level critical optical quality flag", "Formula": "1 if FracWellCritical >= 0.50", "Unit": "0/1", "Where used": "REPLICATES_MEAN", "Shown when": always, "Notes": "Critical flag; points remain reported and used."},
        {"Field": "Robust SD", "Meaning": "Robust dispersion across replicates", "Formula": "1.4826 × MAD, where MAD = median(|x − median(x)|)", "Unit": "same as variable", "Where used": "REPLICATES_MEAN", "Shown when": always, "Notes": "Computed within each replicate group."},
        {"Field": "CIELAB_ref_source", "Meaning": "Reference source used to compute DeltaL/Deltaa/Deltab/DeltaE", "Formula": "zero_calibration = plate-local zero-concentration calibration median", "Unit": "text", "Where used": "RAW, REPLICATES_MEAN", "Shown when": cielab_enabled, "Notes": "Reported only for CIELAB-derived variables because RGB pseudo-absorbance does not use this CIELAB reference."},
        {"Field": "ClipX", "Meaning": "Calibration x values where clipping correction was evaluated", "Formula": "list of calibration concentrations", "Unit": unit_label, "Where used": "FITTING", "Shown when": "RGB calibration", "Notes": "Transferred to StdAdd at the same channel and same x values."},
        {"Field": "ClipDelta", "Meaning": "Positive calibration-derived clipping correction", "Formula": "max(0, y_expected − y_shifted) when deficit exceeds SD threshold", "Unit": "signal", "Where used": "FITTING", "Shown when": "RGB calibration", "Notes": "Computed from calibration after S0 and zero-intercept fit; applied by same channel and same x."},
        {"Field": "S0_calibration", "Meaning": "Channel-wide low-signal offset estimated from raw calibration", "Formula": "max(0, −min(raw calibration response))", "Unit": "signal", "Where used": "FITTING", "Shown when": "RGB calibration", "Notes": "Estimated before forced-zero calibration and transferred to StdAdd/unknown data."},
        {"Field": "beta_k", "Meaning": "Slope ratio for one standard-addition curve", "Formula": "m_std / m_cal", "Unit": "dimensionless", "Where used": "FITTING, METHOD_COMPARISON", "Shown when": stdadd_present, "Notes": "Used to assess whether StdAdd and calibration slopes are coherent."},
        {"Field": "bias_index_k", "Meaning": "Absolute relative slope bias for one standard-addition curve", "Formula": "|beta_k − 1|", "Unit": "dimensionless", "Where used": "FITTING, METHOD_COMPARISON", "Shown when": stdadd_present, "Notes": "0 means identical slopes."},
        {"Field": "flatfield_span", "Meaning": "Image-level slow-field nonuniformity", "Formula": "(P95 − P5) / median of the slow grayscale field", "Unit": "dimensionless", "Where used": "METADATA", "Shown when": always, "Notes": "Part of rule-based image QC; larger values indicate stronger illumination/background gradient."},
        {"Field": "Reliability score", "Meaning": "Overall report-level reliability score", "Formula": "starts at 50 and adds/subtracts rule-based penalties/bonuses for calibration availability, stored calibration, plate QC, critical wells, used fraction, empty-well QC, and ranking separation", "Unit": "0-100", "Where used": "OVERVIEW", "Shown when": always, "Notes": "Not a statistical probability; it is a heuristic audit score."},
        {"Field": "Confidence class", "Meaning": "Qualitative class derived from reliability score", "Formula": "HIGH >= 75; MEDIUM >= 45; LOW < 45; NOT QUANTIFIABLE when no valid quantification is available", "Unit": "class", "Where used": "OVERVIEW", "Shown when": always, "Notes": "Interpret together with the reliability reason."},
        {"Field": "Empty-well QC", "Meaning": "QC based on empty-well robust SD and/or stored empty comparison", "Formula": "warning/watch/ok/not_available from drift score and robust SD thresholds", "Unit": "class", "Where used": "OVERVIEW", "Shown when": always, "Notes": "Flags comparability or background/empty-well drift."},
    ]
    rows.extend([
        {"Term": "Row", "Meaning": "Plate row label", "Formula": "A-based alphabetical row label", "Unit": "well-row label", "Where used": "RAW, REPLICATES_MEAN", "Shown when": always, "Notes": ""},
        {"Term": "Col", "Meaning": "Plate column index", "Formula": "1-based column number", "Unit": "well-column index", "Where used": "RAW, REPLICATES_MEAN", "Shown when": always, "Notes": ""},
        {"Term": "Well", "Meaning": "Human-readable well identifier", "Formula": "Well = Row + Col", "Unit": "well label", "Where used": "RAW, REPLICATES_MEAN", "Shown when": always, "Notes": ""},
        {"Term": "ID", "Meaning": "Sample or reference identifier assigned in the plate map", "Formula": "user/configurator input", "Unit": "text", "Where used": "RAW, REPLICATES_MEAN, FITTING, METHOD_COMPARISON", "Shown when": always, "Notes": ""},
        {"Term": "Type", "Meaning": "Well role in the analytical workflow", "Formula": "C = calibration; A = standard addition; U = unknown; other labels may be reported when present", "Unit": "text", "Where used": "RAW, REPLICATES_MEAN", "Shown when": always, "Notes": ""},
        {"Term": "Conc", "Meaning": "Configured concentration for calibration or added-standard wells", "Formula": "user/configurator input", "Unit": unit_label, "Where used": "RAW, REPLICATES_MEAN, FITTING", "Shown when": always, "Notes": "Unknown wells may have undefined Conc."},
        {"Term": "DF", "Meaning": "Dilution factor used to report original-sample concentration", "Formula": "C0 = DF x concentration in analyzed/diluted well", "Unit": "dimensionless", "Where used": "RAW, REPLICATES_MEAN, FITTING", "Shown when": always, "Notes": ""},
        {"Term": "Channel", "Meaning": "Fitted analytical descriptor", "Formula": "PAbs_Red, PAbs_Green, PAbs_Blue or diagnostic CIELAB/DeltaE descriptor", "Unit": "text", "Where used": "FITTING, METHOD_COMPARISON", "Shown when": always, "Notes": ""},
        {"Term": "FitType", "Meaning": "Type of fit or concentration-estimation row", "Formula": "Calibration, StdAdd, UnknownFromCal, UnknownFromEpsilon, UnknownOnly", "Unit": "text", "Where used": "FITTING", "Shown when": always, "Notes": ""},
        {"Term": "n_points", "Meaning": "Number of finite data points used by the fit", "Formula": "count of finite x,y pairs", "Unit": "count", "Where used": "FITTING", "Shown when": always, "Notes": "All finite points are retained by robust IRLS."},
        {"Term": "sigma_source", "Meaning": "Source of sigma_cal", "Formula": "median_calibration_sd, blank_zero_calibration_sd, or unavailable", "Unit": "text", "Where used": "FITTING", "Shown when": "calibration present", "Notes": ""},
        {"Term": "UsedFraction_median", "Meaning": "Median UsedFraction across replicate wells in a group", "Formula": "median(UsedFraction)", "Unit": "dimensionless", "Where used": "REPLICATES_MEAN", "Shown when": always, "Notes": ""},
        {"Term": "ImageWarning_any", "Meaning": "Group-level indicator that at least one replicate had an image warning", "Formula": "1 if any replicate ImageWarning = 1", "Unit": "0/1", "Where used": "REPLICATES_MEAN", "Shown when": always, "Notes": ""},
        {"Term": "NReplicates", "Meaning": "Number of replicate wells in a summarized group", "Formula": "count of wells in group", "Unit": "count", "Where used": "REPLICATES_MEAN", "Shown when": always, "Notes": ""},
        {"Term": "NWellWarnings/NWellCritical", "Meaning": "Counts of replicate wells with warning or critical optical QC", "Formula": "count over replicate wells", "Unit": "count", "Where used": "REPLICATES_MEAN", "Shown when": always, "Notes": ""},
        {"Term": "FracWellWarnings/FracWellCritical", "Meaning": "Fractions of replicate wells with warning or critical optical QC", "Formula": "count / NReplicates", "Unit": "dimensionless", "Where used": "REPLICATES_MEAN", "Shown when": always, "Notes": ""},
        {"Term": "Family", "Meaning": "Method family used in method comparison", "Formula": "RGB, CIELAB, DeltaCIELAB, or other", "Unit": "text", "Where used": "METHOD_COMPARISON", "Shown when": always, "Notes": ""},
        {"Term": "RankMode", "Meaning": "Data basis available for ranking", "Formula": "calibration_plus_stdadd, calibration_only, stdadd_only, or unavailable", "Unit": "text", "Where used": "METHOD_COMPARISON", "Shown when": always, "Notes": ""},
        {"Term": "BaseScore/FinalScore", "Meaning": "Intermediate/final score fields when present", "Formula": "same common score unless an explicit post-score adjustment is applied", "Unit": "dimensionless", "Where used": "METHOD_COMPARISON", "Shown when": "method comparison", "Notes": "If identical to Score in the exported workbook, Score is the authoritative field."},
        {"Term": "Estimate_value/Estimate_sd", "Meaning": "Representative concentration estimate and associated SD used for method comparison", "Formula": "derived from standard addition, calibration projection, or epsilon mode depending on Estimate_source", "Unit": unit_label, "Where used": "METHOD_COMPARISON", "Shown when": "method comparison", "Notes": ""},
        {"Term": "expected_*/reference_*", "Meaning": "External reference value metadata", "Formula": "user/configurator input", "Unit": unit_label, "Where used": "OVERVIEW, METHOD_COMPARISON, FIGURE_RGB", "Shown when": "reference values configured", "Notes": "External references are checks only and are not used in Score."},
        {"Term": "delta_reference_*", "Meaning": "Difference between estimated concentration and an external reference", "Formula": "C0 - reference_value", "Unit": unit_label, "Where used": "OVERVIEW, METHOD_COMPARISON, FIGURE_RGB", "Shown when": "reference values configured", "Notes": ""},
        {"Term": "recovery_pct_*", "Meaning": "Recovery relative to an external reference", "Formula": "100 x C0 / reference_value", "Unit": "%", "Where used": "OVERVIEW, METHOD_COMPARISON, FIGURE_RGB", "Shown when": "reference values configured", "Notes": ""},
        {"Term": "rel_error_*", "Meaning": "Relative error versus an external reference", "Formula": "100 x (C0 - reference_value) / reference_value", "Unit": "%", "Where used": "OVERVIEW, METHOD_COMPARISON", "Shown when": "reference values configured", "Notes": ""},
        {"Term": "Code implementation", "Meaning": "Main computational libraries used by the analyzer", "Formula": "custom Python code using NumPy/OpenCV/openpyxl/matplotlib", "Unit": "software provenance", "Where used": "all outputs", "Shown when": always, "Notes": "Sources/libraries: NumPy linalg.lstsq for least-squares solves; OpenCV for image masks/color conversion; openpyxl for xlsx export; matplotlib for PNG figures."},
    ])
    rows.extend([
        {"Term":"Field", "Meaning":"Name of a metadata or overview item", "Formula":"label reported in the first column of the sheet", "Unit":"text", "Where used":"02_METADATA, 03_OVERVIEW", "Shown when":"always", "Notes":""},
        {"Term":"Value", "Meaning":"Value associated with a metadata or overview item", "Formula":"reported value for the corresponding Field", "Unit":"mixed", "Where used":"02_METADATA, 03_OVERVIEW", "Shown when":"always", "Notes":""},
        {"Term":"Sheet", "Meaning":"Workbook sheet name", "Formula":"worksheet name", "Unit":"text", "Where used":"01_CONTENTS", "Shown when":"always", "Notes":""},
        {"Term":"Purpose", "Meaning":"Short description of the sheet role", "Formula":"free-text description", "Unit":"text", "Where used":"01_CONTENTS", "Shown when":"always", "Notes":""},
        {"Term":"PAbs", "Meaning":"Image-derived RGB pseudo-absorbance", "Formula":"PAbs = log10(I_BG / I_well) = -log10(I_well / I_BG)", "Unit":"dimensionless", "Where used":"FIGURE_RGB.png, RESULTS_CAPTION.txt, 04_RAW, 05_REPLICATES_MEAN, 06_FITTING, 07_METHOD_COMPARISON", "Shown when":"always", "Notes":"Pseudo-absorbance is not assumed to be spectrophotometric absorbance."},
        {"Term":"PAbs_Red/PAbs_Green/PAbs_Blue", "Meaning":"RGB pseudo-absorbance for the red, green and blue channels", "Formula":"log10(MeanBG_channel / MeanW_channel)", "Unit":"dimensionless", "Where used":"FIGURE_RGB.png, 04_RAW, 05_REPLICATES_MEAN, 06_FITTING, 07_METHOD_COMPARISON", "Shown when":"always", "Notes":"Exported in standard RGB order even though OpenCV stores images internally as BGR."},
        {"Term":"MeanW_Red/MeanW_Green/MeanW_Blue", "Meaning":"Linearized median well intensity for each RGB channel", "Formula":"median ROI channel intensity after gamma linearization", "Unit":"dimensionless", "Where used":"04_RAW", "Shown when":"always", "Notes":""},
        {"Term":"MeanBG_Red/MeanBG_Green/MeanBG_Blue", "Meaning":"Linearized local inter-well background intensity for each RGB channel", "Formula":"2D background surface evaluated at the well and gamma-linearized", "Unit":"dimensionless", "Where used":"04_RAW", "Shown when":"always", "Notes":""},
        {"Term":"SignalT_Red/SignalT_Green/SignalT_Blue", "Meaning":"Transmittance-like ratio before logarithmic conversion", "Formula":"SignalT_channel = MeanW_channel / MeanBG_channel", "Unit":"dimensionless", "Where used":"04_RAW", "Shown when":"always", "Notes":"PAbs_channel = -log10(SignalT_channel)."},
        {"Term":"RGB linearization", "Meaning":"Conversion of gamma-domain RGB values to an intensity-like scale", "Formula":"I_lin = (I/255)^gamma, gamma approximately 2.2", "Unit":"dimensionless", "Where used":"PAbs calculation in 04_RAW and FIGURE_RGB.png", "Shown when":"always", "Notes":"No automatic image correction is applied."},
        {"Term":"RGB to CIELAB", "Meaning":"Diagnostic conversion from image RGB/BGR values to CIE L*a*b* descriptors", "Formula":"OpenCV BGR->Lab for 8-bit sRGB-like images; reference form: linearized sRGB -> XYZ using the standard sRGB-to-XYZ matrix and D65 reference white -> CIE L*a*b*", "Unit":"dimensionless", "Where used":"FIGURE_CIELAB_DELTAE.png, RAW_DATA_DETAILS_CAPTION.txt, 04_RAW, 05_REPLICATES_MEAN, DIAGNOSTICS", "Shown when":"CIELAB/DeltaE outputs present", "Notes":"References: OpenCV color conversions; IEC 61966-2-1:1999 sRGB; CIE 1976 L*a*b*; CIE standard illuminant D65."},
        {"Term":"L/a/b", "Meaning":"CIELAB lightness and opponent-color coordinates", "Formula":"median ROI value after RGB to CIELAB conversion", "Unit":"dimensionless", "Where used":"FIGURE_CIELAB_DELTAE.png, 04_RAW, 05_REPLICATES_MEAN, diagnostics", "Shown when":"CIELAB outputs present", "Notes":""},
        {"Term":"DeltaL/Deltaa/Deltab", "Meaning":"CIELAB coordinate differences from the selected CIELAB reference", "Formula":"DeltaL = L - L_ref; Deltaa = a - a_ref; Deltab = b - b_ref", "Unit":"dimensionless", "Where used":"FIGURE_CIELAB_DELTAE.png, 04_RAW, 05_REPLICATES_MEAN, diagnostics", "Shown when":"Delta outputs present", "Notes":"The reference source is reported in CIELAB_ref_source."},
        {"Term":"DeltaE_ab", "Meaning":"Total CIELAB color difference", "Formula":"sqrt(DeltaL^2 + Deltaa^2 + Deltab^2)", "Unit":"dimensionless", "Where used":"FIGURE_CIELAB_DELTAE.png, RAW_DATA_DETAILS_CAPTION.txt, 04_RAW, 05_REPLICATES_MEAN, 07_METHOD_COMPARISON, diagnostics", "Shown when":"DeltaE outputs present", "Notes":"Reference: CIE 1976 L*a*b* color-difference form."},
        {"Term":"DeltaE_ab_chroma", "Meaning":"Chromatic CIELAB color difference without lightness", "Formula":"sqrt(Deltaa^2 + Deltab^2)", "Unit":"dimensionless", "Where used":"FIGURE_CIELAB_DELTAE.png, RAW_DATA_DETAILS_CAPTION.txt, 04_RAW, 05_REPLICATES_MEAN, diagnostics", "Shown when":"DeltaE outputs present", "Notes":""},
        {"Term":"CIELAB_ref_source", "Meaning":"Source of the CIELAB reference used to compute Delta variables", "Formula":"zero calibration or other reported fallback source", "Unit":"text", "Where used":"04_RAW, 05_REPLICATES_MEAN", "Shown when":"CIELAB/DeltaE outputs present", "Notes":""},
        {"Term":"IRLS", "Meaning":"Iteratively reweighted least-squares robust linear regression with residual-based weights", "Formula":"minimize sum_i w_i (y_i - (m x_i + q))^2; w_i is updated iteratively from residual magnitude", "Unit":"dimensionless", "Where used":"FIGURE_RGB.png, FIGURE_CIELAB_DELTAE.png, RESULTS_CAPTION.txt, RAW_DATA_DETAILS_CAPTION.txt, 06_FITTING, 07_METHOD_COMPARISON, diagnostic fitting tables", "Shown when":"always", "Notes":"Reference: Huber, P. J. (1964), Robust Estimation of a Location Parameter. Implementation: custom NumPy-based IRLS using repeated least-squares solves."},
        {"Term":"m/q", "Meaning":"Slope and intercept of the linear fit", "Formula":"y = m x + q", "Unit":"response / concentration and response", "Where used":"FIGURE_RGB.png, FIGURE_CIELAB_DELTAE.png, 06_FITTING", "Shown when":"fit rows present", "Notes":""},
        {"Term":"R2", "Meaning":"Coefficient of determination", "Formula":"1 - SSE/SST", "Unit":"dimensionless", "Where used":"FIGURE_RGB.png, FIGURE_CIELAB_DELTAE.png, 06_FITTING, 07_METHOD_COMPARISON", "Shown when":"fit rows present", "Notes":""},
        {"Term":"RMSE", "Meaning":"Root mean squared error of the fit", "Formula":"sqrt(mean(residual^2))", "Unit":"response unit", "Where used":"06_FITTING", "Shown when":"fit rows present", "Notes":""},
        {"Term":"LOD/LOQ", "Meaning":"Detection and quantification limits from calibration", "Formula":"LOD = 3 x sigma_cal / |m|; LOQ = 10 x sigma_cal / |m|", "Unit":unit_label, "Where used":"FIGURE_RGB.png, FIGURE_CIELAB_DELTAE.png, 03_OVERVIEW, 06_FITTING, 07_METHOD_COMPARISON", "Shown when":"calibration present", "Notes":""},
        {"Term":"sigma_cal/sigma_source/SNR", "Meaning":"Calibration noise estimate, its source and slope-to-noise ratio", "Formula":"SNR = |m| / sigma_cal", "Unit":"mixed", "Where used":"06_FITTING, 07_METHOD_COMPARISON", "Shown when":"calibration present", "Notes":"SNR is diagnostic only and is not used in the common score."},
        {"Term":"C0/C0_sd", "Meaning":"Estimated original-sample concentration and associated uncertainty", "Formula":"standard addition: DF x q/m; calibration projection: DF x (y - q)/m; epsilon mode: DF x PAbs/(epsilon x l_cm), converted from M to the selected unit", "Unit":unit_label, "Where used":"FIGURE_RGB.png, 03_OVERVIEW, 06_FITTING", "Shown when":"standard addition, unknown or epsilon quantification present", "Notes":"FitType identifies which calculation path produced C0."},
        {"Term":"Score", "Meaning":"Common method-ranking score", "Formula":"for calibration plus standard addition: SlopeAgreement^2 x sqrt(R2_cal x R2_std) x (1/LOQ)", "Unit":"1/" + str(unit_label), "Where used":"FIGURE_RGB.png, METHOD_COMPARISON.png, 07_METHOD_COMPARISON", "Shown when":"method comparison present", "Notes":"Expected/reference values, recovery, SNR and clipping are not used in this score."},
        {"Term":"SlopeAgreement", "Meaning":"Agreement between calibration and standard-addition slope magnitudes", "Formula":"min(|m_cal|, |m_std|) / max(|m_cal|, |m_std|)", "Unit":"dimensionless", "Where used":"METHOD_COMPARISON.png, 07_METHOD_COMPARISON", "Shown when":"calibration plus standard addition present", "Notes":""},
        {"Term":"beta_k/beta_mean", "Meaning":"Standard-addition/calibration slope ratio", "Formula":"m_std / m_cal; beta_mean is the mean across available standard-addition fits", "Unit":"dimensionless", "Where used":"06_FITTING, 07_METHOD_COMPARISON", "Shown when":"calibration plus standard addition present", "Notes":""},
        {"Term":"bias_index_k/bias_index_mean", "Meaning":"Relative slope-bias index", "Formula":"|m_std/m_cal - 1|; bias_index_mean is the mean across available fits", "Unit":"dimensionless", "Where used":"06_FITTING, 07_METHOD_COMPARISON", "Shown when":"calibration plus standard addition present", "Notes":""},
        {"Term":"NClipPoints/ClipX/ClipDelta", "Meaning":"Calibration clipping/baseline diagnostic fields", "Formula":"NClipPoints = number of adjusted points; ClipX = concentrations; ClipDelta = applied response adjustments", "Unit":"count, concentration, response", "Where used":"06_FITTING", "Shown when":"calibration present", "Notes":"Diagnostic only; clipping is not part of the common method score."},
        {"Term":"Method/Family/ComparableGroup/RankMode", "Meaning":"Method-comparison identifiers and comparable-score grouping", "Formula":"text labels defining descriptor, family and score comparability", "Unit":"text", "Where used":"METHOD_COMPARISON.png, 07_METHOD_COMPARISON", "Shown when":"method comparison present", "Notes":"Scores should be compared directly only within the same ComparableGroup."},
        {"Term":"Selected/Rank", "Meaning":"Selected method flag and rank position", "Formula":"Selected = 1 for the selected method; Rank = score order within the highest common-factor group", "Unit":"0/1 and integer", "Where used":"07_METHOD_COMPARISON", "Shown when":"method comparison present", "Notes":""},
        {"Term":"R2_cal/R2_std_mean/m_cal/m_std_mean", "Meaning":"Calibration and standard-addition fit descriptors used in method comparison", "Formula":"R2_cal from calibration fit; R2_std_mean = mean standard-addition R2; m_cal and m_std_mean are fit slopes", "Unit":"mixed", "Where used":"METHOD_COMPARISON.png, 07_METHOD_COMPARISON", "Shown when":"method comparison present", "Notes":""},
        {"Term":"Estimate_value/Estimate_sd/Estimate_source", "Meaning":"Representative concentration estimate, associated SD and source", "Formula":"selected according to Estimate_source", "Unit":unit_label + " and text", "Where used":"METHOD_COMPARISON.png, 07_METHOD_COMPARISON", "Shown when":"method comparison present", "Notes":""},
        {"Term":"expected_*/reference_*", "Meaning":"External reference metadata configured by the user", "Formula":"user/configurator input", "Unit":"text or " + str(unit_label), "Where used":"METHOD_COMPARISON.png, 03_OVERVIEW, 07_METHOD_COMPARISON", "Shown when":"reference values configured", "Notes":"External reference values are checks only and are not used for ranking."},
        {"Term":"delta_expected_*/delta_reference_*", "Meaning":"Difference between estimate and external reference", "Formula":"estimate - reference", "Unit":unit_label, "Where used":"METHOD_COMPARISON.png, 03_OVERVIEW, 07_METHOD_COMPARISON", "Shown when":"reference values configured", "Notes":""},
        {"Term":"recovery_pct_*", "Meaning":"Recovery relative to external reference", "Formula":"100 x estimate / reference", "Unit":"%", "Where used":"FIGURE_RGB.png, METHOD_COMPARISON.png, 03_OVERVIEW, 07_METHOD_COMPARISON", "Shown when":"reference values configured", "Notes":""},
        {"Term":"rel_error_*", "Meaning":"Relative error versus external reference", "Formula":"100 x (estimate - reference) / reference", "Unit":"%", "Where used":"03_OVERVIEW, 07_METHOD_COMPARISON", "Shown when":"reference values configured", "Notes":""},
        {"Term":"ImageWarning", "Meaning":"Well-level optical QC warning flag", "Formula":"1 if optical QC rules flag the well, else 0", "Unit":"0/1", "Where used":"04_RAW", "Shown when":"always", "Notes":""},
        {"Term":"QCFlagged/QCCritical", "Meaning":"Replicate-group QC warning and critical flags", "Formula":"rule-based aggregation of well-level warnings, critical wells and UsedFraction", "Unit":"0/1", "Where used":"05_REPLICATES_MEAN", "Shown when":"always", "Notes":""},
        {"Term":"NReplicates", "Meaning":"Number of replicate wells in a summarized group", "Formula":"count of wells in group", "Unit":"count", "Where used":"05_REPLICATES_MEAN", "Shown when":"always", "Notes":""},
        {"Term":"Geometry and epsilon/path-length quantification", "Meaning":"Assumption used when epsilon-based unknown quantification is enabled", "Formula":"l_cm = (volume_uL / well_bottom_area_mm2) / 10; C_M = PAbs / (epsilon x l_cm)", "Unit":"cm and M", "Where used":"RESULTS_CAPTION.txt, RAW_DATA_DETAILS_CAPTION.txt, REPORT", "Shown when":"epsilon/path-length mode configured", "Notes":"Assumes ANSI/SLAS-compatible flat-bottom microplate geometry; non-flat or non-certified geometries require separate validation."},
        {"Term":"Code implementation", "Meaning":"Main computational libraries and code provenance", "Formula":"custom Python code using NumPy, OpenCV, openpyxl and matplotlib", "Unit":"software provenance", "Where used":"all generated outputs", "Shown when":"always", "Notes":"Sources/libraries: NumPy for least-squares solves; OpenCV for image masks/color conversion; openpyxl for xlsx export; matplotlib for PNG figures."},
    ])
    return _finalize_legend_rows(rows)
def _prepare_output_dirs(run_dir, save_raw_data_details=False):
    results_dir = os.path.join(run_dir, "RESULTS")
    os.makedirs(results_dir, exist_ok=True)
    diagnostics_dir = None
    if bool(save_raw_data_details):
        diagnostics_dir = os.path.join(run_dir, "RAW_DATA_DETAILS")
        os.makedirs(diagnostics_dir, exist_ok=True)
    return results_dir, diagnostics_dir




def _method_family_from_channel(channel):
    ch = str(channel or "")
    if ch.startswith("Signal_"):
        return "RGB", ch.replace("Signal_", "PAbs_", 1)
    if ch.startswith("PAbs_"):
        return "RGB", ch
    if ch in {"L", "a", "b"}:
        return "CIELAB", ch
    if ch.startswith("Delta"):
        return "DeltaCIELAB", ch
    if ch in {"Red", "Green", "Blue"}:
        return "RGB", "PAbs_" + ch
    return "other", ch


def _build_expected_reference_rows(cfg=None):
    cfg = cfg or {}
    rows = []
    seen = set()

    refs = cfg.get("expected_refs", None)
    if isinstance(refs, list):
        for ref in refs:
            if not isinstance(ref, dict):
                continue
            ref_id = str(ref.get("id", ref.get("ref_id", ""))).strip()
            label = str(ref.get("label", "")).strip()
            try:
                value = float(ref.get("value", np.nan))
            except Exception:
                value = np.nan
            try:
                sd = float(ref.get("sd", np.nan))
            except Exception:
                sd = np.nan
            if not np.isfinite(value):
                continue
            key = (ref_id, label, float(value), float(sd) if np.isfinite(sd) else None)
            if key in seen:
                continue
            seen.add(key)
            rows.append({"id": ref_id, "label": label, "value": float(value), "sd": float(sd) if np.isfinite(sd) else np.nan})

    candidates = [
        ("ICP-MS", cfg.get("expected_icpms_value", cfg.get("expected_icpms", np.nan)), cfg.get("expected_icpms_sd", np.nan)),
        ("Colorimetry", cfg.get("expected_colorimetry_value", cfg.get("expected_colorimetry", np.nan)), cfg.get("expected_colorimetry_sd", np.nan)),
        (str(cfg.get("expected_label", "Expected")).strip(), cfg.get("expected_value", np.nan), cfg.get("expected_sd", np.nan)),
    ]
    for label, value, sd in candidates:
        try:
            value = float(value)
        except Exception:
            value = np.nan
        try:
            sd = float(sd)
        except Exception:
            sd = np.nan
        if not np.isfinite(value):
            continue
        key = (str(label), float(value), float(sd) if np.isfinite(sd) else None)
        if key in seen:
            continue
        seen.add(key)
        rows.append({"id": "", "label": str(label), "value": float(value), "sd": float(sd) if np.isfinite(sd) else np.nan})
    return rows


def _format_reference_label(ref, index=None):
    label = str((ref or {}).get("label", "")).strip()
    return label or (f"Reference {index}" if index is not None else "Reference")


def _reference_id(ref):
    return str((ref or {}).get("id", (ref or {}).get("ref_id", ""))).strip()


def _reference_matches_fit_row(ref, fit_row):
    """Match a reference to a result row by Ref ID when Ref ID is provided.

    Legacy references without Ref ID are kept global for backward compatibility.
    """
    rid = _reference_id(ref)
    if not rid:
        return True
    return rid == str((fit_row or {}).get("ID", "")).strip()


def _reference_style(index):
    styles = [
        {"color": "tab:blue", "marker": "D", "facecolor": "white"},
        {"color": "tab:orange", "marker": "s", "facecolor": "white"},
        {"color": "tab:green", "marker": "^", "facecolor": "white"},
        {"color": "tab:red", "marker": "v", "facecolor": "white"},
        {"color": "tab:purple", "marker": "P", "facecolor": "white"},
        {"color": "tab:brown", "marker": "X", "facecolor": "white"},
    ]
    idx = max(0, int(index or 1) - 1) % len(styles)
    return dict(styles[idx])


def _selected_channel_reference_rows(fit_rows, selected_channel, expected_refs=None):
    expected_refs = expected_refs or []
    selected_name = str(selected_channel or "")
    selected_rows = [
        dict(r) for r in (fit_rows or [])
        if str(r.get("Channel", "")) == selected_name
        and str(r.get("FitType", "")) in {"StdAdd", "UnknownFromCal", "UnknownFromEpsilon"}
        and np.isfinite(r.get("C0", np.nan))
    ]
    out = []
    for row in selected_rows:
        c0 = float(row.get("C0", np.nan)) if np.isfinite(row.get("C0", np.nan)) else np.nan
        c0_sd = float(row.get("C0_sd", np.nan)) if np.isfinite(row.get("C0_sd", np.nan)) else np.nan
        fit_type = str(row.get("FitType", ""))
        method_label = "Std add" if fit_type == "StdAdd" else ("Unknown from epsilon" if fit_type == "UnknownFromEpsilon" else "Unknown from calibration")
        base = {
            "Channel": selected_name,
            "FitType": fit_type,
            "MethodLabel": method_label,
            "ID": str(row.get("ID", "")),
            "DF": row.get("DF", np.nan),
            "C0": c0,
            "C0_sd": c0_sd,
        }
        if expected_refs:
            for i_ref, ref in enumerate(expected_refs, start=1):
                if not _reference_matches_fit_row(ref, row):
                    continue
                ref_label = _format_reference_label(ref, i_ref)
                ref_safe = _expected_ref_key_label(ref_label, i_ref)
                ref_val = float(ref.get("value", np.nan)) if np.isfinite(ref.get("value", np.nan)) else np.nan
                ref_sd = _num_or_nan(ref.get("sd", np.nan))
                delta = (c0 - ref_val) if np.isfinite(c0) and np.isfinite(ref_val) else np.nan
                recovery = (100.0 * c0 / ref_val) if np.isfinite(c0) and np.isfinite(ref_val) and abs(ref_val) > 1e-15 else np.nan
                z = (delta / ref_sd) if np.isfinite(delta) and np.isfinite(ref_sd) and ref_sd > 0 else np.nan
                z_combined = (delta / np.sqrt(ref_sd * ref_sd + c0_sd * c0_sd)) if np.isfinite(delta) and np.isfinite(ref_sd) and ref_sd > 0 and np.isfinite(c0_sd) and c0_sd >= 0 else np.nan
                base[f"reference_label_{ref_safe}"] = ref_label
                base[f"reference_value_{ref_safe}"] = ref_val
                base[f"reference_sd_{ref_safe}"] = ref_sd
                base[f"delta_reference_{ref_safe}"] = delta
                base[f"recovery_pct_{ref_safe}"] = recovery
        out.append(base)
    return out


def _project_reference_to_plot_x(ref_value, ref_sd, fit_type, df):
    try:
        rv = float(ref_value)
    except Exception:
        rv = np.nan
    try:
        rsd = float(ref_sd)
    except Exception:
        rsd = np.nan
    try:
        dfv = float(df)
    except Exception:
        dfv = np.nan
    if not np.isfinite(rv):
        return np.nan, np.nan
    if np.isfinite(dfv) and abs(dfv) > 1e-15:
        x_val = rv / dfv
        x_sd = (rsd / abs(dfv)) if np.isfinite(rsd) else np.nan
    else:
        x_val = rv
        x_sd = rsd
    if str(fit_type or "") == "StdAdd":
        x_val = -x_val
    return float(x_val), float(x_sd) if np.isfinite(x_sd) else np.nan


def _reference_positions_for_channel(fit_rows, channel_label, expected_refs=None):
    expected_refs = expected_refs or []
    rows = [
        dict(r) for r in (fit_rows or [])
        if str(r.get("Channel", "")) == str(channel_label or "")
        and str(r.get("FitType", "")) in {"StdAdd", "UnknownFromCal", "UnknownFromEpsilon"}
        and np.isfinite(r.get("C0", np.nan))
    ]
    out = []
    for row in rows:
        fit_type = str(row.get("FitType", ""))
        df = row.get("DF", np.nan)
        item = {
            "FitType": fit_type,
            "ID": str(row.get("ID", "")),
            "DF": df,
            "references": [],
        }
        for i_ref, ref in enumerate(expected_refs, start=1):
            if not _reference_matches_fit_row(ref, row):
                continue
            ref_label = _format_reference_label(ref, i_ref)
            ref_val = float(ref.get("value", np.nan)) if np.isfinite(ref.get("value", np.nan)) else np.nan
            ref_sd = _num_or_nan(ref.get("sd", np.nan))
            x_val, x_sd = _project_reference_to_plot_x(ref_val, ref_sd, fit_type, df)
            if np.isfinite(x_val):
                item["references"].append({
                    "label": ref_label,
                    "value": ref_val,
                    "sd": ref_sd,
                    "x": x_val,
                    "x_sd": x_sd,
                    "index": i_ref,
                    "style": _reference_style(i_ref),
                })
        if item["references"]:
            out.append(item)
    return out


def _reference_info_lines(fit_rows, selected_channel, unit_label, expected_refs=None):
    expected_refs = expected_refs or []
    if not expected_refs:
        return []
    lines = ["REFERENCE VALUES"]
    for i_ref, ref in enumerate(expected_refs, start=1):
        label = _format_reference_label(ref, i_ref)
        val = float(ref.get("value", np.nan)) if np.isfinite(ref.get("value", np.nan)) else np.nan
        sd = _num_or_nan(ref.get("sd", np.nan))
        if np.isfinite(val):
            line = f"{label}: {val:.6g}"
            if np.isfinite(sd):
                line += f" ± {sd:.6g}"
            line += f" {unit_label}"
            lines.append(line)
    return lines


def _expected_ref_key_label(label, index=None):
    base = str(label or "").strip() or (f"Expected_{index}" if index is not None else "Expected")
    safe = []
    for ch in base:
        if ch.isalnum() or ch == "_":
            safe.append(ch)
        elif ch in {" ", "-", "/", "\\"}:
            safe.append("_")
    safe = "".join(safe).strip("_")
    while "__" in safe:
        safe = safe.replace("__", "_")
    return safe or (f"Expected_{index}" if index is not None else "Expected")


def _build_method_comparison_rows(fit_rows, selection_info=None, expected_refs=None):
    selection_info = selection_info or {}
    expected_refs = expected_refs or []
    ranking = selection_info.get("ranking", []) or []
    ranking_map = {str(r.get("Channel", "")): dict(r) for r in ranking}

    fit_map = {}
    for row in fit_rows or []:
        ch = str(row.get("Channel", ""))
        if ch not in fit_map:
            fit_map[ch] = {"Calibration": [], "StdAdd": [], "UnknownFromCal": [], "UnknownFromEpsilon": []}
        fit_map[ch].setdefault(str(row.get("FitType", "")), []).append(row)

    out = []
    for ch in sorted(fit_map.keys()):
        fam, comp = _method_family_from_channel(ch)
        cal_rows = fit_map[ch].get("Calibration", []) or []
        std_rows = [r for r in (fit_map[ch].get("StdAdd", []) or []) if _isfinite_num(r.get("m", np.nan))]
        unk_rows = [r for r in ((fit_map[ch].get("UnknownFromCal", []) or []) + (fit_map[ch].get("UnknownFromEpsilon", []) or [])) if _isfinite_num(r.get("C0", np.nan))]
        cal = cal_rows[0] if cal_rows else {}

        m_cal = _num_or_nan(cal.get("m", np.nan))
        r2_cal = _num_or_nan(cal.get("R2", np.nan))
        lod = _num_or_nan(cal.get("LOD", np.nan))
        loq = _num_or_nan(cal.get("LOQ", np.nan))
        snr = _num_or_nan(cal.get("SNR", np.nan))

        m_std_vals = np.asarray([float(r.get("m", np.nan)) for r in std_rows if np.isfinite(r.get("m", np.nan))], dtype=np.float64)
        r2_std_vals = np.asarray([float(r.get("R2", np.nan)) for r in std_rows if np.isfinite(r.get("R2", np.nan))], dtype=np.float64)
        c0_vals = np.asarray([float(r.get("C0", np.nan)) for r in std_rows if np.isfinite(r.get("C0", np.nan))], dtype=np.float64)
        c0_sd_vals = np.asarray([float(r.get("C0_sd", np.nan)) for r in std_rows if np.isfinite(r.get("C0_sd", np.nan))], dtype=np.float64)
        beta_vals = np.asarray([float(r.get("beta_k", np.nan)) for r in std_rows if np.isfinite(r.get("beta_k", np.nan))], dtype=np.float64)
        bias_vals = np.asarray([float(r.get("bias_index_k", np.nan)) for r in std_rows if np.isfinite(r.get("bias_index_k", np.nan))], dtype=np.float64)

        m_std_mean = float(np.mean(m_std_vals)) if m_std_vals.size else np.nan
        r2_std_mean = float(np.mean(r2_std_vals)) if r2_std_vals.size else np.nan
        c0_mean = float(np.mean(c0_vals)) if c0_vals.size else np.nan
        c0_median = float(np.median(c0_vals)) if c0_vals.size else np.nan
        c0_sd_median = float(np.median(c0_sd_vals)) if c0_sd_vals.size else np.nan
        unk_c0_vals = np.asarray([float(r.get("C0", np.nan)) for r in unk_rows if np.isfinite(r.get("C0", np.nan))], dtype=np.float64)
        unk_c0_sd_vals = np.asarray([float(r.get("C0_sd", np.nan)) for r in unk_rows if np.isfinite(r.get("C0_sd", np.nan))], dtype=np.float64)
        estimate_value = float(np.median(unk_c0_vals)) if unk_c0_vals.size else (float(np.median(c0_vals)) if c0_vals.size else np.nan)
        estimate_sd = float(np.median(unk_c0_sd_vals)) if unk_c0_sd_vals.size else (float(np.median(c0_sd_vals)) if c0_sd_vals.size else np.nan)
        estimate_source = "unknown_from_calibration" if unk_c0_vals.size else ("standard_addition" if c0_vals.size else "")
        beta_mean = float(np.mean(beta_vals)) if beta_vals.size else np.nan
        bias_mean = float(np.mean(bias_vals)) if bias_vals.size else np.nan

        slope_agreement = np.nan
        if np.isfinite(m_cal) and m_std_vals.size:
            vals = []
            for mstd in m_std_vals:
                denom = max(abs(m_cal), abs(mstd), 1e-12)
                vals.append(min(abs(m_cal), abs(mstd)) / denom)
            slope_agreement = float(np.mean(np.asarray(vals, dtype=np.float64))) if vals else np.nan

        rank_row = ranking_map.get(ch, {})
        score = _num_or_nan(rank_row.get("Score", np.nan))
        rank_mode = str(rank_row.get("Mode", ""))
        if not np.isfinite(score):
            if np.isfinite(r2_cal) and np.isfinite(r2_std_mean) and np.isfinite(slope_agreement):
                score = _compute_balanced_score(r2_cal, r2_std_mean, slope_agreement, loq=loq)
                rank_mode = "calibration_plus_stdadd"
            elif np.isfinite(r2_cal) and np.isfinite(m_cal):
                score = float((r2_cal ** 2) * abs(m_cal))
                rank_mode = "calibration_only"
            elif np.isfinite(r2_std_mean) and np.isfinite(m_std_mean):
                score = float((r2_std_mean ** 2) * abs(m_std_mean))
                rank_mode = "stdadd_only"

        out_row = {
            "Method": ch,
            "Family": fam,
            "Component": comp,
            "Score": score,
            "RankMode": rank_mode,
            "R2_cal": r2_cal,
            "R2_std_mean": r2_std_mean,
            "m_cal": m_cal,
            "m_std_mean": m_std_mean,
            "SlopeAgreement": slope_agreement,
            "beta_mean": beta_mean,
            "bias_index_mean": bias_mean,
            "SNR": snr,
            "LOD": lod,
            "LOQ": loq,
            "n_stdadd": int(c0_vals.size),
            "n_unknown": int(len(unk_rows)),
            "C0_mean": c0_mean,
            "C0_median": c0_median,
            "C0_sd_median": c0_sd_median,
            "Estimate_value": estimate_value,
            "Estimate_sd": estimate_sd,
            "Estimate_source": estimate_source,
        }
        for i_ref, ref in enumerate(expected_refs, start=1):
            label = str(ref.get("label", "")).strip() or f"Expected_{i_ref}"
            safe = _expected_ref_key_label(label, i_ref)
            value = _num_or_nan(ref.get("value", np.nan))
            sd = _num_or_nan(ref.get("sd", np.nan))
            out_row[f"expected_label_{safe}"] = label
            out_row[f"expected_id_{safe}"] = _reference_id(ref)
            out_row[f"expected_value_{safe}"] = value
            out_row[f"expected_sd_{safe}"] = sd
            rid = _reference_id(ref)
            if rid:
                matched = [r for r in (unk_rows + std_rows) if _reference_matches_fit_row(ref, r) and np.isfinite(r.get("C0", np.nan))]
                matched_vals = np.asarray([float(r.get("C0", np.nan)) for r in matched], dtype=np.float64)
                ref_basis = float(np.median(matched_vals)) if matched_vals.size else np.nan
            else:
                ref_basis = estimate_value if np.isfinite(estimate_value) else c0_mean
            out_row[f"estimate_for_expected_{safe}"] = ref_basis
            out_row[f"delta_expected_{safe}"] = (ref_basis - value) if np.isfinite(ref_basis) and np.isfinite(value) else np.nan
            out_row[f"recovery_pct_{safe}"] = (100.0 * ref_basis / value) if np.isfinite(ref_basis) and np.isfinite(value) and abs(value) > 1e-15 else np.nan
            out_row[f"rel_error_{safe}"] = ((ref_basis - value) / value) if np.isfinite(ref_basis) and np.isfinite(value) and abs(value) > 1e-15 else np.nan
        out.append(out_row)

    # Common global score for method comparison.  The ranking deliberately uses
    # only factors shared by RGB and CIELAB/Delta methods: R2_cal, R2_std and
    # slope agreement.  RGB-specific SNR, RGB clipping and expected/recovery
    # values are diagnostics only and do not enter Score.
    for r in out:
        if np.isfinite(r.get("R2_cal", np.nan)) and np.isfinite(r.get("R2_std_mean", np.nan)) and np.isfinite(r.get("SlopeAgreement", np.nan)):
            base_score = _compute_fit_base_score(r.get("R2_cal", np.nan), r.get("R2_std_mean", np.nan), r.get("SlopeAgreement", np.nan), loq=r.get("LOQ", np.nan))
            base_score = float(base_score) if np.isfinite(base_score) and base_score > 0 else 0.0
            r["BaseScore"] = base_score
            r["Score"] = base_score
            r["FinalScore"] = base_score
            r["ComparableGroup"] = "calibration_plus_stdadd"
            r["CommonFactorsN"] = 3
            r["ScoreFormula"] = "slope_agreement^2 * sqrt(R2_cal * R2_std) * (1/LOQ)"
        elif np.isfinite(r.get("R2_cal", np.nan)):
            sc = max(float(r.get("R2_cal", 0.0)), 0.0)
            r["BaseScore"] = sc
            r["Score"] = sc
            r["FinalScore"] = sc
            r["ComparableGroup"] = "calibration_only"
            r["CommonFactorsN"] = 1
            r["ScoreFormula"] = "R2_cal"
        elif np.isfinite(r.get("R2_std_mean", np.nan)):
            sc = max(float(r.get("R2_std_mean", 0.0)), 0.0)
            r["BaseScore"] = sc
            r["Score"] = sc
            r["FinalScore"] = sc
            r["ComparableGroup"] = "stdadd_only"
            r["CommonFactorsN"] = 1
            r["ScoreFormula"] = "R2_std_mean"
        else:
            r["BaseScore"] = 0.0
            r["Score"] = 0.0
            r["FinalScore"] = 0.0
            r["ComparableGroup"] = "not_ranked"
            r["CommonFactorsN"] = 0
            r["ScoreFormula"] = "not_ranked"

    def _sort_key(row):
        factors = int(row.get("CommonFactorsN", 0) or 0)
        score = row.get("Score", np.nan)
        score_key = -float(score) if np.isfinite(score) else 1e12
        return (-factors, score_key, str(row.get("Family", "")), str(row.get("Method", "")))
    return sorted(out, key=_sort_key)


def _make_method_comparison_png(path, image_basename, comparison_rows, unit_label, expected_refs=None):
    """
    Compact method-comparison plot.

    The reference-value panel deliberately uses a robust y-scale around the
    external reference(s) and the non-flagged estimates. Very distant estimates
    are not allowed to expand the axis; they are clipped to the panel edge and
    labelled as "out of scale". This avoids the large blank plots generated when
    one channel produces a numerically huge but analytically unreliable C0.
    """
    comparison_rows = [
        dict(r) for r in (comparison_rows or [])
        if np.isfinite(_num_or_nan(r.get("Estimate_value", r.get("C0_mean", np.nan))))
    ]
    expected_refs = expected_refs or []
    if (not path) or len(comparison_rows) == 0:
        return

    matplotlib.rcParams.update({
        "font.family": "DejaVu Sans",
        "font.size": 8.5,
        "axes.titlesize": 10.0,
        "axes.labelsize": 9.0,
        "xtick.labelsize": 7.5,
        "ytick.labelsize": 7.5,
    })

    methods = [str(r.get("Method", "")).replace("Signal_", "PAbs_") for r in comparison_rows]
    x = np.arange(len(methods), dtype=np.float64)
    slope = np.asarray([_num_or_nan(r.get("SlopeAgreement", np.nan)) for r in comparison_rows], dtype=np.float64)
    bias = np.asarray([_num_or_nan(r.get("bias_index_mean", np.nan)) for r in comparison_rows], dtype=np.float64)
    r2_cal = np.asarray([_num_or_nan(r.get("R2_cal", np.nan)) for r in comparison_rows], dtype=np.float64)
    r2_std = np.asarray([_num_or_nan(r.get("R2_std_mean", np.nan)) for r in comparison_rows], dtype=np.float64)
    estimate = np.asarray([_num_or_nan(r.get("Estimate_value", r.get("C0_mean", np.nan))) for r in comparison_rows], dtype=np.float64)
    estimate_sd = np.asarray([_num_or_nan(r.get("Estimate_sd", r.get("C0_sd_median", np.nan))) for r in comparison_rows], dtype=np.float64)

    def _panel_has_data(*arrays):
        for arr in arrays:
            aa = np.asarray(arr, dtype=np.float64)
            if np.any(np.isfinite(aa)):
                return True
        return False

    def _nearest_reference_payload(val):
        best = None
        for i_ref, ref in enumerate(expected_refs, start=1):
            ref_val = _num_or_nan(ref.get('value', np.nan))
            ref_sd = _num_or_nan(ref.get('sd', np.nan))
            if not np.isfinite(val) or not np.isfinite(ref_val):
                continue
            delta = float(val - ref_val)
            ad = abs(delta)
            if best is None or ad < best['abs_delta']:
                best = {
                    'index': i_ref,
                    'label': str(ref.get('label', '')).strip() or f'Reference {i_ref}',
                    'value': ref_val,
                    'sd': ref_sd,
                    'delta': delta,
                    'abs_delta': ad,
                }
        return best

    panel1_ok = _panel_has_data(slope, bias)
    panel2_ok = (len(expected_refs) > 0) and _panel_has_data(estimate)
    panel3_ok = _panel_has_data(r2_cal, r2_std)
    if not (panel1_ok or panel2_ok or panel3_ok):
        return

    panel_specs = []
    if panel1_ok:
        panel_specs.append(('agreement_bias', 1.0))
    if panel2_ok:
        panel_specs.append(('reference_values', 1.05))
    if panel3_ok:
        panel_specs.append(('r2', 1.0))

    # Compact aspect ratio restored: enough space for labels, no excess vertical whitespace.
    fig_h = 2.15 * len(panel_specs) + 0.55
    fig = plt.figure(figsize=(11.5, fig_h), dpi=220)
    gs = fig.add_gridspec(len(panel_specs), 1, height_ratios=[h for _, h in panel_specs], hspace=0.0)
    axes = []
    share_ax = None
    for i, (name, _) in enumerate(panel_specs):
        ax = fig.add_subplot(gs[i, 0], sharex=share_ax)
        if share_ax is None:
            share_ax = ax
        axes.append((name, ax))
    ax_map = {name: ax for name, ax in axes}

    if panel1_ok:
        ax1 = ax_map['agreement_bias']
        mask_slope = np.isfinite(slope)
        mask_bias = np.isfinite(bias)
        if np.any(mask_slope):
            line_slope, = ax1.plot(x[mask_slope], slope[mask_slope], 'o-', label='slope agreement')
            ax1.axhline(1.0, linestyle='--', linewidth=1.0, color=line_slope.get_color(), alpha=0.8)
        if np.any(mask_bias):
            line_bias, = ax1.plot(x[mask_bias], bias[mask_bias], 's--', label='bias index')
            ax1.axhline(0.0, linestyle='--', linewidth=1.0, color=line_bias.get_color(), alpha=0.8)
        ax1.set_ylabel('agreement / bias', fontweight='bold')
        ax1.yaxis.set_label_coords(-0.065, 0.5)
        ax1.grid(True, axis='y', alpha=0.25)
        handles1, labels1 = ax1.get_legend_handles_labels()
        if handles1:
            ax1.legend(handles1, labels1, loc='best', frameon=False)
        ax1.tick_params(labelbottom=False)

    if panel2_ok:
        ax2 = ax_map['reference_values']
        ref_vals = []
        ref_sds = []
        for i_ref, ref in enumerate(expected_refs, start=1):
            val = _num_or_nan(ref.get('value', np.nan))
            sd = _num_or_nan(ref.get('sd', np.nan))
            lab = str(ref.get('label', '')).strip() or f'Reference {i_ref}'
            style = _reference_style(i_ref)
            if np.isfinite(val):
                ref_vals.append(float(val))
                ax2.axhline(val, linestyle='--', linewidth=1.0, color=style['color'], label=f'{lab} = {val:.3g}')
                if np.isfinite(sd) and sd > 0:
                    ref_sds.append(float(sd))
                    ax2.axhspan(val - sd, val + sd, alpha=0.10, color=style['color'])

        low_mask = np.asarray([_comparison_low_reliability(r, expected_refs=expected_refs) for r in comparison_rows], dtype=bool)
        finite_est = np.isfinite(estimate)
        reliable_mask = finite_est & (~low_mask)

        # Robust axis: start from reference(s), include only non-flagged values that are
        # not grossly outside a reference-centered window. This keeps the panel compact.
        if ref_vals:
            ref_arr = np.asarray(ref_vals, dtype=np.float64)
            ref_center = float(np.nanmedian(ref_arr))
            ref_range = float(np.nanmax(ref_arr) - np.nanmin(ref_arr)) if len(ref_arr) > 1 else 0.0
            max_sd = float(np.nanmax(ref_sds)) if ref_sds else max(1.0, 0.03 * max(abs(ref_center), 1.0))
            robust_window = max(20.0, 0.75 * max(abs(ref_center), 1.0), 6.0 * max_sd, 2.0 * ref_range)
            axis_vals = []
            for val, sdv in zip(ref_vals, [max_sd] * len(ref_vals)):
                axis_vals.extend([val - sdv, val, val + sdv])
            for idx in np.where(reliable_mask)[0]:
                yv = float(estimate[idx])
                ev = float(estimate_sd[idx]) if np.isfinite(estimate_sd[idx]) else 0.0
                if abs(yv - ref_center) <= robust_window:
                    axis_vals.extend([yv - ev, yv, yv + ev])
            if len(axis_vals) < 2:
                axis_vals = [ref_center - robust_window, ref_center + robust_window]
            y_min = float(np.nanmin(axis_vals))
            y_max = float(np.nanmax(axis_vals))
            span = max(y_max - y_min, 1e-6)
            pad = 0.16 * span
            y_lo = y_min - pad
            y_hi = y_max + pad
        else:
            finite_vals = estimate[finite_est]
            if finite_vals.size:
                qlo, qhi = np.nanpercentile(finite_vals, [10, 90])
                span = max(float(qhi - qlo), 1.0)
                y_lo = float(qlo - 0.25 * span)
                y_hi = float(qhi + 0.25 * span)
            else:
                y_lo, y_hi = 0.0, 1.0
        if not np.isfinite(y_lo) or not np.isfinite(y_hi) or y_hi <= y_lo:
            y_lo, y_hi = 0.0, 1.0
        ax2.set_ylim(y_lo, y_hi)

        shown_est_legend = False
        shown_unrel_legend = False
        shown_clip_legend = False
        yrange = max(y_hi - y_lo, 1.0)
        text_dy = 0.025 * yrange
        for i, xv in enumerate(x):
            yv = estimate[i]
            if not np.isfinite(yv):
                continue
            ev = float(estimate_sd[i]) if np.isfinite(estimate_sd[i]) else 0.0
            is_rel = bool(np.isfinite(estimate[i]) and not _comparison_low_reliability(comparison_rows[i], expected_refs=expected_refs))
            color = 'tab:blue' if is_rel else 'tab:red'
            est_source = str(comparison_rows[i].get('Estimate_source', '') or '')
            source_txt = 'unknown' if est_source == 'unknown_from_calibration' else ('std add' if est_source == 'standard_addition' else 'estimated value')
            label = None
            if is_rel and not shown_est_legend:
                label = f'{source_txt} (reliable scale)'
                shown_est_legend = True
            elif (not is_rel) and not shown_unrel_legend:
                label = f'{source_txt} (low reliability)'
                shown_unrel_legend = True

            clipped = bool(yv > y_hi or yv < y_lo)
            y_plot = min(max(float(yv), y_lo), y_hi)
            lower = max(y_lo, float(yv) - ev)
            upper = min(y_hi, float(yv) + ev)
            yerr = np.array([[max(0.0, y_plot - lower)], [max(0.0, upper - y_plot)]], dtype=np.float64)
            marker_fmt = 'o'
            ax2.errorbar([xv], [y_plot], yerr=yerr, fmt=marker_fmt, color=color, ecolor=color, capsize=3, label=label)

            ref_payload = _nearest_reference_payload(yv)
            if (ref_payload is not None) and (not clipped):
                delta_txt = f"{ref_payload['delta']:+.2g}"
                y_top_err = upper if np.isfinite(upper) else y_plot
                y_txt = min(y_hi - 0.01 * yrange, y_top_err + text_dy)
                ax2.text(xv, y_txt, delta_txt, ha='center', va='bottom', fontsize=7.0, color=color)

            if clipped:
                clip_label = None
                if not shown_clip_legend:
                    clip_label = 'out of scale'
                    shown_clip_legend = True
                y_edge = y_hi if yv > y_hi else y_lo
                marker = '^' if yv > y_hi else 'v'
                ax2.plot([xv], [y_edge], marker=marker, color=color, linestyle='None', label=clip_label)
                ax2.text(xv, y_hi - 0.035 * yrange if yv > y_hi else y_lo + 0.035 * yrange,
                         'out of scale', ha='center', va='top' if yv > y_hi else 'bottom', fontsize=6.5, color=color)

        ax2.set_ylabel(f'Reference value(s) ({unit_label})', fontweight='bold')
        ax2.yaxis.set_label_coords(-0.065, 0.5)
        ax2.grid(True, axis='y', alpha=0.25)
        handles2, labels2 = ax2.get_legend_handles_labels()
        if handles2:
            uniq = {}
            for h, lab in zip(handles2, labels2):
                if lab not in uniq:
                    uniq[lab] = h
            ax2.legend(list(uniq.values()), list(uniq.keys()), loc='lower left', frameon=False)
        ax2.tick_params(labelbottom=False)

    if panel3_ok:
        ax3 = ax_map['r2']
        mask_r2_cal = np.isfinite(r2_cal)
        mask_r2_std = np.isfinite(r2_std)
        if np.any(mask_r2_cal):
            ax3.plot(x[mask_r2_cal], r2_cal[mask_r2_cal], linestyle='None', marker='o', label='R² calibration')
        if np.any(mask_r2_std):
            ax3.plot(x[mask_r2_std], r2_std[mask_r2_std], linestyle='None', marker='s', label='R² std add')
        ax3.axhline(1.0, linestyle='--', linewidth=1.0, color='black', alpha=0.8)
        ax3.set_ylim(-0.02, 1.05)
        ax3.set_ylabel('R²', fontweight='bold')
        ax3.yaxis.set_label_coords(-0.065, 0.5)
        ax3.set_xlabel('Method', fontweight='bold')
        ax3.grid(True, axis='y', alpha=0.25)
        handles3, labels3 = ax3.get_legend_handles_labels()
        if handles3:
            ax3.legend(handles3, labels3, loc='best', frameon=False)
        ax3.set_xticks(x)
        ax3.set_xticklabels(methods, rotation=40, ha='right')

    bottom_name, bottom_ax = axes[-1]
    if bottom_name != 'r2':
        bottom_ax.set_xlabel('Method', fontweight='bold')
        bottom_ax.set_xticks(x)
        bottom_ax.set_xticklabels(methods, rotation=40, ha='right')
    for _name, _ax in axes:
        _ax.set_xlim(-0.5, max(len(methods) - 0.5, 0.5))
        _ax.xaxis.set_major_locator(FixedLocator(x))
    for name, ax in axes[:-1]:
        ax.tick_params(labelbottom=False)
    bottom_ax.tick_params(labelbottom=True)

    fig.savefig(path, dpi=220, bbox_inches='tight')
    plt.close(fig)

def _write_method_comparison_csv(path, comparison_rows):
    if (not path) or (not comparison_rows):
        return
    headers = list(comparison_rows[0].keys())
    with open(path, 'w', newline='', encoding='utf-8') as f:
        wr = csv.writer(f, delimiter=';')
        wr.writerow(headers)
        for row in comparison_rows:
            wr.writerow([row.get(h, '') for h in headers])


def _append_method_comparison_to_workbook(path, comparison_rows, expected_refs=None):
    """Compatibility hook: method comparison is now written directly by _write_report_workbook."""
    return

def _write_report_workbook(path, image_basename, unit_label, selected_channel, raw_rows, summary_rows, fit_rows, selection_info, fit_options, empty_well_rows=None, method_comparison_rows=None, expected_refs=None, reliability_payload=None, image_qc_info=None):
    def _has_any_finite(rows, keys):
        for row in rows or []:
            for key in keys:
                if np.isfinite(row.get(key, np.nan)):
                    return True
        return False

    def _method_family(channel_name):
        s = str(channel_name or "")
        if s.startswith("PAbs_"):
            return "PAbs"
        if s in {"Red", "Green", "Blue"}:
            return "RGB"
        if s in {"L", "a", "b"}:
            return "CIELAB"
        if s in {"DeltaL", "Deltaa", "Deltab"}:
            return "CIELAB_delta"
        if s in {"DeltaE", "DeltaE_chroma", "DeltaE_ab", "DeltaE_ab_chroma"}:
            return "DeltaE"
        return "Other"

    reliability_payload = reliability_payload or {}

    def _overview_lines():
        mode_flags = _report_mode_flags(fit_rows)
        has_cal = mode_flags.get("has_calibration", False)
        has_std = mode_flags.get("has_stdadd", False)
        has_unk = mode_flags.get("has_unknown", False)
        if has_cal and (not has_std) and (not has_unk):
            mode_line = "Mode: calibration only"
        elif has_cal and has_std and has_unk:
            mode_line = "Mode: calibration + standard addition + unknown"
        elif has_cal and has_std:
            mode_line = "Mode: calibration + standard addition"
        elif has_cal and has_unk:
            mode_line = "Mode: calibration + unknown"
        elif has_std:
            mode_line = "Mode: standard addition only"
        elif mode_flags.get("has_unknown_only", False) or has_unk:
            mode_line = "Mode: unknown only"
        else:
            mode_line = "Mode: no valid analytical fit available"
        weighting_txt = "robust IRLS"

        lines = [
            f"Report: {image_basename}",
            f"Unit: {unit_label}",
            mode_line,
            f"Selected channel: {selected_channel}",
            f"Quantification: {reliability_payload.get('quantification_status', 'available')}",
            f"Reliability score: {reliability_payload.get('reliability_score', np.nan):.1f}" if np.isfinite(reliability_payload.get('reliability_score', np.nan)) else "Reliability score: nan",
            f"Confidence class: {reliability_payload.get('confidence_class', '')}",
            f"Reliability note: {reliability_payload.get('reason', '')}",
            f"Empty-well QC: {reliability_payload.get('empty_qc_status', 'not_available')} (drift score={reliability_payload.get('empty_drift_score', np.nan):.4f})" if np.isfinite(reliability_payload.get('empty_drift_score', np.nan)) else f"Empty-well QC: {reliability_payload.get('empty_qc_status', 'not_available')}",
            f"Configured epsilon: {reliability_payload.get('epsilon', np.nan):.6g} M-1 cm-1" if np.isfinite(reliability_payload.get('epsilon', np.nan)) else "Configured epsilon: not available",
            f"Liquid volume: {reliability_payload.get('liquid_volume_ul', np.nan):.6g} uL" if np.isfinite(reliability_payload.get('liquid_volume_ul', np.nan)) else "Liquid volume: not available",
            f"Calculated path length: {reliability_payload.get('path_length', np.nan):.6g} cm" if np.isfinite(reliability_payload.get('path_length', np.nan)) else "Calculated path length: not available",
            f"Path-length source: {reliability_payload.get('path_length_source', '')}" if str(reliability_payload.get('path_length_source', '')).strip() else "Path-length source: not available",
            "",
            "CHANNEL RANKING",
        ]
        ranking = selection_info.get("ranking", []) or []
        if ranking:
            for i, row in enumerate(ranking[:6], start=1):
                score = row.get("Score", np.nan)
                score_txt = f"{float(score):.4f}" if np.isfinite(score) else "nan"
                lines.append(f"{i}. {row.get('Channel', '')}, score={score_txt} ({row.get('Mode', '')})")
        else:
            lines.append("Ranking not available")

        ref_lines = _reference_info_lines(fit_rows, selected_channel, unit_label, expected_refs=expected_refs)
        if ref_lines:
            lines.append("")
            lines.extend(ref_lines)
        lines.append("")
        lines.append("KEY RESULTS")
        fit_map = {}
        for row in fit_rows or []:
            fit_map.setdefault(row.get("Channel"), []).append(row)
        for label in sorted(fit_map.keys(), key=lambda s: str(s)):
            rows_ch = fit_map.get(label, [])
            cal_rows = [r for r in rows_ch if r.get("FitType") == "Calibration"]
            if cal_rows:
                cal = cal_rows[0]
                if np.isfinite(cal.get("LOD", np.nan)):
                    lines.append(f"{label}, LOD={cal.get('LOD', np.nan):.3g} {unit_label}; LOQ={cal.get('LOQ', np.nan):.3g} {unit_label}")
            for row in rows_ch:
                if row.get("FitType") in {"StdAdd", "UnknownFromCal", "UnknownFromEpsilon"} and np.isfinite(row.get("C0", np.nan)):
                    c0_sd = row.get("C0_sd", np.nan)
                    method = "std add" if row.get("FitType") == "StdAdd" else ("unknown epsilon" if row.get("FitType") == "UnknownFromEpsilon" else "unknown")
                    if np.isfinite(c0_sd):
                        lines.append(f"{label} {method} {row.get('ID', '')} DF={row.get('DF', '')}, C₀={abs(float(row.get('C0', np.nan))):.2f} ± {float(c0_sd):.2f} {unit_label}")
                    else:
                        lines.append(f"{label} {method} {row.get('ID', '')} DF={row.get('DF', '')}, C₀={abs(float(row.get('C0', np.nan))):.2f} {unit_label}")
        return lines

    def _summary_final_rows():
        rows = []
        cmp_rows = list(method_comparison_rows or [])
        if cmp_rows:
            def _score_key(r):
                v = _num_or_nan(r.get("Score", np.nan))
                return v if np.isfinite(v) else -np.inf
            best_cmp = max(cmp_rows, key=_score_key)
        else:
            best_cmp = (selection_info.get("best", {}) or {})

        selected_name = str(best_cmp.get("Method", best_cmp.get("Channel", selected_channel)))
        rows.append({"Field": "Quantification", "Value": reliability_payload.get("quantification_status", "available")})
        rows.append({"Field": "Reliability score", "Value": reliability_payload.get("reliability_score", np.nan)})
        rows.append({"Field": "Confidence class", "Value": reliability_payload.get("confidence_class", "")})
        reason = str(reliability_payload.get("reason", "")).strip()
        if reason:
            rows.append({"Field": "Reliability note", "Value": reason})
        rows.append({"Field": "Empty-well QC status", "Value": reliability_payload.get("empty_qc_status", "not_available")})
        edrift = _num_or_nan(reliability_payload.get("empty_drift_score", np.nan))
        if np.isfinite(edrift):
            rows.append({"Field": "Empty drift score", "Value": float(edrift)})

        eps = _num_or_nan(reliability_payload.get("epsilon", np.nan))
        vol_ul = _num_or_nan(reliability_payload.get("liquid_volume_ul", np.nan))
        path_len = _num_or_nan(reliability_payload.get("path_length", np.nan))
        path_len_mm = _num_or_nan(reliability_payload.get("path_length_mm", np.nan))
        if np.isfinite(eps):
            rows.append({"Field": "Configured epsilon (M-1 cm-1)", "Value": float(eps)})
        if np.isfinite(vol_ul):
            rows.append({"Field": "Liquid volume per well (uL)", "Value": float(vol_ul)})
        if np.isfinite(path_len):
            rows.append({"Field": "Calculated path length (cm)", "Value": float(path_len)})
        if np.isfinite(path_len_mm):
            rows.append({"Field": "Calculated liquid height (mm)", "Value": float(path_len_mm)})
        path_source = str(reliability_payload.get("path_length_source", "")).strip()
        if path_source:
            rows.append({"Field": "Path-length source", "Value": path_source})
        geom_name = str(reliability_payload.get("plate_geometry_name", "")).strip()
        if geom_name:
            rows.append({"Field": "Plate geometry for path length", "Value": geom_name})
        geom_assumption = str(reliability_payload.get("plate_geometry_assumption", "")).strip()
        if geom_assumption:
            rows.append({"Field": "Plate geometry assumption", "Value": geom_assumption})

        rows.append({"Field": "Selected method from rank", "Value": selected_name})
        rows.append({"Field": "Selected family", "Value": best_cmp.get("Family", _method_family(selected_name))})
        rows.append({"Field": "Ranking mode", "Value": best_cmp.get("RankMode", best_cmp.get("Mode", ""))})
        score = _num_or_nan(best_cmp.get("Score", np.nan))
        rows.append({"Field": "Selected method score", "Value": float(score) if np.isfinite(score) else np.nan})

        key_map = [
            ("R2 calibration", "R2_cal"),
            ("R2 std add", "R2_std_mean"),
            ("Slope agreement", "SlopeAgreement"),
            ("C0 median", "C0_median"),
            ("C0 SD median", "C0_sd_median"),
            ("beta (mean)", "beta_mean"),
            ("Bias index (mean)", "bias_index_mean"),
            ("LOD", "LOD"),
            ("LOQ", "LOQ"),
        ]
        for label, key in key_map:
            val = _num_or_nan(best_cmp.get(key, np.nan))
            if np.isfinite(val):
                rows.append({"Field": label, "Value": float(val)})

        # StdAdd/unknown estimates for the selected method, if available.
        selected_rows = [fr for fr in fit_rows or [] if str(fr.get("Channel", "")) == selected_name]
        for fr in selected_rows:
            if fr.get("FitType") in {"StdAdd", "UnknownFromCal", "UnknownFromEpsilon"} and np.isfinite(fr.get("C0", np.nan)):
                lbl = f"{fr.get('FitType')} ID={fr.get('ID', '')} DF={fr.get('DF', '')} C0 ({unit_label})"
                rows.append({"Field": lbl, "Value": float(fr.get("C0"))})
                if np.isfinite(fr.get("C0_sd", np.nan)):
                    rows.append({"Field": f"{fr.get('FitType')} ID={fr.get('ID', '')} DF={fr.get('DF', '')} C0 SD ({unit_label})", "Value": float(fr.get("C0_sd"))})

        selected_ref_rows = _selected_channel_reference_rows(fit_rows, selected_name, expected_refs=expected_refs)
        for row in selected_ref_rows:
            row_tag = f"{row.get('MethodLabel', '')} ID={row.get('ID', '')} DF={row.get('DF', '')}"
            for i_ref, ref in enumerate(expected_refs or [], start=1):
                if not _reference_matches_fit_row(ref, row):
                    continue
                ref_label = _format_reference_label(ref, i_ref)
                ref_safe = _expected_ref_key_label(ref_label, i_ref)
                if f"reference_value_{ref_safe}" not in row:
                    continue
                rows.append({"Field": "Reference label", "Value": ref_label})
                for label, key in [
                    (f"Reference value ({unit_label})", f"reference_value_{ref_safe}"),
                    (f"Reference SD ({unit_label})", f"reference_sd_{ref_safe}"),
                    (f"{row_tag} delta ({unit_label})", f"delta_reference_{ref_safe}"),
                    (f"{row_tag} recovery (%)", f"recovery_pct_{ref_safe}"),
                ]:
                    val = _num_or_nan(row.get(key, np.nan))
                    if np.isfinite(val):
                        rows.append({"Field": label, "Value": float(val)})

        notes_join = "; ".join([str(x) for x in reliability_payload.get("notes", []) if str(x)])
        if notes_join:
            rows.append({"Field": "Status notes", "Value": notes_join})
        return rows

    mode_flags = _report_mode_flags(fit_rows)
    has_cielab = _has_any_finite(raw_rows, ["L", "a", "b"]) or _has_any_finite(summary_rows, ["L_median", "a_median", "b_median"])
    has_deltae = _has_any_finite(raw_rows, ["DeltaE_ab", "DeltaE_ab_chroma", "DeltaL", "Deltaa", "Deltab"]) or _has_any_finite(summary_rows, ["DeltaE_ab_median", "DeltaE_ab_chroma_median", "DeltaL_median", "Deltaa_median", "Deltab_median"])

    wb = Workbook()
    ws_contents = wb.active
    ws_contents.title = "01_CONTENTS"
    result_contents = [
        {"Sheet": "01_CONTENTS", "Purpose": "Index of primary results workbook sheets."},
        {"Sheet": "02_METADATA", "Purpose": "Image-level metadata and rule-based image QC used to audit the analysis."},
        {"Sheet": "03_OVERVIEW", "Purpose": "Final quantitative summary, selected method, reliability and external reference checks when provided."},
        {"Sheet": "04_RAW", "Purpose": "Well-level analytical values used by the fitting pipeline."},
        {"Sheet": "05_REPLICATES_MEAN", "Purpose": "Replicate-group medians, robust SDs and group QC flags."},
        {"Sheet": "06_FITTING", "Purpose": "Calibration, standard-addition and unknown/CRM fit results."},
    ]
    if method_comparison_rows:
        result_contents.append({"Sheet": "07_METHOD_COMPARISON", "Purpose": "Method ranking using only common score factors; expected values are external checks only."})
        result_contents.append({"Sheet": "08_LEGENDS", "Purpose": "Definitions for all fields reported in this workbook and primary RGB figure."})
    else:
        result_contents.append({"Sheet": "07_LEGENDS", "Purpose": "Definitions for all fields reported in this workbook and primary RGB figure."})
    _write_table(ws_contents, 1, ["Sheet", "Purpose"], result_contents)

    ws_meta = wb.create_sheet("02_METADATA")
    _write_table(ws_meta, 1, ["Field", "Value", "Notes"], _build_image_qc_metadata_rows(image_qc_info))

    ws0 = wb.create_sheet("03_OVERVIEW")
    overview_rows = _summary_final_rows()
    # Remove pleonastic implementation flags from the overview; details are in 07_LEGENDS.
    overview_rows = [r for r in overview_rows if str(r.get("Field", "")) not in {"all_points_used", "no SD weighting"}]
    _write_table(ws0, 1, ["Field", "Value"], overview_rows)

    ws1 = wb.create_sheet("04_RAW")
    raw_headers = ["Row", "Col", "Well", "ID", "Type", "Conc", "DF", "MeanW_Red", "MeanW_Green", "MeanW_Blue", "MeanBG_Red", "MeanBG_Green", "MeanBG_Blue", "SignalT_Red", "SignalT_Green", "SignalT_Blue", "PAbs_Red", "PAbs_Green", "PAbs_Blue"]
    if has_cielab:
        raw_headers.extend(["L", "a", "b"])
    if has_deltae:
        raw_headers.extend(["DeltaL", "Deltaa", "Deltab", "DeltaE_ab", "DeltaE_ab_chroma", "CIELAB_ref_source"])
    raw_headers.extend(["ImageWarning"])
    _write_table(ws1, 1, raw_headers, _rows_with_aliases(raw_rows, {"PAbs_Red":"Signal_Red", "PAbs_Green":"Signal_Green", "PAbs_Blue":"Signal_Blue"}))

    ws2 = wb.create_sheet("05_REPLICATES_MEAN")
    summary_headers = ["ID", "DF", "Type", "Conc", "PAbs_Red_median", "PAbs_Red_sd", "PAbs_Green_median", "PAbs_Green_sd", "PAbs_Blue_median", "PAbs_Blue_sd"]
    if has_cielab:
        summary_headers.extend(["L_median", "L_sd", "a_median", "a_sd", "b_median", "b_sd"])
    if has_deltae:
        summary_headers.extend(["DeltaL_median", "DeltaL_sd", "Deltaa_median", "Deltaa_sd", "Deltab_median", "Deltab_sd", "DeltaE_ab_median", "DeltaE_ab_sd", "DeltaE_ab_chroma_median", "DeltaE_ab_chroma_sd", "CIELAB_ref_source"])
    summary_headers.extend(["NReplicates", "QCFlagged", "QCCritical"])
    _write_table(ws2, 1, summary_headers, _rows_with_aliases(summary_rows, {"PAbs_Red_median":"Signal_Red_median", "PAbs_Red_sd":"Signal_Red_sd", "PAbs_Green_median":"Signal_Green_median", "PAbs_Green_sd":"Signal_Green_sd", "PAbs_Blue_median":"Signal_Blue_median", "PAbs_Blue_sd":"Signal_Blue_sd"}))

    ws3 = wb.create_sheet("06_FITTING")
    fit_headers = ["Channel", "FitType", "n_points", "m", "q", "R2", "RMSE"]
    if mode_flags.get("has_calibration", False):
        fit_headers.extend(["sigma_cal", "sigma_source", "SNR", "LOD", "LOQ"])
    if mode_flags.get("has_stdadd", False) or mode_flags.get("has_unknown", False):
        fit_headers.extend(["ID", "DF", "C0", "C0_sd", "beta_k", "bias_index_k"])
    if mode_flags.get("has_calibration", False):
        fit_headers.extend(["S0_calibration", "S0_applied", "NClipPoints", "ClipX", "ClipDelta"])
    _write_table(ws3, 1, fit_headers, _display_rows_with_pabs_names(fit_rows))

    # Method selection is integrated into the method-comparison sheet.
    # This avoids a sparse standalone sheet with columns that are not applicable
    # for calibration-only rows.
    if method_comparison_rows:
        _rows_cmp0 = list(method_comparison_rows)
        max_factors = max([int(r.get("CommonFactorsN", 0) or 0) for r in _rows_cmp0] or [0])
        sorted_cmp = sorted(
            [r for r in _rows_cmp0 if int(r.get("CommonFactorsN", 0) or 0) == max_factors],
            key=lambda r: (_num_or_nan(r.get("Score", np.nan)) if np.isfinite(_num_or_nan(r.get("Score", np.nan))) else -np.inf),
            reverse=True
        )
        rank_map = {str(r.get("Method", r.get("Channel", ""))): i for i, r in enumerate(sorted_cmp, start=1)}
        best_ch = str(sorted_cmp[0].get("Method", sorted_cmp[0].get("Channel", selected_channel))) if sorted_cmp else str(selected_channel)
        cmp_rows = []
        for row in method_comparison_rows:
            rr = dict(row)
            method = str(rr.get("Method", rr.get("Channel", "")))
            rr["Selected"] = int(method == best_ch)
            rr["Rank"] = rank_map.get(method, "")
            cmp_rows.append(rr)
        preferred = ["Selected", "Rank", "Method", "Family", "ComparableGroup", "CommonFactorsN", "Score", "ScoreFormula", "RankMode"]
        all_headers = list(cmp_rows[0].keys())
        cmp_headers = [h for h in preferred if h in all_headers] + [h for h in all_headers if h not in preferred]
        # Drop columns that are completely empty/non-informative.
        kept = []
        for h in cmp_headers:
            vals = [r.get(h, "") for r in cmp_rows]
            has_value = False
            for v in vals:
                if isinstance(v, (int, float, np.integer, np.floating)):
                    if np.isfinite(float(v)):
                        has_value = True; break
                elif str(v).strip() not in {"", "nan", "None"}:
                    has_value = True; break
            if h == "ScoreUsesSNR":
                continue
            if h == "FinalScore":
                continue
            if h == "Component":
                continue
            if has_value or h in {"Selected", "Rank", "Method", "Family", "ComparableGroup", "CommonFactorsN", "Score", "ScoreFormula", "RankMode"}:
                kept.append(h)
        ws_cmp = wb.create_sheet("07_METHOD_COMPARISON")
        _write_table(ws_cmp, 1, kept, _display_rows_with_pabs_names(cmp_rows))

    legend_sheet_name = "08_LEGENDS" if method_comparison_rows else "07_LEGENDS"
    ws_leg = wb.create_sheet(legend_sheet_name)
    legend_headers = ["Term", "Meaning", "Formula", "Unit", "Where used", "Shown when", "Notes"]
    _write_table(ws_leg, 1, legend_headers, _build_legends_rows(unit_label, mode_flags=mode_flags))

    for ws in wb.worksheets:
        _autosize_worksheet(ws)
    wb.save(path)

def _write_methods_txt(path, image_basename, unit_label):
    txt = f"""Legends for {image_basename}

PAbs_Red / PAbs_Green / PAbs_Blue
RGB pseudo-absorbance computed as log10(I_BG / I_well) = -log10(I_well / I_BG).

L, a, b
CIELAB coordinates computed from the RGB image.

DeltaE_ab
Total CIELAB color difference: sqrt(DeltaL^2 + Deltaa^2 + Deltab^2).

LOD
Limit of detection = 3 * sigma_cal / |m|.

LOQ
Limit of quantification = 10 * sigma_cal / |m|.

C0
Original-sample concentration before dilution.
Standard addition: C0 = DF * (q/m), with x = added concentration.
Unknown or CRM from calibration: C0 = DF * ((y_obs - q) / m).

C0_sd
Estimated uncertainty associated with C0.

Units
Reported concentrations are expressed in {unit_label}.
"""
    with open(path, "w", encoding="utf-8") as f:
        f.write(txt)


def _write_figure_caption_txt(path, image_basename, unit_label, figure_kind="ALL", expected_refs=None):
    txt = f"""Figure RGB.
Plate-level RGB/PAbs report for {image_basename}. The left panel shows the analyzed plate image with mouth/floor ROI overlay. The text block reports analysis mode, image/geometry QC, fit-quality channel ranking, reference values, key quantitative results, and model summaries. The right panels show calibration and, when present, standard-addition/unknown fits for the ranked RGB/PseudoAbs channels. Reference values are projected on the plots when available.

Figure CIELAB/DeltaE.
Diagnostic CIELAB and DeltaE report for {image_basename}. The left panel shows the same ROI overlay and a summary based on colorimetric coordinates. The right panels show the corresponding calibration and standard-addition/unknown fits for DeltaE, DeltaL, Deltaa, Deltab and related CIELAB-derived channels when available. This figure is intended for comparison with the RGB/PseudoAbs result, not as a hidden diagnostic.

Method comparison.
Method-comparison summary for {image_basename}. The panels compare methods using slope agreement/bias, estimated reference-scale values, and R² values. Dashed reference lines indicate external reference values when provided. This figure helps identify methods with high fit quality, stable slope agreement, and plausible agreement with reference measurements.

Units.
Reported concentrations are expressed in {unit_label}.
"""
    with open(path, "w", encoding="utf-8") as f:
        f.write(txt)


def _write_results_caption_txt(path, image_basename, unit_label, expected_refs=None):
    txt = f"""RESULTS caption - RGB quantitative output

File scope
This caption applies to the primary RGB outputs in the RESULTS folder for {image_basename}, especially the *_FIGURE_RGB.png and *_REPORT.xlsx files.

Analytical signal
The primary RGB signal is pseudo-absorbance, reported as PAbs_Red, PAbs_Green and PAbs_Blue:
    PAbs = log10(I_BG / I_well) = -log10(I_well / I_BG)
where I_well is the linearized median intensity from the well ROI and I_BG is the linearized local inter-well background predicted for that well. This is an image-derived pseudo-absorbance and is not assumed to be a spectrophotometric absorbance.

Fitting and quantification
Calibration and standard-addition fits use robust residual-based IRLS linear regression. All finite points are included in the fit; outlying points receive lower robust residual weights rather than being excluded. For standard addition, the original-sample concentration is C0 = DF x q/m, where y = m x + q and x is the added concentration.

Ranking score
For methods with both calibration and standard addition, the global score is:
    GlobalScore = slope_agreement^2 x sqrt(R2_cal x R2_std) x (1/LOQ)
with slope_agreement = min(|m_cal|, |m_std|) / max(|m_cal|, |m_std|). Expected/reference values, recovery, SNR and clipping are not used in this score.

Reference values and recovery
External reference values, when provided, are used only for external comparison (Delta and recovery). They are not used to choose the ranked RGB method.

Quality control
Image, plate, geometry and floor-QC messages are alerts on data quality. No automatic image correction is applied.

Geometry and epsilon/path-length quantification
When epsilon-based unknown quantification is used, the optical path length is estimated from the configured liquid volume and the nominal flat-bottom well area of the selected plate geometry. This calculation assumes ANSI/SLAS-compatible flat-bottom microplate geometry; round-bottom, U-bottom, V-bottom, spheroid, deep-well, or non-certified geometries require separate validation.

Units
Reported concentrations are expressed in {unit_label}.
"""
    with open(path, "w", encoding="utf-8") as f:
        f.write(txt)


def _write_raw_data_details_caption_txt(path, image_basename, unit_label):
    txt = f"""RAW_DATA_DETAILS caption - diagnostics and method-development outputs

File scope
This caption applies to diagnostic outputs in RAW_DATA_DETAILS for {image_basename}, including BG_STAT_MASK.png, FIGURE_CIELAB_DELTAE.png, METHOD_COMPARISON.png and the DIAGNOSTICS.xlsx workbook.

BG_STAT_MASK.png
The binary mask shows the pixels accepted as inter-well background after model-based geometric exclusion and statistical filtering. White pixels are used for background sampling; black pixels are rejected. Small holes or fragmented white regions can arise from robust rejection of text, reflections, local artifacts, or implausible background pixels.

FIGURE_CIELAB_DELTAE.png
CIELAB and DeltaE descriptors are derived from the same RGB image information through nonlinear RGB-to-CIELAB transformations. They are reported as diagnostic descriptors for comparison and method development, not as the primary quantitative RGB output. The conversion uses OpenCV BGR→Lab for 8-bit sRGB-like images; the reference form is linearized sRGB to XYZ(D65) using the IEC 61966-2-1:1999 matrix followed by CIE 1976 L*a*b*. DeltaE_ab = sqrt(DeltaL^2 + Deltaa^2 + Deltab^2); DeltaE_ab_chroma = sqrt(Deltaa^2 + Deltab^2).

METHOD_COMPARISON.png
The method-comparison plot summarizes fit quality and reference-scale estimates across RGB and CIELAB-derived descriptors. The comparison score uses only common factors available across method families. SNR, clipping and expected/reference values are not part of the score. Reference values are shown only as external checks.

DIAGNOSTICS.xlsx
The workbook collects the useful diagnostic tables previously exported as separate CSV files. The final LEGENDS sheet defines the diagnostic fields reported in that workbook. These sheets are intended to support auditability, QC, and method development rather than to replace the primary RESULTS report.

Geometry and epsilon/path-length quantification
When epsilon-based unknown quantification is used, the optical path length is estimated from the configured liquid volume and the nominal flat-bottom well area of the selected plate geometry. This calculation assumes ANSI/SLAS-compatible flat-bottom microplate geometry; round-bottom, U-bottom, V-bottom, spheroid, deep-well, or non-certified geometries require separate validation.

Units
Reported concentrations are expressed in {unit_label}.
"""
    with open(path, "w", encoding="utf-8") as f:
        f.write(txt)



def _row_label_from_zero_based(value):
    try:
        idx = int(float(value))
    except Exception:
        return ""
    if idx < 0:
        return ""
    return _row_label_from_index(idx)


def _well_from_zero_based(row_value, col_value):
    try:
        rr = int(float(row_value)); cc = int(float(col_value))
    except Exception:
        return ""
    if rr < 0 or cc < 0:
        return ""
    return _well_name_from_indices(rr, cc)


def _associated_wells_from_bg_cell(cell_r, cell_c):
    """Return the four wells surrounding an inter-well BG cell.

    BG samples are indexed by the inter-well cell between rows r/r+1 and
    columns c/c+1. The index is therefore not a well index.
    """
    try:
        r = int(float(cell_r)); c = int(float(cell_c))
    except Exception:
        return ""
    if r < 0 or c < 0:
        return ""
    return "-".join([
        _well_name_from_indices(r, c),
        _well_name_from_indices(r, c + 1),
        _well_name_from_indices(r + 1, c),
        _well_name_from_indices(r + 1, c + 1),
    ])


def _diagnostic_display_rows(sheet_key, rows):
    """Return workbook-facing diagnostic rows with clear labels and RGB order.

    Internal OpenCV/BGR names and zero-based technical indices are kept out of
    exported workbooks when a clearer analytical name is available.
    """
    out = []
    for row in rows or []:
        r = dict(row)
        rr = {}
        if sheet_key == "BG_SAMPLES":
            rr["BG_Cell_Row"] = r.get("cell_r", "")
            rr["BG_Cell_Col"] = r.get("cell_c", "")
            rr["Associated_Wells"] = _associated_wells_from_bg_cell(r.get("cell_r", ""), r.get("cell_c", ""))
            rr["x"] = r.get("x", "")
            rr["y"] = r.get("y", "")
            rr["area"] = r.get("area", "")
            rr["Red_median_raw"] = r.get("R_med", "")
            rr["Green_median_raw"] = r.get("G_med", "")
            rr["Blue_median_raw"] = r.get("B_med", "")
        elif sheet_key == "BG_WELL_FIT":
            rr["Row"] = _row_label_from_zero_based(r.get("well_r", ""))
            try:
                rr["Col"] = int(float(r.get("well_c", ""))) + 1
            except Exception:
                rr["Col"] = ""
            rr["Well"] = _well_from_zero_based(r.get("well_r", ""), r.get("well_c", ""))
            rr["x"] = r.get("x", "")
            rr["y"] = r.get("y", "")
            rr["BG_Red_raw"] = r.get("R_bg", "")
            rr["BG_Green_raw"] = r.get("G_bg", "")
            rr["BG_Blue_raw"] = r.get("B_bg", "")
        elif sheet_key in {"WELL_ROBUST_STATS", "WELL_BOTTOM", "FLOOR_GEOMETRY_QC"}:
            if "well_r" in r and "well_c" in r:
                rr["Row"] = _row_label_from_zero_based(r.get("well_r", ""))
                try:
                    rr["Col"] = int(float(r.get("well_c", ""))) + 1
                except Exception:
                    rr["Col"] = ""
                rr["Well"] = r.get("Well", "") or _well_from_zero_based(r.get("well_r", ""), r.get("well_c", ""))
            # preserve remaining fields, but export color-channel fields in RGB order
            for k, v in r.items():
                if k in {"well_r", "well_c", "Well"}:
                    continue
                if k.startswith("B_") or k.startswith("G_") or k.startswith("R_"):
                    continue
                rr[k] = v
            for prefix in ["mean", "median", "sd", "p10", "p25", "p50", "p75", "p90", "iqr"]:
                for src, dst in [("R", "Red"), ("G", "Green"), ("B", "Blue")]:
                    key = f"{src}_{prefix}"
                    if key in r:
                        rr[f"{dst}_{prefix}"] = r.get(key, "")
        elif sheet_key == "EMPTY_WELLS":
            order = ["Row", "Col", "Well"]
            for k in order:
                if k in r:
                    rr[k] = r.get(k, "")
            for ch in ["Red", "Green", "Blue"]:
                for stem in ["MeanW", "MeanBG", "PAbs", "Signal"]:
                    key = f"{stem}_{ch}"
                    if key in r:
                        rr[key if stem != "Signal" else f"PAbs_{ch}"] = r.get(key, "")
            for k, v in r.items():
                if k in rr or any(k == f"{stem}_{ch}" for ch in ["Red","Green","Blue"] for stem in ["MeanW","MeanBG","PAbs","Signal"]):
                    continue
                rr[k] = v
        elif sheet_key == "SPATIAL_DIAGNOSTICS":
            n_val = _num_or_nan(r.get("n", np.nan))
            applied = bool(np.isfinite(n_val) and n_val > 0)
            rr["Dataset"] = r.get("dataset", r.get("Dataset", ""))
            rr["Status"] = "applied" if applied else "not_applied"
            rr["Applicability"] = "requires usable unknown or empty wells distributed across rows/columns"
            rr["Reason"] = "" if applied else "no usable wells available for this spatial-trend dataset"
            for k in ["n", "intercept", "slope_col", "slope_row", "R2", "corr_col", "corr_row"]:
                rr[k] = r.get(k, "")
        else:
            rr = r
        out.append(rr)
    return out

def _diagnostics_legend_rows(unit_label, present_sheets=None):
    present_sheets = set(present_sheets or [])
    has_spatial = "09_SPATIAL_DIAGNOSTICS" in present_sheets
    has_method = "10_METHOD_COMPARISON" in present_sheets
    has_cielab_fit = "11_CIELAB_FITTING" in present_sheets
    rows = [
        {"Field":"BG_STAT_MASK", "Meaning":"Binary diagnostic mask of accepted inter-well background pixels", "Formula":"white = used as BG sample; black = rejected", "Unit":"image", "Where used":"BG_STAT_MASK.png", "Notes":"Used to audit the background-sampling step."},
        {"Field":"BG_Cell_Row/BG_Cell_Col", "Meaning":"Inter-well background-cell row/column index", "Formula":"0-based grid index of the inter-well cell, not a well row/column", "Unit":"index", "Where used":"02_BG_SAMPLES", "Notes":"BG_Cell_Row=0 and BG_Cell_Col=0 identify the inter-well region surrounded by A1, A2, B1 and B2."},
        {"Field":"Associated_Wells", "Meaning":"Four wells surrounding an inter-well BG sample cell", "Formula":"well(r,c)-well(r,c+1)-well(r+1,c)-well(r+1,c+1)", "Unit":"well labels", "Where used":"02_BG_SAMPLES", "Notes":"Clarifies that BG samples are inter-well regions rather than wells."},
        {"Field":"Row/Col/Well", "Meaning":"Human-readable well position", "Formula":"Row is A-based; Col is 1-based; Well = Row+Col", "Unit":"well label", "Where used":"03_BG_WELL_FIT, 04_WELL_ROBUST_STATS, 05_GEOMETRY_QC, 06_WELL_BOTTOM, 08_EMPTY_WELLS", "Notes":"Used only for true well-level records, not for inter-well BG cells."},
        {"Field":"x/y", "Meaning":"Image coordinates", "Formula":"pixel coordinate in analyzed image", "Unit":"pixel", "Where used":"02_BG_SAMPLES, 03_BG_WELL_FIT", "Notes":"Coordinate system follows the image array."},
        {"Field":"area", "Meaning":"Accepted mask area", "Formula":"number of accepted pixels in the BG sample mask", "Unit":"pixels", "Where used":"02_BG_SAMPLES", "Notes":"Area after model-based and robust statistical background-mask filtering."},
        {"Field":"Red_median_raw/Green_median_raw/Blue_median_raw", "Meaning":"Median raw channel value in accepted BG-mask pixels", "Formula":"median over accepted mask pixels", "Unit":"raw image intensity", "Where used":"02_BG_SAMPLES", "Notes":"Exported in standard RGB order; internally OpenCV stores images in BGR order."},
        {"Field":"BG_Red_raw/BG_Green_raw/BG_Blue_raw", "Meaning":"Predicted local raw background at a well", "Formula":"2D background surface evaluated at well center", "Unit":"raw image intensity", "Where used":"03_BG_WELL_FIT", "Notes":"Exported in standard RGB order; used to compute per-well MeanBG after linearization."},
        {"Field":"Red_*/Green_*/Blue_*", "Meaning":"Robust statistics of raw well-channel intensities", "Formula":"computed over retained well ROI pixels; *=mean, median, sd, p10, p25, p50, p75, p90, iqr", "Unit":"raw image intensity", "Where used":"04_WELL_ROBUST_STATS", "Notes":"Exported in standard RGB order."},
        {"Field":"Gray_*", "Meaning":"Robust statistics of grayscale intensity", "Formula":"OpenCV grayscale conversion of RGB/BGR image; *=mean, median, sd, p10, p25, p50, p75, p90, iqr", "Unit":"raw image intensity", "Where used":"04_WELL_ROBUST_STATS", "Notes":"Diagnostic descriptor used for highlight/trimming QC, not a primary analytical signal."},
        {"Field":"Purple_*", "Meaning":"Robust statistics of the internal purple-color index", "Formula":"Purple = 0.5 x (Red + Blue) - Green; *=mean, median, sd, p10, p25, p50, p75, p90, iqr", "Unit":"raw image intensity", "Where used":"04_WELL_ROBUST_STATS", "Notes":"Diagnostic color index for purple/blue-red contrast; not a primary quantitative output."},
        {"Field":"L_*/a_*/b_*", "Meaning":"Robust statistics of CIELAB coordinates", "Formula":"computed from OpenCV CIELAB conversion over retained well ROI pixels; *=mean, median, sd, p10, p25, p50, p75, p90, iqr", "Unit":"dimensionless", "Where used":"04_WELL_ROBUST_STATS", "Notes":"CIELAB values are diagnostic descriptors derived from RGB image data."},
        {"Field":"MeanW_*", "Meaning":"Linearized median well intensity", "Formula":"median well ROI intensity after gamma linearization", "Unit":"dimensionless", "Where used":"REPORT", "Notes":"The asterisk denotes Red, Green or Blue."},
        {"Field":"MeanBG_*", "Meaning":"Linearized local background intensity", "Formula":"background surface value after gamma linearization", "Unit":"dimensionless", "Where used":"REPORT", "Notes":"The asterisk denotes Red, Green or Blue."},
        {"Field":"PAbs_*", "Meaning":"RGB pseudo-absorbance", "Formula":"log10(MeanBG_*/MeanW_*)", "Unit":"dimensionless", "Where used":"REPORT, 10_METHOD_COMPARISON", "Notes":"Primary RGB analytical descriptor."},
        {"Field":"IRLS", "Meaning":"Iteratively reweighted least-squares robust linear regression with residual-based weights", "Formula":"minimize sum_i w_i (y_i - (m x_i + q))^2; w_i is updated iteratively from residual magnitude", "Unit":"dimensionless", "Where used":"REPORT fitting tables, 10_METHOD_COMPARISON, 11_CIELAB_FITTING, FIGURE_RGB.png, FIGURE_CIELAB_DELTAE.png", "Notes":"Reference: Huber, P. J. (1964), Robust Estimation of a Location Parameter. Implementation: custom NumPy-based IRLS using repeated least-squares solves."},
        {"Field":"n_roi/n_core/n_used", "Meaning":"Pixel counts used during well ROI filtering", "Formula":"ROI pixels, eroded/fuzzy core pixels, retained pixels", "Unit":"pixels", "Where used":"04_WELL_ROBUST_STATS", "Notes":"Used to audit well-level pixel filtering."},
        {"Field":"UsedFraction", "Meaning":"Fraction of core ROI retained", "Formula":"n_used / n_core", "Unit":"dimensionless", "Where used":"04_WELL_ROBUST_STATS", "Notes":"Low values indicate strong trimming."},
        {"Field":"highlight_fraction_roi/highlight_fraction_core", "Meaning":"Fraction of very bright pixels in ROI/core", "Formula":"fraction of pixels with grayscale >= highlight threshold", "Unit":"dimensionless", "Where used":"04_WELL_ROBUST_STATS", "Notes":"Optical QC descriptor for highlights/specular artifacts."},
        {"Field":"BrightExcludedFraction", "Meaning":"Fraction of core ROI rejected as bright", "Formula":"n_bright_excluded / n_core", "Unit":"dimensionless", "Where used":"04_WELL_ROBUST_STATS", "Notes":"Part of optical quality diagnostics."},
        {"Field":"BrightExcludedMeanGray", "Meaning":"Mean grayscale intensity of bright-excluded pixels", "Formula":"mean(Gray) over pixels excluded as overly bright", "Unit":"raw grayscale intensity", "Where used":"04_WELL_ROBUST_STATS", "Notes":"Blank if no bright pixels were excluded."},
        {"Field":"BrightExcessMeanGray", "Meaning":"Mean excess brightness above the local bright threshold", "Formula":"mean(max(Gray - bright_threshold, 0)) over bright-excluded pixels", "Unit":"raw grayscale intensity", "Where used":"04_WELL_ROBUST_STATS", "Notes":"Used with BrightExcludedFraction to form HighlightIndex."},
        {"Field":"HighlightIndex", "Meaning":"Combined highlight severity index", "Formula":"BrightExcludedFraction x BrightExcessMeanGray", "Unit":"gray-level weighted fraction", "Where used":"04_WELL_ROBUST_STATS", "Notes":"Higher values indicate more severe bright artifacts."},
        {"Field":"ImageWarning/WarningReason", "Meaning":"Well-level optical QC warning", "Formula":"rule-based flags from highlights, trimming and intensity SD", "Unit":"0/1, text", "Where used":"04_WELL_ROBUST_STATS", "Notes":"WarningReason lists the triggering conditions."},
        {"Field":"mouth_* / floor_*", "Meaning":"Detected mouth/floor geometry descriptors", "Formula":"center coordinates, radius and score from geometry detection", "Unit":"pixels or score", "Where used":"06_WELL_BOTTOM", "Notes":"Used for ROI and geometry QC."},
        {"Field":"local_pitch_px / px_per_mm / cyl_r_bg", "Meaning":"Local geometry scale descriptors", "Formula":"local plate pitch, pixel-to-mm conversion and cylindrical background radius", "Unit":"pixels or pixels/mm", "Where used":"06_WELL_BOTTOM", "Notes":"Technical geometry diagnostics."},
        {"Field":"shift_px", "Meaning":"Mouth-to-floor center shift", "Formula":"Euclidean distance between fitted mouth and floor centers", "Unit":"pixels", "Where used":"06_WELL_BOTTOM, 05_GEOMETRY_QC", "Notes":"Large values indicate poor floor/mouth alignment."},
        {"Field":"shift_frac_of_mouth_r", "Meaning":"Relative mouth-to-floor center shift", "Formula":"shift_px / mouth_r", "Unit":"dimensionless", "Where used":"05_GEOMETRY_QC", "Notes":"Large values indicate poor floor/mouth alignment."},
        {"Field":"floor_to_mouth_r_ratio", "Meaning":"Relative floor radius", "Formula":"floor_r / mouth_r", "Unit":"dimensionless", "Where used":"05_GEOMETRY_QC", "Notes":"Used to flag abnormal floor geometry."},
        {"Field":"D_warning/D_critical", "Meaning":"Floor-geometry diagnostic flags", "Formula":"thresholds on shift_frac_of_mouth_r and floor_to_mouth_r_ratio", "Unit":"0/1", "Where used":"05_GEOMETRY_QC", "Notes":"Critical is stricter than warning."},
        {"Field":"RGB to CIELAB", "Meaning":"Diagnostic color-space conversion", "Formula":"OpenCV BGR->Lab for 8-bit sRGB-like images; reference form: linearized sRGB -> XYZ using the standard sRGB-to-XYZ matrix and D65 reference white -> CIE L*a*b*", "Unit":"dimensionless", "Where used":"04_WELL_ROBUST_STATS, 11_CIELAB_FITTING, FIGURE_CIELAB_DELTAE.png", "Notes":"Diagnostic descriptor, not the primary RGB quantitative output. References: OpenCV color conversions; IEC 61966-2-1:1999 sRGB; CIE 1976 L*a*b*; CIE standard illuminant D65."},
        {"Field":"DeltaE_ab", "Meaning":"Total CIELAB color difference", "Formula":"sqrt(DeltaL^2 + Deltaa^2 + Deltab^2)", "Unit":"dimensionless", "Where used":"REPORT, 10_METHOD_COMPARISON, 11_CIELAB_FITTING, FIGURE_CIELAB_DELTAE.png", "Notes":"Reference source is reported in CIELAB_ref_source in the main report."},
        {"Field":"DeltaE_ab_chroma", "Meaning":"Chromatic CIELAB difference", "Formula":"sqrt(Deltaa^2 + Deltab^2)", "Unit":"dimensionless", "Where used":"REPORT, 10_METHOD_COMPARISON, 11_CIELAB_FITTING, FIGURE_CIELAB_DELTAE.png", "Notes":"Excludes the lightness term DeltaL."},
    ]
    if has_spatial:
        rows.extend([
            {"Field":"dataset", "Meaning":"Dataset used for spatial trend", "Formula":"unknown or empty", "Unit":"text", "Where used":"09_SPATIAL_DIAGNOSTICS", "Notes":"Spatial diagnostics are descriptive and do not alter results."},
            {"Field":"slope_col/slope_row", "Meaning":"Spatial trend coefficients", "Formula":"linear trend versus column and row", "Unit":"signal/index", "Where used":"09_SPATIAL_DIAGNOSTICS", "Notes":"Used to identify plate-position effects."},
            {"Field":"corr_col/corr_row", "Meaning":"Correlation with column/row position", "Formula":"Pearson correlation", "Unit":"dimensionless", "Where used":"09_SPATIAL_DIAGNOSTICS", "Notes":"Diagnostic only."},
        ])
    if has_method:
        rows.extend([
            {"Field":"Method", "Meaning":"Compared analytical descriptor", "Formula":"PAbs_*, L/a/b, Delta* or DeltaE*", "Unit":"text", "Where used":"10_METHOD_COMPARISON", "Notes":"PAbs_* are RGB; CIELAB and DeltaCIELAB are derived diagnostics."},
            {"Field":"Family", "Meaning":"Method family", "Formula":"RGB, CIELAB, DeltaCIELAB or other", "Unit":"text", "Where used":"10_METHOD_COMPARISON", "Notes":"Used to separate method families."},
            {"Field":"ComparableGroup", "Meaning":"Set of methods scored with the same formula", "Formula":"calibration_plus_stdadd, calibration_only, stdadd_only, or not_ranked", "Unit":"text", "Where used":"10_METHOD_COMPARISON", "Notes":"Scores are directly comparable only within the same group."},
            {"Field":"CommonFactorsN", "Meaning":"Number of common factors used in the score", "Formula":"4 for R2_cal, R2_std, SlopeAgreement and LOQ; 1 for calibration-only or stdadd-only fallbacks", "Unit":"integer", "Where used":"10_METHOD_COMPARISON", "Notes":"Used to avoid comparing scores obtained from different formulas."},
            {"Field":"Score", "Meaning":"Common ranking score", "Formula":"for calibration+standard addition: SlopeAgreement^2 x sqrt(R2_cal x R2_std) x (1/LOQ); fallback groups use the formula stated in ScoreFormula", "Unit":"1/" + str(unit_label), "Where used":"10_METHOD_COMPARISON, METHOD_COMPARISON.png", "Notes":"Expected/reference values, SNR and clipping are excluded."},
            {"Field":"ScoreFormula", "Meaning":"Formula used to compute Score", "Formula":"text descriptor", "Unit":"text", "Where used":"10_METHOD_COMPARISON", "Notes":"Documents which score formula was applied to the row."},
            {"Field":"RankMode", "Meaning":"Data basis available for ranking", "Formula":"calibration_plus_stdadd, calibration_only, stdadd_only, or unavailable", "Unit":"text", "Where used":"10_METHOD_COMPARISON", "Notes":"Only rows with the same comparable group should be directly compared."},
            {"Field":"R2_cal/R2_std_mean", "Meaning":"Calibration and mean standard-addition coefficient of determination", "Formula":"R2 from linear fits; R2_std_mean is averaged over standard-addition curves", "Unit":"dimensionless", "Where used":"10_METHOD_COMPARISON, 11_CIELAB_FITTING", "Notes":"Higher values indicate better linear fit quality."},
            {"Field":"m_cal/m_std_mean", "Meaning":"Calibration slope and mean standard-addition slope", "Formula":"slope from linear fit y = m x + q", "Unit":"signal/" + str(unit_label), "Where used":"10_METHOD_COMPARISON, 11_CIELAB_FITTING", "Notes":"Used to compute slope agreement."},
            {"Field":"SlopeAgreement", "Meaning":"Agreement between calibration and standard-addition slopes", "Formula":"min(|m_cal|, |m_std|) / max(|m_cal|, |m_std|)", "Unit":"dimensionless", "Where used":"10_METHOD_COMPARISON", "Notes":"1 indicates identical slope magnitude."},
            {"Field":"beta_mean", "Meaning":"Mean standard-addition/calibration slope ratio", "Formula":"mean(m_std / m_cal)", "Unit":"dimensionless", "Where used":"10_METHOD_COMPARISON", "Notes":"1 indicates equal slopes on average."},
            {"Field":"bias_index_mean", "Meaning":"Mean relative slope bias", "Formula":"mean(|m_std/m_cal - 1|)", "Unit":"dimensionless", "Where used":"10_METHOD_COMPARISON", "Notes":"0 indicates no slope bias."},
            {"Field":"C0_median/C0_sd_median", "Meaning":"Median estimated original concentration and median associated SD", "Formula":"median over available standard-addition or unknown estimates", "Unit":str(unit_label), "Where used":"10_METHOD_COMPARISON", "Notes":"Reported for diagnostic comparison; expected/reference values are external checks."},
            {"Field":"Estimate_source", "Meaning":"Source of the representative concentration estimate", "Formula":"standard_addition, unknown_from_calibration, epsilon or unavailable", "Unit":"text", "Where used":"10_METHOD_COMPARISON", "Notes":"Identifies how Estimate_value was obtained."},
        ])
    if has_cielab_fit:
        rows.extend([
            {"Field":"Channel", "Meaning":"Fitted diagnostic descriptor", "Formula":"L, a, b, DeltaL, Deltaa, Deltab, DeltaE or DeltaE_chroma", "Unit":"text", "Where used":"11_CIELAB_FITTING", "Notes":"CIELAB/DeltaE descriptors are diagnostic, not primary quantitative outputs."},
            {"Field":"FitType", "Meaning":"Type of fit row", "Formula":"Calibration, StdAdd, UnknownFromCal, UnknownOnly", "Unit":"text", "Where used":"11_CIELAB_FITTING", "Notes":"Same convention as main fitting output."},
            {"Field":"m/q/R2/RMSE", "Meaning":"Linear-fit parameters and fit quality", "Formula":"y = m x + q; R2 coefficient of determination; RMSE root-mean-square error", "Unit":"descriptor units", "Where used":"11_CIELAB_FITTING", "Notes":"Diagnostic CIELAB/DeltaE fitting only."},
            {"Field":"LOD/LOQ", "Meaning":"Detection and quantification limits", "Formula":"LOD = 3 sigma_cal / |m|; LOQ = 10 sigma_cal / |m|", "Unit":str(unit_label), "Where used":"11_CIELAB_FITTING", "Notes":"Diagnostic for CIELAB/DeltaE fits."},
            {"Field":"C0/C0_sd", "Meaning":"Estimated original concentration and SD", "Formula":"standard-addition intercept or calibration projection depending on FitType", "Unit":str(unit_label), "Where used":"11_CIELAB_FITTING", "Notes":"Diagnostic for CIELAB/DeltaE fits."},
        ])
    rows.extend([
        {"Term":"IRLS", "Meaning":"Iteratively reweighted least-squares robust linear regression with residual-based weights", "Formula":"minimize sum_i w_i (y_i - (m x_i + q))^2; w_i is updated iteratively from residual magnitude", "Unit":"dimensionless", "Where used":"10_METHOD_COMPARISON, 11_CIELAB_FITTING, METHOD_COMPARISON.png, FIGURE_CIELAB_DELTAE.png", "Notes":"Reference: Huber, P. J. (1964), Robust Estimation of a Location Parameter. Implementation: custom NumPy-based IRLS using repeated least-squares solves."},
        {"Term":"Code implementation", "Meaning":"Main computational libraries used by the diagnostic workflow", "Formula":"custom Python code using NumPy/OpenCV/openpyxl/matplotlib", "Unit":"software provenance", "Where used":"DIAGNOSTICS.xlsx and diagnostic PNG files", "Notes":"Sources/libraries: OpenCV for image masks and BGR-to-Lab conversion; NumPy for numerical fitting; openpyxl for xlsx export; matplotlib for PNG figures."},
        {"Term":"B_med/G_med/R_med", "Meaning":"Median raw OpenCV-channel value in accepted BG-mask pixels", "Formula":"median over accepted mask pixels", "Unit":"raw image intensity", "Where used":"02_BG_SAMPLES", "Notes":"Internal OpenCV order is B,G,R; exported diagnostic tables may also provide user-facing Red/Green/Blue aliases where available."},
        {"Term":"B_bg/G_bg/R_bg", "Meaning":"Predicted local raw background at a well in OpenCV channel order", "Formula":"2D background surface evaluated at well center", "Unit":"raw image intensity", "Where used":"03_BG_WELL_FIT", "Notes":"These are diagnostic raw model outputs; main analytical outputs are reported in standard RGB order."},
        {"Term":"B_*/G_*/R_*", "Meaning":"Robust statistics of raw well-channel intensities in OpenCV channel order", "Formula":"computed over retained well ROI pixels; *=mean, median, sd, p10, p25, p50, p75, p90, iqr", "Unit":"raw image intensity", "Where used":"04_WELL_ROBUST_STATS", "Notes":"User-facing RGB summaries are reported as Red/Green/Blue where converted."},
        {"Term":"Status/Applicability/Reason", "Meaning":"Applicability statement for optional spatial diagnostics", "Formula":"applied/not_applied with reason text", "Unit":"text", "Where used":"09_SPATIAL_DIAGNOSTICS", "Notes":"The sheet is retained even when spatial diagnostics cannot be applied."},
        {"Term":"Estimate_value/Estimate_sd", "Meaning":"Representative diagnostic concentration estimate and associated SD", "Formula":"derived according to Estimate_source", "Unit":str(unit_label), "Where used":"10_METHOD_COMPARISON", "Notes":""},
        {"Term":"reference_*/delta_reference_*/recovery_pct_*", "Meaning":"External reference checks used in diagnostic comparison", "Formula":"delta = estimate - reference; recovery = 100 x estimate/reference", "Unit":str(unit_label) + " or %", "Where used":"10_METHOD_COMPARISON, METHOD_COMPARISON.png", "Notes":"Reference values are external checks and are not part of Score."},
    ])
    rows.extend([
        {"Term":"Sheet", "Meaning":"Workbook sheet name", "Formula":"worksheet name", "Unit":"text", "Where used":"01_CONTENTS", "Notes":""},
        {"Term":"Purpose", "Meaning":"Short description of the sheet role", "Formula":"free-text description", "Unit":"text", "Where used":"01_CONTENTS", "Notes":""},
        {"Term":"BG_STAT_MASK", "Meaning":"Binary diagnostic mask of accepted inter-well background pixels", "Formula":"white pixels = accepted for background sampling; black pixels = rejected", "Unit":"image", "Where used":"BG_STAT_MASK.png, RAW_DATA_DETAILS_CAPTION.txt", "Notes":"Useful for auditability and troubleshooting of background sampling."},
        {"Term":"BG_Cell_Row/BG_Cell_Col", "Meaning":"Inter-well background-cell row and column index", "Formula":"0-based index of the inter-well cell, not a well coordinate", "Unit":"index", "Where used":"02_BG_SAMPLES", "Notes":"BG_Cell_Row=0 and BG_Cell_Col=0 are surrounded by A1, A2, B1 and B2."},
        {"Term":"Associated_Wells", "Meaning":"Four wells surrounding an inter-well background cell", "Formula":"well(r,c)-well(r,c+1)-well(r+1,c)-well(r+1,c+1)", "Unit":"well labels", "Where used":"02_BG_SAMPLES", "Notes":""},
        {"Term":"x/y", "Meaning":"Image coordinates of a background sample or well center", "Formula":"pixel coordinate in the analyzed image", "Unit":"pixels", "Where used":"02_BG_SAMPLES, 03_BG_WELL_FIT", "Notes":""},
        {"Term":"area", "Meaning":"Accepted background-mask area", "Formula":"number of accepted pixels in the background sample mask", "Unit":"pixels", "Where used":"02_BG_SAMPLES", "Notes":""},
        {"Term":"B_med/G_med/R_med", "Meaning":"Median raw OpenCV-channel value in accepted BG-mask pixels", "Formula":"median over accepted background-mask pixels", "Unit":"raw image intensity", "Where used":"02_BG_SAMPLES", "Notes":"OpenCV internal channel order is B,G,R."},
        {"Term":"B_bg/G_bg/R_bg", "Meaning":"Predicted local raw background at a well", "Formula":"2D background surface evaluated at well center", "Unit":"raw image intensity", "Where used":"03_BG_WELL_FIT", "Notes":"OpenCV internal channel order is B,G,R."},
        {"Term":"Row/Col/Well", "Meaning":"Human-readable well position", "Formula":"Row is A-based; Col is 1-based; Well = Row + Col", "Unit":"well label", "Where used":"03_BG_WELL_FIT, 04_WELL_ROBUST_STATS, 05_GEOMETRY_QC, 06_WELL_BOTTOM, 08_EMPTY_WELLS", "Notes":"Used for true well-level records, not for inter-well BG cells."},
        {"Term":"n_roi/n_core/n_used", "Meaning":"Pixel counts used during well ROI filtering", "Formula":"ROI pixels, core pixels and retained pixels", "Unit":"pixels", "Where used":"04_WELL_ROBUST_STATS", "Notes":""},
        {"Term":"used_fraction/UsedFraction", "Meaning":"Fraction of ROI core pixels retained after filtering", "Formula":"n_used / n_core", "Unit":"dimensionless", "Where used":"04_WELL_ROBUST_STATS, 08_EMPTY_WELLS", "Notes":""},
        {"Term":"highlight_fraction_roi/highlight_fraction_core", "Meaning":"Fraction of very bright pixels in ROI/core", "Formula":"fraction of pixels with grayscale above the highlight threshold", "Unit":"dimensionless", "Where used":"04_WELL_ROBUST_STATS", "Notes":""},
        {"Term":"B_*/G_*/R_*", "Meaning":"Robust statistics of raw well-channel intensities in OpenCV order", "Formula":"computed over retained well ROI pixels; suffix = mean, median, sd, p10, p25, p50, p75, p90 or iqr", "Unit":"raw image intensity", "Where used":"04_WELL_ROBUST_STATS", "Notes":"OpenCV internal channel order is B,G,R."},
        {"Term":"Gray_*", "Meaning":"Robust statistics of grayscale intensity", "Formula":"computed over retained ROI pixels; suffix = mean, median, sd, p10, p25, p50, p75, p90 or iqr", "Unit":"raw image intensity", "Where used":"04_WELL_ROBUST_STATS", "Notes":"Diagnostic descriptor used for optical QC."},
        {"Term":"Purple_*", "Meaning":"Robust statistics of the internal purple-color index", "Formula":"Purple = 0.5 x (Red + Blue) - Green; suffix = mean, median, sd, p10, p25, p50, p75, p90 or iqr", "Unit":"raw image intensity", "Where used":"04_WELL_ROBUST_STATS", "Notes":"Diagnostic color index, not the primary quantitative signal."},
        {"Term":"L_*/a_*/b_*", "Meaning":"Robust statistics of CIELAB coordinates", "Formula":"computed from OpenCV CIELAB conversion over retained ROI pixels; suffix = mean, median, sd, p10, p25, p50, p75, p90 or iqr", "Unit":"dimensionless", "Where used":"04_WELL_ROBUST_STATS", "Notes":"CIELAB values are diagnostic descriptors derived from RGB image data."},
        {"Term":"BrightExcludedFraction/BrightExcludedMeanGray/BrightExcessMeanGray/HighlightIndex", "Meaning":"Highlight and bright-pixel QC descriptors", "Formula":"HighlightIndex = BrightExcludedFraction x BrightExcessMeanGray", "Unit":"mixed", "Where used":"04_WELL_ROBUST_STATS, 08_EMPTY_WELLS", "Notes":""},
        {"Term":"ImageWarning/WarningReason", "Meaning":"Well-level optical QC warning and triggering reason", "Formula":"rule-based flags from highlights, trimming and intensity SD", "Unit":"0/1 and text", "Where used":"04_WELL_ROBUST_STATS", "Notes":""},
        {"Term":"floor_source/local_pitch_px/px_per_mm/cyl_r_bg", "Meaning":"Geometry-source and local scale descriptors", "Formula":"reported source, local pitch, pixel/mm scale and background distance-transform radius", "Unit":"mixed", "Where used":"05_GEOMETRY_QC, 06_WELL_BOTTOM", "Notes":""},
        {"Term":"mouth_* / floor_*", "Meaning":"Mouth/floor circle center, radius and score descriptors", "Formula":"projected or refined circle parameters", "Unit":"pixels or score", "Where used":"06_WELL_BOTTOM", "Notes":""},
        {"Term":"shift_px/shift_frac_of_mouth_r/floor_to_mouth_r_ratio/floor_to_mouth_area_ratio", "Meaning":"Mouth-to-floor alignment descriptors", "Formula":"shift distance and relative floor/mouth radius or area ratios", "Unit":"pixels or dimensionless", "Where used":"05_GEOMETRY_QC, 06_WELL_BOTTOM", "Notes":""},
        {"Term":"D_warning/D_critical", "Meaning":"Floor-geometry warning and critical flags", "Formula":"rule-based thresholds on floor/mouth geometry descriptors", "Unit":"0/1", "Where used":"05_GEOMETRY_QC", "Notes":""},
        {"Term":"key/value", "Meaning":"Plate-geometry metadata key and value", "Formula":"embedded plate-geometry database entry", "Unit":"mixed", "Where used":"07_PLATE_GEOMETRY", "Notes":"Geometry is assumed to be ANSI/SLAS-compatible flat-bottom geometry unless stated otherwise."},
        {"Term":"PAbs_Red/PAbs_Green/PAbs_Blue", "Meaning":"RGB pseudo-absorbance channels", "Formula":"log10(MeanBG_channel / MeanW_channel)", "Unit":"dimensionless", "Where used":"08_EMPTY_WELLS, 10_METHOD_COMPARISON, METHOD_COMPARISON.png", "Notes":"Primary quantitative RGB descriptor; diagnostic sheets may include it for comparison."},
        {"Term":"MeanW_Red/MeanW_Green/MeanW_Blue", "Meaning":"Linearized median well intensity for each RGB channel", "Formula":"median ROI channel intensity after gamma linearization", "Unit":"dimensionless", "Where used":"08_EMPTY_WELLS", "Notes":""},
        {"Term":"MeanBG_Red/MeanBG_Green/MeanBG_Blue", "Meaning":"Linearized local background intensity for each RGB channel", "Formula":"local background value after gamma linearization", "Unit":"dimensionless", "Where used":"08_EMPTY_WELLS", "Notes":""},
        {"Term":"RGB to CIELAB", "Meaning":"Diagnostic conversion from RGB/BGR to CIE L*a*b*", "Formula":"OpenCV BGR->Lab for 8-bit sRGB-like images; reference form: linearized sRGB -> XYZ using the standard sRGB-to-XYZ matrix and D65 reference white -> CIE L*a*b*", "Unit":"dimensionless", "Where used":"FIGURE_CIELAB_DELTAE.png, RAW_DATA_DETAILS_CAPTION.txt, 04_WELL_ROBUST_STATS, 11_CIELAB_FITTING", "Notes":"References: OpenCV color conversions; IEC 61966-2-1:1999 sRGB; CIE 1976 L*a*b*; CIE standard illuminant D65."},
        {"Term":"DeltaL/Deltaa/Deltab/DeltaE_ab/DeltaE_ab_chroma", "Meaning":"CIELAB difference descriptors", "Formula":"DeltaL = L - L_ref; Deltaa = a - a_ref; Deltab = b - b_ref; DeltaE_ab = sqrt(DeltaL^2 + Deltaa^2 + Deltab^2); DeltaE_ab_chroma = sqrt(Deltaa^2 + Deltab^2)", "Unit":"dimensionless", "Where used":"FIGURE_CIELAB_DELTAE.png, RAW_DATA_DETAILS_CAPTION.txt, 10_METHOD_COMPARISON, 11_CIELAB_FITTING", "Notes":"Reference: CIE 1976 L*a*b* color-difference form."},
        {"Term":"Dataset/Status/Applicability/Reason", "Meaning":"Applicability status for optional spatial diagnostics", "Formula":"applied or not_applied with reason text", "Unit":"text", "Where used":"09_SPATIAL_DIAGNOSTICS", "Notes":"The sheet is retained even when spatial diagnostics cannot be applied."},
        {"Term":"n/intercept/slope_col/slope_row/R2/corr_col/corr_row", "Meaning":"Spatial-trend fit descriptors", "Formula":"linear trend of the diagnostic signal versus row/column position", "Unit":"mixed", "Where used":"09_SPATIAL_DIAGNOSTICS", "Notes":"Diagnostic only; does not alter quantitative results."},
        {"Term":"IRLS", "Meaning":"Iteratively reweighted least-squares robust linear regression with residual-based weights", "Formula":"minimize sum_i w_i (y_i - (m x_i + q))^2; w_i is updated iteratively from residual magnitude", "Unit":"dimensionless", "Where used":"METHOD_COMPARISON.png, FIGURE_CIELAB_DELTAE.png, 10_METHOD_COMPARISON, 11_CIELAB_FITTING", "Notes":"Reference: Huber, P. J. (1964), Robust Estimation of a Location Parameter. Implementation: custom NumPy-based IRLS using repeated least-squares solves."},
        {"Term":"Method/Family/ComparableGroup/CommonFactorsN/RankMode", "Meaning":"Method-comparison identifiers and comparable-score grouping", "Formula":"text labels and number of factors defining score comparability", "Unit":"mixed", "Where used":"10_METHOD_COMPARISON, METHOD_COMPARISON.png", "Notes":"Scores are directly comparable only within the same ComparableGroup."},
        {"Term":"Score/ScoreFormula", "Meaning":"Common method-ranking score and formula descriptor", "Formula":"SlopeAgreement^2 x sqrt(R2_cal x R2_std_mean) x (1/LOQ) for calibration_plus_stdadd rows", "Unit":"1/" + str(unit_label) + " and text", "Where used":"10_METHOD_COMPARISON, METHOD_COMPARISON.png", "Notes":"Expected/reference values, SNR and clipping are not part of Score."},
        {"Term":"R2_cal/R2_std_mean/m_cal/m_std_mean/SlopeAgreement", "Meaning":"Fit-quality and slope-agreement descriptors for method comparison", "Formula":"SlopeAgreement = min(|m_cal|, |m_std_mean|) / max(|m_cal|, |m_std_mean|)", "Unit":"mixed", "Where used":"10_METHOD_COMPARISON, METHOD_COMPARISON.png", "Notes":""},
        {"Term":"beta_mean/bias_index_mean", "Meaning":"Mean slope ratio and mean relative slope bias", "Formula":"beta_mean = mean(m_std/m_cal); bias_index_mean = mean(|m_std/m_cal - 1|)", "Unit":"dimensionless", "Where used":"10_METHOD_COMPARISON", "Notes":""},
        {"Term":"C0_mean/C0_median/C0_sd_median", "Meaning":"Summary of available original-concentration estimates", "Formula":"mean/median of C0 estimates and median of associated C0_sd values", "Unit":str(unit_label), "Where used":"10_METHOD_COMPARISON", "Notes":""},
        {"Term":"Estimate_value/Estimate_sd/Estimate_source", "Meaning":"Representative diagnostic concentration estimate, SD and source", "Formula":"derived according to Estimate_source", "Unit":str(unit_label) + " and text", "Where used":"METHOD_COMPARISON.png, 10_METHOD_COMPARISON", "Notes":""},
        {"Term":"expected_label_*/expected_id_*/expected_value_*/expected_sd_*", "Meaning":"External reference metadata", "Formula":"user/configurator input", "Unit":"text or " + str(unit_label), "Where used":"METHOD_COMPARISON.png, 10_METHOD_COMPARISON", "Notes":"External reference values are checks only and are not part of Score."},
        {"Term":"estimate_for_expected_*/delta_expected_*/recovery_pct_*/rel_error_*", "Meaning":"External-reference comparison metrics", "Formula":"delta = estimate - reference; recovery = 100 x estimate/reference; relative error = 100 x (estimate - reference)/reference", "Unit":str(unit_label) + " or %", "Where used":"METHOD_COMPARISON.png, 10_METHOD_COMPARISON", "Notes":""},
        {"Term":"Channel/FitType/ID/DF/n_points/m/q/R2/RMSE", "Meaning":"Diagnostic fitting identifiers and linear-fit descriptors", "Formula":"y = m x + q; R2 = 1 - SSE/SST; RMSE = sqrt(mean(residual^2))", "Unit":"mixed", "Where used":"11_CIELAB_FITTING", "Notes":""},
        {"Term":"LOD/LOQ", "Meaning":"Detection and quantification limits", "Formula":"LOD = 3 sigma_cal / |m|; LOQ = 10 sigma_cal / |m|", "Unit":str(unit_label), "Where used":"11_CIELAB_FITTING, 10_METHOD_COMPARISON", "Notes":""},
        {"Term":"C0/C0_sd", "Meaning":"Estimated original concentration and associated SD", "Formula":"standard-addition or calibration projection depending on FitType", "Unit":str(unit_label), "Where used":"11_CIELAB_FITTING", "Notes":""},
        {"Term":"sigma_cal/sigma_source/SNR", "Meaning":"Calibration noise estimate, its source and slope-to-noise ratio", "Formula":"SNR = |m|/sigma_cal", "Unit":"mixed", "Where used":"11_CIELAB_FITTING, 10_METHOD_COMPARISON", "Notes":"SNR is diagnostic and is not part of the common Score."},
        {"Term":"beta_k/bias_index_k", "Meaning":"Per-fit slope ratio and relative slope-bias index", "Formula":"beta_k = m_std/m_cal; bias_index_k = |beta_k - 1|", "Unit":"dimensionless", "Where used":"11_CIELAB_FITTING", "Notes":""},
        {"Term":"Geometry and epsilon/path-length quantification", "Meaning":"Assumption used when epsilon-based unknown quantification is enabled", "Formula":"l_cm = (volume_uL / well_bottom_area_mm2) / 10; C_M = PAbs / (epsilon x l_cm)", "Unit":"cm and M", "Where used":"RESULTS_CAPTION.txt, RAW_DATA_DETAILS_CAPTION.txt, REPORT", "Notes":"Assumes ANSI/SLAS-compatible flat-bottom microplate geometry; non-flat or non-certified geometries require separate validation."},
        {"Term":"Code implementation", "Meaning":"Main computational libraries and code provenance", "Formula":"custom Python code using NumPy, OpenCV, openpyxl and matplotlib", "Unit":"software provenance", "Where used":"DIAGNOSTICS.xlsx and diagnostic PNG files", "Notes":"Sources/libraries: OpenCV for masks and BGR-to-Lab conversion; NumPy for numerical fitting; openpyxl for xlsx export; matplotlib for PNG figures."},
    ])
    return _finalize_legend_rows(rows)

def _csv_rows_to_dicts(path, delimiter=';'):
    rows = []
    if not path or not os.path.exists(path):
        return rows
    try:
        with open(path, 'r', encoding='utf-8') as f:
            rd = csv.DictReader(f, delimiter=delimiter)
            for row in rd:
                rows.append(dict(row))
    except Exception:
        return []
    return rows


def _remove_duplicate_csv_files(csv_paths):
    """Remove CSV files that have been copied into DIAGNOSTICS.xlsx.

    This keeps RAW_DATA_DETAILS readable while preserving standalone files that
    are not represented in the diagnostics workbook.
    """
    for pth in (csv_paths or {}).values():
        try:
            if pth and os.path.exists(pth):
                os.remove(pth)
        except Exception:
            pass


def _diagnostic_rows_are_informative(sheet_key, rows):
    if not rows:
        return False
    if sheet_key == "SPATIAL_DIAGNOSTICS":
        # Keep this sheet even when spatial diagnostics are not applicable.
        # The sheet then explicitly reports Status=not_applied and the applicability rule.
        return True
    return True


def _write_diagnostics_workbook(path, image_basename, unit_label, csv_paths=None, method_comparison_rows=None, cielab_fit_rows=None):
    csv_paths = csv_paths or {}
    wb = Workbook()
    ws_contents = wb.active
    ws_contents.title = "01_CONTENTS"

    contents = [
        {"Sheet":"01_CONTENTS", "Purpose":"Index of diagnostic sheets."},
    ]

    sheet_purposes = {
        "02_BG_SAMPLES": "Accepted inter-well background samples used to fit the BG surface.",
        "03_BG_WELL_FIT": "Predicted local background at each well.",
        "04_WELL_ROBUST_STATS": "Well-level robust pixel statistics and optical QC.",
        "05_GEOMETRY_QC": "Floor/mouth geometry quality-control descriptors.",
        "06_WELL_BOTTOM": "Detailed well-bottom and mouth geometry measurements.",
        "07_PLATE_GEOMETRY": "Nominal plate geometry parameters used by the analyzer.",
        "08_EMPTY_WELLS": "Empty-well diagnostic values when empty wells are present.",
        "09_SPATIAL_DIAGNOSTICS": "Spatial trends across row/column positions.",
        "10_METHOD_COMPARISON": "Cross-method diagnostic comparison using common score factors.",
        "11_CIELAB_FITTING": "CIELAB/DeltaE diagnostic fit rows.",
        "12_LEGENDS": "Definitions for diagnostic workbook fields and figures.",
    }

    csv_sheet_map = [
        ("02_BG_SAMPLES", "BG_SAMPLES"),
        ("03_BG_WELL_FIT", "BG_WELL_FIT"),
        ("04_WELL_ROBUST_STATS", "WELL_ROBUST_STATS"),
        ("05_GEOMETRY_QC", "FLOOR_GEOMETRY_QC"),
        ("06_WELL_BOTTOM", "WELL_BOTTOM"),
        ("07_PLATE_GEOMETRY", "PLATE_GEOMETRY"),
        ("08_EMPTY_WELLS", "EMPTY_WELLS"),
        ("09_SPATIAL_DIAGNOSTICS", "SPATIAL_DIAGNOSTICS"),
    ]
    for sheet_name, key in csv_sheet_map:
        rows = _csv_rows_to_dicts(csv_paths.get(key))
        if not _diagnostic_rows_are_informative(key, rows):
            continue
        wsx = wb.create_sheet(sheet_name)
        rows = _diagnostic_display_rows(key, rows)
        if not rows:
            continue
        headers = list(rows[0].keys())
        _write_table(wsx, 1, headers, rows)
        contents.append({"Sheet": sheet_name, "Purpose": sheet_purposes.get(sheet_name, "Diagnostic data.")})

    if method_comparison_rows:
        wsx = wb.create_sheet("10_METHOD_COMPARISON")
        rows = _display_rows_with_pabs_names(method_comparison_rows)
        cleaned_rows = []
        for row in rows or []:
            rr = dict(row)
            # BaseScore and FinalScore are currently identical to Score in this
            # common-factor diagnostic ranking; remove them to avoid redundancy.
            rr.pop("BaseScore", None)
            rr.pop("FinalScore", None)
            rr.pop("Component", None)
            cleaned_rows.append(rr)
        rows = cleaned_rows
        if rows:
            headers = [h for h in ["Method", "Family", "ComparableGroup", "CommonFactorsN", "Score", "ScoreFormula", "RankMode", "R2_cal", "R2_std_mean", "m_cal", "m_std_mean", "SlopeAgreement", "beta_mean", "bias_index_mean", "C0_median", "C0_sd_median", "Estimate_value", "Estimate_sd", "Estimate_source", "LOD", "LOQ"] if h in rows[0]]
            headers += [h for h in rows[0].keys() if h not in headers]
            _write_table(wsx, 1, headers, rows)
            contents.append({"Sheet":"10_METHOD_COMPARISON", "Purpose":sheet_purposes["10_METHOD_COMPARISON"]})

    if cielab_fit_rows:
        wsx = wb.create_sheet("11_CIELAB_FITTING")
        rows = _display_rows_with_pabs_names(cielab_fit_rows)
        if rows:
            headers = ["Channel", "FitType", "ID", "DF", "n_points", "m", "q", "R2", "RMSE", "LOD", "LOQ", "C0", "C0_sd"]
            headers = [h for h in headers if h in rows[0]] + [h for h in rows[0].keys() if h not in headers and not h.startswith("__")]
            _write_table(wsx, 1, headers, rows)
            contents.append({"Sheet":"11_CIELAB_FITTING", "Purpose":sheet_purposes["11_CIELAB_FITTING"]})

    contents.append({"Sheet":"12_LEGENDS", "Purpose":sheet_purposes["12_LEGENDS"]})
    _write_table(ws_contents, 1, ["Sheet", "Purpose"], contents)

    wsleg = wb.create_sheet("12_LEGENDS")
    _write_table(wsleg, 1, ["Term", "Meaning", "Formula", "Unit", "Where used", "Notes"], _diagnostics_legend_rows(unit_label, present_sheets=wb.sheetnames))
    for wsx in wb.worksheets:
        _autosize_worksheet(wsx)
    wb.save(path)

def _draw_dashed_polyline(img, pts, color=(0, 0, 0), thickness=1, dash=8, gap=6):
    pts = np.asarray(pts, dtype=np.int32).reshape(-1, 2)
    if len(pts) < 2:
        return
    draw = True
    remain = dash
    for i in range(len(pts)):
        p0 = pts[i]
        p1 = pts[(i + 1) % len(pts)]
        seg = p1.astype(np.float64) - p0.astype(np.float64)
        seg_len = float(np.hypot(seg[0], seg[1]))
        if seg_len < 1e-9:
            continue
        u = seg / seg_len
        pos = 0.0
        while pos < seg_len:
            step = min(remain, seg_len - pos)
            a = p0.astype(np.float64) + u * pos
            b = p0.astype(np.float64) + u * (pos + step)
            if draw:
                cv2.line(img, tuple(np.round(a).astype(int)), tuple(np.round(b).astype(int)), color, thickness, cv2.LINE_AA)
            pos += step
            remain -= step
            if remain <= 1e-9:
                draw = not draw
                remain = dash if draw else gap


def _build_well_annular_masks(img_shape, well_bottom_shapes, inner_frac=1.02, outer_frac=1.22):
    h, w = img_shape[:2]
    yy, xx = np.mgrid[0:h, 0:w]

    masks = [[None for _ in range(len(well_bottom_shapes[0]))] for _ in range(len(well_bottom_shapes))]

    for r in range(len(well_bottom_shapes)):
        for c in range(len(well_bottom_shapes[r])):
            s = well_bottom_shapes[r][c]
            if s is None:
                continue

            cx = float(s.get("mouth_cx", s.get("cx", np.nan)))
            cy = float(s.get("mouth_cy", s.get("cy", np.nan)))
            rr = float(s.get("mouth_r", np.nan))

            if not (np.isfinite(cx) and np.isfinite(cy) and np.isfinite(rr) and rr > 2.0):
                continue

            r_in = float(inner_frac * rr)
            r_out = float(outer_frac * rr)

            dist2 = (xx - cx) ** 2 + (yy - cy) ** 2
            ring = (dist2 >= r_in * r_in) & (dist2 <= r_out * r_out)

            m = np.zeros((h, w), dtype=np.uint8)
            m[ring] = 255
            masks[r][c] = m

    return masks


def _overlay_annular_masks_on_plate(img_bgr, annular_masks, color=(0, 255, 255), alpha=0.35):
    vis = img_bgr.copy()
    overlay = vis.copy()

    for r in range(len(annular_masks)):
        for c in range(len(annular_masks[r])):
            m = annular_masks[r][c]
            if m is None or np.count_nonzero(m) == 0:
                continue

            overlay[m > 0] = color

            cnts, _ = cv2.findContours(m, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_NONE)
            if cnts:
                cv2.drawContours(vis, cnts, -1, color, 1, cv2.LINE_AA)

    vis = cv2.addWeighted(overlay, alpha, vis, 1 - alpha, 0)
    return vis

def _overlay_well_centers_on_plate(img_bgr, centers, color=(0, 255, 255), radius=4, cross_half=10, thickness=1):
    vis = img_bgr.copy()
    nrow, ncol = centers.shape[:2]
    for r in range(nrow):
        for c in range(ncol):
            cx = int(round(float(centers[r, c, 0])))
            cy = int(round(float(centers[r, c, 1])))
            cv2.circle(vis, (cx, cy), int(radius), color, thickness, cv2.LINE_AA)
            cv2.line(vis, (cx - int(cross_half), cy), (cx + int(cross_half), cy), color, thickness, cv2.LINE_AA)
            cv2.line(vis, (cx, cy - int(cross_half)), (cx, cy + int(cross_half)), color, thickness, cv2.LINE_AA)
    return vis


def _overlay_mouth_floor_roi_on_plate(img_bgr, well_bottom_masks, well_bottom_shapes,
                                      mouth_color=(0, 255, 0), floor_color=(255, 255, 0),
                                      roi_color=(0, 0, 0), roi_alpha=0.18):
    vis = img_bgr.copy()
    overlay = vis.copy()

    for r in range(len(well_bottom_shapes)):
        for c in range(len(well_bottom_shapes[r])):
            s = well_bottom_shapes[r][c]
            if s is None:
                continue

            mouth_cx = float(s.get("mouth_cx", s.get("cx", np.nan)))
            mouth_cy = float(s.get("mouth_cy", s.get("cy", np.nan)))
            mouth_r = float(s.get("mouth_r", np.nan))
            floor_cx = float(s.get("floor_cx", np.nan))
            floor_cy = float(s.get("floor_cy", np.nan))
            floor_r = float(s.get("floor_r", np.nan))

            if np.isfinite(mouth_cx) and np.isfinite(mouth_cy) and np.isfinite(mouth_r) and mouth_r > 1.0:
                cv2.circle(vis, (int(round(mouth_cx)), int(round(mouth_cy))), int(round(mouth_r)), mouth_color, 1, cv2.LINE_AA)

            if np.isfinite(floor_cx) and np.isfinite(floor_cy) and np.isfinite(floor_r) and floor_r > 1.0:
                cv2.circle(vis, (int(round(floor_cx)), int(round(floor_cy))), int(round(floor_r)), floor_color, 1, cv2.LINE_AA)

            m = well_bottom_masks[r][c] if r < len(well_bottom_masks) and c < len(well_bottom_masks[r]) else None
            if m is not None and np.count_nonzero(m) > 0:
                
                cnts, _ = cv2.findContours(m, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
                if cnts:
                    cv2.drawContours(vis, cnts, -1, roi_color, 1, cv2.LINE_AA)

    return vis

def _build_floor_geometry_qc_rows(well_bottom_shapes):
    rows = []
    shift_fracs = []
    radius_ratios = []
    area_fracs = []
    for r in range(len(well_bottom_shapes)):
        for c in range(len(well_bottom_shapes[r])):
            s = well_bottom_shapes[r][c]
            if s is None:
                continue
            mouth_r = float(s.get("mouth_r", np.nan))
            floor_r = float(s.get("floor_r", np.nan))
            shift_px = float(s.get("shift_px", np.nan))
            local_pitch = float(s.get("local_pitch_px", np.nan))
            floor_src = str(s.get("floor_source", ""))
            shift_frac = shift_px / mouth_r if np.isfinite(shift_px) and np.isfinite(mouth_r) and mouth_r > 1e-9 else np.nan
            radius_ratio = floor_r / mouth_r if np.isfinite(floor_r) and np.isfinite(mouth_r) and mouth_r > 1e-9 else np.nan
            area_frac = (floor_r * floor_r) / (mouth_r * mouth_r) if np.isfinite(floor_r) and np.isfinite(mouth_r) and mouth_r > 1e-9 else np.nan
            is_warning = int(
                (np.isfinite(shift_frac) and shift_frac > 0.38) or
                (np.isfinite(radius_ratio) and (radius_ratio < 0.55 or radius_ratio > 1.02))
            )
            is_critical = int(
                (np.isfinite(shift_frac) and shift_frac > 0.50) or
                (np.isfinite(radius_ratio) and (radius_ratio < 0.45 or radius_ratio > 1.08))
            )
            rows.append({
                "well_r": int(r),
                "well_c": int(c),
                "Well": _well_name_from_indices(int(r), int(c)),
                "floor_source": floor_src,
                "local_pitch_px": local_pitch,
                "mouth_r": mouth_r,
                "floor_r": floor_r,
                "shift_px": shift_px,
                "shift_frac_of_mouth_r": shift_frac,
                "floor_to_mouth_r_ratio": radius_ratio,
                "floor_to_mouth_area_ratio": area_frac,
                "D_warning": is_warning,
                "D_critical": is_critical,
            })
            if np.isfinite(shift_frac):
                shift_fracs.append(shift_frac)
            if np.isfinite(radius_ratio):
                radius_ratios.append(radius_ratio)
            if np.isfinite(area_frac):
                area_fracs.append(area_frac)

    shift_med = float(np.nanmedian(np.asarray(shift_fracs, dtype=np.float64))) if shift_fracs else np.nan
    radius_med = float(np.nanmedian(np.asarray(radius_ratios, dtype=np.float64))) if radius_ratios else np.nan
    area_med = float(np.nanmedian(np.asarray(area_fracs, dtype=np.float64))) if area_fracs else np.nan
    n_warn = int(sum(int(r.get("D_warning", 0)) for r in rows))
    n_crit = int(sum(int(r.get("D_critical", 0)) for r in rows))
    n_tot = int(len(rows))
    status = "OK"
    if n_crit > 0:
        status = "WARNING"
    elif n_warn > 0:
        status = "WATCH"
    payload = {
        "status": status,
        "n_wells": n_tot,
        "warning_wells": n_warn,
        "critical_wells": n_crit,
        "shift_frac_median": shift_med,
        "radius_ratio_median": radius_med,
        "area_ratio_median": area_med,
    }
    return rows, payload


def _add_panel_label(fig, ax, letter, dx=0.028, dy=0.010, fontsize=16):
    """Place panel letter in figure coordinates, aligned to the top-left of the subplot area."""
    bb = ax.get_position()
    x = max(0.001, bb.x0 - dx)
    y = min(0.995, bb.y1 + dy)
    fig.text(x, y, letter, ha="left", va="top", fontsize=fontsize, fontweight="bold")


def _nice_step(v):
    if v <= 0 or not np.isfinite(v):
        return 1.0
    exp = np.floor(np.log10(v))
    frac = v / (10 ** exp)
    if frac <= 1:
        base = 1
    elif frac <= 2:
        base = 2
    elif frac <= 2.5:
        base = 2.5
    elif frac <= 5:
        base = 5
    else:
        base = 10
    return float(base * (10 ** exp))


def _infer_x_axis_label(plot_payload, unit_label):
    has_stdadd = False
    has_unknown = False
    has_cal = False
    for _payload in (plot_payload or {}).values():
        if (_payload or {}).get("stdadd_groups"):
            has_stdadd = True
        if (_payload or {}).get("unknown_points"):
            has_unknown = True
        if (_payload or {}).get("calibration_points"):
            has_cal = True
    if has_stdadd:
        return f"Added concentration ({unit_label})"
    if has_unknown and has_cal:
        return f"Concentration from calibration ({unit_label})"
    return f"Concentration ({unit_label})"



def _pretty_channel_label(label):
    s = str(label or "")
    s = s.replace("Signal_", "")
    s = s.replace("PAbs_", "")
    s = s.replace("DeltaE_ab_chroma", "ΔE_ab,chrom")
    s = s.replace("DeltaE_ab", "ΔE_ab")
    s = s.replace("DeltaL", "ΔL")
    s = s.replace("Deltaa", "Δa")
    s = s.replace("Deltab", "Δb")
    return s


_BOLD_LINE_PREFIX = "__BOLD_LINE__"

def _bold_unicode_text(text):
    # Marker only. Actual bold rendering is handled line-by-line in the PNG writer.
    # Do not use mathematical-bold Unicode: some systems render it as square boxes.
    return _BOLD_LINE_PREFIX + str(text)


def _round_sig_value(value, sig=2):
    try:
        value = float(value)
    except Exception:
        return np.nan
    if not np.isfinite(value):
        return np.nan
    if abs(value) < 1e-15:
        return 0.0
    decimals = int(sig - 1 - math.floor(math.log10(abs(value))))
    decimals = max(0, decimals)
    return float(round(value, decimals))


def _fmt_sig(value, sig=3, max_decimals=4):
    try:
        value = float(value)
    except Exception:
        return "NA"
    if not np.isfinite(value):
        return "NA"
    if abs(value) < 1e-15:
        return "0"
    decimals = int(sig - 1 - math.floor(math.log10(abs(value))))
    decimals = max(0, min(int(max_decimals), decimals))
    txt = f"{value:.{decimals}f}"
    if "." in txt:
        txt = txt.rstrip("0").rstrip(".")
    return txt


def _fmt_measure(value, sd=None, sig_sd=1, sig_value=3):
    try:
        value = float(value)
    except Exception:
        return "NA"
    if not np.isfinite(value):
        return "NA"
    if sd is None or not np.isfinite(sd):
        return _fmt_sig(value, sig=sig_value, max_decimals=3)
    sd = abs(float(sd))
    if sd <= 0 or not np.isfinite(sd):
        return _fmt_sig(value, sig=sig_value, max_decimals=3)
    sd_r = _round_sig_value(sd, sig=sig_sd)
    if not np.isfinite(sd_r) or sd_r <= 0:
        return _fmt_sig(value, sig=sig_value, max_decimals=3)
    decimals = int(max(0, -math.floor(math.log10(sd_r)))) if sd_r < 1 else 0
    value_r = round(value, decimals)
    fmt = f"{{:.{decimals}f}}"
    return f"{fmt.format(value_r)} ± {fmt.format(sd_r)}"


def _decimals_for_sig(values, sig=3, max_decimals=5):
    vals = []
    for v in values or []:
        try:
            x = float(v)
        except Exception:
            continue
        if np.isfinite(x) and abs(x) > 1e-15:
            vals.append(abs(x))
    if not vals:
        return 0
    vmax = max(vals)
    dec = int(sig - 1 - math.floor(math.log10(vmax)))
    return max(0, min(int(max_decimals), dec))


def _fmt_fixed_dec(value, decimals):
    try:
        x = float(value)
    except Exception:
        return "NA"
    if not np.isfinite(x):
        return "NA"
    return f"{x:.{int(decimals)}f}"


def _fmt_integer(value):
    try:
        x = float(value)
    except Exception:
        return "NA"
    if not np.isfinite(x):
        return "NA"
    return f"{round(x):.0f}"


def _fmt_c0_pm_aligned(values_sds, sig_sd=1):
    entries = []
    for value, sd in values_sds or []:
        try:
            v = float(value)
        except Exception:
            v = np.nan
        try:
            e = abs(float(sd))
        except Exception:
            e = np.nan
        if not np.isfinite(v):
            entries.append(("NA", "", ""))
            continue
        if not np.isfinite(e) or e <= 0:
            txt = _fmt_sig(v, sig=3, max_decimals=3)
            entries.append((txt, "", ""))
            continue
        e_r = _round_sig_value(e, sig=sig_sd)
        if not np.isfinite(e_r) or e_r <= 0:
            txt = _fmt_sig(v, sig=3, max_decimals=3)
            entries.append((txt, "", ""))
            continue
        decimals = int(max(0, -math.floor(math.log10(e_r)))) if e_r < 1 else 0
        v_r = round(v, decimals)
        fmt = f"{{:.{decimals}f}}"
        entries.append((fmt.format(v_r), "±", fmt.format(e_r)))
    wv = max([len(a) for a, _, _ in entries] + [1])
    ws = max([len(c) for _, _, c in entries] + [1])
    out = []
    for a, pm, c in entries:
        if pm:
            out.append(f"{a:>{wv}} {pm} {c:>{ws}}")
        else:
            out.append(f"{a:>{wv}}")
    return out


def _format_c0_text(c0, c0_sd=None, unit_label=""):
    if not np.isfinite(c0):
        return "C₀=NA"
    txt = "C₀=" + _fmt_measure(c0, c0_sd, sig_sd=1, sig_value=3)
    if unit_label:
        txt += f" {unit_label}"
    return txt




def _single_column_figsize_from_image(reference_img=None, width_cm=9.0, height_cm=6.0):
    """Return a fixed single-column matplotlib figsize.

    The output canvas is deliberately independent of the source image pixel
    dimensions, so all exported single-column PNGs have exactly the same
    physical size and pixel dimensions at the requested DPI.
    The default 9 x 6 cm canvas follows the plate ROI overlay proportions
    used for the single-column figure export.
    """
    width_in = float(width_cm) / 2.54
    height_in = float(height_cm) / 2.54
    if not np.isfinite(width_in) or width_in <= 0:
        width_in = 9.0 / 2.54
    if not np.isfinite(height_in) or height_in <= 0:
        height_in = 6.0 / 2.54
    return width_in, height_in


def _save_plate_roi_overlay_single_column_png(path, overlay_img, width_cm=9.0, height_cm=6.0, dpi=300):
    """Save ROI overlay as a fixed-size single-column PNG with embedded DPI."""
    if not path or overlay_img is None:
        return
    try:
        os.makedirs(os.path.dirname(path), exist_ok=True)
        fig_w, fig_h = _single_column_figsize_from_image(overlay_img, width_cm=width_cm, height_cm=height_cm)
        fig = plt.figure(figsize=(fig_w, fig_h), dpi=int(dpi))
        ax = fig.add_axes([0.0, 0.0, 1.0, 1.0])
        ax.imshow(cv2.cvtColor(overlay_img, cv2.COLOR_BGR2RGB), aspect='equal')
        ax.set_axis_off()
        fig.savefig(path, dpi=int(dpi), bbox_inches=None, pad_inches=0, facecolor='white')
        plt.close(fig)
    except Exception:
        try:
            plt.close(fig)
        except Exception:
            pass


def _make_a4_report_png(path, image_basename, overlay_img, plot_payload, fit_rows, selected_channel, selection_info, unit_label, fit_options=None, plate_qc_info=None, comparison_payload=None, cluster_estimate=None, fit_rows_bg2d=None, bg2d_scale_summary=None, expected_refs=None, best_channel_plot_path=None):
    matplotlib.rcParams.update({
        "font.family": "DejaVu Sans",
        "font.size": 8.8,
        "axes.titlesize": 10.0,
        "axes.labelsize": 9.4,
        "xtick.labelsize": 7.8,
        "ytick.labelsize": 7.8,
        "mathtext.default": "regular",
    })

    from matplotlib.ticker import AutoMinorLocator

    fit_options = fit_options or {}
    selection_info = selection_info or {}
    plot_payload = plot_payload or {}
    fit_rows = fit_rows or []
    plate_qc_info = plate_qc_info or {}
    expected_refs = expected_refs or []

    def _wrap_lines(lines, width):
        out = []
        for line in lines:
            s = str(line)
            is_bold = s.startswith(_BOLD_LINE_PREFIX)
            if is_bold:
                s = s[len(_BOLD_LINE_PREFIX):]
            if s == "":
                out.append("")
                continue
            wrapped = textwrap.wrap(s, width=width, break_long_words=False, break_on_hyphens=False)
            if not wrapped:
                wrapped = [""]
            for j, wline in enumerate(wrapped):
                out.append((_BOLD_LINE_PREFIX if is_bold else "") + wline)
        return out

    def _clean_limits(vmin, vmax, frac=0.10, min_span=1e-6):
        if not np.isfinite(vmin) or not np.isfinite(vmax):
            return -1.0, 1.0
        if vmax < vmin:
            vmin, vmax = vmax, vmin
        span = max(vmax - vmin, min_span)
        pad = frac * span
        if abs(vmin) < 1e-12 and abs(vmax) < 1e-12:
            pad = 0.1
        return vmin - pad, vmax + pad

    def _nice_tick_step(span, target_intervals=5):
        span = abs(float(span)) if np.isfinite(span) and abs(float(span)) > 1e-15 else 1.0
        raw = span / max(1.0, float(target_intervals))
        exp = math.floor(math.log10(raw)) if raw > 0 else 0
        base = 10.0 ** exp
        for m in (1.0, 2.0, 5.0, 10.0):
            step = m * base
            if step >= raw:
                return float(step)
        return float(10.0 * base)

    def _clean_axis_with_ticks(vmin, vmax, nticks=5, frac=0.08, min_span=1e-6):
        lo, hi = _clean_limits(vmin, vmax, frac=frac, min_span=min_span)
        span = max(hi - lo, min_span)
        step = _nice_tick_step(span, target_intervals=max(4, int(nticks)))
        lo = math.floor(lo / step) * step
        hi = math.ceil(hi / step) * step
        if hi <= lo:
            hi = lo + step * max(2, int(nticks))
        ticks = np.arange(lo + step, hi, step, dtype=np.float64)
        # Remove endpoint ticks robustly: labels on panel borders collide with adjacent axes.
        eps = max(abs(step) * 1e-7, 1e-12)
        ticks = ticks[(ticks > lo + eps) & (ticks < hi - eps)]
        if ticks.size == 0:
            mid = 0.5 * (lo + hi)
            if mid > lo + eps and mid < hi - eps:
                ticks = np.asarray([mid], dtype=np.float64)
            else:
                ticks = np.asarray([], dtype=np.float64)
        return float(lo), float(hi), np.round(ticks, 10)

    def _plain_line(line):
        s = str(line)
        return s[len(_BOLD_LINE_PREFIX):] if s.startswith(_BOLD_LINE_PREFIX) else s

    def _text_table(headers, rows):
        headers = [str(h) for h in headers]
        clean_rows = []
        bold_rows = []
        for row in rows:
            row = list(row)
            is_bold = False
            if row and str(row[0]).startswith(_BOLD_LINE_PREFIX):
                is_bold = True
                row[0] = str(row[0])[len(_BOLD_LINE_PREFIX):]
            clean_rows.append([str(c) for c in row])
            bold_rows.append(is_bold)
        widths = [len(h) for h in headers]
        for row in clean_rows:
            for i, cell in enumerate(row):
                if i < len(widths):
                    widths[i] = max(widths[i], len(cell))
        def fmt(row):
            cells = []
            for i, cell in enumerate(row):
                text = str(cell)
                if i == 0:
                    cells.append(text.ljust(widths[i]))
                else:
                    cells.append(text.rjust(widths[i]))
            return "  ".join(cells)
        out = [fmt(headers), fmt(["-" * w for w in widths])]
        for row, is_bold in zip(clean_rows, bold_rows):
            line = fmt(row)
            out.append((_BOLD_LINE_PREFIX if is_bold else "") + line)
        return out

    def _safe_txt(v, fmt="{:.2f}"):
        if not np.isfinite(v):
            return "NA"
        try:
            return fmt.format(float(v))
        except Exception:
            return _fmt_sig(v, sig=3, max_decimals=4)

    def _plate_label_from_total(total_wells):
        try:
            total_wells = int(total_wells)
        except Exception:
            return "plate"
        mapping = {96: "96-well", 384: "384-well", 1536: "1536-well"}
        return mapping.get(total_wells, f"{total_wells}-well")

    def _format_fit_equation(fit):
        if fit is None or not np.isfinite(fit.get("m", np.nan)) or not np.isfinite(fit.get("q", np.nan)):
            return "not available"
        m = float(fit["m"])
        q = float(fit["q"])
        r2 = float(fit.get("R2", np.nan))
        sign = "+" if q >= 0 else "−"
        return f"y = {m:.5g}x {sign} {abs(q):.5g}, R² = {r2:.4f}"

    selected_channel_upper = str(selected_channel).upper()
    if selected_channel_upper == "CIELAB_DELTAE":
        axis_labels = ["DeltaE_ab", "DeltaE_ab_chroma", "DeltaL", "Deltaa", "Deltab"]
    elif selected_channel_upper == "CIELAB":
        axis_labels = ["DeltaL", "Deltaa", "Deltab"]
    elif selected_channel_upper == "DELTAE":
        axis_labels = ["DeltaE_ab", "DeltaE_ab_chroma", "DeltaL"]
    elif selected_channel_upper.startswith("SIGNAL_") or selected_channel_upper == "SIGNAL":
        axis_labels = ["Signal_Red", "Signal_Green", "Signal_Blue"]
    elif selected_channel_upper.startswith("PABS_") or selected_channel_upper == "PABS":
        axis_labels = ["PAbs_Red", "PAbs_Green", "PAbs_Blue"]
    else:
        axis_labels = ["Signal_Red", "Signal_Green", "Signal_Blue"]

    channel_color = {
        "Red": "red",
        "Green": "green",
        "Blue": "blue",
        "Signal_Red": "red",
        "Signal_Green": "green",
        "Signal_Blue": "blue",
        "PAbs_Red": "red",
        "PAbs_Green": "green",
        "PAbs_Blue": "blue",
        "DeltaE_ab": "tab:gray",
        "DeltaE_ab_chroma": "tab:brown",
        "DeltaL": "tab:olive",
        "Deltaa": "tab:purple",
        "Deltab": "tab:cyan",
        "L": "tab:olive",
        "a": "tab:purple",
        "b": "tab:cyan",
    }
    y_label_map = {
        "Red": "Signal (Red)",
        "Green": "Signal (Green)",
        "Blue": "Signal (Blue)",
        "Signal_Red": "PAbs_Red",
        "Signal_Green": "PAbs_Green",
        "Signal_Blue": "PAbs_Blue",
        "PAbs_Red": "PAbs_Red",
        "PAbs_Green": "PAbs_Green",
        "PAbs_Blue": "PAbs_Blue",
        "DeltaL": r"$\Delta L$",
        "Deltaa": r"$\Delta a$",
        "Deltab": r"$\Delta b$",
        "DeltaE_ab": r"$\Delta E_{ab}$",
        "DeltaE_ab_chroma": r"$\Delta E_{ab,chrom}$",
        "DeltaL": r"$\Delta L$",
    }

    def _stdadd_group_key(sample_id, df):
        return (str(sample_id or ""), str(df if df is not None else ""))

    stdadd_marker_cycle = ["s", "D", "^", "v", "P", "X", "o"]
    stdadd_marker_by_key = {}
    for _label in axis_labels:
        _payload = plot_payload.get(_label, {}) or {}
        for _grp in (_payload.get("stdadd_groups", []) or []):
            _key = _stdadd_group_key(_grp.get("ID", ""), _grp.get("DF", np.nan))
            if _key not in stdadd_marker_by_key:
                stdadd_marker_by_key[_key] = stdadd_marker_cycle[len(stdadd_marker_by_key) % len(stdadd_marker_cycle)]

    def _stdadd_marker(sample_id, df):
        key = _stdadd_group_key(sample_id, df)
        if key not in stdadd_marker_by_key:
            stdadd_marker_by_key[key] = stdadd_marker_cycle[len(stdadd_marker_by_key) % len(stdadd_marker_cycle)]
        return stdadd_marker_by_key[key]

    fit_map = {}
    for row in fit_rows:
        ch = row.get("Channel")
        if ch not in fit_map:
            fit_map[ch] = {"Calibration": [], "StdAdd": [], "UnknownFromCal": [], "UnknownFromEpsilon": []}
        fit_map[ch].setdefault(row.get("FitType"), []).append(row)

    if selected_channel_upper == "CIELAB" and isinstance(plot_payload.get("__selection_info__", None), dict):
        selection_info = plot_payload.get("__selection_info__", selection_info)

    ranking_rows = selection_info.get("ranking", []) or []
    weighting_txt = "Fit: robust IRLS"
    mode_flags = _report_mode_flags(fit_rows)
    has_stdadd = bool(mode_flags.get("has_stdadd", False))
    has_unknown = bool(mode_flags.get("has_unknown", False))
    has_calibration = bool(mode_flags.get("has_calibration", False))

    if has_calibration and (not has_stdadd) and (not has_unknown):
        mode_line = "Mode: calibration only"
        ranking_formula_math = "fit quality score"
    elif has_calibration and has_stdadd and has_unknown:
        mode_line = "Mode: calibration + standard addition + unknown"
        ranking_formula_math = "fit quality score"
    elif has_calibration and has_stdadd:
        mode_line = "Mode: calibration + standard addition"
        ranking_formula_math = "fit quality score"
    elif has_calibration and has_unknown:
        mode_line = "Mode: calibration + unknown"
        ranking_formula_math = "fit quality score"
    elif has_stdadd:
        mode_line = "Mode: standard addition only"
        ranking_formula_math = "fit quality score"
    elif has_unknown:
        mode_line = "Mode: unknown only"
        ranking_formula_math = "fit quality score"
    else:
        mode_line = "Mode: no valid analytical fit available"
        ranking_formula_math = "fit quality score"

    total_wells = plate_qc_info.get("total_wells", np.nan)
    plate_label = _plate_label_from_total(total_wells)
    plate_status_raw = str(plate_qc_info.get("status", "NA"))
    plate_status = "Passed" if plate_status_raw.strip().upper() == "OK" else plate_status_raw
    flagged = plate_qc_info.get("flagged_wells", 0)
    critical = plate_qc_info.get("critical_wells", 0)
    total_txt = int(total_wells) if np.isfinite(total_wells) else "NA"

    best_channel_raw = str((ranking_rows[0].get("Channel", "") if ranking_rows else "")).strip()

    def _is_best_channel(label):
        return str(label or "").strip() == best_channel_raw

    def _maybe_bold_line(label, line):
        return _bold_unicode_text(line) if _is_best_channel(label) else line

    def _clip_factor_from_calibration_row(cal_row):
        if cal_row is None:
            return np.nan
        clip_x = str(cal_row.get("ClipX", "") or "")
        clip_delta = str(cal_row.get("ClipDelta", "") or "")
        xs = [v for v in clip_x.split(",") if str(v).strip() != ""]
        ds = []
        for v in clip_delta.split(","):
            try:
                ds.append(float(v))
            except Exception:
                pass
        n_total = len(xs) if len(xs) > 0 else len(ds)
        if n_total <= 0:
            return 1.0
        n_clip = int(sum(1 for d in ds if np.isfinite(d) and d > 0))
        return float(max(0.0, min(1.0, (n_total - n_clip) / max(1, n_total))))

    def _calibration_score_eff_for_label(label):
        cal_rows = fit_map.get(label, {}).get("Calibration", []) or []
        if not cal_rows:
            return np.nan
        cal_row = cal_rows[0]
        mcal = abs(_num_or_nan(cal_row.get("m", np.nan)))
        r2_cal = _num_or_nan(cal_row.get("R2", np.nan))
        if not (np.isfinite(mcal) and np.isfinite(r2_cal)):
            return np.nan
        return float((max(r2_cal, 0.0) ** 2) * mcal)

    def _stdadd_score_for_label(label, std_row):
        cal_rows = fit_map.get(label, {}).get("Calibration", []) or []
        if not cal_rows or std_row is None:
            return np.nan
        cal_row = cal_rows[0]
        mcal = abs(_num_or_nan(cal_row.get("m", np.nan)))
        mstd = abs(_num_or_nan(std_row.get("m", np.nan)))
        r2_cal = _num_or_nan(cal_row.get("R2", np.nan))
        r2_std = _num_or_nan(std_row.get("R2", np.nan))
        if not (np.isfinite(mcal) and np.isfinite(mstd) and mcal > 0 and mstd > 0):
            return np.nan
        slope_agreement = min(mcal, mstd) / max(mcal, mstd)
        loq = _num_or_nan(cal_row.get("LOQ", np.nan))
        score = _compute_fit_base_score(r2_cal, r2_std, slope_agreement, loq=loq)
        return float(score) if np.isfinite(score) and score > 0 else 0.0

    # Only sample/application results are reported in this section. Calibration
    # LOD/LOQ values are reported once, together with the calibration fit
    # parameters in the model table below, to avoid duplicate calibration blocks.
    key_lines = []

    fit_groups = []
    seen_groups = set()
    for label in axis_labels:
        fit_rows_ch = fit_map.get(label, {})
        for row in ((fit_rows_ch.get("StdAdd", []) or []) + (fit_rows_ch.get("UnknownFromCal", []) or []) + (fit_rows_ch.get("UnknownFromEpsilon", []) or []))[:4]:
            key = (str(row.get("FitType", "")), str(row.get("ID", "")), str(row.get("DF", "")))
            if key not in seen_groups:
                seen_groups.add(key)
                fit_groups.append(key)
    for idx_group, (fit_type, sample_id, df_txt) in enumerate(fit_groups[:4]):
        if key_lines:
            key_lines.append("")
        method_txt = "Std Add" if fit_type == "StdAdd" else ("Unknown epsilon" if fit_type == "UnknownFromEpsilon" else "Unknown")
        key_lines.append(f"{method_txt} | ID: {sample_id} | DF={df_txt}")
        result_rows_raw = []
        for label in axis_labels:
            pretty = _pretty_channel_label(label)
            rows_for_label = [r for r in (fit_map.get(label, {}).get(fit_type, []) or []) if str(r.get("ID", "")) == sample_id and str(r.get("DF", "")) == df_txt]
            for row in rows_for_label[:1]:
                c0 = row.get("C0", np.nan)
                c0_sd = row.get("C0_sd", np.nan)
                score_val = _stdadd_score_for_label(label, row)
                recovery_txt = ""
                delta_txt = ""
                for i_ref, ref in enumerate(expected_refs or [], start=1):
                    if not _reference_matches_fit_row(ref, row):
                        continue
                    ref_val = _num_or_nan(ref.get('value', np.nan))
                    if np.isfinite(ref_val) and abs(ref_val) > 1e-15 and np.isfinite(c0):
                        delta = float(c0) - ref_val
                        recovery = 100.0 * float(c0) / ref_val
                        delta_txt = _fmt_sig(delta, sig=3, max_decimals=2)
                        recovery_txt = _fmt_sig(recovery, sig=3, max_decimals=1)
                        break
                if np.isfinite(c0):
                    result_rows_raw.append([
                        label, pretty,
                        _format_c0_text(c0, c0_sd, "").replace("C₀=", ""),
                        score_val if np.isfinite(score_val) else np.nan,
                        _num_or_nan(delta_txt) if delta_txt else np.nan,
                        _num_or_nan(recovery_txt) if recovery_txt else np.nan,
                        c0, c0_sd,
                    ])
        if result_rows_raw:
            finite_scores = [_num_or_nan(r[3]) for r in result_rows_raw if np.isfinite(_num_or_nan(r[3]))]
            best_score = max(finite_scores) if finite_scores else np.nan
            score_dec = _decimals_for_sig(finite_scores, sig=3, max_decimals=5)
            c0_texts = _fmt_c0_pm_aligned([(r[6], r[7]) for r in result_rows_raw], sig_sd=1)
            delta_dec = _decimals_for_sig([_num_or_nan(r[4]) for r in result_rows_raw], sig=3, max_decimals=2)
            result_rows = []
            for idx_r, (label, pretty, _c0txt, scoretxt, deltatxt, rectxt, c0_val_raw, c0_sd_raw) in enumerate(result_rows_raw):
                score_val = _num_or_nan(scoretxt)
                if np.isfinite(score_val) and np.isfinite(best_score) and abs(score_val - best_score) <= max(1e-15, 0.5 * 10 ** (-score_dec)):
                    pretty = _BOLD_LINE_PREFIX + pretty
                delta_val = _num_or_nan(deltatxt)
                recovery_val = _num_or_nan(rectxt)
                result_rows.append([
                    pretty,
                    c0_texts[idx_r],
                    _fmt_fixed_dec(score_val, score_dec) if np.isfinite(score_val) else "NA",
                    _fmt_fixed_dec(delta_val, delta_dec) if np.isfinite(delta_val) else "NA",
                    _fmt_integer(recovery_val) if np.isfinite(recovery_val) else "NA",
                ])
            key_lines.extend(_text_table(["Channel", f"C0 ({unit_label})", "Score", f"Delta ({unit_label})", "Recovery (%)"], result_rows))

    summary_lines = []
    model_rows_raw = []
    for label in axis_labels:
        payload = plot_payload.get(label, {})
        cal_fit = payload.get("calibration_fit")
        cal_row = (fit_map.get(label, {}).get("Calibration", []) or [None])[0]
        if cal_fit is not None and np.isfinite(cal_fit.get("m", np.nan)):
            lod_val = _num_or_nan(cal_row.get("LOD", np.nan)) if cal_row is not None else np.nan
            loq_val = _num_or_nan(cal_row.get("LOQ", np.nan)) if cal_row is not None else np.nan
            model_rows_raw.append([
                _pretty_channel_label(label),
                cal_fit.get("m", np.nan),
                cal_fit.get("q", np.nan),
                cal_fit.get("R2", np.nan),
                lod_val,
                loq_val,
            ])
    if model_rows_raw:
        summary_lines.append("Calibration")
        lod_dec_model = _decimals_for_sig([r[4] for r in model_rows_raw], sig=3, max_decimals=3)
        loq_dec_model = _decimals_for_sig([r[5] for r in model_rows_raw], sig=3, max_decimals=3)
        model_rows = []
        for pretty, slope, intercept, r2, lod_val, loq_val in model_rows_raw:
            model_rows.append([
                pretty,
                _fmt_sig(slope, sig=4, max_decimals=6),
                _fmt_sig(intercept, sig=4, max_decimals=6),
                _fmt_sig(r2, sig=4, max_decimals=4),
                _fmt_fixed_dec(lod_val, lod_dec_model) if np.isfinite(lod_val) else "NA",
                _fmt_fixed_dec(loq_val, loq_dec_model) if np.isfinite(loq_val) else "NA",
            ])
        summary_lines.extend(_text_table(["Channel", "Slope", "Intercept", "R2", f"LOD ({unit_label})", f"LOQ ({unit_label})"], model_rows))
    summary_groups = []
    seen_summary_groups = set()
    for label in axis_labels:
        payload = plot_payload.get(label, {})
        for grp in (payload.get("stdadd_groups", []) or [])[:2]:
            key = (str(grp.get("ID", "")), str(grp.get("DF", "")))
            if key not in seen_summary_groups:
                seen_summary_groups.add(key)
                summary_groups.append(key)
    for sample_id, df_txt in summary_groups[:4]:
        summary_lines.append("")
        summary_lines.append(f"Std Add | ID: {sample_id} | DF={df_txt}")
        std_model_rows = []
        for label in axis_labels:
            payload = plot_payload.get(label, {})
            for grp in (payload.get("stdadd_groups", []) or []):
                if str(grp.get("ID", "")) == sample_id and str(grp.get("DF", "")) == df_txt:
                    fit = grp.get("fit")
                    if fit is not None and np.isfinite(fit.get("m", np.nan)):
                        std_model_rows.append([
                            _pretty_channel_label(label),
                            _fmt_sig(fit.get("m", np.nan), sig=4, max_decimals=6),
                            _fmt_sig(fit.get("q", np.nan), sig=4, max_decimals=6),
                            _fmt_sig(fit.get("R2", np.nan), sig=4, max_decimals=4),
                        ])
                    break
        if std_model_rows:
            summary_lines.extend(_text_table(["Channel", "Slope", "Intercept", "R2"], std_model_rows))

    if selected_channel_upper == "CIELAB_DELTAE":
        formula_lines = [
            "DeltaE_ab = sqrt((L-Lref)^2 + (a-aref)^2 + (b-bref)^2)",
            "DeltaE_ab,chrom = sqrt((a-aref)^2 + (b-bref)^2)",
            "",
        ]
    elif selected_channel_upper == "DELTAE":
        formula_lines = [
            "DeltaE_ab = sqrt((L-Lref)^2 + (a-aref)^2 + (b-bref)^2)",
            "",
        ]
    else:
        formula_lines = [
            "Pseudo-absorbance = log10(I_BG/I_well)",
            "",
        ]

    info_lines = formula_lines + [
        mode_line,
        weighting_txt,
        f"Plate: {plate_label} | QC: {plate_status}",
        f"Plate QC: wells flagged {flagged}/{total_txt} | wells critical {critical}/{total_txt}",
        f"Floor D QC: {str(plate_qc_info.get('floor_qc_status', 'NA'))} | D warnings {int(plate_qc_info.get('floor_warning_wells', 0) or 0)}/{total_txt} | D critical {int(plate_qc_info.get('floor_critical_wells', 0) or 0)}/{total_txt}",
        "",
    ]
    ref_lines = _reference_info_lines(fit_rows, selected_channel, unit_label, expected_refs=expected_refs)
    if ref_lines:
        info_lines.extend(ref_lines)
        info_lines.append("")
    info_lines.extend(key_lines)
    info_lines.append("")
    info_lines.extend(summary_lines)
    wrapped_info_lines = _wrap_lines(info_lines, width=66)

    # True adaptive layout in physical units (inches), without compressing right panels.
    fig_w = 8.27
    left_x = 0.055
    left_w = 0.43
    gap = 0.055
    right_x = left_x + left_w + gap
    right_w = 0.41

    top_margin_in = 0.22
    bottom_margin_in = 0.55
    img_h_in = 1.92
    img_to_text_gap_in = 0.20
    formula_block_in = 0.0
    line_h_in = 0.126
    n_panels = max(1, len(axis_labels))
    panel_h_in = 1.86
    panel_gap_in = 0.0
    panel_stack_h_in = n_panels * panel_h_in + max(0, n_panels - 1) * panel_gap_in
    info_h_in = formula_block_in + max(1, len(wrapped_info_lines)) * line_h_in
    left_stack_h_in = img_h_in + img_to_text_gap_in + info_h_in
    fig_h = max(top_margin_in + panel_stack_h_in + bottom_margin_in,
                top_margin_in + left_stack_h_in + bottom_margin_in)

    fig = plt.figure(figsize=(fig_w, fig_h), dpi=300)

    def _fy(inches):
        return inches / fig_h

    top_anchor = 1.0 - _fy(top_margin_in)
    panel_h = _fy(panel_h_in)
    panel_gap = _fy(panel_gap_in)
    img_h = _fy(img_h_in)
    img_gap = _fy(img_to_text_gap_in)
    info_h = _fy(info_h_in)
    bottom_margin = _fy(bottom_margin_in)

    img_y = top_anchor - img_h
    info_top = img_y - img_gap
    info_y = max(bottom_margin, info_top - info_h)
    info_h = info_top - info_y

    ax_img = fig.add_axes([left_x, img_y, left_w, img_h])
    ax_info = fig.add_axes([left_x, info_y, left_w, info_h])

    axes = []
    share_ax = None
    for idx_panel, axis_label in enumerate(axis_labels):
        y_panel = top_anchor - (idx_panel + 1) * panel_h - idx_panel * panel_gap
        ax_panel = fig.add_axes([right_x, y_panel, right_w, panel_h], sharex=share_ax)
        if share_ax is None:
            share_ax = ax_panel
        axes.append((ax_panel, axis_label))

    if overlay_img is not None:
        ax_img.imshow(cv2.cvtColor(overlay_img, cv2.COLOR_BGR2RGB), aspect='equal')
    ax_img.set_anchor('W')
    ax_img.axis("off")

    ax_info.axis("off")
    ax_info.set_xlim(0, 1)
    ax_info.set_ylim(0, 1)
    info_top_y = 0.995

    # Render report text line-by-line so selected-channel rows can be bold
    # without relying on unsupported mathematical-bold Unicode characters.
    base_line_step = (line_h_in / max(info_h_in, 1e-9)) * 0.92
    fit_line_step = (info_top_y - 0.012) / max(1, len(wrapped_info_lines))
    line_step = min(base_line_step, fit_line_step)
    text_fontsize = 6.35 if len(wrapped_info_lines) <= 58 else 5.95
    y_line = info_top_y
    for raw_line in wrapped_info_lines:
        if y_line < 0.002:
            break
        is_bold = str(raw_line).startswith(_BOLD_LINE_PREFIX)
        line = str(raw_line)[len(_BOLD_LINE_PREFIX):] if is_bold else str(raw_line)
        ax_info.text(
            0.00, y_line, line,
            va="top", ha="left", fontsize=text_fontsize,
            fontfamily="DejaVu Sans Mono",
            fontweight=("bold" if is_bold else "normal"),
        )
        y_line -= line_step

    all_x = []
    all_intercepts = []
    for label in axis_labels:
        payload = plot_payload.get(label, {})
        for p in payload.get("calibration_points", []) or []:
            xv = float(p.get("x", np.nan))
            if np.isfinite(xv):
                all_x.append(xv)
        for grp in payload.get("stdadd_groups", []) or []:
            for xv in np.asarray(grp.get("x", []), dtype=np.float64):
                if np.isfinite(xv):
                    all_x.append(float(xv))
            fit = grp.get("fit")
            if fit is not None and np.isfinite(fit.get("m", np.nan)) and np.isfinite(fit.get("q", np.nan)) and abs(float(fit["m"])) > 1e-15:
                all_intercepts.append(float(-fit["q"] / fit["m"]))
        cal_pts_local = payload.get("calibration_points", []) or []
        cal_x_local = [float(p.get("x", np.nan)) for p in cal_pts_local if np.isfinite(p.get("x", np.nan))]
        cal_x_min_local = min(cal_x_local) if len(cal_x_local) > 0 else np.nan
        cal_x_max_local = max(cal_x_local) if len(cal_x_local) > 0 else np.nan
        for p in payload.get("unknown_points", []) or []:
            xv = float(p.get("x", np.nan))
            if not np.isfinite(xv):
                continue
            out_of_range = False
            if np.isfinite(cal_x_min_local) and np.isfinite(cal_x_max_local):
                out_of_range = (xv < cal_x_min_local) or (xv > cal_x_max_local)
            if out_of_range:
                continue
            all_x.append(xv)
            xerr = float(p.get("xerr", np.nan))
            if np.isfinite(xerr):
                all_x.extend([xv - xerr, xv + xerr])
        for p in payload.get("excluded_points", []) or []:
            xv = float(p.get("x", np.nan))
            if np.isfinite(xv):
                all_x.append(xv)

    if len(all_x) == 0:
        all_x = [0.0, 1.0]
    ref_x_vals = []
    for label in axis_labels:
        for ref_item in _reference_positions_for_channel(fit_rows, label, expected_refs=expected_refs):
            for ref in ref_item.get("references", []):
                rv = float(ref.get("x", np.nan)) if np.isfinite(ref.get("x", np.nan)) else np.nan
                rsd = float(ref.get("x_sd", np.nan)) if np.isfinite(ref.get("x_sd", np.nan)) else np.nan
                if np.isfinite(rv):
                    ref_x_vals.append(rv)
                    if np.isfinite(rsd) and rsd > 0:
                        ref_x_vals.extend([rv - rsd, rv + rsd])
    x_min_data = min(all_x + all_intercepts + ref_x_vals) if (len(all_intercepts) or len(ref_x_vals)) else min(all_x)
    x_max_data = max(all_x + ref_x_vals) if len(ref_x_vals) else max(all_x)
    x_left, x_right, x_ticks_major = _clean_axis_with_ticks(x_min_data, x_max_data, nticks=5, frac=0.08, min_span=1.0)

    def _render_channel_axis(ax, label, idx=0, n_axes=1, force_xlabel=False, monochrome=False):
        payload = plot_payload.get(label, {
            "calibration_points": [],
            "stdadd_groups": [],
            "unknown_points": [],
            "excluded_points": [],
            "calibration_fit": None,
        })
        point_color = "black" if monochrome else channel_color.get(label, "black")
        excluded_color = "black" if monochrome else "red"
        panel_y_data = []

        ax.tick_params(direction="in", top=True, right=True, which="major", length=4)
        ax.tick_params(direction="in", top=True, right=True, which="minor", length=2.5)
        if (not force_xlabel) and idx < (n_axes - 1):
            ax.tick_params(labelbottom=False)

        cal_pts = payload.get("calibration_points", []) or []
        cal_fit = payload.get("calibration_fit")
        if len(cal_pts) > 0:
            x = np.asarray([p.get("x", np.nan) for p in cal_pts], dtype=np.float64)
            y = np.asarray([p.get("y", np.nan) for p in cal_pts], dtype=np.float64)
            yerr = np.nan_to_num(np.asarray([p.get("yerr", np.nan) for p in cal_pts], dtype=np.float64), nan=0.0)
            ax.errorbar(x, y, yerr=yerr, fmt="none", ecolor=point_color, elinewidth=1.1, capsize=2.8, capthick=1.1, zorder=3)
            ax.plot(x, y, "o", ms=4.8, mfc="white", mec=point_color, linestyle="None", zorder=4)
            panel_y_data.extend([float(v) for v in y if np.isfinite(v)])
            panel_y_data.extend([float(v - e) for v, e in zip(y, yerr) if np.isfinite(v) and np.isfinite(e)])
            panel_y_data.extend([float(v + e) for v, e in zip(y, yerr) if np.isfinite(v) and np.isfinite(e)])
            if cal_fit is not None and np.isfinite(cal_fit.get("m", np.nan)) and np.isfinite(cal_fit.get("q", np.nan)):
                xx = np.linspace(x_left, x_right, 300)
                yy = cal_fit["m"] * xx + cal_fit["q"]
                ax.plot(xx, yy, "--", color=point_color, lw=1.2, zorder=2)
                panel_y_data.extend([float(v) for v in yy if np.isfinite(v)])

        for grp_idx, grp in enumerate(payload.get("stdadd_groups", []) or []):
            x = np.asarray(grp.get("x", []), dtype=np.float64)
            y = np.asarray(grp.get("y", []), dtype=np.float64)
            yerr = np.nan_to_num(np.asarray(grp.get("yerr", []), dtype=np.float64), nan=0.0)
            if x.size == 0:
                continue
            marker = _stdadd_marker(grp.get("ID", ""), grp.get("DF", np.nan))
            ax.errorbar(x, y, yerr=yerr, fmt="none", ecolor=point_color, elinewidth=1.0, capsize=2.6, capthick=1.0, zorder=3)
            ax.plot(x, y, marker=marker, ms=4.8, mfc=point_color, mec=point_color, linestyle="None", zorder=4)
            panel_y_data.extend([float(v) for v in y if np.isfinite(v)])
            panel_y_data.extend([float(v - e) for v, e in zip(y, yerr) if np.isfinite(v) and np.isfinite(e)])
            panel_y_data.extend([float(v + e) for v, e in zip(y, yerr) if np.isfinite(v) and np.isfinite(e)])
            fit = grp.get("fit")
            if fit is not None and np.isfinite(fit.get("m", np.nan)) and np.isfinite(fit.get("q", np.nan)):
                xx = np.linspace(x_left, x_right, 300)
                yy = fit["m"] * xx + fit["q"]
                ax.plot(xx, yy, "-", color=point_color, lw=1.35, zorder=2)
                panel_y_data.extend([float(v) for v in yy if np.isfinite(v)])

        cal_x_vals = [float(p.get("x", np.nan)) for p in cal_pts if np.isfinite(p.get("x", np.nan))]
        cal_x_min = min(cal_x_vals) if len(cal_x_vals) > 0 else np.nan
        cal_x_max = max(cal_x_vals) if len(cal_x_vals) > 0 else np.nan
        x_span_panel = max(float(x_right - x_left), 1e-12)
        for grp in payload.get("unknown_points", []) or []:
            xv = float(grp.get("x", np.nan))
            yv = float(grp.get("y", np.nan))
            if not (np.isfinite(xv) and np.isfinite(yv)):
                continue
            xerr = float(grp.get("xerr", np.nan))
            yerr = float(grp.get("yerr", np.nan))
            out_of_range = False
            if np.isfinite(cal_x_min) and np.isfinite(cal_x_max):
                out_of_range = (xv < cal_x_min) or (xv > cal_x_max)
            if out_of_range:
                x_plot = x_left + 0.03 * x_span_panel if xv < cal_x_min else x_right - 0.03 * x_span_panel
                ax.errorbar([x_plot], [yv], xerr=None, yerr=[[yerr]] if np.isfinite(yerr) else None,
                            fmt="s", ms=5.0, mfc=point_color, mec=point_color, ecolor=point_color,
                            elinewidth=1.0, capsize=2.6, capthick=1.0, linestyle="None", zorder=5)
                ax.text(x_plot, yv, " out of range", color=point_color, fontsize=6.8,
                        ha="left" if xv < cal_x_min else "right", va="bottom", zorder=6)
            else:
                ax.errorbar([xv], [yv], xerr=[[xerr]] if np.isfinite(xerr) else None, yerr=[[yerr]] if np.isfinite(yerr) else None,
                            fmt="s", ms=5.0, mfc=point_color, mec=point_color, ecolor=point_color,
                            elinewidth=1.0, capsize=2.6, capthick=1.0, linestyle="None", zorder=5)
            panel_y_data.append(yv)
            if np.isfinite(yerr):
                panel_y_data.extend([yv - yerr, yv + yerr])

        exc = payload.get("excluded_points", []) or []
        if len(exc) > 0:
            ex = np.asarray([p.get("x", np.nan) for p in exc], dtype=np.float64)
            ey = np.asarray([p.get("y", np.nan) for p in exc], dtype=np.float64)
            ax.plot(ex, ey, marker="*", color=excluded_color, ms=6.0, linestyle="None", zorder=6)
            panel_y_data.extend([float(v) for v in ey if np.isfinite(v)])

        ref_handles = []
        seen_ref_labels = set()
        for ref_item in _reference_positions_for_channel(fit_rows, label, expected_refs=expected_refs):
            fit_type_here = str(ref_item.get("FitType", ""))
            id_here = str(ref_item.get("ID", ""))
            df_here = ref_item.get("DF", np.nan)
            for ref in ref_item.get("references", []):
                rv = float(ref.get("x", np.nan)) if np.isfinite(ref.get("x", np.nan)) else np.nan
                rsd = float(ref.get("x_sd", np.nan)) if np.isfinite(ref.get("x_sd", np.nan)) else np.nan
                if not np.isfinite(rv):
                    continue
                base_label = str(ref.get("label", "Reference"))
                marker = _stdadd_marker(id_here, df_here) if fit_type_here == "StdAdd" else _reference_style(ref.get("index", 1)).get("marker", "D")
                color = point_color
                facecolor = "white"
                pos_label = base_label
                if fit_type_here == "StdAdd":
                    pos_label += f" (ref ID={id_here or 'Std add'}, DF={df_here})"
                    ax.errorbar([rv], [0.0], xerr=[[rsd]] if np.isfinite(rsd) and rsd > 0 else None,
                                yerr=None, fmt=marker, ms=5.4, mfc=facecolor, mec=color, ecolor=color,
                                elinewidth=1.0, capsize=2.8, capthick=1.0, linestyle="None", zorder=7)
                    panel_y_data.append(0.0)
                    if pos_label not in seen_ref_labels:
                        ref_handles.append(Line2D([0], [0], marker=marker, ms=5.4, mfc=facecolor, mec=color,
                                                  color=color, linestyle="None", label=pos_label))
                        seen_ref_labels.add(pos_label)
                else:
                    pos_label += f" ({id_here or 'Unknown'}, DF={df_here}, x=ref/DF)"
                    ax.axvline(rv, linestyle="--", linewidth=0.9, color=color, zorder=1)
                    if np.isfinite(rsd) and rsd > 0:
                        ax.axvspan(rv - rsd, rv + rsd, color=color, alpha=0.14, zorder=0)
                    if pos_label not in seen_ref_labels:
                        ref_handles.append(Line2D([0], [0], color=color, linestyle="--", lw=0.9, marker=marker,
                                                  mfc=facecolor, mec=color, label=pos_label))
                        seen_ref_labels.add(pos_label)
        finite_y = [v for v in panel_y_data if np.isfinite(v)]
        if len(finite_y) == 0:
            finite_y = [0.0, 1.0]
        y_min_data = min(finite_y)
        y_max_data = max(finite_y)
        y_bottom, y_top, y_ticks_major = _clean_axis_with_ticks(y_min_data, y_max_data, nticks=5, frac=0.08, min_span=0.05)
        if y_bottom <= 0.0 <= y_top:
            ax.axhline(0.0, color="black", lw=0.5, alpha=0.5)

        ax.set_xlim(x_left, x_right)
        ax.set_ylim(y_bottom, y_top)
        ax.set_ylabel(y_label_map.get(label, label), fontweight="bold")
        ax.xaxis.set_major_locator(FixedLocator(x_ticks_major))
        ax.yaxis.set_major_locator(FixedLocator(y_ticks_major))
        ax.xaxis.set_minor_locator(AutoMinorLocator(2))
        ax.yaxis.set_minor_locator(AutoMinorLocator(2))
        ax.xaxis.set_major_formatter(FormatStrFormatter('%.0f'))
        if selected_channel_upper not in {"CIELAB", "DELTAE", "CIELAB_DELTAE"}:
            ax.yaxis.set_major_formatter(FormatStrFormatter('%.2f'))
        ax.grid(False)

        handles = []
        if len(cal_pts) > 0:
            handles.append(Line2D([0], [0], color=point_color, linestyle='--', marker='o', mfc='white', mec=point_color, markersize=4.6, lw=1.1, label='calibration'))
        for grp in (payload.get("stdadd_groups", []) or []):
            _gid = str(grp.get("ID", ""))
            _gdf = grp.get("DF", np.nan)
            _marker = _stdadd_marker(_gid, _gdf)
            _label = f"std add ID={_gid}, DF={_gdf}"
            handles.append(Line2D([0], [0], color=point_color, linestyle='-', marker=_marker, mfc=point_color, mec=point_color, markersize=4.6, lw=1.2, label=_label))
        if len(payload.get("unknown_points", []) or []) > 0:
            handles.append(Line2D([0], [0], color=point_color, linestyle='None', marker='s', mfc=point_color, mec=point_color, markersize=4.8, label='unknown'))
        if len(exc) > 0:
            handles.append(Line2D([0], [0], color=excluded_color, linestyle='None', marker='*', markersize=6.0, label='excluded'))
        handles.extend(ref_handles)
        if handles:
            uniq = {}
            for h, lab in zip(handles, [h.get_label() for h in handles]):
                if lab not in uniq:
                    uniq[lab] = h
            ax.legend(list(uniq.values()), list(uniq.keys()), loc='upper left', frameon=False, fontsize=7.1, handlelength=1.8, borderaxespad=0.25)

        if idx == 0:
            ax.spines['bottom'].set_linewidth(0.8)
        elif idx == 1:
            ax.spines['top'].set_linewidth(0.8)
            ax.spines['bottom'].set_linewidth(0.8)
        else:
            ax.spines['top'].set_linewidth(0.8)


    for idx, (ax, label) in enumerate(axes):
        _render_channel_axis(ax, label, idx=idx, n_axes=len(axes), force_xlabel=False)

    if axes:
        axes[-1][0].set_xlabel(_infer_x_axis_label(plot_payload, unit_label), fontweight="bold")

    if best_channel_plot_path:
        def _equiv_channel_names(name):
            name = str(name or "").strip()
            names = {name}
            if name.startswith("Signal_"):
                names.add("PAbs_" + name.split("_", 1)[1])
            if name.startswith("PAbs_"):
                names.add("Signal_" + name.split("_", 1)[1])
            if name in {"Red", "Green", "Blue"}:
                names.add("Signal_" + name)
                names.add("PAbs_" + name)
            return {n for n in names if n}

        target_names = set()
        target_names.update(_equiv_channel_names(selected_channel))
        target_names.update(_equiv_channel_names(best_channel_raw))
        best_label = None
        for _ax, _label in axes:
            if str(_label) in target_names:
                best_label = _label
                break
        if best_label is not None:
            try:
                os.makedirs(os.path.dirname(best_channel_plot_path), exist_ok=True)
                single_figsize = _single_column_figsize_from_image(overlay_img, width_cm=9.0, height_cm=6.0)
                single_fig, single_ax = plt.subplots(figsize=single_figsize, dpi=300)
                _render_channel_axis(single_ax, best_label, idx=0, n_axes=1, force_xlabel=True, monochrome=True)
                single_ax.set_xlabel(_infer_x_axis_label(plot_payload, unit_label), fontweight="bold")
                single_fig.subplots_adjust(left=0.18, right=0.985, bottom=0.20, top=0.97)
                single_fig.savefig(best_channel_plot_path, dpi=300, bbox_inches=None, facecolor="white")
                plt.close(single_fig)
            except Exception:
                try:
                    plt.close(single_fig)
                except Exception:
                    pass

    fig.savefig(path, dpi=300, bbox_inches=None, facecolor="white")
    plt.close(fig)

def _find_spatial_clusters(raw_rows, base, nrow, ncol, min_size=2, max_size=4):
    """
    Diagnostic-only scan of contiguous rectangular clusters of unknown wells.
    Ranked by SD, then CV, then larger n.
    """
    grid = [[np.nan for _ in range(int(ncol))] for _ in range(int(nrow))]
    well_name_grid = [[None for _ in range(int(ncol))] for _ in range(int(nrow))]

    for row in raw_rows:
        typ = str(row.get("Type", "")).upper()
        if typ not in {"U", "UNK", "UNKNOWN"}:
            continue
        rr = int(_row_index_from_label(str(row.get("Row", "A"))))
        cc = int(row.get("Col", 1)) - 1
        if rr < 0 or rr >= int(nrow) or cc < 0 or cc >= int(ncol):
            continue
        v = row.get(base, np.nan)
        if np.isfinite(v):
            grid[rr][cc] = float(v)
            well_name_grid[rr][cc] = str(row.get("Well", _well_name_from_indices(rr, cc)))

    out = []
    for h in range(int(min_size), int(max_size) + 1):
        for w in range(int(min_size), int(max_size) + 1):
            for r0 in range(int(nrow) - h + 1):
                for c0 in range(int(ncol) - w + 1):
                    vals = []
                    wells = []
                    for dr in range(h):
                        for dc in range(w):
                            v = grid[r0 + dr][c0 + dc]
                            if np.isfinite(v):
                                vals.append(float(v))
                                wells.append(well_name_grid[r0 + dr][c0 + dc] or _well_name_from_indices(r0 + dr, c0 + dc))
                    min_required = max(4, int(math.ceil(0.70 * h * w)))
                    if len(vals) < min_required:
                        continue
                    arr = np.asarray(vals, dtype=np.float64)
                    sd = float(np.std(arr, ddof=1)) if arr.size > 1 else np.nan
                    mean = float(np.mean(arr)) if arr.size > 0 else np.nan
                    cv = float(sd / abs(mean)) if np.isfinite(sd) and np.isfinite(mean) and abs(mean) > 1e-15 else np.nan
                    out.append({
                        "base": base,
                        "r0": int(r0),
                        "c0": int(c0),
                        "h": int(h),
                        "w": int(w),
                        "n": int(arr.size),
                        "mean": mean,
                        "sd": sd,
                        "cv": cv,
                        "top_left": _well_name_from_indices(r0, c0),
                        "bottom_right": _well_name_from_indices(r0 + h - 1, c0 + w - 1),
                        "wells": ", ".join(wells),
                    })

    out.sort(key=lambda d: (
        np.inf if not np.isfinite(d.get("sd", np.nan)) else float(d["sd"]),
        np.inf if not np.isfinite(d.get("cv", np.nan)) else float(d["cv"]),
        -int(d.get("n", 0)),
        int(d.get("r0", 0)),
        int(d.get("c0", 0)),
    ))
    return out


def _write_cluster_analysis_csv(path, clusters):
    with open(path, "w", newline="", encoding="utf-8") as f:
        wr = csv.writer(f, delimiter=';')
        wr.writerow(["base", "top_left", "bottom_right", "h", "w", "n", "mean", "sd", "cv", "wells"])
        for row in clusters:
            wr.writerow([
                row.get("base", ""),
                row.get("top_left", ""),
                row.get("bottom_right", ""),
                row.get("h", ""),
                row.get("w", ""),
                row.get("n", ""),
                row.get("mean", ""),
                row.get("sd", ""),
                row.get("cv", ""),
                row.get("wells", ""),
            ])


def _estimate_unknown_from_cluster(cluster_row, fit_rows, selected_channel, unit_label="mM"):
    """
    Diagnostic-only concentration estimate from the best spatial cluster,
    using the selected-channel calibration fit already computed.
    """
    if not isinstance(cluster_row, dict):
        return None
    mean_y = float(cluster_row.get("mean", np.nan))
    sd_y = float(cluster_row.get("sd", np.nan))
    if not (np.isfinite(mean_y) and np.isfinite(sd_y)):
        return None

    cal_row = None
    for row in fit_rows:
        if row.get("FitType") == "Calibration" and str(row.get("Channel")) == str(selected_channel):
            cal_row = row
            break
    if cal_row is None:
        return None

    m = float(cal_row.get("m", np.nan))
    q = float(cal_row.get("q", np.nan))
    if not (np.isfinite(m) and np.isfinite(q)) or abs(m) < 1e-15:
        return None

    c_val = (mean_y - q) / m
    c_sd = abs(sd_y / m) if np.isfinite(sd_y) else np.nan
    return {
        "Channel": str(selected_channel),
        "Unit": str(unit_label),
        "ClusterTopLeft": str(cluster_row.get("top_left", "")),
        "ClusterBottomRight": str(cluster_row.get("bottom_right", "")),
        "ClusterSize": f"{int(cluster_row.get('h', 0))}x{int(cluster_row.get('w', 0))}",
        "N": int(cluster_row.get("n", 0)),
        "AbsMean": mean_y,
        "AbsSD": sd_y,
        "C_est": float(c_val) if np.isfinite(c_val) else np.nan,
        "C_est_sd": float(c_sd) if np.isfinite(c_sd) else np.nan,
        "Wells": str(cluster_row.get("wells", "")),
    }


def _append_cluster_sheet_to_workbook(path, selected_channel, cluster_rows, cluster_estimate):
    from openpyxl import load_workbook

    wb = load_workbook(path)
    if "06_CLUSTER_ANALYSIS" in wb.sheetnames:
        del wb["06_CLUSTER_ANALYSIS"]
    ws = wb.create_sheet("06_CLUSTER_ANALYSIS")

    lines = [
        "Spatial cluster diagnostics (diagnostic only; does not replace the main analytical result).",
        f"Selected channel for clustering: {selected_channel}",
        "",
    ]
    for i, line in enumerate(lines, start=1):
        ws.cell(i, 1, line)

    row0 = 5
    if isinstance(cluster_estimate, dict):
        est_headers = ["Channel", "ClusterTopLeft", "ClusterBottomRight", "ClusterSize", "N", "AbsMean", "AbsSD", "C_est", "C_est_sd", "Unit", "Wells"]
        _write_table(ws, row0, est_headers, [cluster_estimate])
        row0 += 4
    else:
        ws.cell(row0, 1, "No cluster-based estimate available.")
        row0 += 2

    top_rows = cluster_rows[:50] if isinstance(cluster_rows, list) else []
    headers = ["base", "top_left", "bottom_right", "h", "w", "n", "mean", "sd", "cv", "wells"]
    _write_table(ws, row0, headers, top_rows)
    _autosize_worksheet(ws)
    wb.save(path)


def _highlight_best_cluster_on_overlay(overlay_img, best_cluster, centers, color=(0, 215, 255)):
    if overlay_img is None or not isinstance(best_cluster, dict):
        return overlay_img
    vis = overlay_img.copy()
    wells_txt = str(best_cluster.get("wells", "")).strip()
    if wells_txt == "":
        return vis

    wells = [w.strip() for w in wells_txt.split(",") if w.strip()]
    pts = []
    for w in wells:
        try:
            m = re.match(r"^([A-Za-z]+)(\d+)$", w)
            if not m:
                continue
            rr = _row_index_from_label(m.group(1))
            cc = int(m.group(2)) - 1
        except Exception:
            continue
        if rr < 0 or rr >= centers.shape[0] or cc < 0 or cc >= centers.shape[1]:
            continue
        pts.append(centers[rr, cc])

    if len(pts) == 0:
        return vis

    pts = np.asarray(pts, dtype=np.float32)
    pad = 20
    x0 = max(0, int(np.floor(np.min(pts[:, 0]) - pad)))
    x1 = min(vis.shape[1] - 1, int(np.ceil(np.max(pts[:, 0]) + pad)))
    y0 = max(0, int(np.floor(np.min(pts[:, 1]) - pad)))
    y1 = min(vis.shape[0] - 1, int(np.ceil(np.max(pts[:, 1]) + pad)))
    cv2.rectangle(vis, (x0, y0), (x1, y1), color, 2, cv2.LINE_AA)
    label = f"Best cluster {best_cluster.get('h', '?')}x{best_cluster.get('w', '?')}"
    cv2.putText(vis, label, (x0, max(18, y0 - 8)), cv2.FONT_HERSHEY_SIMPLEX, 0.5, color, 1, cv2.LINE_AA)
    return vis




def _build_empty_well_rows(well_stats_rows, well_bg_rows, metadata_rows):
    """
    Build diagnostic rows for wells with no metadata entry (structurally empty wells).
    This is diagnostic only and does not enter the analytical pipeline.

    CIELAB-derived Delta values are computed using the same reference-selection
    logic used for the main plate report via _augment_raw_rows_with_deltae(...).
    """
    meta_keys = {(int(r["well_r"]), int(r["well_c"])) for r in metadata_rows}
    stats_map = {(int(r["well_r"]), int(r["well_c"])): r for r in well_stats_rows}
    bg_map = {(int(r["well_r"]), int(r["well_c"])): r for r in well_bg_rows}

    out = []
    for key, s in stats_map.items():
        if key in meta_keys:
            continue
        bg = bg_map.get(key)
        if bg is None:
            continue

        meanw_blue = _safe_gamma_linearize_scalar(s.get("B_p50", s.get("B_median")))
        meanw_green = _safe_gamma_linearize_scalar(s.get("G_p50", s.get("G_median")))
        meanw_red = _safe_gamma_linearize_scalar(s.get("R_p50", s.get("R_median")))

        meanbg_blue = _safe_gamma_linearize_scalar(bg.get("B_bg"))
        meanbg_green = _safe_gamma_linearize_scalar(bg.get("G_bg"))
        meanbg_red = _safe_gamma_linearize_scalar(bg.get("R_bg"))

        def _abs_from_pair(bg_v, w_v):
            if not np.isfinite(bg_v) or not np.isfinite(w_v) or bg_v <= 0.0 or w_v <= 0.0:
                return np.nan
            return float(np.log10(bg_v / w_v))

        rr, cc = key
        out.append({
            "Row": _row_label_from_index(rr),
            "Col": int(cc) + 1,
            "Well": _well_name_from_indices(rr, cc),
            "MeanW_Blue": meanw_blue,
            "MeanBG_Blue": meanbg_blue,
            "Signal_Blue": _abs_from_pair(meanbg_blue, meanw_blue),
            "MeanW_Green": meanw_green,
            "MeanBG_Green": meanbg_green,
            "Signal_Green": _abs_from_pair(meanbg_green, meanw_green),
            "MeanW_Red": meanw_red,
            "MeanBG_Red": meanbg_red,
            "Signal_Red": _abs_from_pair(meanbg_red, meanw_red),
            "L": float(s.get("L_median", np.nan)),
            "a": float(s.get("a_median", np.nan)),
            "b": float(s.get("b_median", np.nan)),
            "UsedFraction": float(s.get("used_fraction", np.nan)),
            "BrightExcludedFraction": float(s.get("BrightExcludedFraction", np.nan)),
            "HighlightIndex": float(s.get("HighlightIndex", np.nan)),
        })

    out.sort(key=lambda d: (str(d["Row"]), int(d["Col"])))

    # Apply the same CIELAB reference logic used for the main plate rows.
    out = _augment_raw_rows_with_deltae(out)

    return out
def _summarize_spatial_axis(rows, value_key, axis="row"):
    groups = {}
    for row in rows:
        v = float(row.get(value_key, np.nan))
        if not np.isfinite(v):
            continue
        key = str(row.get("Row")) if axis == "row" else int(row.get("Col"))
        groups.setdefault(key, []).append(v)

    out = []
    for key in sorted(groups.keys(), key=lambda x: (isinstance(x, str), x)):
        arr = np.asarray(groups[key], dtype=np.float64)
        med, sd, n, mad = _median_and_robust_sd(arr)
        out.append({
            "Axis": axis,
            "Level": key,
            "n": int(arr.size),
            "mean": float(np.mean(arr)) if arr.size else np.nan,
            "median": med,
            "sd": float(np.std(arr, ddof=1)) if arr.size > 1 else np.nan,
            "robust_sd": sd,
            "mad": mad,
            "min": float(np.min(arr)) if arr.size else np.nan,
            "max": float(np.max(arr)) if arr.size else np.nan,
        })
    return out


def _compute_spatial_linear_trend(rows, value_key):
    xs = []
    ys = []
    zs = []
    for row in rows:
        v = float(row.get(value_key, np.nan))
        if not np.isfinite(v):
            continue
        try:
            rr = _row_index_from_label(str(row.get("Row", "A")))
            cc = int(row.get("Col", 1)) - 1
        except Exception:
            continue
        xs.append(float(cc))
        ys.append(float(rr))
        zs.append(v)

    if len(zs) < 4:
        return {
            "n": int(len(zs)),
            "intercept": np.nan,
            "slope_col": np.nan,
            "slope_row": np.nan,
            "R2": np.nan,
            "corr_col": np.nan,
            "corr_row": np.nan,
        }

    x = np.asarray(xs, dtype=np.float64)
    y = np.asarray(ys, dtype=np.float64)
    z = np.asarray(zs, dtype=np.float64)

    A = np.column_stack([np.ones_like(x), x, y])
    coef, _, _, _ = np.linalg.lstsq(A, z, rcond=None)
    zhat = A @ coef
    sse = float(np.sum((z - zhat) ** 2))
    sst = float(np.sum((z - np.mean(z)) ** 2))
    r2 = float(1.0 - sse / sst) if sst > 1e-15 else np.nan

    def _safe_corr(a, b):
        if len(a) < 2:
            return np.nan
        aa = np.asarray(a, dtype=np.float64)
        bb = np.asarray(b, dtype=np.float64)
        if np.std(aa) < 1e-15 or np.std(bb) < 1e-15:
            return np.nan
        return float(np.corrcoef(aa, bb)[0, 1])

    return {
        "n": int(len(z)),
        "intercept": float(coef[0]),
        "slope_col": float(coef[1]),
        "slope_row": float(coef[2]),
        "R2": r2,
        "corr_col": _safe_corr(x, z),
        "corr_row": _safe_corr(y, z),
    }


def _build_spatial_diagnostic_payload(raw_rows, empty_rows, selected_channel):
    base_map = {
        "PAbs_Red": "PAbs_Red", "PAbs_Green": "PAbs_Green", "PAbs_Blue": "PAbs_Blue",
        "Red": "PAbs_Red", "Green": "PAbs_Green", "Blue": "PAbs_Blue",
    }
    base = base_map.get(selected_channel, "PAbs_Green")

    unknown_rows = [
        dict(r) for r in raw_rows
        if str(r.get("Type", "")).upper() in {"U", "UNK", "UNKNOWN"} and np.isfinite(r.get(base, np.nan))
    ]
    empty_rows = [dict(r) for r in empty_rows if np.isfinite(r.get(base, np.nan))]

    payload = {
        "selected_channel": selected_channel,
        "selected_base": base,
        "unknown_rows": unknown_rows,
        "empty_rows": empty_rows,
        "unknown_by_row": _summarize_spatial_axis(unknown_rows, base, axis="row"),
        "unknown_by_col": _summarize_spatial_axis(unknown_rows, base, axis="col"),
        "empty_by_row": _summarize_spatial_axis(empty_rows, base, axis="row"),
        "empty_by_col": _summarize_spatial_axis(empty_rows, base, axis="col"),
        "unknown_trend": _compute_spatial_linear_trend(unknown_rows, base),
        "empty_trend": _compute_spatial_linear_trend(empty_rows, base),
    }
    return payload


def _write_empty_well_csv(path, rows):
    headers = [
        "Row", "Col", "Well",
        "MeanW_Blue", "MeanBG_Blue", "Signal_Blue",
        "MeanW_Green", "MeanBG_Green", "Signal_Green",
        "MeanW_Red", "MeanBG_Red", "Signal_Red",
        "L", "a", "b",
        "UsedFraction", "BrightExcludedFraction", "HighlightIndex",
    ]
    with open(path, "w", newline="", encoding="utf-8") as f:
        wr = csv.writer(f, delimiter=';')
        wr.writerow(headers)
        for row in rows:
            wr.writerow([row.get(h, "") for h in headers])


def _append_spatial_diagnostics_to_workbook(path, spatial_payload):
    from openpyxl import load_workbook

    wb = load_workbook(path)

    for name in ["07_SPATIAL_DIAGNOSTICS", "08_EMPTY_WELLS"]:
        if name in wb.sheetnames:
            del wb[name]

    ws = wb.create_sheet("07_SPATIAL_DIAGNOSTICS")
    ws2 = wb.create_sheet("08_EMPTY_WELLS")

    selected_channel = spatial_payload.get("selected_channel", "")
    selected_base = spatial_payload.get("selected_base", "")

    intro = [
        "Spatial diagnostics (diagnostic only; no automatic correction applied).",
        f"Selected channel: {selected_channel}",
        f"Selected relative intensity signal field: {selected_base}",
        "",
    ]
    for i, line in enumerate(intro, start=1):
        ws.cell(i, 1, line)

    row0 = 6
    trend_headers = ["Dataset", "n", "intercept", "slope_col", "slope_row", "R2", "corr_col", "corr_row"]
    trend_rows = [
        {"Dataset": "Unknown wells", **(spatial_payload.get("unknown_trend", {}) or {})},
        {"Dataset": "Empty wells", **(spatial_payload.get("empty_trend", {}) or {})},
    ]
    _write_table(ws, row0, trend_headers, trend_rows)

    row0 += 5
    ws.cell(row0, 1, "Unknown wells by row")
    row0 += 1
    _write_table(ws, row0, ["Axis", "Level", "n", "mean", "median", "sd", "robust_sd", "mad", "min", "max"], spatial_payload.get("unknown_by_row", []))

    row0 += max(3, len(spatial_payload.get("unknown_by_row", [])) + 3)
    ws.cell(row0, 1, "Unknown wells by column")
    row0 += 1
    _write_table(ws, row0, ["Axis", "Level", "n", "mean", "median", "sd", "robust_sd", "mad", "min", "max"], spatial_payload.get("unknown_by_col", []))

    row0 += max(3, len(spatial_payload.get("unknown_by_col", [])) + 3)
    ws.cell(row0, 1, "Empty wells by row")
    row0 += 1
    _write_table(ws, row0, ["Axis", "Level", "n", "mean", "median", "sd", "robust_sd", "mad", "min", "max"], spatial_payload.get("empty_by_row", []))

    row0 += max(3, len(spatial_payload.get("empty_by_row", [])) + 3)
    ws.cell(row0, 1, "Empty wells by column")
    row0 += 1
    _write_table(ws, row0, ["Axis", "Level", "n", "mean", "median", "sd", "robust_sd", "mad", "min", "max"], spatial_payload.get("empty_by_col", []))

    intro2 = [
        "Structurally empty wells (diagnostic only).",
        f"Selected channel: {selected_channel}",
        "",
    ]
    for i, line in enumerate(intro2, start=1):
        ws2.cell(i, 1, line)

    empty_headers = [
        "Row", "Col", "Well",
        "MeanW_Blue", "MeanBG_Blue", "Signal_Blue",
        "MeanW_Green", "MeanBG_Green", "Signal_Green",
        "MeanW_Red", "MeanBG_Red", "Signal_Red",
        "UsedFraction", "BrightExcludedFraction", "HighlightIndex",
    ]
    _write_table(ws2, 5, empty_headers, spatial_payload.get("empty_rows", []))

    _autosize_worksheet(ws)
    _autosize_worksheet(ws2)
    wb.save(path)


def _compute_blank_from_raw(raw_rows, base):
    """
    Compute blank from RAW rows (not summary).
    Uses calibration points at Conc == 0.
    Returns (blank_median, blank_sd) or (nan, nan)
    """
    type_cal = {"CAL", "STD", "STANDARD", "CALIBRATION", "C"}
    vals = []
    for r in raw_rows:
        if str(r.get("Type", "")).upper() in type_cal:
            if np.isfinite(r.get("Conc", np.nan)) and abs(float(r.get("Conc"))) < 1e-12:
                v = r.get(base, np.nan)
                if np.isfinite(v):
                    vals.append(float(v))
    if len(vals) == 0:
        return np.nan, np.nan
    arr = np.asarray(vals, dtype=float)
    med = float(np.nanmedian(arr))
    sd = float(np.nanstd(arr, ddof=1)) if len(arr) > 1 else 0.0
    return med, sd



def _overlay_selected_pixels_on_plate(
    img_bgr, centers, well_debug_rows, outline_color=(0, 0, 0),
    selected_color=None, bright_excl_color=(255, 255, 255),
    dark_excl_color=(0, 0, 0), alpha=0.60
):
    """
    Build a plate-scale overlay showing:
    - geometric ROI outline
    - selected pixels actually used in stats
    - bright excluded pixels
    - dark excluded pixels
    """
    vis = img_bgr.copy()
    h, w = vis.shape[:2]
    overlay = vis.copy()

    by_key = {(int(r["row"]), int(r["col"])): r for r in well_debug_rows}

    for rr in range(centers.shape[0]):
        for cc in range(centers.shape[1]):
            key = (rr, cc)
            if key not in by_key:
                continue
            dbg = by_key[key]

            roi_u8 = dbg.get("roi_u8")
            used_u8 = dbg.get("used_u8")
            bright_u8 = dbg.get("bright_excluded_u8")
            dark_u8 = dbg.get("dark_excluded_u8")
            if roi_u8 is None:
                continue

            cnts, _ = cv2.findContours(roi_u8, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
            if cnts:
                cv2.drawContours(vis, cnts, -1, outline_color, 1, cv2.LINE_AA)

            if bright_u8 is not None:
                overlay[bright_u8 > 0] = bright_excl_color
            if dark_u8 is not None:
                overlay[dark_u8 > 0] = dark_excl_color

    vis = cv2.addWeighted(overlay, alpha, vis, 1 - alpha, 0)
    return vis




def _selected_rgb_channel_index(selected_channel):
    """Return OpenCV BGR index for an RGB/PAbs/Signal selected channel, or None."""
    s = str(selected_channel or "").strip()
    mapping = {
        "Blue": 0, "Signal_Blue": 0, "PAbs_Blue": 0,
        "Green": 1, "Signal_Green": 1, "PAbs_Green": 1,
        "Red": 2, "Signal_Red": 2, "PAbs_Red": 2,
    }
    return mapping.get(s)


def _normalize_selected_channel_label(selected_channel):
    """Return the plot-payload label for a selected RGB/PAbs/Signal channel."""
    s = str(selected_channel or "").strip()
    mapping = {
        "Red": "Signal_Red", "Green": "Signal_Green", "Blue": "Signal_Blue",
        "Signal_Red": "Signal_Red", "Signal_Green": "Signal_Green", "Signal_Blue": "Signal_Blue",
        "PAbs_Red": "PAbs_Red", "PAbs_Green": "PAbs_Green", "PAbs_Blue": "PAbs_Blue",
    }
    return mapping.get(s, s)


def _save_best_channel_plot(path, plot_payload, fit_rows, selected_channel, unit_label="mM", expected_refs=None):
    """Save only the graph for the selected best channel.

    Export-only helper: it reuses the already computed plot_payload/fit_rows and
    does not modify calibration, fitting, QC, stored calibration, or tables.
    """
    if not path:
        return False
    plot_payload = plot_payload or {}
    fit_rows = fit_rows or []
    expected_refs = expected_refs or []
    label = _normalize_selected_channel_label(selected_channel)
    if label not in plot_payload and label.startswith("PAbs_"):
        label = "Signal_" + label.split("_", 1)[1]
    if label not in plot_payload:
        return False
    payload = plot_payload.get(label, {}) or {}

    color_map = {
        "Signal_Red": "red", "Signal_Green": "green", "Signal_Blue": "blue",
        "PAbs_Red": "red", "PAbs_Green": "green", "PAbs_Blue": "blue",
        "Red": "red", "Green": "green", "Blue": "blue",
    }
    y_label_map = {
        "Signal_Red": "PAbs_Red", "Signal_Green": "PAbs_Green", "Signal_Blue": "PAbs_Blue",
        "PAbs_Red": "PAbs_Red", "PAbs_Green": "PAbs_Green", "PAbs_Blue": "PAbs_Blue",
        "Red": "PAbs_Red", "Green": "PAbs_Green", "Blue": "PAbs_Blue",
    }
    point_color = color_map.get(label, "black")

    def _safe_array(values):
        return np.asarray(values if values is not None else [], dtype=np.float64)

    def _stdadd_marker(sample_id, df):
        keys = []
        for grp in payload.get("stdadd_groups", []) or []:
            k = (str(grp.get("ID", "")), str(grp.get("DF", "")))
            if k not in keys:
                keys.append(k)
        cycle = ["s", "D", "^", "v", "P", "X", "o"]
        key = (str(sample_id or ""), str(df if df is not None else ""))
        if key in keys:
            return cycle[keys.index(key) % len(cycle)]
        return cycle[len(keys) % len(cycle)]

    x_data = []
    y_data = []
    cal_pts = payload.get("calibration_points", []) or []
    for pnt in cal_pts:
        xv = _num_or_nan(pnt.get("x", np.nan))
        yv = _num_or_nan(pnt.get("y", np.nan))
        ye = _num_or_nan(pnt.get("yerr", np.nan))
        if np.isfinite(xv) and np.isfinite(yv):
            x_data.append(xv)
            y_data.extend([yv, yv - ye if np.isfinite(ye) else yv, yv + ye if np.isfinite(ye) else yv])
    for grp in payload.get("stdadd_groups", []) or []:
        x = _safe_array(grp.get("x", []))
        y = _safe_array(grp.get("y", []))
        ye = _safe_array(grp.get("yerr", []))
        x_data.extend([float(v) for v in x if np.isfinite(v)])
        for i, v in enumerate(y):
            if np.isfinite(v):
                e = ye[i] if i < ye.size and np.isfinite(ye[i]) else 0.0
                y_data.extend([float(v), float(v - e), float(v + e)])
    for grp in payload.get("unknown_points", []) or []:
        xv = _num_or_nan(grp.get("x", np.nan))
        yv = _num_or_nan(grp.get("y", np.nan))
        ye = _num_or_nan(grp.get("yerr", np.nan))
        if np.isfinite(xv) and np.isfinite(yv):
            x_data.append(xv)
            y_data.extend([yv, yv - ye if np.isfinite(ye) else yv, yv + ye if np.isfinite(ye) else yv])
    for pnt in payload.get("excluded_points", []) or []:
        xv = _num_or_nan(pnt.get("x", np.nan))
        yv = _num_or_nan(pnt.get("y", np.nan))
        if np.isfinite(xv) and np.isfinite(yv):
            x_data.append(xv)
            y_data.append(yv)
    for ref_item in _reference_positions_for_channel(fit_rows, label, expected_refs=expected_refs):
        for ref in ref_item.get("references", []):
            xv = _num_or_nan(ref.get("x", np.nan))
            if np.isfinite(xv):
                x_data.append(xv)
                y_data.append(0.0)

    if len(x_data) == 0:
        return False
    x_min, x_max = float(np.nanmin(x_data)), float(np.nanmax(x_data))
    x_span = max(x_max - x_min, 1e-6)
    x_left, x_right = x_min - 0.10 * x_span, x_max + 0.10 * x_span
    y_finite = [v for v in y_data if np.isfinite(v)] or [0.0, 1.0]
    y_min, y_max = min(y_finite), max(y_finite)
    y_span = max(y_max - y_min, 0.05)
    y_bottom, y_top = y_min - 0.10 * y_span, y_max + 0.10 * y_span

    fig, ax = plt.subplots(figsize=(7.2, 5.0), dpi=220)
    if y_bottom <= 0.0 <= y_top:
        ax.axhline(0.0, color="black", lw=0.6, alpha=0.5)

    cal_x = []
    cal_y = []
    cal_yerr = []
    for pnt in cal_pts:
        xv = _num_or_nan(pnt.get("x", np.nan))
        yv = _num_or_nan(pnt.get("y", np.nan))
        ye = _num_or_nan(pnt.get("yerr", np.nan))
        if np.isfinite(xv) and np.isfinite(yv):
            cal_x.append(xv)
            cal_y.append(yv)
            cal_yerr.append(ye if np.isfinite(ye) else 0.0)
    if cal_x:
        ax.errorbar(cal_x, cal_y, yerr=cal_yerr, fmt="none", ecolor=point_color, elinewidth=1.1, capsize=3.0, capthick=1.0)
        ax.plot(cal_x, cal_y, "o", ms=5.2, mfc="white", mec=point_color, linestyle="None", label="calibration")
        cal_fit = payload.get("calibration_fit")
        if cal_fit is not None and np.isfinite(cal_fit.get("m", np.nan)) and np.isfinite(cal_fit.get("q", np.nan)):
            xx = np.linspace(x_left, x_right, 300)
            ax.plot(xx, cal_fit["m"] * xx + cal_fit["q"], "--", color=point_color, lw=1.3)

    for grp in payload.get("stdadd_groups", []) or []:
        x = _safe_array(grp.get("x", []))
        y = _safe_array(grp.get("y", []))
        yerr = np.nan_to_num(_safe_array(grp.get("yerr", [])), nan=0.0)
        if x.size == 0 or y.size == 0:
            continue
        marker = _stdadd_marker(grp.get("ID", ""), grp.get("DF", np.nan))
        label_txt = f"std add ID={grp.get('ID', '')}, DF={grp.get('DF', np.nan)}"
        ax.errorbar(x, y, yerr=yerr if yerr.size == y.size else None, fmt="none", ecolor=point_color, elinewidth=1.0, capsize=3.0, capthick=1.0)
        ax.plot(x, y, marker=marker, ms=5.2, mfc=point_color, mec=point_color, linestyle="None", label=label_txt)
        fit = grp.get("fit")
        if fit is not None and np.isfinite(fit.get("m", np.nan)) and np.isfinite(fit.get("q", np.nan)):
            xx = np.linspace(x_left, x_right, 300)
            ax.plot(xx, fit["m"] * xx + fit["q"], "-", color=point_color, lw=1.4)

    for grp in payload.get("unknown_points", []) or []:
        xv = _num_or_nan(grp.get("x", np.nan))
        yv = _num_or_nan(grp.get("y", np.nan))
        xe = _num_or_nan(grp.get("xerr", np.nan))
        ye = _num_or_nan(grp.get("yerr", np.nan))
        if np.isfinite(xv) and np.isfinite(yv):
            ax.errorbar([xv], [yv], xerr=[[xe]] if np.isfinite(xe) else None, yerr=[[ye]] if np.isfinite(ye) else None,
                        fmt="s", ms=5.2, mfc=point_color, mec=point_color, ecolor=point_color, capsize=3.0, linestyle="None", label="unknown")

    exc = payload.get("excluded_points", []) or []
    if exc:
        ex = [_num_or_nan(p.get("x", np.nan)) for p in exc]
        ey = [_num_or_nan(p.get("y", np.nan)) for p in exc]
        ok = [np.isfinite(a) and np.isfinite(b) for a, b in zip(ex, ey)]
        if any(ok):
            ax.plot([a for a, good in zip(ex, ok) if good], [b for b, good in zip(ey, ok) if good], marker="*", color="red", ms=7.0, linestyle="None", label="excluded")

    for ref_item in _reference_positions_for_channel(fit_rows, label, expected_refs=expected_refs):
        fit_type_here = str(ref_item.get("FitType", ""))
        id_here = str(ref_item.get("ID", ""))
        df_here = ref_item.get("DF", np.nan)
        for ref in ref_item.get("references", []):
            rv = _num_or_nan(ref.get("x", np.nan))
            rsd = _num_or_nan(ref.get("x_sd", np.nan))
            if not np.isfinite(rv):
                continue
            marker = _stdadd_marker(id_here, df_here) if fit_type_here == "StdAdd" else _reference_style(ref.get("index", 1)).get("marker", "D")
            ref_label = f"{ref.get('label', 'Reference')} (ref ID={id_here}, DF={df_here})" if fit_type_here == "StdAdd" else str(ref.get("label", "Reference"))
            if fit_type_here == "StdAdd":
                ax.errorbar([rv], [0.0], xerr=[[rsd]] if np.isfinite(rsd) and rsd > 0 else None, fmt=marker,
                            ms=5.5, mfc="white", mec=point_color, ecolor=point_color, capsize=3.0, linestyle="None", label=ref_label)
            else:
                ax.axvline(rv, linestyle="--", linewidth=0.9, color=point_color, label=ref_label)
                if np.isfinite(rsd) and rsd > 0:
                    ax.axvspan(rv - rsd, rv + rsd, color=point_color, alpha=0.14)

    ax.set_xlim(x_left, x_right)
    ax.set_ylim(y_bottom, y_top)
    ax.set_xlabel(_infer_x_axis_label(plot_payload, unit_label), fontweight="bold")
    ax.set_ylabel(y_label_map.get(label, label), fontweight="bold")
    ax.set_title(f"Best channel: {y_label_map.get(label, label)}")
    ax.grid(False)
    handles, labels = ax.get_legend_handles_labels()
    if handles:
        uniq = {}
        for h, lab in zip(handles, labels):
            if lab not in uniq:
                uniq[lab] = h
        ax.legend(list(uniq.values()), list(uniq.keys()), loc="upper left", frameon=False, fontsize=8.0)
    fig.tight_layout()
    fig.savefig(path, dpi=220, bbox_inches="tight", facecolor="white")
    plt.close(fig)
    return True


def _save_best_channel_image(path, img_bgr, selected_channel):
    """Legacy fallback: save the image plane corresponding to the selected channel."""
    if not path or img_bgr is None:
        return False
    idx = _selected_rgb_channel_index(selected_channel)
    if idx is None:
        return False
    channel_img = img_bgr[:, :, idx]
    if channel_img.dtype != np.uint8:
        arr = channel_img.astype(np.float32)
        finite = np.isfinite(arr)
        if not np.any(finite):
            return False
        lo = float(np.nanmin(arr[finite]))
        hi = float(np.nanmax(arr[finite]))
        if hi > lo:
            arr = (arr - lo) * (255.0 / (hi - lo))
        channel_img = np.clip(arr, 0, 255).astype(np.uint8)
    return bool(cv2.imwrite(str(path), channel_img))

def _save_selected_pixel_debug_images(out_dir, image_basename, img_bgr, centers, well_debug_rows,
                                      bg_debug_can_u8=None, H_can_to_img=None, out_shape=None):
    """
    Save diagnostic images showing the ACTUAL selected pixels, not only the geometric ROI.
    """
    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    # Plate overlay for well-selected pixels
    well_vis = _overlay_selected_pixels_on_plate(img_bgr, centers, well_debug_rows)
    cv2.imwrite(str(out_dir / f"{image_basename}_WELL_SELECTED_PIXELS_V27.png"), well_vis)

    # Background selected pixels, back-projected to image if possible
    if bg_debug_can_u8 is not None and H_can_to_img is not None and out_shape is not None:
        bg_u8 = cv2.warpPerspective(
            bg_debug_can_u8,
            H_can_to_img,
            (out_shape[1], out_shape[0]),
            flags=cv2.INTER_NEAREST
        )
        bg_vis = img_bgr.copy()
        overlay = bg_vis.copy()
        overlay[bg_u8 > 0] = (0, 255, 0)
        bg_vis = cv2.addWeighted(overlay, 0.45, bg_vis, 0.55, 0)
        cv2.imwrite(str(out_dir / f"{image_basename}_BG_SELECTED_PIXELS_V27.png"), bg_vis)



def _build_cielab_report_payload(summary_rows, fit_options=None, stored_calibration_bundle=None):
    """
    Build a diagnostic CIELAB payload independent from the RGB analytical pipeline.
    This restores the dedicated A4 CIELAB report without changing RGB selection,
    workbook logic or BG2D analysis.
    """
    fit_options = fit_options or {}
    type_cal = {"CAL", "STD", "STANDARD", "CALIBRATION", "C"}
    type_stdadd = {"A", "SA", "STDADD", "STANDARD_ADDITION", "ADDITION"}
    type_unk = {"UNK", "UNKNOWN", "U"}
    channels = [("DeltaL", "DeltaL"), ("Deltaa", "Deltaa"), ("Deltab", "Deltab")]

    fit_rows = []
    plot_payload = {}
    has_any = False

    for label, base in channels:
        centered_summary_rows, cal_ref_value, cal_ref_source = _center_summary_rows_for_calibration(summary_rows, base)
        payload = {
            "calibration_points": [],
            "stdadd_groups": [],
            "unknown_points": [],
            "excluded_points": [],
            "calibration_fit": None,
            "calibration_ref": cal_ref_value,
            "calibration_ref_source": cal_ref_source,
        }

        cal_all = [
            r for r in centered_summary_rows
            if str(r.get("Type", "")).upper() in type_cal
            and np.isfinite(r.get(f"{base}_median", np.nan))
            and np.isfinite(r.get("Conc", np.nan))
        ]
        cal_used = list(cal_all)
        cal_exc = []
        payload["excluded_points"].extend([
            {"x": float(r["Conc"]), "y": float(r[f"{base}_median"])} for r in cal_exc
        ])

        cal_fit = None
        sigma_cal, sigma_source = _estimate_sigma_for_lod(cal_used, base)
        use_stored_calibration = bool(fit_options.get("use_stored_calibration", False))
        stored_channels = stored_calibration_bundle.get("channels", {}) or {} if isinstance(stored_calibration_bundle, dict) else {}
        if use_stored_calibration and label in stored_channels:
            cal_fit = _fit_from_stored_channel_row(stored_channels.get(label))
            stored_row = stored_channels.get(label, {})
            sigma_cal = float(stored_row["sigma_cal"]) if stored_row.get("sigma_cal", None) is not None else np.nan
            sigma_source = stored_row.get("sigma_source", "stored_calibration")
        elif len(cal_used) >= 2:
            x = np.asarray([float(r["Conc"]) for r in cal_used], dtype=np.float64)
            y = np.asarray([float(r[f"{base}_median"]) for r in cal_used], dtype=np.float64)
            w = np.asarray([_row_weight(r, base) for r in cal_used], dtype=np.float64)
            cal_fit = _fit_line_with_covariance(x, y, w=w, force_zero=False)
        payload["calibration_fit"] = cal_fit
        if use_stored_calibration and label in stored_channels:
            stored_pts = stored_channels.get(label, {}).get("calibration_points", []) or []
            payload["calibration_points"] = [{
                "x": float(p.get("x", np.nan)),
                "y": float(p.get("y", np.nan)),
                "yerr": float(p.get("yerr", np.nan)) if p.get("yerr", None) is not None else np.nan,
                "excluded": bool(p.get("excluded", False)),
            } for p in stored_pts if np.isfinite(p.get("x", np.nan)) and np.isfinite(p.get("y", np.nan))]
        else:
            payload["calibration_points"] = [{
                "x": float(r["Conc"]),
                "y": float(r[f"{base}_median"]),
                "yerr": float(r.get(f"{base}_sd", np.nan)),
                "excluded": False,
            } for r in cal_used]

        fit_rows.append({
            "Channel": label, "FitType": "Calibration", "ID": None, "DF": None,
            "n_points": cal_fit["n_points"] if cal_fit else len(cal_used),
            "m": cal_fit["m"] if cal_fit else np.nan,
            "q": cal_fit["q"] if cal_fit else np.nan,
            "R2": cal_fit["R2"] if cal_fit else np.nan,
            "RMSE": cal_fit["RMSE"] if cal_fit else np.nan,
            "sigma_cal": sigma_cal,
            "sigma_source": sigma_source,
            "SNR": float(abs(cal_fit["m"]) / sigma_cal) if cal_fit and np.isfinite(sigma_cal) and sigma_cal > 0 else np.nan,
            "LOD": float(3.0 * sigma_cal / abs(cal_fit["m"])) if cal_fit and np.isfinite(sigma_cal) and sigma_cal > 0 and abs(cal_fit["m"]) > 1e-15 else np.nan,
            "LOQ": float(10.0 * sigma_cal / abs(cal_fit["m"])) if cal_fit and np.isfinite(sigma_cal) and sigma_cal > 0 and abs(cal_fit["m"]) > 1e-15 else np.nan,
            "C0": np.nan,
            "C0_sd": np.nan,
            "C0": np.nan,
            "C0_sd": np.nan,
            "beta_k": np.nan,
            "bias_index_k": np.nan,
            "__cov_mq": cal_fit["cov_mq"] if cal_fit else None,
        })

        stdadd_all = [
            r for r in summary_rows
            if str(r.get("Type", "")).upper() in type_stdadd
            and np.isfinite(r.get(f"{base}_median", np.nan))
            and np.isfinite(r.get("Conc", np.nan))
        ]
        groups = {}
        for row in stdadd_all:
            groups.setdefault((str(row["ID"]), float(row["DF"])), []).append(row)
        for (sample_id, dilution_factor), rows in sorted(groups.items(), key=lambda kv: (kv[0][0], kv[0][1])):
            rows_sorted = sorted(rows, key=lambda r: float(r["Conc"]))
            used_rows = list(rows_sorted)
            exc_rows = []
            payload["excluded_points"].extend([
                {"x": float(r["Conc"]), "y": float(r[f"{base}_median"])} for r in exc_rows
            ])
            fit = None
            c0_orig = np.nan
            c0_orig_sd = np.nan
            if len(used_rows) >= 2:
                x = np.asarray([float(r["Conc"]) for r in used_rows], dtype=np.float64)
                y = np.asarray([float(r[f"{base}_median"]) for r in used_rows], dtype=np.float64)
                w = np.asarray([_row_weight(r, base) for r in used_rows], dtype=np.float64)
                fit = _fit_line_with_covariance(x, y, w=w)
                c0_dil, c0_dil_sd = _stdadd_c0_sd_from_fit(fit)
                c0_orig_stdadd = float(dilution_factor * c0_dil) if np.isfinite(c0_dil) else np.nan
                c0_orig_stdadd_sd = float(dilution_factor * c0_dil_sd) if np.isfinite(c0_dil_sd) else np.nan

                # When a stored calibration is used, quantify the sample from the
                # standard-addition intercept and the STORED calibration slope:
                #   C0_original = DF * ( q_std / m_cal_stored )
                # This is the intended meaning of "apply stored calibration".
                # The ordinary standard-addition estimate DF*(-q_std/m_std) is
                # still kept in auxiliary fields below for traceability.
                c0_orig = c0_orig_stdadd
                c0_orig_sd = c0_orig_stdadd_sd
                if use_stored_calibration and cal_fit is not None and fit is not None:
                    c0_corr, c0_corr_sd = _corrected_c0_from_calibration_slope(
                        cal_fit=cal_fit,
                        std_fit=fit,
                        dilution_factor=dilution_factor,
                    )
                    if np.isfinite(c0_corr):
                        c0_orig = float(c0_corr)
                        c0_orig_sd = float(c0_corr_sd) if np.isfinite(c0_corr_sd) else np.nan
            payload["stdadd_groups"].append({
                "ID": sample_id,
                "DF": dilution_factor,
                "x": np.asarray([float(r["Conc"]) for r in used_rows], dtype=np.float64),
                "y": np.asarray([float(r[f"{base}_median"]) for r in used_rows], dtype=np.float64),
                "yerr": np.asarray([float(r.get(f"{base}_sd", np.nan)) for r in used_rows], dtype=np.float64),
                "fit": fit,
                "c0_orig": c0_orig,
                "c0_orig_sd": c0_orig_sd,
                "c0_orig_stdadd_only": c0_orig_stdadd if 'c0_orig_stdadd' in locals() else np.nan,
                "c0_orig_stdadd_only_sd": c0_orig_stdadd_sd if 'c0_orig_stdadd_sd' in locals() else np.nan,
                "c0_source": "stored_calibration_slope" if (use_stored_calibration and cal_fit is not None and fit is not None and np.isfinite(c0_orig)) else "stdadd_intercept",
            })
            fit_rows.append({
                "Channel": label, "FitType": "StdAdd", "ID": sample_id, "DF": dilution_factor,
                "n_points": fit["n_points"] if fit else len(used_rows),
                "m": fit["m"] if fit else np.nan,
                "q": fit["q"] if fit else np.nan,
                "R2": fit["R2"] if fit else np.nan,
                "RMSE": fit["RMSE"] if fit else np.nan,
                "sigma_cal": np.nan,
                "sigma_source": "unavailable",
                "SNR": np.nan,
                "LOD": np.nan,
                "LOQ": np.nan,
                "C0": c0_orig,
                "C0_sd": c0_orig_sd,
                "C0_stdadd_only": c0_orig_stdadd if 'c0_orig_stdadd' in locals() else np.nan,
                "C0_stdadd_only_sd": c0_orig_stdadd_sd if 'c0_orig_stdadd_sd' in locals() else np.nan,
                "C0_source": "stored_calibration_slope" if (use_stored_calibration and cal_fit is not None and fit is not None and np.isfinite(c0_orig)) else "stdadd_intercept",
                "beta_k": np.nan,
                "bias_index_k": np.nan,
                "__cov_mq": fit["cov_mq"] if fit else None,
            })

        unk_points = [
            r for r in summary_rows
            if str(r.get("Type", "")).upper() in type_unk and np.isfinite(r.get(f"{base}_median", np.nan))
        ]
        if cal_fit is not None and len(unk_points) > 0 and abs(cal_fit["m"]) > 1e-15:
            for row in unk_points:
                y_obs = float(row[f"{base}_median"])
                y_obs_cal = float(y_obs - cal_ref_value) if np.isfinite(cal_ref_value) else y_obs
                c_dil = (y_obs_cal - cal_fit["q"]) / cal_fit["m"]
                df_val = float(row["DF"]) if np.isfinite(row.get("DF", np.nan)) else np.nan
                c_orig = df_val * c_dil if np.isfinite(df_val) else c_dil
                sigma_y = float(row.get(f"{base}_sd", np.nan))
                c_dil_sd = float(abs(sigma_y / cal_fit["m"])) if np.isfinite(sigma_y) else np.nan
                c_sd = (df_val * c_dil_sd) if np.isfinite(df_val) and np.isfinite(c_dil_sd) else c_dil_sd
                n_rep = int(row.get("NReplicates", 1))
                fit_rows.append({
                    "Channel": label, "FitType": "UnknownFromCal", "ID": str(row["ID"]), "DF": df_val,
                    "n_points": n_rep, "m": cal_fit["m"], "q": cal_fit["q"], "R2": cal_fit["R2"], "RMSE": cal_fit["RMSE"],
                    "sigma_cal": np.nan, "sigma_source": "unavailable", "SNR": np.nan, "LOD": np.nan, "LOQ": np.nan,
                    "C0": c_orig, "C0_sd": c_sd,
                    "C0": c_orig, "C0_sd": c_sd,
                    "__cov_mq": None,
                })
                payload["unknown_points"].append({
                    "ID": str(row["ID"]),
                    "DF": df_val,
                    "n_points": n_rep,
                    "x": float(c_dil) if np.isfinite(c_dil) else np.nan,
                    "xerr": float(c_dil_sd) if np.isfinite(c_dil_sd) else np.nan,
                    "y": y_obs_cal,
                    "yerr": sigma_y if np.isfinite(sigma_y) else np.nan,
                    "y_raw": y_obs,
                    "c_orig": float(c_orig) if np.isfinite(c_orig) else np.nan,
                    "c_orig_sd": float(c_sd) if np.isfinite(c_sd) else np.nan,
                })
        elif len(unk_points) > 0:
            # Unknown-only CIELAB values are descriptive diagnostics. Without a
            # valid calibration/stored calibration, no concentration or plotted
            # unknown-point coordinate is generated for CIELAB components.
            pass

        if payload["calibration_points"] or payload["stdadd_groups"] or payload["unknown_points"]:
            has_any = True
        plot_payload[label] = payload

    ranking = []
    for label, base in channels:
        st = _channel_stats_from_summary(summary_rows, base)
        has_cal = st.get("cal_fit") is not None and np.isfinite(st.get("r2_cal", np.nan))
        has_std = len(st.get("std_fits", [])) > 0 and np.isfinite(st.get("r2_std", np.nan))

        if has_cal and has_std:
            mode = "calibration_plus_stdadd"
            score = _compute_balanced_score(st.get("r2_cal", np.nan), st.get("r2_std", np.nan), st.get("slope_agreement", np.nan), loq=st.get("loq", np.nan))
            std_best = None
            std_fits = [f for f in st.get("std_fits", []) if np.isfinite(f.get("R2", np.nan))]
            if len(std_fits) > 0:
                std_best = max(std_fits, key=lambda f: f.get("R2", -np.inf))
            c0_sd = np.nan
            if std_best is not None:
                try:
                    _, c0_sd = _stdadd_c0_sd_from_fit(std_best)
                except Exception:
                    c0_sd = np.nan
            ranking.append({
                "Channel": label,
                "Base": base,
                "Mode": mode,
                "R2_cal": st["r2_cal"],
                "R2_std": st["r2_std"],
                "SlopeAgreement": st["slope_agreement"],
                "Score": score,
                "SNR": st["snr"],
                "LOQ": st.get("loq", np.nan),
                "|mstd|": float(abs(std_best["m"])) if std_best is not None and np.isfinite(std_best.get("m", np.nan)) else np.nan,
                "C0_sd": c0_sd,
            })
        elif has_std:
            mode = "stdadd_only"
            std_best = None
            std_fits = [f for f in st.get("std_fits", []) if np.isfinite(f.get("R2", np.nan))]
            if len(std_fits) > 0:
                std_best = max(std_fits, key=lambda f: f.get("R2", -np.inf))
            if std_best is not None and np.isfinite(std_best.get("R2", np.nan)) and np.isfinite(std_best.get("m", np.nan)):
                score = float((std_best["R2"] ** 2) * abs(std_best["m"]))
                try:
                    _, c0_sd = _stdadd_c0_sd_from_fit(std_best)
                except Exception:
                    c0_sd = np.nan
            else:
                score = np.nan
                c0_sd = np.nan
            ranking.append({
                "Channel": label,
                "Base": base,
                "Mode": mode,
                "R2_cal": st["r2_cal"],
                "R2_std": st["r2_std"],
                "SlopeAgreement": st["slope_agreement"],
                "Score": score,
                "SNR": st["snr"],
                "LOQ": st.get("loq", np.nan),
                "|mstd|": float(abs(std_best["m"])) if std_best is not None and np.isfinite(std_best.get("m", np.nan)) else np.nan,
                "C0_sd": c0_sd,
            })
        elif has_cal:
            mode = "calibration_only"
            cal_fit = st.get("cal_fit")
            if cal_fit is not None and np.isfinite(cal_fit.get("R2", np.nan)) and np.isfinite(cal_fit.get("m", np.nan)):
                score = float((cal_fit["R2"] ** 2) * abs(cal_fit["m"]))
            else:
                score = np.nan
            ranking.append({
                "Channel": label,
                "Base": base,
                "Mode": mode,
                "R2_cal": st["r2_cal"],
                "R2_std": st["r2_std"],
                "SlopeAgreement": st["slope_agreement"],
                "Score": score,
                "SNR": st["snr"],
                "LOQ": st.get("loq", np.nan),
                "|mstd|": np.nan,
                "C0_sd": np.nan,
            })
        else:
            ranking.append({
                "Channel": label,
                "Base": base,
                "Mode": "unavailable",
                "R2_cal": st["r2_cal"],
                "R2_std": st["r2_std"],
                "SlopeAgreement": st["slope_agreement"],
                "Score": np.nan,
                "SNR": st["snr"],
                "LOQ": st.get("loq", np.nan),
                "|mstd|": np.nan,
                "C0_sd": np.nan,
            })

    ranking_sorted = sorted(ranking, key=lambda d: (-d["Score"] if np.isfinite(d.get("Score", np.nan)) else np.inf))
    best = ranking_sorted[0] if ranking_sorted else {"Channel": "", "Mode": "unavailable", "Score": np.nan}

    reason = [
        f"Selected CIELAB component: {best['Channel']}",
        "Selection criterion:",
        "Balanced fit-quality ranking:",
        "if calibration + std add: GlobalScore = (slope_agreement^2) x sqrt(R²cal x R²std) x (1/LOQ); SNR, clipping and expected values are not used",
        "if std add only: score = (R²std^2) x |mstd|",
        "if calibration only: score = (R²cal^2) x |mcal|",
        "slope_agreement = min(|mcal|,|madd|)/max(|mcal|,|madd|)",
        "",
        "Ranking:",
    ]
    for r in ranking_sorted:
        score_txt = f"{_num_or_nan(r.get('Score', np.nan)):.4f}" if _isfinite_num(r.get('Score', np.nan)) else "nan"
        r2cal_txt = f"{_num_or_nan(r.get('R2_cal', np.nan)):.4f}" if _isfinite_num(r.get('R2_cal', np.nan)) else "nan"
        r2std_txt = f"{_num_or_nan(r.get('R2_std', np.nan)):.4f}" if _isfinite_num(r.get('R2_std', np.nan)) else "nan"
        slope_txt = f"{_num_or_nan(r.get('SlopeAgreement', np.nan)):.4f}" if _isfinite_num(r.get('SlopeAgreement', np.nan)) else "nan"
        snr_txt = f"{_num_or_nan(r.get('SNR', np.nan)):.3f}" if _isfinite_num(r.get('SNR', np.nan)) else "nan"
        mstd_txt = f"{_num_or_nan(r.get('|mstd|', np.nan)):.5f}" if _isfinite_num(r.get('|mstd|', np.nan)) else "nan"
        loq_txt = f"{_num_or_nan(r.get('LOQ', np.nan)):.5g}" if _isfinite_num(r.get('LOQ', np.nan)) else "nan"
        c0sd_txt = f"{_num_or_nan(r.get('C0_sd', np.nan)):.3f}" if _isfinite_num(r.get('C0_sd', np.nan)) else "nan"
        reason.append(f"{r['Channel']}: mode={r['Mode']}  score={score_txt}  R²cal={r2cal_txt}  R²std={r2std_txt}  slope={slope_txt}  LOQ={loq_txt}  |mstd|={mstd_txt}  SNR={snr_txt}  C0sd={c0sd_txt}")

    plot_payload["__selection_info__"] = {
        "best": best,
        "ranking": ranking_sorted,
        "reason_lines": reason,
    }

    return fit_rows, plot_payload, has_any


def _build_deltae_report_payload(summary_rows, fit_options=None, stored_calibration_bundle=None):
    """Dedicated DeltaE analytical report.
    Monovariate fits are built only for DeltaE_ab.
    Unknown/CRM points are propagated exactly as in the RGB analytical payload.
    """
    fit_options = fit_options or {}
    type_cal = {"CAL", "STD", "STANDARD", "CALIBRATION", "C"}
    type_stdadd = {"A", "SA", "STDADD", "STANDARD_ADDITION", "ADDITION"}
    type_unk = {"UNK", "UNKNOWN", "U"}
    channels = [("DeltaE_ab", "DeltaE_ab"), ("DeltaE_ab_chroma", "DeltaE_ab_chroma"), ("DeltaL", "DeltaL")]

    fit_rows = []
    plot_payload = {}
    has_any = False

    for label, base in channels:
        centered_summary_rows, cal_ref_value, cal_ref_source = _center_summary_rows_for_calibration(summary_rows, base)
        payload = {
            "calibration_points": [],
            "stdadd_groups": [],
            "unknown_points": [],
            "excluded_points": [],
            "calibration_fit": None,
            "calibration_ref": cal_ref_value,
            "calibration_ref_source": cal_ref_source,
        }

        cal_all = [
            r for r in centered_summary_rows
            if str(r.get("Type", "")).upper() in type_cal
            and np.isfinite(r.get(f"{base}_median", np.nan))
            and np.isfinite(r.get("Conc", np.nan))
        ]
        cal_used = list(cal_all)
        cal_exc = []
        payload["excluded_points"].extend(
            [{"x": float(r["Conc"]), "y": float(r[f"{base}_median"])} for r in cal_exc]
        )

        cal_fit = None
        sigma_cal, sigma_source = _estimate_sigma_for_lod(cal_used, base)
        use_stored_calibration = bool(fit_options.get("use_stored_calibration", False))
        stored_channels = stored_calibration_bundle.get("channels", {}) or {} if isinstance(stored_calibration_bundle, dict) else {}
        if use_stored_calibration and label in stored_channels:
            cal_fit = _fit_from_stored_channel_row(stored_channels.get(label))
            stored_row = stored_channels.get(label, {})
            sigma_cal = float(stored_row["sigma_cal"]) if stored_row.get("sigma_cal", None) is not None else np.nan
            sigma_source = stored_row.get("sigma_source", "stored_calibration")
        elif len(cal_used) >= 2:
            x = np.asarray([float(r["Conc"]) for r in cal_used], dtype=np.float64)
            y = np.asarray([float(r[f"{base}_median"]) for r in cal_used], dtype=np.float64)
            w = np.asarray(
                [_row_weight(r, base) for r in cal_used],
                dtype=np.float64
            )
            cal_fit = _fit_line_with_covariance(x, y, w=w, force_zero=False)

        payload["calibration_fit"] = cal_fit
        if use_stored_calibration and label in stored_channels:
            stored_pts = stored_channels.get(label, {}).get("calibration_points", []) or []
            payload["calibration_points"] = [{
                "x": float(p.get("x", np.nan)),
                "y": float(p.get("y", np.nan)),
                "yerr": float(p.get("yerr", np.nan)) if p.get("yerr", None) is not None else np.nan,
                "excluded": bool(p.get("excluded", False)),
            } for p in stored_pts if np.isfinite(p.get("x", np.nan)) and np.isfinite(p.get("y", np.nan))]
        else:
            payload["calibration_points"] = [
                {
                    "x": float(r["Conc"]),
                    "y": float(r[f"{base}_median"]),
                    "yerr": float(r.get(f"{base}_sd", np.nan)),
                    "excluded": False,
                }
                for r in cal_used
            ]

        fit_rows.append({
            "Channel": label,
            "FitType": "Calibration",
            "ID": None,
            "DF": None,
            "n_points": cal_fit["n_points"] if cal_fit else len(cal_used),
            "m": cal_fit["m"] if cal_fit else np.nan,
            "q": cal_fit["q"] if cal_fit else np.nan,
            "R2": cal_fit["R2"] if cal_fit else np.nan,
            "RMSE": cal_fit["RMSE"] if cal_fit else np.nan,
            "sigma_cal": sigma_cal,
            "sigma_source": sigma_source,
            "SNR": float(abs(cal_fit["m"]) / sigma_cal) if cal_fit and np.isfinite(sigma_cal) and sigma_cal > 0 else np.nan,
            "LOD": float(3.0 * sigma_cal / abs(cal_fit["m"])) if cal_fit and np.isfinite(sigma_cal) and sigma_cal > 0 and abs(cal_fit["m"]) > 1e-15 else np.nan,
            "LOQ": float(10.0 * sigma_cal / abs(cal_fit["m"])) if cal_fit and np.isfinite(sigma_cal) and sigma_cal > 0 and abs(cal_fit["m"]) > 1e-15 else np.nan,
            "C0": np.nan,
            "C0_sd": np.nan,
            "__cov_mq": cal_fit["cov_mq"] if cal_fit else None,
        })

        stdadd_all = [
            r for r in summary_rows
            if str(r.get("Type", "")).upper() in type_stdadd
            and np.isfinite(r.get(f"{base}_median", np.nan))
            and np.isfinite(r.get("Conc", np.nan))
        ]
        groups = {}
        for row in stdadd_all:
            groups.setdefault((str(row["ID"]), float(row["DF"])), []).append(row)

        for (sample_id, dilution_factor), rows in sorted(groups.items(), key=lambda kv: (kv[0][0], kv[0][1])):
            rows_sorted = sorted(rows, key=lambda r: float(r["Conc"]))
            used_rows = list(rows_sorted)
            exc_rows = []
            payload["excluded_points"].extend(
                [{"x": float(r["Conc"]), "y": float(r[f"{base}_median"])} for r in exc_rows]
            )

            fit = None
            c0_orig = np.nan
            c0_orig_sd = np.nan
            if len(used_rows) >= 2:
                x = np.asarray([float(r["Conc"]) for r in used_rows], dtype=np.float64)
                y = np.asarray([float(r[f"{base}_median"]) for r in used_rows], dtype=np.float64)
                w = np.asarray(
                    [_row_weight(r, base) for r in used_rows],
                    dtype=np.float64
                )
                fit = _fit_line_with_covariance(x, y, w=w)
                c0_dil, c0_dil_sd = _stdadd_c0_sd_from_fit(fit)
                c0_orig_stdadd = float(dilution_factor * c0_dil) if np.isfinite(c0_dil) else np.nan
                c0_orig_stdadd_sd = float(dilution_factor * c0_dil_sd) if np.isfinite(c0_dil_sd) else np.nan

                # When a stored calibration is used, quantify the sample from the
                # standard-addition intercept and the STORED calibration slope:
                #   C0_original = DF * ( q_std / m_cal_stored )
                # This is the intended meaning of "apply stored calibration".
                # The ordinary standard-addition estimate DF*(-q_std/m_std) is
                # still kept in auxiliary fields below for traceability.
                c0_orig = c0_orig_stdadd
                c0_orig_sd = c0_orig_stdadd_sd
                if use_stored_calibration and cal_fit is not None and fit is not None:
                    c0_corr, c0_corr_sd = _corrected_c0_from_calibration_slope(
                        cal_fit=cal_fit,
                        std_fit=fit,
                        dilution_factor=dilution_factor,
                    )
                    if np.isfinite(c0_corr):
                        c0_orig = float(c0_corr)
                        c0_orig_sd = float(c0_corr_sd) if np.isfinite(c0_corr_sd) else np.nan

            fit_rows.append({
                "Channel": label,
                "FitType": "StdAdd",
                "ID": sample_id,
                "DF": dilution_factor,
                "n_points": fit["n_points"] if fit else len(used_rows),
                "m": fit["m"] if fit else np.nan,
                "q": fit["q"] if fit else np.nan,
                "R2": fit["R2"] if fit else np.nan,
                "RMSE": fit["RMSE"] if fit else np.nan,
                "sigma_cal": np.nan,
                "sigma_source": "unavailable",
                "SNR": np.nan,
                "LOD": np.nan,
                "LOQ": np.nan,
                "C0": c0_orig,
                "C0_sd": c0_orig_sd,
                "__cov_mq": fit["cov_mq"] if fit else None,
            })

            payload["stdadd_groups"].append({
                "ID": sample_id,
                "DF": dilution_factor,
                "x": np.asarray([float(r["Conc"]) for r in used_rows], dtype=np.float64),
                "y": np.asarray([float(r[f"{base}_median"]) for r in used_rows], dtype=np.float64),
                "yerr": np.asarray([float(r.get(f"{base}_sd", np.nan)) for r in used_rows], dtype=np.float64),
                "fit": fit,
                "c0_orig": c0_orig,
                "c0_orig_sd": c0_orig_sd,
                "c0_orig_stdadd_only": c0_orig_stdadd if 'c0_orig_stdadd' in locals() else np.nan,
                "c0_orig_stdadd_only_sd": c0_orig_stdadd_sd if 'c0_orig_stdadd_sd' in locals() else np.nan,
                "c0_source": "stored_calibration_slope" if (use_stored_calibration and cal_fit is not None and fit is not None and np.isfinite(c0_orig)) else "stdadd_intercept",
            })

        unk_points = [
            r for r in summary_rows
            if str(r.get("Type", "")).upper() in type_unk
            and np.isfinite(r.get(f"{base}_median", np.nan))
        ]

        if cal_fit is not None and len(unk_points) > 0 and abs(cal_fit["m"]) > 1e-15:
            for row in unk_points:
                y_obs = float(row[f"{base}_median"])
                y_obs_cal = float(y_obs - cal_ref_value) if np.isfinite(cal_ref_value) else y_obs
                c_dil = (y_obs_cal - cal_fit["q"]) / cal_fit["m"]
                df_val = float(row["DF"]) if np.isfinite(row.get("DF", np.nan)) else np.nan
                c_orig = df_val * c_dil if np.isfinite(df_val) else c_dil
                sigma_y = float(row.get(f"{base}_sd", np.nan))
                c_dil_sd = float(abs(sigma_y / cal_fit["m"])) if np.isfinite(sigma_y) else np.nan
                c_sd = (df_val * c_dil_sd) if np.isfinite(df_val) and np.isfinite(c_dil_sd) else c_dil_sd
                n_rep = int(row.get("NReplicates", 1))

                fit_rows.append({
                    "Channel": label,
                    "FitType": "UnknownFromCal",
                    "ID": str(row["ID"]),
                    "DF": df_val,
                    "n_points": n_rep,
                    "m": cal_fit["m"],
                    "q": cal_fit["q"],
                    "R2": cal_fit["R2"],
                    "RMSE": cal_fit["RMSE"],
                    "sigma_cal": np.nan,
                    "sigma_source": "unavailable",
                    "SNR": np.nan,
                    "LOD": np.nan,
                    "LOQ": np.nan,
                    "C0": c_orig,
                    "C0_sd": c_sd,
                    "__cov_mq": None,
                })

                payload["unknown_points"].append({
                    "ID": str(row["ID"]),
                    "DF": df_val,
                    "n_points": n_rep,
                    "x": float(c_dil) if np.isfinite(c_dil) else np.nan,
                    "xerr": float(c_dil_sd) if np.isfinite(c_dil_sd) else np.nan,
                    "y": y_obs_cal,
                    "yerr": sigma_y if np.isfinite(sigma_y) else np.nan,
                    "y_raw": y_obs,
                    "c_orig": float(c_orig) if np.isfinite(c_orig) else np.nan,
                    "c_orig_sd": float(c_sd) if np.isfinite(c_sd) else np.nan,
                })
        elif len(unk_points) > 0:
            # Unknown-only DeltaE/CIELAB values are descriptive diagnostics. Without a
            # valid calibration/stored calibration, no concentration or plotted
            # unknown-point coordinate is generated.
            pass

        if payload["calibration_points"] or payload["stdadd_groups"] or payload["unknown_points"]:
            has_any = True
        plot_payload[label] = payload

    return fit_rows, plot_payload, has_any



def _build_channel_ranking_from_fit_rows(fit_rows, channels, selection_header="Ranking"):
    ranking = []
    fit_rows = list(fit_rows or [])

    for channel in channels:
        cal_rows = [
            r for r in fit_rows
            if str(r.get("Channel", "")) == str(channel)
            and str(r.get("FitType", "")) == "Calibration"
        ]
        std_rows = [
            r for r in fit_rows
            if str(r.get("Channel", "")) == str(channel)
            and str(r.get("FitType", "")) == "StdAdd"
            and np.isfinite(r.get("R2", np.nan))
        ]

        cal = cal_rows[0] if cal_rows else None
        has_cal = cal is not None and np.isfinite(cal.get("m", np.nan)) and np.isfinite(cal.get("R2", np.nan))
        has_std = len(std_rows) > 0
        best_std = max(std_rows, key=lambda r: float(r.get("R2", -np.inf))) if has_std else None

        slope = np.nan
        r2_cal = float(cal.get("R2", np.nan)) if cal is not None else np.nan
        r2_std = float(best_std.get("R2", np.nan)) if best_std is not None else np.nan
        snr = _num_or_nan(cal.get("SNR", np.nan)) if cal is not None else np.nan

        if has_cal and has_std:
            mcal = abs(float(cal.get("m", np.nan)))
            mstd = abs(float(best_std.get("m", np.nan))) if best_std is not None and np.isfinite(best_std.get("m", np.nan)) else np.nan
            slope = min(mcal, mstd) / max(mcal, mstd, 1e-12) if np.isfinite(mstd) else np.nan
            loq = _num_or_nan(cal.get("LOQ", np.nan)) if cal is not None else np.nan
            base_score = _compute_fit_base_score(r2_cal, r2_std, slope, loq=loq)
            score = base_score
            mode = "calibration_plus_stdadd"
        elif has_cal:
            score = float((float(cal.get("R2", 0.0)) ** 2) * abs(float(cal.get("m", 0.0))))
            mode = "calibration_only"
            base_score = np.nan
        elif has_std:
            score = float((float(best_std.get("R2", 0.0)) ** 2) * abs(float(best_std.get("m", 0.0))))
            mode = "stdadd_only"
            base_score = np.nan
        else:
            score = np.nan
            mode = "unavailable"
            base_score = np.nan

        ranking.append({
            "Channel": str(channel),
            "Mode": mode,
            "Score": score,
            "BaseScore": base_score,
            "R2_cal": r2_cal,
            "R2_std": r2_std,
            "SlopeAgreement": slope,
            "SNR": snr,
            "LOQ": _num_or_nan(cal.get("LOQ", np.nan)) if cal is not None else np.nan,
        })

    for row in ranking:
        if str(row.get("Mode", "")) == "calibration_plus_stdadd":
            row["Score"] = _score_candidate_for_ranking(row, use_snr=False)

    ranking_sorted = sorted(
        ranking,
        key=lambda d: (-float(d.get("Score")) if np.isfinite(d.get("Score", np.nan)) else np.inf, str(d.get("Channel", "")))
    )
    best = ranking_sorted[0] if ranking_sorted else {}

    formula_line = "Ranking formula: calibration + std add = (slope_agreement^2) x sqrt(R²cal x R²std) x (1/LOQ). SNR, clipping and expected values are excluded."

    reason_lines = [selection_header, formula_line, "Ranking:"]
    for row in ranking_sorted:
        score = row.get("Score", np.nan)
        base_score = row.get("BaseScore", np.nan)
        score_txt = f"{float(score):.4f}" if np.isfinite(score) else "nan"
        base_txt = f"{float(base_score):.4f}" if np.isfinite(base_score) else "nan"
        r2c_txt = f"{float(row.get('R2_cal', np.nan)):.4f}" if np.isfinite(row.get("R2_cal", np.nan)) else "nan"
        r2s_txt = f"{float(row.get('R2_std', np.nan)):.4f}" if np.isfinite(row.get("R2_std", np.nan)) else "nan"
        slope_txt = f"{float(row.get('SlopeAgreement', np.nan)):.4f}" if np.isfinite(row.get("SlopeAgreement", np.nan)) else "nan"
        snr_txt = f"{float(row.get('SNR', np.nan)):.4f}" if np.isfinite(row.get("SNR", np.nan)) else "nan"
        loq_txt = f"{float(row.get('LOQ', np.nan)):.5g}" if np.isfinite(row.get("LOQ", np.nan)) else "nan"
        reason_lines.append(
            f"{row['Channel']}: mode={row['Mode']}  score={score_txt}  base={base_txt}  "
            f"R²cal={r2c_txt}  R²std={r2s_txt}  slope={slope_txt}  LOQ={loq_txt}  SNR={snr_txt}"
        )

    return {"best": best, "ranking": ranking_sorted, "reason_lines": reason_lines}


class Analyzer:
    def __init__(self, image_path, cfg, img_bgr=None):
        self.image_path = image_path
        self.cfg = cfg

        if img_bgr is None:
            self.img_bgr = cv2.imread(image_path)
        else:
            self.img_bgr = img_bgr
        if self.img_bgr is None:
            raise RuntimeError("Image not loaded.")

        self.image_qc_payload = _compute_initial_image_qc_payload(self.img_bgr)
        try:
            self.cfg["__image_qc_payload"] = dict(self.image_qc_payload)
        except Exception:
            pass

        self.img_lin = to_linear_rgb01(self.img_bgr)
        self.nrow = int(cfg.get("nrow", 8))
        self.ncol = int(cfg.get("ncol", 12))

        g = load_all_config() or {}
        self.origin = np.array(g.get("d_origin_img"), dtype=np.float32)
        self.vrow = np.array(g.get("d_vrow_img"), dtype=np.float32)
        self.vcol = np.array(g.get("d_vcol_img"), dtype=np.float32)
        self.well_r0 = float(g.get("dr", 22.0))
        self.plate_geom = _get_plate_geometry(self.nrow, self.ncol)
        self.pathlength_payload = _compute_pathlength_payload_from_volume(
            cfg.get("liquid_volume_ul", np.nan), self.plate_geom
        )
        if np.isfinite(self.pathlength_payload.get("path_length", np.nan)):
            self.cfg["path_length"] = float(self.pathlength_payload["path_length"])
            self.cfg["path_length_source"] = self.pathlength_payload.get("path_length_source", "")
            self.cfg["path_length_mm"] = self.pathlength_payload.get("path_length_mm", np.nan)
            self.cfg["liquid_volume_ul"] = self.pathlength_payload.get("liquid_volume_ul", np.nan)
            self.cfg["well_bottom_diam_mm_for_pathlength"] = self.pathlength_payload.get("well_bottom_diam_mm", np.nan)
            self.cfg["well_bottom_area_mm2_for_pathlength"] = self.pathlength_payload.get("well_bottom_area_mm2", np.nan)
            self.cfg["plate_geometry_assumption"] = self.pathlength_payload.get("plate_geometry_assumption", "")
            self.cfg["plate_geometry_name"] = self.pathlength_payload.get("plate_geometry_name", "")
            self.cfg["path_length_warning"] = self.pathlength_payload.get("path_length_warning", "")
        self.fourcorner_geometry = _load_fourcorner_geometry(self.image_path, cfg)

    def compute_centers(self):
        if self.fourcorner_geometry is not None:
            return _bilinear_grid_from_four_corners(
                self.fourcorner_geometry['corner_a1'],
                self.fourcorner_geometry['corner_a12'],
                self.fourcorner_geometry['corner_h12'],
                self.fourcorner_geometry['corner_h1'],
                self.nrow,
                self.ncol,
            )
        if self.origin is None or self.vrow is None or self.vcol is None:
            raise RuntimeError("No 4-point or affine grid saved. Run alignment first.")
        return grid_centers_affine(self.origin, self.vrow, self.vcol, ncol=self.ncol, nrow=self.nrow)

    def run(self):
        centers = self.compute_centers()
        well_r = float(self.well_r0)
        use_background_subtraction = bool(self.cfg.get("use_background_subtraction", True))

        iqc = getattr(self, "image_qc_payload", {}) or {}
        print(
            "Image QC: "
            f"{iqc.get('initial_image_qc', 'NA')} | "
            f"class={iqc.get('image_qc_class', 'NA')} | "
            "analysis_only=image_qc_only"
        )
        if str(iqc.get("image_qc_messages", "")).strip():
            print(f"Image QC messages: {iqc.get('image_qc_messages')}")

        if use_background_subtraction:
            per_cell_masks, union_mask_raw, global_exclusion_mask = build_bg_masks_physical(
                centers=centers,
                img_bgr=self.img_bgr,
                well_r_px=well_r,
                pitch_mm=float(self.plate_geom.get("pitch_mm", 9.0)),
                inner_diam_mm=float(self.plate_geom.get("inner_diam_mm", 6.90)),
                outer_diam_mm=float(self.plate_geom.get("outer_diam_mm", 7.75)),
                bridge_width_mm=float(self.plate_geom.get("bridge_width_mm", 0.60)),
                extra_optical_margin_mm=float(self.plate_geom.get("extra_optical_margin_mm", 0.20)),
                canonical_size=220,
                erode_px=1,
            )

            per_cell_masks_stat = _extract_bg_masks_statistical(
                per_cell_masks=per_cell_masks,
                img_bgr=self.img_bgr,
                k_mad=2.5,
                open_close=True,
            )

            bg_samples = _extract_bg_samples(per_cell_masks_stat, self.img_bgr)
            if len(bg_samples) < 6:
                raise RuntimeError("Not enough interspace BG samples for 2D fit.")

            x_s = np.array([s["x"] for s in bg_samples], dtype=np.float64)
            y_s = np.array([s["y"] for s in bg_samples], dtype=np.float64)
            b_s = np.array([s["B"] for s in bg_samples], dtype=np.float64)
            g_s = np.array([s["G"] for s in bg_samples], dtype=np.float64)
            r_s = np.array([s["R"] for s in bg_samples], dtype=np.float64)

            model_B = _fit_poly2_robust(x_s, y_s, b_s, max_iter=6, clip_k=2.5)
            model_G = _fit_poly2_robust(x_s, y_s, g_s, max_iter=6, clip_k=2.5)
            model_R = _fit_poly2_robust(x_s, y_s, r_s, max_iter=6, clip_k=2.5)
            well_bg_rows = _predict_bg_for_wells(centers, model_B, model_G, model_R)

            union_mask_stat = np.zeros(self.img_bgr.shape[:2], dtype=np.uint8)
            for r in range(self.nrow - 1):
                for c in range(self.ncol - 1):
                    union_mask_stat = cv2.bitwise_or(union_mask_stat, per_cell_masks_stat[r][c])
        else:
            per_cell_masks = [[np.zeros(self.img_bgr.shape[:2], dtype=np.uint8) for _ in range(self.ncol - 1)] for _ in range(self.nrow - 1)]
            per_cell_masks_stat = [[np.zeros(self.img_bgr.shape[:2], dtype=np.uint8) for _ in range(self.ncol - 1)] for _ in range(self.nrow - 1)]
            union_mask_raw = np.zeros(self.img_bgr.shape[:2], dtype=np.uint8)
            global_exclusion_mask = np.zeros(self.img_bgr.shape[:2], dtype=np.uint8)
            union_mask_stat = np.zeros(self.img_bgr.shape[:2], dtype=np.uint8)
            bg_samples = []
            well_bg_rows = _empty_bg_rows_for_wells(centers)

        well_bottom_masks, well_bottom_shapes, bg_cyl_dist = _build_well_bottom_masks_v12(
            centers=centers,
            img_bgr=self.img_bgr,
            union_mask_stat=union_mask_stat,
            pitch_mm=float(self.plate_geom.get("pitch_mm", 9.0)),
            mouth_diam_mm=float(self.plate_geom.get("mouth_diam_mm", 6.90)),
            floor_diam_mm=float(self.plate_geom.get("floor_diam_mm", 6.48)),
            well_depth_mm=float(self.plate_geom.get("well_depth_mm", 10.9)),
            interface_rel_to_mouth=0.860,
            floor_rel_to_interface=0.840,
            max_shift_frac_of_interface_r=0.090,
            roi_mode=str(self.cfg.get("roi_mode", "circle")),
            floor_corner_circles={
                "floor_a1_circle_img": self.cfg.get("floor_a1_circle_img"),
                "floor_a12_circle_img": self.cfg.get("floor_a12_circle_img"),
                "floor_h12_circle_img": self.cfg.get("floor_h12_circle_img"),
                "floor_h1_circle_img": self.cfg.get("floor_h1_circle_img"),
            },
        )

        centers_overlay = _overlay_projected_mouth_circles(
            self.img_bgr,
            centers,
            mouth_diam_mm=float(self.plate_geom.get("mouth_diam_mm", 6.90)),
            pitch_mm=float(self.plate_geom.get("pitch_mm", 9.0)),
            color=(0, 255, 0),
            thickness=1,
        )

        well_stats_rows, well_debug_rows = _compute_well_robust_statistics(self.img_bgr, well_bottom_masks, return_debug=True)

        floor_qc_rows, floor_qc_payload = _build_floor_geometry_qc_rows(well_bottom_shapes)
        n_warn = int(sum(int(row.get("is_image_quality_warning", 0)) for row in well_stats_rows))
        warn_frac = n_warn / max(1, len(well_stats_rows))
        critical_wells = sum(int((float(r.get("used_fraction", 1.0)) < 0.40) or int(r.get("is_image_quality_warning", 0)) == 1 and float(r.get("used_fraction", 1.0)) < 0.55) for r in well_stats_rows)
        plate_status = "OK" if warn_frac < 0.20 and critical_wells == 0 else "WARNING"
        print(f"Plate image QC: {plate_status}. Flagged wells={n_warn}/{len(well_stats_rows)}, critical wells={critical_wells}/{len(well_stats_rows)}.")
        print(f"Floor D QC: {floor_qc_payload.get('status', 'NA')}. Flagged wells={floor_qc_payload.get('warning_wells', 0)}/{floor_qc_payload.get('n_wells', 0)}, critical wells={floor_qc_payload.get('critical_wells', 0)}/{floor_qc_payload.get('n_wells', 0)}.")

        metadata_rows = _load_well_metadata(self.cfg, self.nrow, self.ncol)
        print("Loaded metadata rows:", len(metadata_rows))
        if len(metadata_rows) == 0:
            raise RuntimeError("No well metadata found. Report workbook and A4 report cannot be generated.")

        fit_options_preview = _prompt_fitting_options([], self.cfg)
        stored_calibration_path = _get_stored_calibration_path(self.image_path, self.cfg)
        stored_calibration_bundle = None
        if bool(fit_options_preview.get("use_stored_calibration", False)):
            stored_calibration_bundle = _load_stored_calibration_bundle(stored_calibration_path)
            if stored_calibration_bundle is None:
                raise RuntimeError(f"Stored calibration file not found or invalid: {stored_calibration_path}")

        raw_report_rows_base = _build_raw_report_rows(
            well_stats_rows,
            well_bg_rows,
            metadata_rows,
            alpha_bg=1.0,
            stored_calibration_bundle=stored_calibration_bundle,
        )

        summary_tmp = _build_summary_rows([dict(r) for r in raw_report_rows_base])

        if stored_calibration_bundle is not None:
            selected_channel = _get_stored_selected_channel(stored_calibration_bundle) or "Signal_Green"
            selection_info = stored_calibration_bundle.get("selection_info", {}) or {}
            reason_lines = list(selection_info.get("reason_lines", []))
            reason_lines = _sanitize_rgb_references_in_lines(reason_lines)
            reason_lines = [f"Stored calibration loaded: {Path(stored_calibration_path).name}"] + reason_lines
            selection_info["reason_lines"] = reason_lines
        else:
            selected_channel, selection_info = _select_best_channel(summary_tmp)

        if stored_calibration_bundle is not None:
            raw_report_rows, blank_info_to_store = _compute_raw_rows_for_stored_background_mode(
                [dict(r) for r in raw_report_rows_base],
                                use_background_subtraction=bool(fit_options_preview.get("use_background_subtraction", True)),
                blank_mode=fit_options_preview.get("blank_mode", "none"),
                stored_calibration_bundle=stored_calibration_bundle,
            )
        else:
            raw_report_rows, blank_info_to_store = _compute_raw_rows_for_background_mode(
                [dict(r) for r in raw_report_rows_base],
                                use_background_subtraction=bool(fit_options_preview.get("use_background_subtraction", True)),
                blank_mode=fit_options_preview.get("blank_mode", "none"),
            )

        for _base in ["Signal_Red", "Signal_Green", "Signal_Blue"]:
            _b = blank_info_to_store[_base].get("blank", np.nan)
            _sd = blank_info_to_store[_base].get("blank_sd", np.nan)
            if np.isfinite(_b):
                prefix = "Stored blank" if stored_calibration_bundle is not None else "Blank"
                print(f"{prefix} (RAW) for {_base}: {_b:.5f} (sd={_sd:.5f})")
            else:
                msg = "No stored blank available" if stored_calibration_bundle is not None and bool(fit_options_preview.get("use_background_subtraction", True)) else "No blank available"
                if not bool(fit_options_preview.get("use_background_subtraction", True)):
                    msg = "Background subtraction OFF"
                print(f"{msg} for {_base}")

        raw_report_rows = _augment_raw_rows_with_deltae(raw_report_rows, stored_calibration_bundle=stored_calibration_bundle)
        summary_rows = _build_summary_rows(raw_report_rows)
        fit_options = fit_options_preview
        fit_options["bg_correction_alpha"] = 1.0
        fit_options["epsilon"] = _cfg_optional_float(self.cfg, "epsilon")
        fit_options["path_length"] = _cfg_optional_float(self.cfg, "path_length")
        fit_options["unit_label"] = str(self.cfg.get("unit", "mM"))
        fit_options["liquid_volume_ul"] = _cfg_optional_float(self.cfg, "liquid_volume_ul")
        fit_options["path_length_mm"] = _cfg_optional_float(self.cfg, "path_length_mm")
        fit_options["path_length_source"] = str(self.cfg.get("path_length_source", ""))
        fit_rows, plot_payload = _build_fitting_rows(
            summary_rows, selected_channel, fit_options=fit_options, stored_calibration_bundle=stored_calibration_bundle
        )
        fit_rows = _augment_fit_rows_with_bias_metrics(fit_rows)
        fit_rows_cielab, plot_payload_cielab, has_cielab_payload = _build_cielab_report_payload(
            summary_rows=summary_rows,
            fit_options=fit_options,
            stored_calibration_bundle=stored_calibration_bundle,
        )
        fit_rows_deltae, plot_payload_deltae, has_deltae_payload = _build_deltae_report_payload(
            summary_rows=summary_rows,
            fit_options=fit_options,
            stored_calibration_bundle=stored_calibration_bundle,
        )


        comparison_payload = None
        if bool(fit_options.get("show_bg_comparison_in_plots", False)):
            alt_use_bg = not bool(fit_options.get("use_background_subtraction", True))
            alt_fit_options = dict(fit_options)
            alt_fit_options["use_background_subtraction"] = alt_use_bg
            if stored_calibration_bundle is not None:
                alt_raw_rows, _ = _compute_raw_rows_for_stored_background_mode(
                    [dict(r) for r in raw_report_rows_base],
                                        use_background_subtraction=alt_use_bg,
                    blank_mode=alt_fit_options.get("blank_mode", "none"),
                    stored_calibration_bundle=stored_calibration_bundle,
                )
            else:
                alt_raw_rows, _ = _compute_raw_rows_for_background_mode(
                    [dict(r) for r in raw_report_rows_base],
                                        use_background_subtraction=alt_use_bg,
                    blank_mode=alt_fit_options.get("blank_mode", "none"),
                )
            alt_raw_rows = _augment_raw_rows_with_deltae(alt_raw_rows, stored_calibration_bundle=stored_calibration_bundle)
            alt_summary_rows = _build_summary_rows(alt_raw_rows)
            _, comparison_payload = _build_fitting_rows(
                alt_summary_rows, selected_channel, fit_options=alt_fit_options, stored_calibration_bundle=stored_calibration_bundle
            )

        selected_base_map = {"Signal_Red": "Signal_Red", "Signal_Green": "Signal_Green", "Signal_Blue": "Signal_Blue", "PAbs_Red": "PAbs_Red", "PAbs_Green": "PAbs_Green", "PAbs_Blue": "PAbs_Blue", "Red": "Signal_Red", "Green": "Signal_Green", "Blue": "Signal_Blue"}
        selected_base = selected_base_map.get(selected_channel, "PAbs_Green")
        cluster_rows = []
        best_cluster = None
        cluster_estimate = None
        empty_well_rows = _build_empty_well_rows(
            well_stats_rows=well_stats_rows,
            well_bg_rows=well_bg_rows,
            metadata_rows=metadata_rows,
        )
        spatial_payload = _build_spatial_diagnostic_payload(
            raw_rows=raw_report_rows,
            empty_rows=empty_well_rows,
            selected_channel=selected_channel,
        )
        empty_well_payload = _build_empty_well_payload(empty_well_rows)
        stored_empty_comparison_rows = _compare_empty_well_payloads(
            current_payload=empty_well_payload,
            stored_calibration_bundle=stored_calibration_bundle,
        ) if stored_calibration_bundle is not None else []
        empty_transfer_model_rows = _fit_affine_transfer_from_paired_empty_wells(
            current_empty_rows=empty_well_rows,
            stored_calibration_bundle=stored_calibration_bundle,
        ) if stored_calibration_bundle is not None else []
        bg_transfer_payload = _fit_bg_transfer_model(
            current_well_bg_rows=well_bg_rows,
            stored_calibration_bundle=stored_calibration_bundle,
        ) if stored_calibration_bundle is not None else {"summary_rows": [], "paired_rows": []}
        bg2d_correction_rows = _build_interplate_bg2d_correction_rows(
            raw_rows=raw_report_rows,
            current_well_bg_rows=well_bg_rows,
            stored_calibration_bundle=stored_calibration_bundle,
            centers=centers,
        ) if stored_calibration_bundle is not None else []
        bg2d_scale_summary_rows = _summarize_bg2d_scale_factors(bg2d_correction_rows) if bg2d_correction_rows else []
        raw_report_rows_bg2d = _build_bg2d_corrected_raw_rows(raw_report_rows, bg2d_correction_rows) if bg2d_correction_rows else []
        raw_report_rows_bg2d = _augment_raw_rows_with_deltae(raw_report_rows_bg2d, stored_calibration_bundle=stored_calibration_bundle) if raw_report_rows_bg2d else []
        summary_rows_bg2d = _build_summary_rows(raw_report_rows_bg2d) if raw_report_rows_bg2d else []
        fit_rows_bg2d, plot_payload_bg2d = _build_fitting_rows(
            summary_rows_bg2d, selected_channel, fit_options=fit_options, stored_calibration_bundle=stored_calibration_bundle
        ) if summary_rows_bg2d else ([], {})
        fit_rows_bg2d = _augment_fit_rows_with_bias_metrics(fit_rows_bg2d) if fit_rows_bg2d else []
        cielab_reference_payload = _extract_cielab_reference_payload(raw_report_rows)
        if not bool(fit_options.get("use_stored_calibration", False)):
            rgb_reference_payload = {}
            for _ch in ["Blue", "Green", "Red"]:
                _vals = np.asarray([_num_or_nan(r.get(f"MeanBG_{_ch}", np.nan)) for r in raw_report_rows_base], dtype=np.float64)
                _vals = _vals[np.isfinite(_vals) & (_vals > 0)]
                rgb_reference_payload[_ch] = {"value": float(np.nanmedian(_vals)) if _vals.size else np.nan, "source": "full_bg"}
            _save_stored_calibration_bundle(
                _get_stored_calibration_path(self.image_path, self.cfg),
                image_basename=Path(self.image_path).stem,
                unit_label=str(self.cfg.get("unit", "mM")),
                selected_channel=selected_channel,
                fit_rows=fit_rows,
                selection_info=selection_info,
                blank_info=blank_info_to_store,
                plot_payload=plot_payload,
                empty_well_payload=empty_well_payload,
                empty_well_rows=empty_well_rows,
                well_bg_rows=well_bg_rows,
                bg_samples=bg_samples,
                bg_models={"B": _bg_model_to_jsonable(model_B), "G": _bg_model_to_jsonable(model_G), "R": _bg_model_to_jsonable(model_R)} if use_background_subtraction else {},
                fit_rows_cielab=fit_rows_cielab,
                plot_payload_cielab=plot_payload_cielab,
                fit_rows_deltae=fit_rows_deltae,
                plot_payload_deltae=plot_payload_deltae,
                cielab_reference=cielab_reference_payload,
                rgb_reference=rgb_reference_payload,
            )

        run_dir, base = _make_run_dir(self.image_path)
        save_raw_data_details = bool(self.cfg.get("save_raw_data_details", self.cfg.get("save_diagnostics", False)))
        results_dir, diagnostics_dir = _prepare_output_dirs(run_dir, save_raw_data_details=save_raw_data_details)

        out_report_xlsx = os.path.join(results_dir, f"{base}_REPORT.xlsx")
        out_report_png = os.path.join(results_dir, f"{base}_FIGURE_RGB.png")
        out_plate_roi_png = os.path.join(results_dir, f"{base}_PLATE_ROI_OVERLAY.png")
        out_best_channel_png = os.path.join(results_dir, f"{base}_BEST_CHANNEL.png")
        # CIELAB/DeltaE and method-comparison figures are diagnostic/raw-data details,
        # not primary quantitative RESULTS. If diagnostics are disabled, they are not written.
        out_method_cmp_png = os.path.join(diagnostics_dir, f"{base}_METHOD_COMPARISON.png") if diagnostics_dir else None
        out_results_caption_txt = os.path.join(results_dir, f"{base}_RESULTS_CAPTION.txt")
        out_raw_caption_txt = os.path.join(diagnostics_dir, f"{base}_RAW_DATA_DETAILS_CAPTION.txt") if diagnostics_dir else None
        out_diagnostics_xlsx = os.path.join(diagnostics_dir, f"{base}_DIAGNOSTICS.xlsx") if diagnostics_dir else None
        out_report_cielab_deltae_png = os.path.join(diagnostics_dir, f"{base}_FIGURE_CIELAB_DELTAE.png") if diagnostics_dir else None

        out_mask_stat = os.path.join(diagnostics_dir, f"{base}_BG_STAT_MASK.png") if diagnostics_dir else None
        out_mask_raw = None
        out_excl = None
        out_bg_samples = os.path.join(diagnostics_dir, f"{base}_BG_SAMPLES.csv") if diagnostics_dir else None
        out_bg_wells = os.path.join(diagnostics_dir, f"{base}_BG_WELL_FIT.csv") if diagnostics_dir else None
        out_well_bottom_csv = os.path.join(diagnostics_dir, f"{base}_WELL_BOTTOM.csv") if diagnostics_dir else None
        out_plate_geom_csv = os.path.join(diagnostics_dir, f"{base}_PLATE_GEOMETRY.csv") if diagnostics_dir else None
        out_well_stats_csv = os.path.join(diagnostics_dir, f"{base}_WELL_ROBUST_STATS.csv") if diagnostics_dir else None
        out_floor_qc_csv = os.path.join(diagnostics_dir, f"{base}_FLOOR_GEOMETRY_QC.csv") if diagnostics_dir else None
        out_report_cielab_png = None
        out_report_deltae_png = None
        out_cluster_csv = None
        out_empty_wells_csv = os.path.join(diagnostics_dir, f"{base}_EMPTY_WELLS.csv") if diagnostics_dir else None
        out_spatial_diag_csv = os.path.join(diagnostics_dir, f"{base}_SPATIAL_DIAGNOSTICS.csv") if diagnostics_dir else None
        out_stored_empty_cmp_csv = os.path.join(diagnostics_dir, f"{base}_STORED_CAL_EMPTY_COMPARISON.csv") if diagnostics_dir else None
        out_empty_transfer_model_csv = os.path.join(diagnostics_dir, f"{base}_EMPTY_TRANSFER_MODEL.csv") if diagnostics_dir else None
        out_bg_transfer_summary_csv = os.path.join(diagnostics_dir, f"{base}_BG_TRANSFER_MODEL.csv") if diagnostics_dir else None
        out_bg_transfer_pairs_csv = os.path.join(diagnostics_dir, f"{base}_BG_TRANSFER_PAIRS.csv") if diagnostics_dir else None
        out_bg2d_correction_csv = os.path.join(diagnostics_dir, f"{base}_BG2D_CORRECTION.csv") if diagnostics_dir else None

        # Main report overlay: mouth/floor/intersection geometry, not pixel-level trimming debug.
        vis_overlay = _overlay_mouth_floor_roi_on_plate(
            self.img_bgr,
            well_bottom_masks,
            well_bottom_shapes,
            mouth_color=(0, 255, 0),
            floor_color=(255, 255, 0),
            roi_color=(0, 0, 0),
            roi_alpha=0.12,
        )

        _save_plate_roi_overlay_single_column_png(out_plate_roi_png, vis_overlay, width_cm=9.0, height_cm=6.0, dpi=300)

        if diagnostics_dir:
            cv2.imwrite(out_mask_stat, union_mask_stat)
            _write_bg_samples_csv(out_bg_samples, bg_samples)
            _write_well_bg_csv(out_bg_wells, well_bg_rows)
            _write_well_bottom_csv(out_well_bottom_csv, well_bottom_shapes)
            _write_plate_geometry_csv(out_plate_geom_csv, self.plate_geom)
            _write_well_statistics_csv(out_well_stats_csv, well_stats_rows)
            _write_generic_csv(out_floor_qc_csv, floor_qc_rows)
            _write_bg2d_correction_csv(out_bg2d_correction_csv, bg2d_correction_rows)

        plate_qc_payload = {
            "status": plate_status,
            "flagged_wells": n_warn,
            "critical_wells": critical_wells,
            "total_wells": len(well_stats_rows),
            "bright_excl_frac_med": float(np.nanmedian(np.asarray([r.get("BrightExcludedFraction", np.nan) for r in well_stats_rows], dtype=np.float64))),
            "bright_mean_gray_med": float(np.nanmedian(np.asarray([r.get("BrightExcludedMeanGray", np.nan) for r in well_stats_rows], dtype=np.float64))),
            "highlight_index_med": float(np.nanmedian(np.asarray([r.get("HighlightIndex", np.nan) for r in well_stats_rows], dtype=np.float64))),
            "floor_qc_status": floor_qc_payload.get("status", "NA"),
            "floor_warning_wells": int(floor_qc_payload.get("warning_wells", 0) or 0),
            "floor_critical_wells": int(floor_qc_payload.get("critical_wells", 0) or 0),
            "floor_shift_frac_med": float(floor_qc_payload.get("shift_frac_median", np.nan)),
            "floor_radius_ratio_med": float(floor_qc_payload.get("radius_ratio_median", np.nan)),
        }

        expected_refs = _build_expected_reference_rows(self.cfg)
        image_qc_info = _compute_image_qc_info(
            img_bgr=self.img_bgr,
            centers=centers,
            well_bottom_masks=well_bottom_masks,
            image_path=self.image_path,
            cfg=self.cfg,
        )
        metadata_counts = _count_metadata_types(metadata_rows)
        empty_qc_payload = _compute_empty_well_qc_status(empty_well_payload, stored_empty_comparison_rows=stored_empty_comparison_rows)
        reliability_payload = _build_reliability_payload(
            metadata_counts=metadata_counts,
            fit_rows=fit_rows,
            fit_options=fit_options,
            plate_qc_payload=plate_qc_payload,
            empty_qc_payload=empty_qc_payload,
            selection_info=selection_info,
            cfg=self.cfg,
        )
        method_comparison_rows = _build_method_comparison_rows(
            fit_rows=fit_rows,
            selection_info=selection_info,
            expected_refs=expected_refs,
        )
        bg_channel_model_rows = _build_bg_channel_model_rows(
            raw_rows=raw_report_rows,
            fit_rows=fit_rows,
            expected_refs=expected_refs,
        )

        _write_report_workbook(
            path=out_report_xlsx,
            image_basename=base,
            unit_label=str(self.cfg.get("unit", "mM")),
            selected_channel=selected_channel,
            raw_rows=raw_report_rows,
            summary_rows=summary_rows,
            fit_rows=fit_rows,
            selection_info=selection_info,
            fit_options=fit_options,
            empty_well_rows=empty_well_rows,
            method_comparison_rows=method_comparison_rows,
            expected_refs=expected_refs,
            reliability_payload=reliability_payload,
            image_qc_info=image_qc_info,
        )
        if out_method_cmp_png:
            _make_method_comparison_png(
                path=out_method_cmp_png,
                image_basename=base,
                comparison_rows=method_comparison_rows,
                unit_label=str(self.cfg.get("unit", "mM")),
                expected_refs=expected_refs,
            )
        _make_a4_report_png(
            path=out_report_png,
            image_basename=base,
            overlay_img=vis_overlay,
            plot_payload=plot_payload,
            fit_rows=fit_rows,
            selected_channel=selected_channel,
            selection_info=selection_info,
            unit_label=str(self.cfg.get("unit", "mM")),
            fit_options=fit_options,
            plate_qc_info=plate_qc_payload,
            comparison_payload=comparison_payload,
            cluster_estimate=None,
            fit_rows_bg2d=fit_rows_bg2d,
            bg2d_scale_summary=bg2d_scale_summary_rows,
            expected_refs=expected_refs,
            best_channel_plot_path=out_best_channel_png,
        )
        if out_report_cielab_deltae_png and (has_cielab_payload or has_deltae_payload):
            plot_payload_cielab_deltae = {}
            plot_payload_cielab_deltae.update(plot_payload_deltae or {})
            plot_payload_cielab_deltae.update(plot_payload_cielab or {})
            fit_rows_cielab_deltae_raw = list(fit_rows_deltae or []) + list(fit_rows_cielab or [])
            fit_rows_cielab_deltae = []
            _seen_fit_keys = set()
            for _fr in fit_rows_cielab_deltae_raw:
                _key = (
                    str(_fr.get("Channel", "")),
                    str(_fr.get("FitType", "")),
                    str(_fr.get("ID", "")),
                    str(_fr.get("DF", "")),
                )
                if _key in _seen_fit_keys:
                    continue
                _seen_fit_keys.add(_key)
                fit_rows_cielab_deltae.append(_fr)
            _make_a4_report_png(
                path=out_report_cielab_deltae_png,
                image_basename=base,
                overlay_img=vis_overlay,
                plot_payload=plot_payload_cielab_deltae,
                fit_rows=fit_rows_cielab_deltae,
                selected_channel="CIELAB_DELTAE",
                selection_info=_build_channel_ranking_from_fit_rows(
                    fit_rows_cielab_deltae,
                    ["DeltaE_ab", "DeltaE_ab_chroma", "DeltaL", "Deltaa", "Deltab"],
                    selection_header="Dedicated combined CIELAB / DeltaE report."
                ),
                unit_label=str(self.cfg.get("unit", "mM")),
                fit_options=fit_options,
                plate_qc_info=plate_qc_payload,
                comparison_payload=None,
                cluster_estimate=None,
                fit_rows_bg2d=None,
                bg2d_scale_summary=None,
                expected_refs=expected_refs,
            )
        _write_results_caption_txt(out_results_caption_txt, base, str(self.cfg.get("unit", "mM")), expected_refs=expected_refs)

        if diagnostics_dir:
            _write_raw_data_details_caption_txt(out_raw_caption_txt, base, str(self.cfg.get("unit", "mM")))
            _write_empty_well_csv(out_empty_wells_csv, empty_well_rows)

            with open(out_spatial_diag_csv, "w", newline="", encoding="utf-8") as f:
                wr = csv.writer(f, delimiter=';')
                wr.writerow(["dataset", "n", "intercept", "slope_col", "slope_row", "R2", "corr_col", "corr_row"])
                for dataset_name, trend in [
                    ("unknown", spatial_payload.get("unknown_trend", {})),
                    ("empty", spatial_payload.get("empty_trend", {})),
                ]:
                    wr.writerow([
                        dataset_name,
                        trend.get("n", ""),
                        trend.get("intercept", ""),
                        trend.get("slope_col", ""),
                        trend.get("slope_row", ""),
                        trend.get("R2", ""),
                        trend.get("corr_col", ""),
                        trend.get("corr_row", ""),
                    ])
            if stored_calibration_bundle is not None:
                _write_stored_empty_comparison_csv(out_stored_empty_cmp_csv, stored_empty_comparison_rows)
                _write_empty_transfer_model_csv(out_empty_transfer_model_csv, empty_transfer_model_rows)
                _write_bg_transfer_summary_csv(out_bg_transfer_summary_csv, bg_transfer_payload.get("summary_rows", []))
                _write_bg_transfer_pairs_csv(out_bg_transfer_pairs_csv, bg_transfer_payload.get("paired_rows", []))

            _diagnostic_csv_paths = {
                "BG_SAMPLES": out_bg_samples,
                "BG_WELL_FIT": out_bg_wells,
                "WELL_ROBUST_STATS": out_well_stats_csv,
                "FLOOR_GEOMETRY_QC": out_floor_qc_csv,
                "WELL_BOTTOM": out_well_bottom_csv,
                "PLATE_GEOMETRY": out_plate_geom_csv,
                "EMPTY_WELLS": out_empty_wells_csv,
                "SPATIAL_DIAGNOSTICS": out_spatial_diag_csv,
                "STORED_CAL_EMPTY_COMPARISON": out_stored_empty_cmp_csv,
                "EMPTY_TRANSFER_MODEL": out_empty_transfer_model_csv,
                "BG_TRANSFER_MODEL": out_bg_transfer_summary_csv,
                "BG_TRANSFER_PAIRS": out_bg_transfer_pairs_csv,
                "BG2D_CORRECTION": out_bg2d_correction_csv,
            }
            _write_diagnostics_workbook(
                path=out_diagnostics_xlsx,
                image_basename=base,
                unit_label=str(self.cfg.get("unit", "mM")),
                csv_paths=_diagnostic_csv_paths,
                method_comparison_rows=method_comparison_rows,
                cielab_fit_rows=(list(fit_rows_cielab or []) + list(fit_rows_deltae or [])),
            )
            _remove_duplicate_csv_files(_diagnostic_csv_paths)

        if best_cluster is not None:
            print(f"\nBEST SPATIAL CLUSTER ({selected_channel} channel):")
            print(f"size={best_cluster['h']}x{best_cluster['w']}, n={best_cluster['n']}")
            print(f"{selected_base} mean={best_cluster['mean']:.5f}")
            print(f"{selected_base} sd={best_cluster['sd']:.5f}")
            print(f"wells: {best_cluster['wells']}")
            if isinstance(cluster_estimate, dict) and np.isfinite(cluster_estimate.get("C_est", np.nan)):
                print(
                    f"Cluster-based estimate ({selected_channel}): "
                    f"{abs(float(cluster_estimate['C_est'])):.2f} +/- "
                    f"{float(cluster_estimate['C_est_sd']):.2f} {cluster_estimate.get('Unit', str(self.cfg.get('unit', 'mM')))}"
                )

        unknown_trend = spatial_payload.get("unknown_trend", {}) or {}
        empty_trend = spatial_payload.get("empty_trend", {}) or {}
        print(f"Spatial trend (unknown, {selected_channel}): corr_col={unknown_trend.get('corr_col', np.nan):.3f}, corr_row={unknown_trend.get('corr_row', np.nan):.3f}, R²={unknown_trend.get('R2', np.nan):.3f}")
        if len(empty_well_rows) > 0:
            print(f"Spatial trend (empty, {selected_channel}): corr_col={empty_trend.get('corr_col', np.nan):.3f}, corr_row={empty_trend.get('corr_row', np.nan):.3f}, R²={empty_trend.get('R2', np.nan):.3f}")
        if stored_calibration_bundle is not None and len(stored_empty_comparison_rows) > 0:
            print("Stored calibration empty-well comparison:")
            for _row in stored_empty_comparison_rows:
                print(
                    f"{_row['Channel']}: delta_median={_row.get('Delta_median', np.nan):+.5f}, "
                    f"ratio_median={_row.get('Ratio_median', np.nan):.5f}, "
                    f"delta_mean={_row.get('Delta_mean', np.nan):+.5f}, "
                    f"ratio_mean={_row.get('Ratio_mean', np.nan):.5f}"
                )
        if stored_calibration_bundle is not None and len(empty_transfer_model_rows) > 0:
            print("Empty-based transfer model (diagnostic only):")
            for _row in empty_transfer_model_rows:
                print(
                    f"{_row['Channel']}: mode={_row.get('Mode', 'na')}, "
                    f"a={_row.get('a', np.nan):.5f}, "
                    f"b={_row.get('b', np.nan):+.5f}, "
                    f"R²={_row.get('R2', np.nan):.5f}, "
                    f"n_pairs={int(_row.get('n_pairs', 0) or 0)}"
                )
        if stored_calibration_bundle is not None and len(bg_transfer_payload.get("summary_rows", [])) > 0:
            print("BG-vs-BG transfer model (diagnostic only):")
            for _row in bg_transfer_payload.get("summary_rows", []):
                print(
                    f"{_row['Channel']}: a={_row.get('a', np.nan):.5f}, "
                    f"b={_row.get('b', np.nan):+.5f}, "
                    f"R²={_row.get('R2', np.nan):.5f}, "
                    f"n_pairs={int(_row.get('n_pairs', 0) or 0)}"
                )

        print(f"Quantification: {reliability_payload.get('quantification_status', 'available')}")
        print(f"Reliability score: {reliability_payload.get('reliability_score', np.nan):.1f}")
        print(f"Confidence class: {reliability_payload.get('confidence_class', '')}")
        print(f"Reliability note: {reliability_payload.get('reason', '')}")
        def _print_saved_if_exists(_path):
            if _path and os.path.exists(_path):
                print("Saved:", _path)

        _print_saved_if_exists(out_report_xlsx)
        _print_saved_if_exists(out_report_png)
        _print_saved_if_exists(out_method_cmp_png)
        _print_saved_if_exists(out_results_caption_txt)
        _print_saved_if_exists(out_report_cielab_deltae_png)
        if diagnostics_dir:
            _print_saved_if_exists(out_raw_caption_txt)
            _print_saved_if_exists(out_diagnostics_xlsx)
            _print_saved_if_exists(out_report_cielab_png)
            _print_saved_if_exists(out_report_deltae_png)
            _print_saved_if_exists(out_cluster_csv)
            _print_saved_if_exists(out_empty_wells_csv)
            _print_saved_if_exists(out_spatial_diag_csv)
            if stored_calibration_bundle is not None:
                _print_saved_if_exists(out_stored_empty_cmp_csv)
                _print_saved_if_exists(out_empty_transfer_model_csv)
                _print_saved_if_exists(out_bg_transfer_summary_csv)
                _print_saved_if_exists(out_bg_transfer_pairs_csv)


def _safe_solve_least_squares(X, y):
    try:
        coef, _, _, _ = np.linalg.lstsq(X, y, rcond=None)
        return coef
    except Exception:
        return None


def _predict_from_linear_model(model, feature_row):
    if not isinstance(model, dict):
        return np.nan
    features = model.get("features", []) or []
    coef = np.asarray(model.get("coef", []), dtype=np.float64)
    if coef.size != len(features) + 1:
        return np.nan
    vals = []
    for feat in features:
        v = feature_row.get(feat, np.nan)
        if not np.isfinite(v):
            return np.nan
        vals.append(float(v))
    x = np.asarray([1.0] + vals, dtype=np.float64)
    pred = float(np.dot(x, coef))
    return pred if np.isfinite(pred) else np.nan
