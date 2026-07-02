#!/usr/bin/env python3
"""Direct XLSX/TXT parity audit for Python and web TIPICA exports.

This tool intentionally compares the actual exported files instead of relying
on higher-level comparator summaries. Python output is the reference.
"""

from __future__ import annotations

import argparse
import hashlib
import math
import re
import sys
import zipfile
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from io import BytesIO
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


STATUS_TRUE = "true"
STATUS_FALSE = "false"


@dataclass
class SourceFiles:
    root: Path
    input_path: str
    kind: str
    files: dict[str, bytes]
    sha256: str | None = None


@dataclass
class CellDiff:
    coordinate: str
    classification: str
    python_value: Any
    web_value: Any
    python_type: str
    web_type: str


@dataclass
class SheetAudit:
    workbook_path: str
    sheet_name: str
    python_shape: tuple[int, int] | None
    web_shape: tuple[int, int] | None
    shape_match: bool
    cell_value_match: bool
    n_differences: int = 0
    n_text_differences: int = 0
    n_numeric_differences: int = 0
    n_blank_pattern_differences: int = 0
    n_type_differences: int = 0
    first_diffs: list[CellDiff] = field(default_factory=list)
    is_legend: bool = False


@dataclass
class WorkbookAudit:
    path: str
    present_python: bool
    present_web: bool
    sheet_count_match: bool = False
    sheet_names_match: bool = False
    sheet_order_match: bool = False
    python_sheets: list[str] = field(default_factory=list)
    web_sheets: list[str] = field(default_factory=list)
    sheet_audits: list[SheetAudit] = field(default_factory=list)
    load_error: str = ""

    @property
    def full_match(self) -> bool:
        return (
            self.present_python
            and self.present_web
            and not self.load_error
            and self.sheet_count_match
            and self.sheet_names_match
            and self.sheet_order_match
            and all(sheet.shape_match and sheet.cell_value_match for sheet in self.sheet_audits)
        )


@dataclass
class TextDiffLine:
    index: int
    python_text: str
    web_text: str


@dataclass
class TextAudit:
    path: str
    present_python: bool
    present_web: bool
    exact_match: bool = False
    normalized_line_ending_match: bool = False
    line_count_python: int = 0
    line_count_web: int = 0
    paragraph_count_python: int = 0
    paragraph_count_web: int = 0
    classification: str = "missing"
    first_line_diffs: list[TextDiffLine] = field(default_factory=list)
    first_paragraph_diffs: list[TextDiffLine] = field(default_factory=list)


def canonical_member_name(name: str) -> str:
    parts = [part for part in name.replace("\\", "/").split("/") if part]
    if len(parts) >= 3 and parts[0] not in {"RESULTS", "RAW_DATA_DETAILS"} and parts[1] in {"RESULTS", "RAW_DATA_DETAILS"}:
        parts = parts[1:]
    return "/".join(parts)


def read_source(path: Path, input_path: str | None = None) -> SourceFiles:
    display_path = input_path or str(path)
    if path.is_file():
        files: dict[str, bytes] = {}
        with zipfile.ZipFile(path) as zf:
            for name in zf.namelist():
                if name.endswith("/"):
                    continue
                files[canonical_member_name(name)] = zf.read(name)
        return SourceFiles(path, display_path, "zip", files, hashlib.sha256(path.read_bytes()).hexdigest())
    if path.is_dir():
        files = {
            item.relative_to(path).as_posix(): item.read_bytes()
            for item in path.rglob("*")
            if item.is_file()
        }
        return SourceFiles(path, display_path, "dir", files, None)
    raise FileNotFoundError(f"Input path does not exist or is not a file/directory: {path}")


def parse_warning_substrings(value: str) -> list[str]:
    return [item.strip() for item in value.split(",") if item.strip()]


def matching_python_name_warnings(path: Path, substrings: Iterable[str]) -> list[str]:
    filename = path.name
    return [substring for substring in substrings if substring in filename]


def group_by_extension(names: Iterable[str]) -> dict[str, list[str]]:
    grouped: dict[str, list[str]] = defaultdict(list)
    for name in sorted(names):
        suffix = Path(name).suffix.lower() or "other"
        if suffix not in {".xlsx", ".txt", ".png"}:
            suffix = "other"
        grouped[suffix].append(name)
    return grouped


def is_blank(value: Any) -> bool:
    return value is None or value == ""


def is_numeric(value: Any) -> bool:
    return isinstance(value, (int, float)) and not isinstance(value, bool) and math.isfinite(float(value))


def value_type_name(value: Any) -> str:
    if value is None:
        return "blank"
    return type(value).__name__


def cell_values_equal(py_value: Any, web_value: Any, numeric_tolerance: float) -> bool:
    if is_blank(py_value) and is_blank(web_value):
        return True
    if is_numeric(py_value) and is_numeric(web_value):
        return abs(float(py_value) - float(web_value)) <= numeric_tolerance
    return py_value == web_value


def classify_cell_diff(py_value: Any, web_value: Any, numeric_tolerance: float) -> str:
    if is_blank(py_value) and is_blank(web_value):
        return "both blank"
    if is_blank(py_value) != is_blank(web_value):
        return "blank vs nonblank"
    if is_numeric(py_value) and is_numeric(web_value):
        return "numeric mismatch" if abs(float(py_value) - float(web_value)) > numeric_tolerance else "both blank"
    if type(py_value) is not type(web_value):
        return "type mismatch"
    if isinstance(py_value, str) or isinstance(web_value, str):
        return "text mismatch"
    return "type mismatch" if py_value != web_value else "both blank"


def format_value(value: Any, limit: int = 160) -> str:
    if value is None:
        return ""
    text = str(value).replace("\r", "\\r").replace("\n", "\\n")
    if len(text) > limit:
        return text[: limit - 3] + "..."
    return text


def excel_coordinate(row: int, col: int) -> str:
    return f"{get_column_letter(col)}{row}"


def load_xlsx(data: bytes):
    return load_workbook(BytesIO(data), data_only=False, read_only=False)


def audit_sheet(py_ws, web_ws, workbook_path: str, sheet_name: str, numeric_tolerance: float, max_diffs: int) -> SheetAudit:
    py_shape = (py_ws.max_row, py_ws.max_column)
    web_shape = (web_ws.max_row, web_ws.max_column)
    max_row = max(py_ws.max_row, web_ws.max_row)
    max_col = max(py_ws.max_column, web_ws.max_column)
    audit = SheetAudit(
        workbook_path=workbook_path,
        sheet_name=sheet_name,
        python_shape=py_shape,
        web_shape=web_shape,
        shape_match=py_shape == web_shape,
        cell_value_match=True,
        is_legend="LEGEND" in sheet_name.upper(),
    )
    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            py_value = py_ws.cell(row=row, column=col).value
            web_value = web_ws.cell(row=row, column=col).value
            if cell_values_equal(py_value, web_value, numeric_tolerance):
                continue
            classification = classify_cell_diff(py_value, web_value, numeric_tolerance)
            audit.cell_value_match = False
            audit.n_differences += 1
            if classification == "text mismatch":
                audit.n_text_differences += 1
            elif classification == "numeric mismatch":
                audit.n_numeric_differences += 1
            elif classification == "blank vs nonblank":
                audit.n_blank_pattern_differences += 1
            elif classification == "type mismatch":
                audit.n_type_differences += 1
            if len(audit.first_diffs) < max_diffs:
                audit.first_diffs.append(CellDiff(
                    coordinate=excel_coordinate(row, col),
                    classification=classification,
                    python_value=py_value,
                    web_value=web_value,
                    python_type=value_type_name(py_value),
                    web_type=value_type_name(web_value),
                ))
    return audit


def audit_workbook(path: str, py_data: bytes | None, web_data: bytes | None, numeric_tolerance: float, max_diffs: int) -> WorkbookAudit:
    audit = WorkbookAudit(path=path, present_python=py_data is not None, present_web=web_data is not None)
    if py_data is None or web_data is None:
        return audit
    try:
        py_wb = load_xlsx(py_data)
        web_wb = load_xlsx(web_data)
    except Exception as exc:
        audit.load_error = str(exc)
        return audit
    audit.python_sheets = list(py_wb.sheetnames)
    audit.web_sheets = list(web_wb.sheetnames)
    audit.sheet_count_match = len(audit.python_sheets) == len(audit.web_sheets)
    audit.sheet_names_match = set(audit.python_sheets) == set(audit.web_sheets)
    audit.sheet_order_match = audit.python_sheets == audit.web_sheets
    for sheet_name in audit.python_sheets:
        if sheet_name not in web_wb.sheetnames:
            audit.sheet_audits.append(SheetAudit(
                workbook_path=path,
                sheet_name=sheet_name,
                python_shape=(py_wb[sheet_name].max_row, py_wb[sheet_name].max_column),
                web_shape=None,
                shape_match=False,
                cell_value_match=False,
                n_differences=1,
                is_legend="LEGEND" in sheet_name.upper(),
            ))
            continue
        audit.sheet_audits.append(audit_sheet(py_wb[sheet_name], web_wb[sheet_name], path, sheet_name, numeric_tolerance, max_diffs))
    for sheet_name in audit.web_sheets:
        if sheet_name not in py_wb.sheetnames:
            audit.sheet_audits.append(SheetAudit(
                workbook_path=path,
                sheet_name=sheet_name,
                python_shape=None,
                web_shape=(web_wb[sheet_name].max_row, web_wb[sheet_name].max_column),
                shape_match=False,
                cell_value_match=False,
                n_differences=1,
                is_legend="LEGEND" in sheet_name.upper(),
            ))
    return audit


def normalize_line_endings(text: str) -> str:
    return text.replace("\r\n", "\n").replace("\r", "\n")


def split_paragraphs(text: str) -> list[str]:
    normalized = normalize_line_endings(text).strip()
    if not normalized:
        return []
    return [paragraph.strip() for paragraph in re.split(r"\n\s*\n+", normalized)]


def first_sequence_diffs(py_items: list[str], web_items: list[str], max_diffs: int) -> list[TextDiffLine]:
    diffs: list[TextDiffLine] = []
    max_len = max(len(py_items), len(web_items))
    for index in range(max_len):
        py_text = py_items[index] if index < len(py_items) else ""
        web_text = web_items[index] if index < len(web_items) else ""
        if py_text == web_text:
            continue
        diffs.append(TextDiffLine(index=index + 1, python_text=py_text, web_text=web_text))
        if len(diffs) >= max_diffs:
            break
    return diffs


def audit_text(path: str, py_data: bytes | None, web_data: bytes | None, max_diffs: int) -> TextAudit:
    audit = TextAudit(path=path, present_python=py_data is not None, present_web=web_data is not None)
    if py_data is None or web_data is None:
        audit.classification = "missing"
        return audit
    py_text = py_data.decode("utf-8", errors="replace")
    web_text = web_data.decode("utf-8", errors="replace")
    audit.exact_match = py_text == web_text
    py_norm = normalize_line_endings(py_text)
    web_norm = normalize_line_endings(web_text)
    audit.normalized_line_ending_match = py_norm == web_norm
    py_lines = py_norm.split("\n")
    web_lines = web_norm.split("\n")
    py_paragraphs = split_paragraphs(py_text)
    web_paragraphs = split_paragraphs(web_text)
    audit.line_count_python = len(py_lines)
    audit.line_count_web = len(web_lines)
    audit.paragraph_count_python = len(py_paragraphs)
    audit.paragraph_count_web = len(web_paragraphs)
    audit.first_line_diffs = first_sequence_diffs(py_lines, web_lines, max_diffs)
    audit.first_paragraph_diffs = first_sequence_diffs(py_paragraphs, web_paragraphs, max_diffs)
    if audit.exact_match:
        audit.classification = "exact match"
    elif audit.normalized_line_ending_match:
        audit.classification = "line-ending-only difference"
    else:
        audit.classification = "text difference"
    return audit


def md_escape(text: Any) -> str:
    return str(text).replace("|", "\\|")


def markdown_list(items: Iterable[str]) -> list[str]:
    values = list(items)
    return [f"- `{item}`" for item in values] if values else ["- none"]


def append_file_list_section(report: list[str], py_files: dict[str, bytes], web_files: dict[str, bytes]) -> None:
    py_names = set(py_files)
    web_names = set(web_files)
    common = sorted(py_names & web_names)
    only_py = sorted(py_names - web_names)
    only_web = sorted(web_names - py_names)
    report.extend([
        "## File-list Parity",
        f"- Python files: {len(py_names)}",
        f"- Web files: {len(web_names)}",
        f"- Common files: {len(common)}",
        f"- Present only in Python: {len(only_py)}",
        f"- Present only in web: {len(only_web)}",
        "",
        "### Present only in Python",
        *markdown_list(only_py),
        "",
        "### Present only in Web",
        *markdown_list(only_web),
        "",
        "### Common Files by Extension",
    ])
    grouped = group_by_extension(common)
    for extension in [".xlsx", ".txt", ".png", "other"]:
        report.append(f"- {extension}: {len(grouped.get(extension, []))}")
        for name in grouped.get(extension, []):
            report.append(f"  - `{name}`")
    report.append("")


def append_reference_provenance(
    report: list[str],
    py_source: SourceFiles,
    web_source: SourceFiles,
    reference_label: str,
    python_name_warnings: list[str],
) -> None:
    report.extend([
        "## Reference Provenance",
        f"- Reference label: `{reference_label}`",
        f"- Python input path: `{py_source.input_path}`",
        f"- Python source kind: `{py_source.kind}`",
        f"- Python total file count: {len(py_source.files)}",
        f"- Python ZIP SHA-256: `{py_source.sha256 or 'n/a'}`",
        f"- Web input path: `{web_source.input_path}`",
        f"- Web source kind: `{web_source.kind}`",
        f"- Web total file count: {len(web_source.files)}",
        f"- Web ZIP SHA-256: `{web_source.sha256 or 'n/a'}`",
    ])
    if python_name_warnings:
        report.extend([
            "",
            "### WARNING",
            f"- WARNING: Python input filename `{py_source.root.name}` contains listed warning substring(s): "
            f"{', '.join(f'`{substring}`' for substring in python_name_warnings)}",
        ])
    report.append("")


def append_workbook_summary(report: list[str], audits: list[WorkbookAudit]) -> None:
    report.extend([
        "## XLSX Workbook Summary",
        "| Workbook | Present Python | Present Web | Sheet count | Sheet names | Sheet order | Cell values | Full match |",
        "|---|---:|---:|---:|---:|---:|---:|---:|",
    ])
    for audit in audits:
        cell_match = all(sheet.shape_match and sheet.cell_value_match for sheet in audit.sheet_audits) and not audit.load_error
        report.append(
            f"| `{audit.path}` | {audit.present_python} | {audit.present_web} | {audit.sheet_count_match} | "
            f"{audit.sheet_names_match} | {audit.sheet_order_match} | {cell_match} | {audit.full_match} |"
        )
    if not audits:
        report.append("| none | false | false | false | false | false | false | false |")
    report.append("")


def append_sheet_details(report: list[str], audits: list[WorkbookAudit], max_diffs: int) -> None:
    report.append("## Detailed XLSX Sheet Differences")
    for audit in audits:
        report.append(f"### `{audit.path}`")
        if audit.load_error:
            report.append(f"- load_error: `{audit.load_error}`")
            continue
        report.append(f"- python_sheets={audit.python_sheets}")
        report.append(f"- web_sheets={audit.web_sheets}")
        if not audit.sheet_count_match or not audit.sheet_names_match or not audit.sheet_order_match:
            report.append(
                f"- sheet_structure_mismatch: count_match={audit.sheet_count_match}, "
                f"names_match={audit.sheet_names_match}, order_match={audit.sheet_order_match}"
            )
        report.append("| Sheet | Legend | Python shape | Web shape | shape_match | cell_value_match | n_differences | text | numeric | blank-pattern | type |")
        report.append("|---|---:|---:|---:|---:|---:|---:|---:|---:|---:|---:|")
        for sheet in audit.sheet_audits:
            report.append(
                f"| `{sheet.sheet_name}` | {sheet.is_legend} | {sheet.python_shape} | {sheet.web_shape} | "
                f"{sheet.shape_match} | {sheet.cell_value_match} | {sheet.n_differences} | "
                f"{sheet.n_text_differences} | {sheet.n_numeric_differences} | "
                f"{sheet.n_blank_pattern_differences} | {sheet.n_type_differences} |"
            )
        legend_sheets = [sheet for sheet in audit.sheet_audits if sheet.is_legend]
        if legend_sheets:
            report.append("#### Legend Sheets")
            for sheet in legend_sheets:
                report.append(
                    f"- `{sheet.sheet_name}`: shape_match={sheet.shape_match}, "
                    f"cell_value_match={sheet.cell_value_match}, n_differences={sheet.n_differences}"
                )
        for sheet in audit.sheet_audits:
            if not sheet.first_diffs and sheet.cell_value_match:
                continue
            report.append(f"#### `{sheet.sheet_name}` first {max_diffs} differing cells")
            if not sheet.first_diffs:
                report.append("- sheet missing on one side or no stored cell examples")
                continue
            report.append("| Cell | Class | Python type | Web type | Python value | Web value |")
            report.append("|---|---|---|---|---|---|")
            for diff in sheet.first_diffs:
                report.append(
                    f"| `{diff.coordinate}` | {diff.classification} | {diff.python_type} | {diff.web_type} | "
                    f"`{md_escape(format_value(diff.python_value))}` | `{md_escape(format_value(diff.web_value))}` |"
                )
        report.append("")


def append_text_summary(report: list[str], audits: list[TextAudit]) -> None:
    report.extend([
        "## TXT Summary",
        "| TXT file | Present Python | Present Web | Classification | Exact | Normalized line endings | Python lines | Web lines | Python paragraphs | Web paragraphs |",
        "|---|---:|---:|---|---:|---:|---:|---:|---:|---:|",
    ])
    for audit in audits:
        report.append(
            f"| `{audit.path}` | {audit.present_python} | {audit.present_web} | {audit.classification} | "
            f"{audit.exact_match} | {audit.normalized_line_ending_match} | {audit.line_count_python} | "
            f"{audit.line_count_web} | {audit.paragraph_count_python} | {audit.paragraph_count_web} |"
        )
    if not audits:
        report.append("| none | false | false | missing | false | false | 0 | 0 | 0 | 0 |")
    report.append("")


def append_text_details(report: list[str], audits: list[TextAudit]) -> None:
    report.append("## Detailed TXT Differences")
    for audit in audits:
        report.append(f"### `{audit.path}`")
        report.append(f"- classification={audit.classification}")
        report.append(f"- exact_match={audit.exact_match}")
        report.append(f"- normalized_line_ending_match={audit.normalized_line_ending_match}")
        report.append("#### First differing lines")
        if audit.first_line_diffs:
            report.append("| Line | Python | Web |")
            report.append("|---:|---|---|")
            for diff in audit.first_line_diffs:
                report.append(f"| {diff.index} | `{md_escape(format_value(diff.python_text, 220))}` | `{md_escape(format_value(diff.web_text, 220))}` |")
        else:
            report.append("- none")
        report.append("#### First differing paragraphs")
        if audit.first_paragraph_diffs:
            report.append("| Paragraph | Python | Web |")
            report.append("|---:|---|---|")
            for diff in audit.first_paragraph_diffs:
                report.append(f"| {diff.index} | `{md_escape(format_value(diff.python_text, 260))}` | `{md_escape(format_value(diff.web_text, 260))}` |")
        else:
            report.append("- none")
        report.append("")


def audit_sources(py_source: SourceFiles, web_source: SourceFiles, numeric_tolerance: float, max_diffs: int) -> tuple[list[WorkbookAudit], list[TextAudit]]:
    py_files = py_source.files
    web_files = web_source.files
    xlsx_paths = sorted(name for name in set(py_files) | set(web_files) if name.lower().endswith(".xlsx"))
    txt_paths = sorted(name for name in set(py_files) | set(web_files) if name.lower().endswith(".txt"))
    workbook_audits = [
        audit_workbook(path, py_files.get(path), web_files.get(path), numeric_tolerance, max_diffs)
        for path in xlsx_paths
    ]
    text_audits = [
        audit_text(path, py_files.get(path), web_files.get(path), max_diffs)
        for path in txt_paths
    ]
    return workbook_audits, text_audits


def build_report(
    py_source: SourceFiles,
    web_source: SourceFiles,
    workbook_audits: list[WorkbookAudit],
    text_audits: list[TextAudit],
    numeric_tolerance: float,
    max_diffs: int,
    include_styles: bool,
    reference_label: str,
    python_name_warnings: list[str],
) -> str:
    xlsx_full_parity = all(audit.full_match for audit in workbook_audits) and bool(workbook_audits)
    txt_full_parity = all(audit.exact_match for audit in text_audits) and bool(text_audits)
    safe_to_claim = xlsx_full_parity and txt_full_parity
    xlsx_diff_counts = Counter()
    for workbook in workbook_audits:
        for sheet in workbook.sheet_audits:
            xlsx_diff_counts["cells"] += sheet.n_differences
            xlsx_diff_counts["text"] += sheet.n_text_differences
            xlsx_diff_counts["numeric"] += sheet.n_numeric_differences
            xlsx_diff_counts["blank_pattern"] += sheet.n_blank_pattern_differences
            xlsx_diff_counts["type"] += sheet.n_type_differences
    report = [
        "# Direct XLSX/TXT Parity Audit",
        "",
    ]
    append_reference_provenance(
        report,
        py_source,
        web_source,
        reference_label=reference_label,
        python_name_warnings=python_name_warnings,
    )
    report.extend([
        "## Executive Summary",
        f"- Python input: `{py_source.input_path}` ({py_source.kind})",
        f"- Web input: `{web_source.input_path}` ({web_source.kind})",
        f"- Numeric tolerance: `{numeric_tolerance}`",
        f"- Max differences per sheet/TXT section: `{max_diffs}`",
        f"- Include styles: `{include_styles}` (style comparison is not implemented in this milestone)",
        f"- XLSX workbooks audited: {len(workbook_audits)}",
        f"- TXT files audited: {len(text_audits)}",
        f"- XLSX_FULL_PARITY = {STATUS_TRUE if xlsx_full_parity else STATUS_FALSE}",
        f"- TXT_FULL_PARITY = {STATUS_TRUE if txt_full_parity else STATUS_FALSE}",
        f"- SAFE_TO_CLAIM_XLSX_TXT_MATCH = {STATUS_TRUE if safe_to_claim else STATUS_FALSE}",
        f"- Total XLSX cell differences: {xlsx_diff_counts['cells']} "
        f"(text={xlsx_diff_counts['text']}, numeric={xlsx_diff_counts['numeric']}, "
        f"blank_pattern={xlsx_diff_counts['blank_pattern']}, type={xlsx_diff_counts['type']})",
        "",
    ])
    append_file_list_section(report, py_source.files, web_source.files)
    append_workbook_summary(report, workbook_audits)
    append_sheet_details(report, workbook_audits, max_diffs)
    append_text_summary(report, text_audits)
    append_text_details(report, text_audits)
    report.extend([
        "## Explicit Final Status",
        f"- XLSX_FULL_PARITY = {STATUS_TRUE if xlsx_full_parity else STATUS_FALSE}",
        f"- TXT_FULL_PARITY = {STATUS_TRUE if txt_full_parity else STATUS_FALSE}",
        f"- SAFE_TO_CLAIM_XLSX_TXT_MATCH = {STATUS_TRUE if safe_to_claim else STATUS_FALSE}",
        "",
    ])
    return "\n".join(report)


def parse_bool(value: str) -> bool:
    normalized = value.strip().lower()
    if normalized in {"1", "true", "yes", "y"}:
        return True
    if normalized in {"0", "false", "no", "n"}:
        return False
    raise argparse.ArgumentTypeError(f"Expected boolean value, got {value!r}")


def main() -> int:
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--python", required=True, help="Python reference ZIP or extracted output directory")
    parser.add_argument("--web", required=True, help="Web ZIP or extracted output directory")
    parser.add_argument("--report", required=True, type=Path, help="Output Markdown report path")
    parser.add_argument("--numeric-tolerance", type=float, default=0.0)
    parser.add_argument("--max-diffs-per-sheet", type=int, default=50)
    parser.add_argument("--include-styles", type=parse_bool, default=False)
    parser.add_argument("--reference-label", default="unspecified", help="Human-readable label for the Python reference source")
    parser.add_argument(
        "--warn-if-python-name-contains",
        default="",
        help="Comma-separated filename substrings that should trigger a visible report warning for the Python input",
    )
    args = parser.parse_args()

    python_path = Path(args.python)
    web_path = Path(args.web)
    py_source = read_source(python_path, args.python)
    web_source = read_source(web_path, args.web)
    python_name_warnings = matching_python_name_warnings(
        python_path,
        parse_warning_substrings(args.warn_if_python_name_contains),
    )
    workbook_audits, text_audits = audit_sources(
        py_source,
        web_source,
        numeric_tolerance=args.numeric_tolerance,
        max_diffs=max(0, args.max_diffs_per_sheet),
    )
    report = build_report(
        py_source,
        web_source,
        workbook_audits,
        text_audits,
        numeric_tolerance=args.numeric_tolerance,
        max_diffs=max(0, args.max_diffs_per_sheet),
        include_styles=args.include_styles,
        reference_label=args.reference_label,
        python_name_warnings=python_name_warnings,
    )
    args.report.parent.mkdir(parents=True, exist_ok=True)
    args.report.write_text(report, encoding="utf-8")
    print(report)
    print(f"Wrote {args.report}")
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except ImportError as exc:
        print(f"Missing required dependency: {exc}", file=sys.stderr)
        raise SystemExit(2)
